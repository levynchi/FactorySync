"""Stickers tab: Label printing system with paths configuration and print queue."""
import os
import json
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd

# SumatraPDF path for 1:1 scale printing
SUMATRA_PDF_PATH = r"C:\Users\levyn\AppData\Local\SumatraPDF\SumatraPDF.exe"


class StickersTabMixin:
    def _create_stickers_tab(self):
        """Create the 'מדבקות' tab UI with printing and paths configuration."""
        # Notebook must be defined by MainWindow
        self.stickers_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.stickers_tab, text="מדבקות")

        container = self.stickers_tab

        # Sub-notebook inside the Stickers tab
        self.stickers_sub_notebook = ttk.Notebook(container)
        self.stickers_sub_notebook.pack(fill="both", expand=True)

        # First sub-tab: Print (הדפסה)
        self._create_print_tab()

        # Second sub-tab: Paths Settings (הגדרת נתיבים)
        self._create_paths_tab()

        # Third sub-tab: Label generator from Rivhit products (יצירת מדבקות)
        self._create_label_generator_tab()

        # Load saved data
        self._label_paths_load()

        # Refresh print products when switching tabs
        self.stickers_sub_notebook.bind('<<NotebookTabChanged>>', lambda e: self._on_stickers_tab_change())

    def _on_stickers_tab_change(self):
        """Handle tab change - refresh data as needed."""
        try:
            current_tab = self.stickers_sub_notebook.index(self.stickers_sub_notebook.select())
            if current_tab == 0:  # Print tab
                self._refresh_print_products()
            elif current_tab == 1:  # Paths tab
                self._refresh_path_comboboxes()
            elif current_tab == 2:  # Label generator tab
                self._refresh_label_gen_products()
        except Exception:
            pass

    # ==================== PATHS TAB ====================
    def _create_paths_tab(self):
        """Create the paths configuration tab."""
        paths_tab = ttk.Frame(self.stickers_sub_notebook)
        self.stickers_sub_notebook.add(paths_tab, text="הגדרת נתיבים")

        # Header
        tk.Label(paths_tab, text="הגדרת נתיבים למדבקות", font=('Arial', 14, 'bold')).pack(pady=(10, 5))

        # Mapping input frame
        input_frame = ttk.LabelFrame(paths_tab, text="הוספת מיפוי", padding=10)
        input_frame.pack(fill="x", padx=15, pady=10)

        # Variables for input
        self._path_product_var = tk.StringVar()
        self._path_size_var = tk.StringVar()
        self._path_fabric_var = tk.StringVar()
        self._path_file_var = tk.StringVar()

        # Product
        ttk.Label(input_frame, text="מוצר:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self._path_product_cb = ttk.Combobox(input_frame, textvariable=self._path_product_var, width=20)
        self._path_product_cb.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # Size
        ttk.Label(input_frame, text="מידה:").grid(row=0, column=2, padx=5, pady=5, sticky="e")
        self._path_size_cb = ttk.Combobox(input_frame, textvariable=self._path_size_var, width=15)
        self._path_size_cb.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        # Fabric type
        ttk.Label(input_frame, text="סוג בד:").grid(row=0, column=4, padx=5, pady=5, sticky="e")
        self._path_fabric_cb = ttk.Combobox(input_frame, textvariable=self._path_fabric_var, width=15)
        self._path_fabric_cb.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        # PDF file
        ttk.Label(input_frame, text="קובץ PDF:").grid(row=0, column=6, padx=5, pady=5, sticky="e")
        ttk.Entry(input_frame, textvariable=self._path_file_var, width=20).grid(row=0, column=7, padx=5, pady=5, sticky="w")
        ttk.Button(input_frame, text="בחר...", command=self._browse_pdf_file).grid(row=0, column=8, padx=5, pady=5)

        # Add button
        tk.Button(input_frame, text="➕ הוסף מיפוי", bg="#27ae60", fg="white", command=self._add_path_mapping).grid(row=0, column=9, padx=10, pady=5)

        # Mappings table
        table_frame = ttk.LabelFrame(paths_tab, text="טבלת מיפויים", padding=10)
        table_frame.pack(fill="both", expand=True, padx=15, pady=10)

        cols = ("מוצר", "מידה", "סוג בד", "קובץ PDF")
        self._paths_tree = ttk.Treeview(table_frame, columns=cols, show="headings")
        for c in cols:
            self._paths_tree.heading(c, text=c)
            self._paths_tree.column(c, width=150, anchor="center")

        vs = ttk.Scrollbar(table_frame, orient="vertical", command=self._paths_tree.yview)
        self._paths_tree.configure(yscrollcommand=vs.set)
        self._paths_tree.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Double-click to edit path
        self._paths_tree.bind('<Double-1>', self._edit_path_mapping)

        # Actions
        actions = tk.Frame(paths_tab)
        actions.pack(fill="x", padx=15, pady=(0, 10))
        tk.Button(actions, text="🗑️ מחק נבחר", bg="#e67e22", fg="white", command=self._delete_path_mapping).pack(side="left", padx=5)
        tk.Button(actions, text="💾 שמור", bg="#3498db", fg="white", command=self._label_paths_save).pack(side="left", padx=5)
        tk.Button(actions, text="📤 ייצא לאקסל", bg="#9b59b6", fg="white", command=self._export_paths_to_excel).pack(side="left", padx=5)
        tk.Button(actions, text="📥 ייבא מאקסל", bg="#1abc9c", fg="white", command=self._import_paths_from_excel).pack(side="left", padx=5)

        # Initialize data
        self._label_paths_data = {"mappings": []}

        # Populate comboboxes
        self._refresh_path_comboboxes()

    def _browse_pdf_file(self):
        """Browse for PDF file."""
        file_path = filedialog.askopenfilename(
            title="בחר קובץ PDF",
            initialdir=r"G:\My Drive\מדבקות\מדבקות PDF מוכן להדפסה",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_path:
            self._path_file_var.set(file_path)

    def _refresh_path_comboboxes(self):
        """Refresh product, size, and fabric comboboxes from catalog."""
        products = []
        sizes = []
        fabrics = []

        try:
            dp = getattr(self, 'data_processor', None)
            if dp:
                # Get products from model names or catalog
                model_names = getattr(dp, 'product_model_names', []) or []
                for r in model_names:
                    n = (r.get('name') or '').strip()
                    if n:
                        products.append(n)

                if not products:
                    catalog = getattr(dp, 'products_catalog', []) or []
                    for r in catalog:
                        n = (r.get('name') or '').strip()
                        if n:
                            products.append(n)

                # Get sizes from catalog
                catalog = getattr(dp, 'products_catalog', []) or []
                for r in catalog:
                    s = (r.get('size') or '').strip()
                    if s:
                        sizes.append(s)

                # Get fabric categories (קטגוריות בדים)
                fabric_categories = getattr(dp, 'product_fabric_categories', []) or []
                for f in fabric_categories:
                    n = (f.get('name') or f.get('category') or '').strip()
                    if n:
                        fabrics.append(n)
        except Exception:
            pass

        # Deduplicate and sort
        products = sorted(set(products))
        sizes = sorted(set(sizes))
        fabrics = sorted(set(fabrics))

        self._path_product_cb['values'] = products
        self._path_size_cb['values'] = sizes
        self._path_fabric_cb['values'] = fabrics

    def _add_path_mapping(self):
        """Add a new path mapping."""
        product = self._path_product_var.get().strip()
        size = self._path_size_var.get().strip()
        fabric = self._path_fabric_var.get().strip()
        pdf_file = self._path_file_var.get().strip()

        if not product:
            messagebox.showwarning("אזהרה", "אנא בחר מוצר")
            return
        if not size:
            messagebox.showwarning("אזהרה", "אנא בחר מידה")
            return
        if not fabric:
            messagebox.showwarning("אזהרה", "אנא בחר סוג בד")
            return
        if not pdf_file:
            messagebox.showwarning("אזהרה", "אנא בחר קובץ PDF")
            return

        # Check for duplicates
        for m in self._label_paths_data.get("mappings", []):
            if m.get("product") == product and m.get("size") == size and m.get("fabric") == fabric:
                messagebox.showwarning("אזהרה", "מיפוי זה כבר קיים")
                return

        # Add mapping
        self._label_paths_data["mappings"].append({
            "product": product,
            "size": size,
            "fabric": fabric,
            "file": pdf_file
        })

        self._refresh_paths_tree()
        self._label_paths_save()

        # Clear inputs
        self._path_product_var.set("")
        self._path_size_var.set("")
        self._path_fabric_var.set("")
        self._path_file_var.set("")

    def _refresh_paths_tree(self):
        """Refresh the paths table."""
        for item in self._paths_tree.get_children():
            self._paths_tree.delete(item)

        for m in self._label_paths_data.get("mappings", []):
            self._paths_tree.insert("", "end", values=(
                m.get("product", ""),
                m.get("size", ""),
                m.get("fabric", ""),
                m.get("file", "")
            ))

    def _delete_path_mapping(self):
        """Delete selected path mapping."""
        sel = self._paths_tree.selection()
        if not sel:
            messagebox.showwarning("אזהרה", "אנא בחר שורה למחיקה")
            return

        if not messagebox.askyesno("אישור", "למחוק את המיפוי הנבחר?"):
            return

        vals = self._paths_tree.item(sel[0], 'values')
        if vals:
            new_mappings = []
            deleted = False
            for m in self._label_paths_data.get("mappings", []):
                if not deleted and m.get("product") == vals[0] and m.get("size") == vals[1] and m.get("fabric") == vals[2]:
                    deleted = True
                    continue
                new_mappings.append(m)
            self._label_paths_data["mappings"] = new_mappings
            self._refresh_paths_tree()
            self._label_paths_save()

    def _edit_path_mapping(self, event):
        """Edit PDF path on double-click."""
        sel = self._paths_tree.selection()
        if not sel:
            return

        vals = self._paths_tree.item(sel[0], 'values')
        if not vals:
            return

        product = vals[0]
        size = vals[1]
        fabric = vals[2]
        current_file = vals[3]

        # Find the mapping index
        mapping_idx = None
        for idx, m in enumerate(self._label_paths_data.get("mappings", [])):
            if m.get("product") == product and m.get("size") == size and m.get("fabric") == fabric:
                mapping_idx = idx
                break

        if mapping_idx is None:
            return

        # Create edit dialog
        dialog = tk.Toplevel(self.stickers_tab)
        dialog.title("עריכת נתיב PDF")
        dialog.grab_set()
        dialog.resizable(False, False)

        frm = ttk.Frame(dialog, padding=15)
        frm.pack(fill='both', expand=True)

        # Info labels
        ttk.Label(frm, text=f"מוצר: {product}", font=('Arial', 10, 'bold')).grid(row=0, column=0, columnspan=3, sticky='w', pady=2)
        ttk.Label(frm, text=f"מידה: {size}  |  סוג בד: {fabric}", font=('Arial', 10)).grid(row=1, column=0, columnspan=3, sticky='w', pady=(0, 10))

        # Path entry
        ttk.Label(frm, text="נתיב:").grid(row=2, column=0, sticky='e', padx=5, pady=5)
        path_var = tk.StringVar(value=current_file)
        path_entry = ttk.Entry(frm, textvariable=path_var, width=50)
        path_entry.grid(row=2, column=1, sticky='we', padx=5, pady=5)

        def browse():
            file_path = filedialog.askopenfilename(
                title="בחר קובץ PDF",
                initialdir=r"G:\My Drive\מדבקות\מדבקות PDF מוכן להדפסה",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )
            if file_path:
                path_var.set(file_path)

        ttk.Button(frm, text="בחר...", command=browse).grid(row=2, column=2, padx=5, pady=5)

        # Buttons
        btn_frame = ttk.Frame(frm)
        btn_frame.grid(row=3, column=0, columnspan=3, pady=(15, 0))

        def save():
            new_path = path_var.get().strip()
            if not new_path:
                messagebox.showwarning("אזהרה", "אנא הזן נתיב")
                return
            self._label_paths_data["mappings"][mapping_idx]["file"] = new_path
            self._refresh_paths_tree()
            self._label_paths_save()
            dialog.destroy()

        ttk.Button(btn_frame, text="שמור", command=save).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="ביטול", command=dialog.destroy).pack(side='left', padx=5)

    def _label_paths_file_path(self) -> str:
        return os.path.join(os.getcwd(), "label_paths.json")

    def _label_paths_load(self):
        """Load paths data from JSON file."""
        try:
            p = self._label_paths_file_path()
            if os.path.exists(p):
                with open(p, 'r', encoding='utf-8') as f:
                    self._label_paths_data = json.load(f)
                self._refresh_paths_tree()
        except Exception:
            self._label_paths_data = {"mappings": []}

    def _label_paths_save(self):
        """Save paths data to JSON file."""
        try:
            p = self._label_paths_file_path()
            with open(p, 'w', encoding='utf-8') as f:
                json.dump(self._label_paths_data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בשמירה: {e}")

    def _export_paths_to_excel(self):
        """Export path mappings to Excel file."""
        mappings = self._label_paths_data.get("mappings", [])
        if not mappings:
            messagebox.showwarning("אזהרה", "אין מיפויים לייצוא")
            return

        # Ask for save location
        file_path = filedialog.asksaveasfilename(
            title="שמור קובץ אקסל",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="label_paths_export.xlsx"
        )
        if not file_path:
            return

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import numbers
            
            # Create DataFrame with Hebrew column names
            data = []
            for m in mappings:
                data.append({
                    "מוצר": m.get("product", ""),
                    "מידה": m.get("size", ""),
                    "סוג בד": m.get("fabric", ""),
                    "קובץ PDF": m.get("file", "")
                })
            
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False, engine='openpyxl')
            
            # Format size column as text to prevent Excel date conversion
            wb = load_workbook(file_path)
            ws = wb.active
            # Column B is "מידה" (size)
            for cell in ws['B']:
                cell.number_format = numbers.FORMAT_TEXT
            wb.save(file_path)
            
            messagebox.showinfo("הצלחה", f"יוצאו {len(mappings)} מיפויים לקובץ:\n{file_path}")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בייצוא: {e}")

    def _import_paths_from_excel(self):
        """Import path mappings from Excel file."""
        # Ask for file to import
        file_path = filedialog.askopenfilename(
            title="בחר קובץ אקסל לייבוא",
            filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
        )
        if not file_path:
            return

        try:
            # Read Excel file with size column as string to prevent date conversion
            df = pd.read_excel(file_path, engine='openpyxl', dtype={'מידה': str})
            
            # Validate columns
            required_cols = {"מוצר", "מידה", "סוג בד", "קובץ PDF"}
            if not required_cols.issubset(set(df.columns)):
                messagebox.showerror("שגיאה", f"הקובץ חייב להכיל את העמודות:\n{', '.join(required_cols)}")
                return

            # Create import dialog
            dialog = tk.Toplevel(self.stickers_tab)
            dialog.title("ייבוא מאקסל")
            dialog.grab_set()
            dialog.resizable(False, False)

            frm = ttk.Frame(dialog, padding=20)
            frm.pack(fill='both', expand=True)

            ttk.Label(frm, text=f"נמצאו {len(df)} שורות בקובץ", font=('Arial', 11, 'bold')).pack(pady=(0, 15))
            ttk.Label(frm, text="בחר אופן ייבוא:").pack(pady=(0, 10))

            import_mode = tk.StringVar(value="add")

            ttk.Radiobutton(frm, text="הוספה - הוסף מיפויים חדשים (דלג על כפילויות)", 
                           variable=import_mode, value="add").pack(anchor='w', pady=2)
            ttk.Radiobutton(frm, text="דריסה - החלף את כל המיפויים הקיימים", 
                           variable=import_mode, value="overwrite").pack(anchor='w', pady=2)

            btn_frame = ttk.Frame(frm)
            btn_frame.pack(pady=(20, 0))

            def do_import():
                mode = import_mode.get()
                new_mappings = []
                
                for _, row in df.iterrows():
                    product = str(row.get("מוצר", "")).strip()
                    size = str(row.get("מידה", "")).strip()
                    fabric = str(row.get("סוג בד", "")).strip()
                    file = str(row.get("קובץ PDF", "")).strip()
                    
                    if product and size and fabric and file:
                        new_mappings.append({
                            "product": product,
                            "size": size,
                            "fabric": fabric,
                            "file": file
                        })

                if mode == "overwrite":
                    self._label_paths_data["mappings"] = new_mappings
                    added = len(new_mappings)
                    skipped = 0
                else:  # add mode
                    existing = self._label_paths_data.get("mappings", [])
                    existing_keys = {(m["product"], m["size"], m["fabric"]) for m in existing}
                    
                    added = 0
                    skipped = 0
                    for m in new_mappings:
                        key = (m["product"], m["size"], m["fabric"])
                        if key not in existing_keys:
                            existing.append(m)
                            existing_keys.add(key)
                            added += 1
                        else:
                            skipped += 1
                    
                    self._label_paths_data["mappings"] = existing

                self._refresh_paths_tree()
                self._label_paths_save()
                dialog.destroy()

                msg = f"יובאו {added} מיפויים"
                if skipped > 0:
                    msg += f"\nדולגו {skipped} כפילויות"
                messagebox.showinfo("הצלחה", msg)

            ttk.Button(btn_frame, text="ייבא", command=do_import).pack(side='left', padx=5)
            ttk.Button(btn_frame, text="ביטול", command=dialog.destroy).pack(side='left', padx=5)

        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בייבוא: {e}")

    # ==================== PRINT TAB ====================
    def _create_print_tab(self):
        """Create the print tab."""
        print_tab = ttk.Frame(self.stickers_sub_notebook)
        self.stickers_sub_notebook.add(print_tab, text="הדפסה")

        # Header
        tk.Label(print_tab, text="הדפסת מדבקות", font=('Arial', 14, 'bold')).pack(pady=(10, 5))

        # Selection frame
        select_frame = ttk.LabelFrame(print_tab, text="בחירת פריטים להדפסה", padding=10)
        select_frame.pack(fill="x", padx=15, pady=10)

        # Variables
        self._print_product_var = tk.StringVar()
        self._print_fabric_var = tk.StringVar()
        self._print_qty_var = tk.StringVar(value="1")

        # Row 1: Product and Fabric
        row1 = tk.Frame(select_frame)
        row1.pack(fill="x", pady=5)

        ttk.Label(row1, text="מוצר:").pack(side="left", padx=5)
        self._print_product_cb = ttk.Combobox(row1, textvariable=self._print_product_var, width=25, state='readonly')
        self._print_product_cb.pack(side="left", padx=5)
        self._print_product_cb.bind('<<ComboboxSelected>>', lambda e: self._on_print_product_change())

        ttk.Label(row1, text="סוג בד:").pack(side="left", padx=(20, 5))
        self._print_fabric_cb = ttk.Combobox(row1, textvariable=self._print_fabric_var, width=15, state='readonly')
        self._print_fabric_cb.pack(side="left", padx=5)
        self._print_fabric_cb.bind('<<ComboboxSelected>>', lambda e: self._on_print_fabric_change())

        ttk.Label(row1, text="כמות:").pack(side="left", padx=(20, 5))
        ttk.Entry(row1, textvariable=self._print_qty_var, width=8).pack(side="left", padx=5)

        tk.Button(row1, text="➕ הוסף לרשימה", bg="#27ae60", fg="white", command=self._add_to_print_queue).pack(side="left", padx=20)

        # Row 2: Sizes (checkboxes)
        sizes_frame = ttk.LabelFrame(select_frame, text="בחר מידות", padding=5)
        sizes_frame.pack(fill="x", pady=10)

        self._print_size_vars = {}  # Will be populated dynamically
        self._print_sizes_container = tk.Frame(sizes_frame)
        self._print_sizes_container.pack(fill="x")

        # Print queue table
        queue_frame = ttk.LabelFrame(print_tab, text="רשימת הדפסה", padding=10)
        queue_frame.pack(fill="both", expand=True, padx=15, pady=10)

        cols = ("מוצר", "סוג בד", "מידה", "כמות", "קובץ PDF")
        self._print_queue_tree = ttk.Treeview(queue_frame, columns=cols, show="headings")
        for c in cols:
            self._print_queue_tree.heading(c, text=c)
            w = 80 if c == "כמות" else 150
            self._print_queue_tree.column(c, width=w, anchor="center")

        vs = ttk.Scrollbar(queue_frame, orient="vertical", command=self._print_queue_tree.yview)
        self._print_queue_tree.configure(yscrollcommand=vs.set)
        self._print_queue_tree.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        queue_frame.grid_rowconfigure(0, weight=1)
        queue_frame.grid_columnconfigure(0, weight=1)

        # Double-click to edit quantity
        self._print_queue_tree.bind('<Double-1>', self._edit_print_qty)

        # Actions
        actions = tk.Frame(print_tab)
        actions.pack(fill="x", padx=15, pady=(0, 10))

        tk.Button(actions, text="🗑️ מחק נבחר", bg="#e67e22", fg="white", command=self._delete_from_print_queue).pack(side="left", padx=5)
        tk.Button(actions, text="🧹 נקה הכל", bg="#95a5a6", fg="white", command=self._clear_print_queue).pack(side="left", padx=5)
        tk.Button(actions, text="🖨️ הדפס הכל", bg="#2980b9", fg="white", font=('Arial', 11, 'bold'), command=self._print_all).pack(side="right", padx=5)

        # Print queue data
        self._print_queue_data = []

    def _on_print_product_change(self):
        """When product changes, update fabric options."""
        product = self._print_product_var.get()
        self._print_fabric_var.set("")

        # Clear sizes
        for widget in self._print_sizes_container.winfo_children():
            widget.destroy()
        self._print_size_vars.clear()

        # Get available fabrics for this product
        fabrics = set()
        for m in self._label_paths_data.get("mappings", []):
            if m.get("product") == product:
                fabrics.add(m.get("fabric", ""))

        self._print_fabric_cb['values'] = sorted(fabrics)

    def _on_print_fabric_change(self):
        """When fabric changes, update size checkboxes."""
        product = self._print_product_var.get()
        fabric = self._print_fabric_var.get()

        # Clear existing checkboxes
        for widget in self._print_sizes_container.winfo_children():
            widget.destroy()
        self._print_size_vars.clear()

        # Get available sizes for this product+fabric
        sizes = []
        for m in self._label_paths_data.get("mappings", []):
            if m.get("product") == product and m.get("fabric") == fabric:
                sizes.append(m.get("size", ""))

        # Create checkboxes
        for size in sorted(sizes):
            var = tk.BooleanVar(value=False)
            self._print_size_vars[size] = var
            cb = ttk.Checkbutton(self._print_sizes_container, text=size, variable=var)
            cb.pack(side="left", padx=10)

    def _refresh_print_products(self):
        """Refresh product options from mappings."""
        products = set()
        for m in self._label_paths_data.get("mappings", []):
            products.add(m.get("product", ""))
        self._print_product_cb['values'] = sorted(products)

    def _add_to_print_queue(self):
        """Add selected items to print queue."""
        product = self._print_product_var.get()
        fabric = self._print_fabric_var.get()
        qty_str = self._print_qty_var.get().strip()

        if not product:
            messagebox.showwarning("אזהרה", "אנא בחר מוצר")
            return
        if not fabric:
            messagebox.showwarning("אזהרה", "אנא בחר סוג בד")
            return

        try:
            qty = int(qty_str)
            if qty <= 0:
                raise ValueError()
        except ValueError:
            messagebox.showwarning("אזהרה", "כמות חייבת להיות מספר חיובי")
            return

        # Get selected sizes
        selected_sizes = [size for size, var in self._print_size_vars.items() if var.get()]
        if not selected_sizes:
            messagebox.showwarning("אזהרה", "אנא בחר לפחות מידה אחת")
            return

        # Find PDF files and add to queue
        for size in selected_sizes:
            pdf_file = None
            for m in self._label_paths_data.get("mappings", []):
                if m.get("product") == product and m.get("size") == size and m.get("fabric") == fabric:
                    pdf_file = m.get("file", "")
                    break

            if pdf_file:
                self._print_queue_data.append({
                    "product": product,
                    "fabric": fabric,
                    "size": size,
                    "qty": qty,
                    "file": pdf_file
                })

        self._refresh_print_queue_tree()

        # Reset size checkboxes
        for var in self._print_size_vars.values():
            var.set(False)

    def _refresh_print_queue_tree(self):
        """Refresh the print queue table."""
        for item in self._print_queue_tree.get_children():
            self._print_queue_tree.delete(item)

        for idx, item in enumerate(self._print_queue_data):
            self._print_queue_tree.insert("", "end", iid=str(idx), values=(
                item.get("product", ""),
                item.get("fabric", ""),
                item.get("size", ""),
                item.get("qty", 1),
                item.get("file", "")
            ))

    def _edit_print_qty(self, event):
        """Edit quantity on double-click."""
        sel = self._print_queue_tree.selection()
        if not sel:
            return

        idx = int(sel[0])
        current_qty = self._print_queue_data[idx].get("qty", 1)

        # Create edit dialog
        dialog = tk.Toplevel(self.stickers_tab)
        dialog.title("עריכת כמות")
        dialog.grab_set()
        dialog.resizable(False, False)

        frm = ttk.Frame(dialog, padding=15)
        frm.pack()

        ttk.Label(frm, text="כמות חדשה:").pack(side="left", padx=5)
        qty_var = tk.StringVar(value=str(current_qty))
        entry = ttk.Entry(frm, textvariable=qty_var, width=10)
        entry.pack(side="left", padx=5)
        entry.focus()
        entry.select_range(0, tk.END)

        def save():
            try:
                new_qty = int(qty_var.get())
                if new_qty <= 0:
                    raise ValueError()
                self._print_queue_data[idx]["qty"] = new_qty
                self._refresh_print_queue_tree()
                dialog.destroy()
            except ValueError:
                messagebox.showwarning("אזהרה", "כמות חייבת להיות מספר חיובי")

        ttk.Button(frm, text="שמור", command=save).pack(side="left", padx=5)
        entry.bind('<Return>', lambda e: save())

    def _delete_from_print_queue(self):
        """Delete selected item from print queue."""
        sel = self._print_queue_tree.selection()
        if not sel:
            messagebox.showwarning("אזהרה", "אנא בחר שורה למחיקה")
            return

        idx = int(sel[0])
        del self._print_queue_data[idx]
        self._refresh_print_queue_tree()

    def _clear_print_queue(self):
        """Clear all items from print queue."""
        if not self._print_queue_data:
            return
        if messagebox.askyesno("אישור", "לנקות את כל רשימת ההדפסה?"):
            self._print_queue_data.clear()
            self._refresh_print_queue_tree()

    def _print_pdf_sumatra(self, pdf_path: str) -> bool:
        """
        Print PDF at 1:1 scale (actual size) using SumatraPDF.
        Returns True if successful, False otherwise.
        """
        try:
            if not os.path.exists(SUMATRA_PDF_PATH):
                return False
            
            # SumatraPDF with noscale = actual size (1:1)
            subprocess.run([
                SUMATRA_PDF_PATH,
                "-print-to-default",
                "-print-settings", "noscale",
                pdf_path
            ], check=True)
            return True
            
        except Exception as e:
            print(f"SumatraPDF print error: {e}")
            return False

    def _print_all(self):
        """Print all items in the queue."""
        if not self._print_queue_data:
            messagebox.showwarning("אזהרה", "רשימת ההדפסה ריקה")
            return

        errors = []
        printed = 0

        for item in self._print_queue_data:
            pdf_file = item.get("file", "")
            qty = item.get("qty", 1)

            if not os.path.exists(pdf_file):
                errors.append(f"קובץ לא נמצא: {pdf_file}")
                continue

            try:
                # Print the file (qty times)
                for _ in range(qty):
                    if os.name == 'nt':  # Windows
                        # Try SumatraPDF for 1:1 scale printing
                        if not self._print_pdf_sumatra(pdf_file):
                            # Fallback to default method
                            os.startfile(pdf_file, "print")
                    else:  # Linux/Mac
                        subprocess.run(['lpr', pdf_file], check=True)
                printed += qty
            except Exception as e:
                errors.append(f"שגיאה בהדפסת {pdf_file}: {e}")

        if errors:
            messagebox.showwarning("אזהרות", "\n".join(errors))

        if printed > 0:
            messagebox.showinfo("הצלחה", f"נשלחו {printed} מדבקות להדפסה")
            self._print_queue_data.clear()
            self._refresh_print_queue_tree()

    # ==================== LABEL GENERATOR TAB ====================
    def _create_label_generator_tab(self):
        """סאב-טאב ליצירת מדבקות מתבנית, ממוצרי ריווחית."""
        gen_tab = ttk.Frame(self.stickers_sub_notebook)
        self.stickers_sub_notebook.add(gen_tab, text="יצירת מדבקות")

        tk.Label(gen_tab, text="יצירת דף מדבקות מתבנית", font=('Arial', 14, 'bold')).pack(pady=(10, 5))

        # Selection frame
        sel = ttk.LabelFrame(gen_tab, text="בחירת מוצר מריווחית", padding=10)
        sel.pack(fill="x", padx=15, pady=10)

        row1 = tk.Frame(sel)
        row1.pack(fill="x", pady=5)
        ttk.Label(row1, text="מוצר:").pack(side="right", padx=5)
        self._lg_product_var = tk.StringVar()
        self._lg_product_cb = ttk.Combobox(row1, textvariable=self._lg_product_var, width=45)
        self._lg_product_cb.pack(side="right", padx=5)
        self._lg_product_cb.bind('<<ComboboxSelected>>', lambda e: self._on_lg_product_change())
        self._lg_product_cb.bind('<KeyRelease>', lambda e: self._filter_lg_products())

        ttk.Label(row1, text="כמות:").pack(side="right", padx=(20, 5))
        self._lg_qty_var = tk.StringVar(value="1")
        ttk.Entry(row1, textvariable=self._lg_qty_var, width=8).pack(side="right", padx=5)

        tk.Button(row1, text="➕ הוסף לתור", bg="#27ae60", fg="white", command=self._lg_add_to_queue).pack(side="right", padx=20)

        # Preview of label fields for the selected product
        self._lg_preview_var = tk.StringVar(value="בחר מוצר כדי לראות את שדות המדבקה")
        tk.Label(sel, textvariable=self._lg_preview_var, fg="#7f8c8d", justify="right", anchor="e").pack(fill="x", pady=(6, 0))
        tk.Label(sel, text="טיפ: לעריכת שדות המדבקה (שם להדפסה/מידה/סוג בד/מארז) - בטאב ריווחית, דאבל-קליק על מוצר.",
                 fg="#95a5a6", font=('Arial', 8), justify="right", anchor="e").pack(fill="x")

        # Queue table
        queue_frame = ttk.LabelFrame(gen_tab, text="תור מדבקות", padding=10)
        queue_frame.pack(fill="both", expand=True, padx=15, pady=10)

        cols = ("print_name", "size", "fabric", "pack", "barcode", "qty")
        headers = {"print_name": "שם להדפסה", "size": "מידה", "fabric": "סוג בד",
                   "pack": "מארז", "barcode": "ברקוד", "qty": "כמות"}
        self._lg_tree = ttk.Treeview(queue_frame, columns=cols, show="headings")
        for c in cols:
            self._lg_tree.heading(c, text=headers[c])
            w = 70 if c in ("size", "pack", "qty") else 160
            self._lg_tree.column(c, width=w, anchor="center")
        vs = ttk.Scrollbar(queue_frame, orient="vertical", command=self._lg_tree.yview)
        self._lg_tree.configure(yscrollcommand=vs.set)
        self._lg_tree.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        queue_frame.grid_rowconfigure(0, weight=1)
        queue_frame.grid_columnconfigure(0, weight=1)

        # Actions
        actions = tk.Frame(gen_tab)
        actions.pack(fill="x", padx=15, pady=(0, 10))
        tk.Button(actions, text="🗑️ מחק נבחר", bg="#e67e22", fg="white", command=self._lg_delete_selected).pack(side="left", padx=5)
        tk.Button(actions, text="🧹 נקה הכל", bg="#95a5a6", fg="white", command=self._lg_clear).pack(side="left", padx=5)
        tk.Button(actions, text="🖨️ הדפס", bg="#2980b9", fg="white", font=('Arial', 11, 'bold'), command=self._lg_print).pack(side="right", padx=5)
        tk.Button(actions, text="🧾 צור דף מדבקות (PDF)", bg="#27ae60", fg="white", font=('Arial', 11, 'bold'), command=self._lg_generate_pdf).pack(side="right", padx=5)

        # State
        self._lg_queue_data = []          # [{print_name,size,fabric,pack_qty,barcode,qty}]
        self._lg_all_products = []        # [(label, product_dict)]
        self._lg_last_pdf = None
        self._refresh_label_gen_products()

    def _rivhit_products(self):
        dp = getattr(self, 'data_processor', None)
        return list(getattr(dp, 'rivhit_products', []) or []) if dp else []

    def _refresh_label_gen_products(self):
        """רענון רשימת המוצרים בבורר מתוך מוצרי ריווחית."""
        items = self._rivhit_products()
        pairs = []
        for p in items:
            name = str(p.get('item_name', '')).strip()
            bc = str(p.get('item_part_num', '')).strip()
            if not name and not bc:
                continue
            label = f"{name} | {bc}" if bc else name
            pairs.append((label, p))
        self._lg_all_products = pairs
        if hasattr(self, '_lg_product_cb'):
            self._lg_product_cb['values'] = [lbl for lbl, _ in pairs]

    def _filter_lg_products(self):
        """סינון הבורר לפי הטקסט שהוקלד (שם או ברקוד)."""
        q = (self._lg_product_var.get() or '').strip().lower()
        if not q:
            values = [lbl for lbl, _ in self._lg_all_products]
        else:
            terms = q.split()
            values = [lbl for lbl, _ in self._lg_all_products if all(t in lbl.lower() for t in terms)]
        self._lg_product_cb['values'] = values

    def _selected_lg_product(self):
        """מחזיר את מילון המוצר שנבחר בבורר (לפי הלייבל)."""
        label = (self._lg_product_var.get() or '').strip()
        for lbl, p in self._lg_all_products:
            if lbl == label:
                return p
        return None

    def _on_lg_product_change(self):
        p = self._selected_lg_product()
        if not p:
            self._lg_preview_var.set("בחר מוצר כדי לראות את שדות המדבקה")
            return
        bc = str(p.get('item_part_num', '')).strip()
        f = self.data_processor.get_rivhit_label_fields(bc, product=p)
        parts = [f"שם: {f.get('print_name','')}"]
        if f.get('size'):
            parts.append(f"מידה: {f.get('size')}")
        if f.get('fabric'):
            parts.append(f"בד: {f.get('fabric')}")
        if f.get('pack_qty', 1) and int(f.get('pack_qty', 1)) > 1:
            parts.append(f"מארז: {f.get('pack_qty')}")
        self._lg_preview_var.set("   |   ".join(parts))

    def _lg_add_to_queue(self):
        p = self._selected_lg_product()
        if not p:
            messagebox.showwarning("אזהרה", "אנא בחר מוצר מהרשימה")
            return
        try:
            qty = int(self._lg_qty_var.get())
            if qty <= 0:
                raise ValueError()
        except ValueError:
            messagebox.showwarning("אזהרה", "כמות חייבת להיות מספר חיובי")
            return
        bc = str(p.get('item_part_num', '')).strip()
        if not bc:
            messagebox.showwarning("אין ברקוד", "למוצר זה אין מק\"ט/ברקוד; לא ניתן ליצור מדבקה סרוקה")
            return
        f = self.data_processor.get_rivhit_label_fields(bc, product=p)
        self._lg_queue_data.append({
            "print_name": f.get('print_name', ''),
            "size": f.get('size', ''),
            "size_unit": f.get('size_unit', ''),
            "fabric": f.get('fabric', ''),
            "pack_qty": int(f.get('pack_qty', 1) or 1),
            "image": f.get('image', ''),
            "barcode": bc,
            "qty": qty,
        })
        self._refresh_lg_tree()

    def _refresh_lg_tree(self):
        for item in self._lg_tree.get_children():
            self._lg_tree.delete(item)
        for idx, it in enumerate(self._lg_queue_data):
            self._lg_tree.insert("", "end", iid=str(idx), values=(
                it.get("print_name", ""), it.get("size", ""), it.get("fabric", ""),
                it.get("pack_qty", 1), it.get("barcode", ""), it.get("qty", 1),
            ))

    def _lg_delete_selected(self):
        sel = self._lg_tree.selection()
        if not sel:
            messagebox.showwarning("אזהרה", "אנא בחר שורה למחיקה")
            return
        idx = int(sel[0])
        del self._lg_queue_data[idx]
        self._refresh_lg_tree()

    def _lg_clear(self):
        if not self._lg_queue_data:
            return
        if messagebox.askyesno("אישור", "לנקות את כל התור?"):
            self._lg_queue_data.clear()
            self._refresh_lg_tree()

    def _lg_expand_items(self):
        """מרחיב את התור לרשימת מדבקות בודדות לפי כמות."""
        expanded = []
        for it in self._lg_queue_data:
            for _ in range(int(it.get("qty", 1) or 1)):
                expanded.append({
                    "print_name": it.get("print_name", ""),
                    "size": it.get("size", ""),
                    "size_unit": it.get("size_unit", ""),
                    "fabric": it.get("fabric", ""),
                    "pack_qty": it.get("pack_qty", 1),
                    "image": it.get("image", ""),
                    "barcode": it.get("barcode", ""),
                })
        return expanded

    def _lg_generate_pdf(self):
        if not self._lg_queue_data:
            messagebox.showwarning("אזהרה", "התור ריק")
            return
        from datetime import datetime
        from optitex_analyzer.core.label_sheet import build_label_sheet_pdf
        out_dir = os.path.join(os.getcwd(), "exports", "labels")
        file_path = os.path.join(out_dir, f"labels_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        try:
            items = self._lg_expand_items()
            count = build_label_sheet_pdf(items, file_path)
            self._lg_last_pdf = file_path
            messagebox.showinfo("הצלחה", f"נוצרו {count} מדבקות בקובץ:\n{file_path}")
            try:
                os.startfile(file_path)  # תצוגה מקדימה
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה ביצירת הקובץ:\n{e}")

    def _lg_print(self):
        """מייצר (אם צריך) ומדפיס את דף המדבקות בקנה מידה 1:1."""
        if not self._lg_queue_data:
            messagebox.showwarning("אזהרה", "התור ריק")
            return
        from datetime import datetime
        from optitex_analyzer.core.label_sheet import build_label_sheet_pdf
        out_dir = os.path.join(os.getcwd(), "exports", "labels")
        file_path = os.path.join(out_dir, f"labels_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        try:
            items = self._lg_expand_items()
            build_label_sheet_pdf(items, file_path)
            self._lg_last_pdf = file_path
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה ביצירת הקובץ:\n{e}")
            return
        try:
            if os.name == 'nt':
                if not self._print_pdf_sumatra(file_path):
                    os.startfile(file_path, "print")
            else:
                subprocess.run(['lpr', file_path], check=True)
            messagebox.showinfo("הצלחה", "דף המדבקות נשלח להדפסה")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בהדפסה:\n{e}")
