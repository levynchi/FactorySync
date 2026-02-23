import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import re
import os
import json

class ProductsCatalogMethodsMixin:
    """All event handlers, loaders, and helpers for Products Catalog tab."""
    # ===== products section builders =====
    def _build_products_section(self, parent):
        form = ttk.LabelFrame(parent, text="הוספת פריט", padding=10)
        form.pack(fill='x', padx=10, pady=6)
        self.prod_name_var = tk.StringVar(); self.prod_size_var = tk.StringVar(); self.prod_fabric_type_var = tk.StringVar(); self.prod_fabric_color_var = tk.StringVar(); self.prod_print_name_var = tk.StringVar()
        self.prod_category_var = tk.StringVar(); self.prod_ticks_var = tk.StringVar(); self.prod_elastic_var = tk.StringVar(); self.prod_ribbon_var = tk.StringVar(); self.prod_unit_type_var = tk.StringVar()
        self.prod_fabric_category_var = tk.StringVar(); self.prod_main_category_var = tk.StringVar(); self.prod_square_area_var = tk.StringVar()

        # --- Main Category selector (drives field visibility) ---
        self._products_field_widgets = {}
        tk.Label(form, text="קטגוריה ראשית:", font=('Arial',10,'bold')).grid(row=0, column=0, sticky='w', padx=4, pady=(0,6))
        main_cat_names = [c.get('name','') for c in getattr(self.data_processor, 'main_categories', [])]
        self.prod_main_category_combobox = ttk.Combobox(form, textvariable=self.prod_main_category_var, values=main_cat_names, state='readonly', width=16, justify='right')
        self.prod_main_category_combobox.grid(row=0, column=1, sticky='w', padx=2, pady=(0,6))
        self.prod_main_category_combobox.bind('<<ComboboxSelected>>', lambda e: self._apply_main_category_field_visibility())

        # shift existing rows by +1 to make space for main category row
        tk.Label(form, text="שם הדגם:", font=('Arial',10,'bold')).grid(row=1, column=0, sticky='w', padx=4, pady=4)
        model_names = [r.get('name') for r in getattr(self.data_processor, 'product_model_names', [])]
        self.model_name_combobox = ttk.Combobox(
            form,
            textvariable=self.prod_name_var,
            values=model_names,
            state='readonly',
            width=16,
            justify='right'
        )
        self.model_name_combobox.grid(row=1, column=1, sticky='w', padx=2, pady=4)
        # map widgets for field visibility: model_name
        self._products_field_widgets['model_name'] = []
        self._products_field_widgets['model_name'].append(self.model_name_combobox)

        # sub category (תת קטגוריה) - multi-select like sizes/types/colors/prints
        tk.Label(form, text="תת קטגוריה:", font=('Arial',10,'bold')).grid(row=1, column=2, sticky='w', padx=4, pady=4)
        cat_names = [c.get('name','') for c in getattr(self.data_processor, 'categories', [])]
        # Group sub-category picker + display + clear into one frame to keep them adjacent
        self.subcat_frame = ttk.Frame(form)
        self.subcat_frame.grid(row=1, column=3, columnspan=3, sticky='w', padx=2, pady=4)
        self.category_combobox = ttk.Combobox(
            self.subcat_frame,
            values=cat_names,
            state='readonly',
            width=12,
            justify='right'
        )
        self.category_combobox.pack(side='left', padx=(0, 4))
        self.category_combobox.bind('<<ComboboxSelected>>', lambda e: self._on_attr_select('sub_category'))
        # readonly display of selected sub-categories and clear button
        self.subcat_selected_entry = tk.Entry(self.subcat_frame, textvariable=self.prod_category_var, width=20, state='readonly')
        self.subcat_selected_entry.pack(side='left', padx=(0, 4))
        self.btn_clear_subcat = tk.Button(self.subcat_frame, text='נקה', command=lambda: self._clear_attr('sub_category'), width=4)
        self.btn_clear_subcat.pack(side='left')
        # sub_category mapping (label+combo will be toggled together)
        # We'll collect after creating label widgets too using winfo_children search if needed

        # fabric category
        tk.Label(form, text="קטגוריית בד:", font=('Arial',10,'bold')).grid(row=1, column=6, sticky='w', padx=4, pady=4)
        fabric_cat_names = [r.get('name') for r in getattr(self.data_processor, 'product_fabric_categories', [])]
        self.fabric_category_combobox = ttk.Combobox(
            form,
            textvariable=self.prod_fabric_category_var,
            values=fabric_cat_names,
            state='readonly',
            width=12,
            justify='right'
        )
        self.fabric_category_combobox.grid(row=1, column=7, sticky='w', padx=2, pady=4)

        # multiselect helpers state
        self.selected_sub_categories = []
        self.selected_sizes = []
        self.selected_fabric_types = []
        self.selected_fabric_colors = []
        self.selected_print_names = []

    # size (new row for more space)
        tk.Label(form, text="מידות:", font=('Arial',10,'bold')).grid(row=2, column=0, sticky='w', padx=4, pady=4)
        self.size_picker = ttk.Combobox(form, values=[r.get('name') for r in getattr(self.data_processor,'product_sizes',[])], state='readonly', width=12, justify='right')
        self.size_picker.grid(row=2, column=1, sticky='w', padx=2, pady=4)
        self.size_picker.bind('<<ComboboxSelected>>', lambda e: self._on_attr_select('size'))
        tk.Entry(form, textvariable=self.prod_size_var, width=20, state='readonly').grid(row=2, column=2, sticky='w', padx=2, pady=4)
        self.btn_clear_size = tk.Button(form, text='נקה', command=lambda: self._clear_attr('size'), width=4)
        self.btn_clear_size.grid(row=2, column=3, padx=2)

        # types (move to its own row for more space)
        tk.Label(form, text="סוגי בד:", font=('Arial',10,'bold')).grid(row=3, column=0, sticky='w', padx=4, pady=4)
        self.ftype_picker = ttk.Combobox(form, values=[r.get('name') for r in getattr(self.data_processor,'product_fabric_types',[])], state='readonly', width=12, justify='right')
        self.ftype_picker.grid(row=3, column=1, sticky='w', padx=2, pady=4)
        self.ftype_picker.bind('<<ComboboxSelected>>', lambda e: self._on_attr_select('fabric_type'))
        tk.Entry(form, textvariable=self.prod_fabric_type_var, width=20, state='readonly').grid(row=3, column=2, sticky='w', padx=2, pady=4)
        self.btn_clear_ftype = tk.Button(form, text='נקה', command=lambda: self._clear_attr('fabric_type'), width=4)
        self.btn_clear_ftype.grid(row=3, column=3, padx=2)

        # colors (keep on their own row)
        tk.Label(form, text="צבעי בד:", font=('Arial',10,'bold')).grid(row=4, column=0, sticky='w', padx=4, pady=4)
        self.fcolor_picker = ttk.Combobox(form, values=[r.get('name') for r in getattr(self.data_processor,'product_fabric_colors',[])], state='readonly', width=12, justify='right')
        self.fcolor_picker.grid(row=4, column=1, sticky='w', padx=2, pady=4)
        self.fcolor_picker.bind('<<ComboboxSelected>>', lambda e: self._on_attr_select('fabric_color'))
        tk.Entry(form, textvariable=self.prod_fabric_color_var, width=20, state='readonly').grid(row=4, column=2, sticky='w', padx=2, pady=4)
        self.btn_clear_fcolor = tk.Button(form, text='נקה', command=lambda: self._clear_attr('fabric_color'), width=4)
        self.btn_clear_fcolor.grid(row=4, column=3, padx=2)

        # prints (move to a dedicated row)
        tk.Label(form, text="שמות פרינט:", font=('Arial',10,'bold')).grid(row=5, column=0, sticky='w', padx=4, pady=4)
        self.pname_picker = ttk.Combobox(form, values=[r.get('name') for r in getattr(self.data_processor,'product_print_names',[])], state='readonly', width=12, justify='right')
        self.pname_picker.grid(row=5, column=1, sticky='w', padx=2, pady=4)
        self.pname_picker.bind('<<ComboboxSelected>>', lambda e: self._on_attr_select('print_name'))
        tk.Entry(form, textvariable=self.prod_print_name_var, width=20, state='readonly').grid(row=5, column=2, sticky='w', padx=2, pady=4)
        self.btn_clear_pname = tk.Button(form, text='נקה', command=lambda: self._clear_attr('print_name'), width=4)
        self.btn_clear_pname.grid(row=5, column=3, padx=2, pady=4)

        # accessories quantities (move down to separate row)
        tk.Label(form, text="טיקטקים:", font=('Arial',10,'bold')).grid(row=6, column=0, sticky='w', padx=4, pady=4)
        tk.Entry(form, textvariable=self.prod_ticks_var, width=10).grid(row=6, column=1, sticky='w', padx=2, pady=4)
        tk.Label(form, text="גומי:", font=('Arial',10,'bold')).grid(row=6, column=2, sticky='w', padx=4, pady=4)
        tk.Entry(form, textvariable=self.prod_elastic_var, width=10).grid(row=6, column=3, sticky='w', padx=2, pady=4)
        tk.Label(form, text="סרט:", font=('Arial',10,'bold')).grid(row=6, column=4, sticky='w', padx=4, pady=4)
        tk.Entry(form, textvariable=self.prod_ribbon_var, width=10).grid(row=6, column=5, sticky='w', padx=2, pady=4)
        # Unit Type (סוג יחידה)
        tk.Label(form, text="סוג יחידה:", font=('Arial',10,'bold')).grid(row=6, column=6, sticky='w', padx=4, pady=4)
        tk.Entry(form, textvariable=self.prod_unit_type_var, width=12).grid(row=6, column=7, sticky='w', padx=2, pady=4)
        
        # Square Area (שטח רבוע)
        tk.Label(form, text="שטח רבוע:", font=('Arial',10,'bold')).grid(row=7, column=0, sticky='w', padx=4, pady=4)
        tk.Entry(form, textvariable=self.prod_square_area_var, width=12).grid(row=7, column=1, sticky='w', padx=2, pady=4)

        # actions moved to their own row to avoid horizontal clipping
        tk.Button(form, text="➕ הוסף", command=self._add_product_catalog_entry, bg='#27ae60', fg='white').grid(row=8, column=0, padx=12, pady=6, sticky='w')
        tk.Button(form, text="✏️ ערוך נבחר", command=self._edit_selected_product, bg='#3498db', fg='white').grid(row=8, column=1, padx=4, pady=6, sticky='w')
        tk.Button(form, text="🗑️ מחק נבחר", command=self._delete_selected_product_entry, bg='#e67e22', fg='white').grid(row=8, column=2, padx=4, pady=6, sticky='w')
        tk.Button(form, text="💾 ייצוא ל-Excel", command=self._export_products_catalog, bg='#2c3e50', fg='white').grid(row=8, column=3, padx=4, pady=6, sticky='w')
        tk.Button(form, text="⬆️ יבוא מקובץ", command=self._import_products_catalog_dialog, bg='#34495e', fg='white').grid(row=8, column=4, padx=4, pady=6, sticky='w')

        # סרגל סינון לפי שם הדגם + טוגל תצוגה
        filter_frame = ttk.Frame(parent, padding=4)
        filter_frame.pack(fill='x', padx=10, pady=(0, 2))
        
        # טוגל תצוגת עלויות (בצד שמאל)
        self.cost_view_mode = tk.BooleanVar(value=False)
        self.toggle_view_btn = tk.Button(
            filter_frame, 
            text='💰 תצוגת עלויות', 
            command=self._toggle_cost_view,
            bg='#3498db', 
            fg='white', 
            font=('Arial', 10, 'bold'),
            width=14
        )
        self.toggle_view_btn.pack(side='left', padx=4)
        
        # טוגל שיטת חישוב עלות בד (מ"ר או משקל)
        self.fabric_cost_method = tk.StringVar(value='weight')  # 'sqm' = לפי מ"ר, 'weight' = לפי משקל
        self.toggle_fabric_method_btn = tk.Button(
            filter_frame, 
            text='⚖️ עלות בד: לפי משקל', 
            command=self._toggle_fabric_cost_method,
            bg='#e67e22', 
            fg='white', 
            font=('Arial', 10, 'bold'),
            width=18
        )
        self.toggle_fabric_method_btn.pack(side='left', padx=4)
        
        # סינון לפי קטגוריית בד (באמצע השורה)
        tk.Label(filter_frame, text="קטגוריית בד:", font=('Arial', 10, 'bold'), bg='#f7f9fa').pack(side='left', padx=(20, 4))
        self.filter_fabric_category_var = tk.StringVar(value="הכל")
        fabric_categories = ["הכל"] + [c.get('name') for c in getattr(self.data_processor, 'product_fabric_categories', [])]
        self.filter_fabric_category_combo = ttk.Combobox(filter_frame, textvariable=self.filter_fabric_category_var, values=fabric_categories, state='readonly', width=15)
        self.filter_fabric_category_combo.pack(side='left', padx=2)
        self.filter_fabric_category_combo.bind('<<ComboboxSelected>>', lambda e: self._filter_products_tree())
        
        # כפתור נקה סינון
        tk.Button(filter_frame, text='נקה סינון', command=self._clear_product_filters, bg='#95a5a6', fg='white', width=10).pack(side='left', padx=4)
        
        # חיפוש (בצד ימין)
        tk.Label(filter_frame, text="🔍 חיפוש לפי שם הדגם:", font=('Arial', 10, 'bold'), bg='#f7f9fa').pack(side='right', padx=(8, 4))
        self.product_filter_var = tk.StringVar()
        self.product_filter_var.trace_add('write', lambda *args: self._filter_products_tree())
        filter_entry = tk.Entry(filter_frame, textvariable=self.product_filter_var, width=30, font=('Arial', 10))
        filter_entry.pack(side='right', padx=2)
        tk.Button(filter_frame, text='🗑️ נקה', command=lambda: self.product_filter_var.set(''), bg='#e74c3c', fg='white', width=6).pack(side='right', padx=4)

        # Frame לטבלה (שומר רפרנס לשימוש בהחלפת תצוגה)
        self.products_tree_frame = ttk.LabelFrame(parent, text="פריטים", padding=6)
        self.products_tree_frame.pack(fill='both', expand=True, padx=10, pady=6)
        
        # הגדרות עמודות - תצוגה רגילה
        self.regular_cols = ('barcode','id','name','main_category','category','size','fabric_type','fabric_color','print_name','fabric_category','square_area','ticks_qty','elastic_qty','ribbon_qty','created_at')
        self.regular_headers = {
            'barcode':'מק"ט',
            'id':'ID',
            'name':'שם הדגם',
            'main_category':'קטגוריה ראשית',
            'category':'תת קטגוריה',
            'size':'מידה',
            'fabric_type':'סוג בד',
            'fabric_color':'צבע בד',
            'print_name':'שם פרינט',
            'fabric_category':'קטגוריית בד',
            'square_area':'שטח רבוע',
            'ticks_qty':'טיקטקים',
            'elastic_qty':'גומי',
            'ribbon_qty':'סרט',
            'created_at':'נוצר'
        }
        self.regular_widths = {
            'barcode':120, 'id':40, 'name':140, 'main_category':110, 'category':100,
            'size':70, 'fabric_type':110, 'fabric_color':110, 'print_name':110,
            'fabric_category':120, 'square_area':100, 'ticks_qty':70, 'elastic_qty':60,
            'ribbon_qty':60, 'created_at':140
        }
        
        # הגדרות עמודות - תצוגת עלויות
        self.cost_cols = ('barcode','name','size','fabric_category','fabric_color','print_name','fabric_cost','ticks_cost','elastic_cost','ribbon_cost','sewing_cost','total_cost')
        self.cost_headers = {
            'barcode':'מק"ט',
            'name':'שם הדגם',
            'size':'מידה',
            'fabric_category':'קטגוריית בד',
            'fabric_color':'צבע בד',
            'print_name':'פרינט',
            'fabric_cost':'עלות בד',
            'ticks_cost':'עלות טיקטקים',
            'elastic_cost':'עלות גומי',
            'ribbon_cost':'עלות סרט',
            'sewing_cost':'עלות תפירה',
            'total_cost':'סה"כ עלות'
        }
        self.cost_widths = {
            'barcode':110, 'name':130, 'size':60, 'fabric_category':100, 'fabric_color':90,
            'print_name':80, 'fabric_cost':85, 'ticks_cost':90, 'elastic_cost':80,
            'ribbon_cost':80, 'sewing_cost':85, 'total_cost':90
        }
        
        # יצירת הטבלה בתצוגה רגילה
        self._create_products_tree(self.regular_cols, self.regular_headers, self.regular_widths)
        self._load_products_catalog_into_tree()

        # collect field widgets for toggling (labels + inputs/buttons)
        # Helper to find label widgets by text; if not found we still toggle inputs
        def _try_find_label(text):
            for w in form.grid_slaves():
                if isinstance(w, tk.Label) and str(w.cget('text')) == text:
                    return w
            return None

        # Helper to find a widget by grid row/column, optionally class and text
        def _get_grid_widget(row: int, col: int, cls=None, text: str | None = None):
            for w in form.grid_slaves():
                info = w.grid_info()
                try:
                    r = int(info.get('row'))
                    c = int(info.get('column'))
                except Exception:
                    continue
                if r == row and c == col:
                    if cls and not isinstance(w, cls):
                        continue
                    if text is not None and str(getattr(w, 'cget', lambda *_: '')('text')) != text:
                        continue
                    return w
            return None

        # Map labels explicitly
        lbl_model = _try_find_label("שם הדגם:")
        lbl_subcat = _try_find_label("תת קטגוריה:")
        lbl_fabric_cat = _try_find_label("קטגוריית בד:")
        lbl_sizes = _try_find_label("מידות:")
        lbl_ftype = _try_find_label("סוגי בד:")
        lbl_fcolor = _try_find_label("צבעי בד:")
        lbl_pname = _try_find_label("שמות פרינט:")
        lbl_ticks = _try_find_label("טיקטקים:")
        lbl_elastic = _try_find_label("גומי:")
        lbl_ribbon = _try_find_label("סרט:")
        lbl_unit_type = _try_find_label("סוג יחידה:")

        self._products_field_widgets['model_name'] = [x for x in [lbl_model, self.model_name_combobox] if x]
        # include readonly entry and clear for sub_category
        # Toggle the entire frame to show/hide sub-category widgets together
        subcat_widgets = [lbl_subcat]
        if hasattr(self, 'subcat_frame'):
            subcat_widgets.append(self.subcat_frame)
        self._products_field_widgets['sub_category'] = [x for x in subcat_widgets if x]
        self._products_field_widgets['fabric_category'] = [x for x in [lbl_fabric_cat, self.fabric_category_combobox] if x]

        # sizes group includes picker, readonly entry, clear button
        sizes_widgets = [lbl_sizes, self.size_picker]
        for w in form.grid_slaves():
            if isinstance(w, tk.Entry) and w.cget('state') == 'readonly' and w.cget('textvariable'):
                if str(w.cget('textvariable')) == str(self.prod_size_var):
                    sizes_widgets.append(w)
        btn_clear_size = getattr(self, 'btn_clear_size', None)
        if btn_clear_size:
            sizes_widgets.append(btn_clear_size)
        self._products_field_widgets['sizes'] = [w for w in sizes_widgets if w]

        # fabric type group
        ftype_widgets = [lbl_ftype, self.ftype_picker]
        for w in form.grid_slaves():
            if isinstance(w, tk.Entry) and w.cget('state') == 'readonly' and w.cget('textvariable'):
                if str(w.cget('textvariable')) == str(self.prod_fabric_type_var):
                    ftype_widgets.append(w)
        btn_clear_ftype = getattr(self, 'btn_clear_ftype', None)
        if btn_clear_ftype:
            ftype_widgets.append(btn_clear_ftype)
        self._products_field_widgets['fabric_type'] = [w for w in ftype_widgets if w]

        # fabric color group
        fcolor_widgets = [lbl_fcolor, self.fcolor_picker]
        for w in form.grid_slaves():
            if isinstance(w, tk.Entry) and w.cget('state') == 'readonly' and w.cget('textvariable'):
                if str(w.cget('textvariable')) == str(self.prod_fabric_color_var):
                    fcolor_widgets.append(w)
        btn_clear_fcolor = getattr(self, 'btn_clear_fcolor', None)
        if btn_clear_fcolor:
            fcolor_widgets.append(btn_clear_fcolor)
        self._products_field_widgets['fabric_color'] = [w for w in fcolor_widgets if w]

        # print name group
        pname_widgets = [lbl_pname, self.pname_picker]
        for w in form.grid_slaves():
            if isinstance(w, tk.Entry) and w.cget('state') == 'readonly' and w.cget('textvariable'):
                if str(w.cget('textvariable')) == str(self.prod_print_name_var):
                    pname_widgets.append(w)
        btn_clear_pname = getattr(self, 'btn_clear_pname', None)
        if btn_clear_pname:
            pname_widgets.append(btn_clear_pname)
        self._products_field_widgets['print_name'] = [w for w in pname_widgets if w]

        # quantities
        ticks_widgets = [x for x in [lbl_ticks] if x]
        elastic_widgets = [x for x in [lbl_elastic] if x]
        ribbon_widgets = [x for x in [lbl_ribbon] if x]
        for w in form.grid_slaves():
            if isinstance(w, tk.Entry) and w.cget('state') != 'readonly' and w.cget('textvariable'):
                tv = str(w.cget('textvariable'))
                if tv == str(self.prod_ticks_var):
                    ticks_widgets.append(w)
                elif tv == str(self.prod_elastic_var):
                    elastic_widgets.append(w)
                elif tv == str(self.prod_ribbon_var):
                    ribbon_widgets.append(w)
        self._products_field_widgets['ticks_qty'] = ticks_widgets
        self._products_field_widgets['elastic_qty'] = elastic_widgets
        self._products_field_widgets['ribbon_qty'] = ribbon_widgets
        # unit type group
        unit_widgets = [x for x in [lbl_unit_type] if x]
        for w in form.grid_slaves():
            if isinstance(w, tk.Entry) and w.cget('state') != 'readonly' and w.cget('textvariable'):
                if str(w.cget('textvariable')) == str(self.prod_unit_type_var):
                    unit_widgets.append(w)
        self._products_field_widgets['unit_type'] = unit_widgets

        # initial visibility
        self._refresh_main_categories_for_products()
        self._apply_main_category_field_visibility()

    # ===== accessories builders =====
    def _build_accessories_section(self, parent):
        print("🏗️  DEBUG: בונה טאב אביזרי תפירה...")
        self.acc_name_var = tk.StringVar(); self.acc_unit_var = tk.StringVar()
        acc_form = ttk.LabelFrame(parent, text="הוספת אביזר תפירה", padding=10)
        acc_form.pack(fill='x', padx=10, pady=6)
        tk.Label(acc_form, text="שם אביזר:", font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4, sticky='w')
        tk.Entry(acc_form, textvariable=self.acc_name_var, width=20).grid(row=0, column=1, padx=4, pady=4)
        tk.Label(acc_form, text="יחידת מדידה:", font=('Arial',10,'bold')).grid(row=0, column=2, padx=4, pady=4, sticky='w')
        tk.Entry(acc_form, textvariable=self.acc_unit_var, width=12).grid(row=0, column=3, padx=4, pady=4)
        tk.Button(acc_form, text="➕ הוסף", command=self._add_sewing_accessory, bg='#27ae60', fg='white').grid(row=0, column=4, padx=8)
        print("🔴 DEBUG: יוצר כפתור 'מחק נבחר' עם command=self._delete_selected_sewing_accessory")
        delete_btn = tk.Button(acc_form, text="🗑️ מחק נבחר", command=self._delete_selected_sewing_accessory, bg='#e67e22', fg='white')
        delete_btn.grid(row=0, column=5, padx=4)
        print("✅ DEBUG: כפתור 'מחק נבחר' נוצר בהצלחה!")
        print(f"🔍 DEBUG: הכפתור נוצר: {delete_btn}")
        print(f"🔍 DEBUG: הפונקציה: {self._delete_selected_sewing_accessory}")
        
        # הוספת פונקציה פשוטה לבדיקה
        def test_button_click():
            print("🧪 TEST: כפתור נלחץ - פונקציה פשוטה!")
            self._delete_selected_sewing_accessory()
        
        # יצירת כפתור נוסף לבדיקה
        test_btn = tk.Button(acc_form, text="🧪 בדיקה", command=test_button_click, bg='#3498db', fg='white')
        test_btn.grid(row=0, column=6, padx=4)
        print("🧪 DEBUG: כפתור בדיקה נוסף נוצר!")

        acc_tree_frame = ttk.LabelFrame(parent, text="אביזרים", padding=6)
        acc_tree_frame.pack(fill='both', expand=True, padx=10, pady=6)
        acc_cols = ('id','name','unit','created_at')
        self.sewing_accessories_tree = ttk.Treeview(acc_tree_frame, columns=acc_cols, show='headings', height=10)
        acc_headers = {'id':'ID','name':'שם','unit':'יחידה','created_at':'נוצר'}
        acc_widths = {'id':50,'name':160,'unit':100,'created_at':140}
        for c in acc_cols:
            self.sewing_accessories_tree.heading(c, text=acc_headers[c])
            self.sewing_accessories_tree.column(c, width=acc_widths[c], anchor='center')
        acc_vs = ttk.Scrollbar(acc_tree_frame, orient='vertical', command=self.sewing_accessories_tree.yview)
        self.sewing_accessories_tree.configure(yscroll=acc_vs.set)
        self.sewing_accessories_tree.pack(side='left', fill='both', expand=True)
        acc_vs.pack(side='right', fill='y')
        self._load_accessories_into_tree()

    # ===== categories builders =====
    def _build_categories_section(self, parent):
        self.cat_name_var = tk.StringVar()
        cat_form = ttk.LabelFrame(parent, text="הוספת תת קטגוריה", padding=10)
        cat_form.pack(fill='x', padx=10, pady=6)
        tk.Label(cat_form, text="שם תת קטגוריה:", font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4, sticky='w')
        tk.Entry(cat_form, textvariable=self.cat_name_var, width=22).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(cat_form, text="➕ הוסף", command=self._add_category, bg='#27ae60', fg='white').grid(row=0, column=2, padx=8)
        tk.Button(cat_form, text="🗑️ מחק נבחר", command=self._delete_selected_category, bg='#e67e22', fg='white').grid(row=0, column=3, padx=4)

        cat_tree_frame = ttk.LabelFrame(parent, text="תת קטגוריות", padding=6)
        cat_tree_frame.pack(fill='both', expand=True, padx=10, pady=6)
        cat_cols = ('id','name','created_at')
        self.categories_tree = ttk.Treeview(cat_tree_frame, columns=cat_cols, show='headings', height=10)
        cat_headers = {'id':'ID','name':'שם תת קטגוריה','created_at':'נוצר'}
        cat_widths = {'id':60,'name':180,'created_at':140}
        for c in cat_cols:
            self.categories_tree.heading(c, text=cat_headers[c])
            self.categories_tree.column(c, width=cat_widths[c], anchor='center')
        cat_vs = ttk.Scrollbar(cat_tree_frame, orient='vertical', command=self.categories_tree.yview)
        self.categories_tree.configure(yscroll=cat_vs.set)
        self.categories_tree.pack(side='left', fill='both', expand=True)
        cat_vs.pack(side='right', fill='y')
        self._load_categories_into_tree()

    # ===== main categories builders =====
    def _build_main_categories_section(self, parent):
        # Notebook inside the Main Categories tab
        mcat_nb = ttk.Notebook(parent)
        mcat_nb.pack(fill='both', expand=True, padx=8, pady=6)

        manage_tab = tk.Frame(mcat_nb, bg='#f7f9fa')
        fields_tab = tk.Frame(mcat_nb, bg='#f7f9fa')
        mcat_nb.add(manage_tab, text='קטגוריות')
        mcat_nb.add(fields_tab, text='שדות לקטגוריה')

        # Manage tab: add/delete + list
        self.main_cat_name_var = tk.StringVar()
        mcat_form = ttk.LabelFrame(manage_tab, text="הוספת קטגוריה ראשית", padding=10)
        mcat_form.pack(fill='x', padx=10, pady=6)
        tk.Label(mcat_form, text="שם קטגוריה ראשית:", font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4, sticky='w')
        tk.Entry(mcat_form, textvariable=self.main_cat_name_var, width=22).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(mcat_form, text="➕ הוסף", command=self._add_main_category, bg='#27ae60', fg='white').grid(row=0, column=2, padx=8)
        tk.Button(mcat_form, text="🗑️ מחק נבחר", command=self._delete_selected_main_category, bg='#e67e22', fg='white').grid(row=0, column=3, padx=4)

        mcat_tree_frame = ttk.LabelFrame(manage_tab, text="קטגוריות ראשיות", padding=6)
        mcat_tree_frame.pack(fill='both', expand=True, padx=10, pady=6)
        mcat_cols = ('id','name','created_at')
        self.main_categories_tree = ttk.Treeview(mcat_tree_frame, columns=mcat_cols, show='headings', height=10)
        mcat_headers = {'id':'ID','name':'שם קטגוריה','created_at':'נוצר'}
        mcat_widths = {'id':60,'name':180,'created_at':140}
        for c in mcat_cols:
            self.main_categories_tree.heading(c, text=mcat_headers[c])
            self.main_categories_tree.column(c, width=mcat_widths[c], anchor='center')
        mcat_vs = ttk.Scrollbar(mcat_tree_frame, orient='vertical', command=self.main_categories_tree.yview)
        self.main_categories_tree.configure(yscroll=mcat_vs.set)
        self.main_categories_tree.pack(side='left', fill='both', expand=True)
        mcat_vs.pack(side='right', fill='y')
        self._load_main_categories_into_tree()

        # Fields tab: assign item fields to main category
        self._build_main_category_fields_tab(fields_tab)

    def _get_item_fields_defs(self) -> list[tuple[str, str]]:
        # key, display label mapping for item fields in 'פריטים > הוספת פריט'
        return [
            ('model_name', 'שם הדגם'),
            ('sub_category', 'תת קטגוריה'),
            ('fabric_category', 'קטגוריית בד'),
            ('unit_type', 'סוג יחידה'),
            ('sizes', 'מידות'),
            ('fabric_type', 'סוגי בד'),
            ('fabric_color', 'צבעי בד'),
            ('print_name', 'שמות פרינט'),
            ('ticks_qty', 'טיקטקים'),
            ('elastic_qty', 'גומי'),
            ('ribbon_qty', 'סרט'),
        ]

    def _build_main_category_fields_tab(self, parent):
        frm = ttk.LabelFrame(parent, text='שיוך שדות לקטגוריה ראשית', padding=10)
        frm.pack(fill='both', expand=True, padx=10, pady=8)

        # Category selector
        tk.Label(frm, text='קטגוריה ראשית:', font=('Arial',10,'bold')).grid(row=0, column=0, sticky='w', padx=4, pady=4)
        self.main_cat_fields_var = tk.StringVar()
        self.main_cat_fields_combo = ttk.Combobox(frm, textvariable=self.main_cat_fields_var, state='readonly', width=24, justify='right')
        self.main_cat_fields_combo.grid(row=0, column=1, sticky='w', padx=4, pady=4)
        self.main_cat_fields_combo.bind('<<ComboboxSelected>>', lambda e: self._on_select_main_category_for_fields())

        # Checkboxes
        fields_defs = self._get_item_fields_defs()
        self.main_cat_field_vars = {}
        chk_container = ttk.Frame(frm)
        chk_container.grid(row=1, column=0, columnspan=3, sticky='w', padx=2, pady=4)
        col = 0; row = 0
        for key, label in fields_defs:
            var = tk.BooleanVar(value=False)
            self.main_cat_field_vars[key] = var
            cb = ttk.Checkbutton(chk_container, text=label, variable=var)
            cb.grid(row=row, column=col, padx=8, pady=4, sticky='w')
            col += 1
            if col >= 3:
                col = 0; row += 1

        # Action buttons
        tk.Button(frm, text='💾 שמור שדות', command=self._save_main_category_fields, bg='#2c3e50', fg='white').grid(row=2, column=0, padx=4, pady=8, sticky='w')
        tk.Button(frm, text='אפס בחירה', command=self._reset_main_category_fields).grid(row=2, column=1, padx=4, pady=8, sticky='w')

        # Load categories into combo
        self._load_main_categories_for_fields()

    def _load_main_categories_for_fields(self):
        names = [c.get('name','') for c in getattr(self.data_processor, 'main_categories', [])]
        if hasattr(self, 'main_cat_fields_combo'):
            self.main_cat_fields_combo['values'] = names
            # keep selection if still valid
            cur = self.main_cat_fields_var.get()
            if cur not in names:
                self.main_cat_fields_var.set('')

    def _find_main_category_by_name(self, name: str):
        for c in getattr(self.data_processor, 'main_categories', []):
            if c.get('name','') == name:
                return c
        return None

    def _on_select_main_category_for_fields(self):
        name = self.main_cat_fields_var.get().strip()
        rec = self._find_main_category_by_name(name)
        if not rec:
            self._reset_main_category_fields()
            return
        fields = rec.get('fields') or self.data_processor.get_main_category_fields(rec.get('id'))
        fields = fields or []
        for key, var in self.main_cat_field_vars.items():
            var.set(key in fields)

    def _reset_main_category_fields(self):
        for var in self.main_cat_field_vars.values():
            var.set(False)

    def _save_main_category_fields(self):
        name = self.main_cat_fields_var.get().strip()
        rec = self._find_main_category_by_name(name)
        if not rec:
            messagebox.showerror('שגיאה', 'בחר קטגוריה ראשית לשיוך שדות')
            return
        selected_keys = [k for k,v in self.main_cat_field_vars.items() if v.get()]
        ok = self.data_processor.set_main_category_fields(int(rec.get('id')), selected_keys)
        if ok:
            messagebox.showinfo('הצלחה', 'השדות נשמרו לקטגוריה')
        else:
            messagebox.showerror('שגיאה', 'שמירת השדות נכשלה')

    # ===== attributes builders =====
    def _build_attributes_section(self, attributes_tab):
        attr_nb = ttk.Notebook(attributes_tab)
        attr_nb.pack(fill='both', expand=True, padx=8, pady=6)

        sizes_tab = tk.Frame(attr_nb, bg='#f7f9fa')
        ftypes_tab = tk.Frame(attr_nb, bg='#f7f9fa')
        fcolors_tab = tk.Frame(attr_nb, bg='#f7f9fa')
        prints_tab = tk.Frame(attr_nb, bg='#f7f9fa')
        fcats_tab = tk.Frame(attr_nb, bg='#f7f9fa')
        modelnames_tab = tk.Frame(attr_nb, bg='#f7f9fa')
        fabric_prices_tab = tk.Frame(attr_nb, bg='#f7f9fa')
        cost_settings_tab = tk.Frame(attr_nb, bg='#f7f9fa')
        attr_nb.add(sizes_tab, text='מידות')
        attr_nb.add(ftypes_tab, text='סוגי בד')
        attr_nb.add(fcolors_tab, text='צבעי בד')
        attr_nb.add(prints_tab, text='שמות פרינט')
        attr_nb.add(fcats_tab, text='קטגוריות בדים')
        attr_nb.add(modelnames_tab, text='שם הדגם')
        attr_nb.add(fabric_prices_tab, text='מחירי בדים')
        attr_nb.add(cost_settings_tab, text='הגדרות עלויות')

        # bind vars
        self.attr_size_var = tk.StringVar(); self.attr_fabric_type_var = tk.StringVar(); self.attr_fabric_color_var = tk.StringVar(); self.attr_print_name_var = tk.StringVar(); self.attr_fabric_category_var = tk.StringVar(); self.attr_model_name_var = tk.StringVar()

        # Sizes
        sz_form = ttk.LabelFrame(sizes_tab, text='הוספת מידה', padding=8)
        sz_form.pack(fill='x', padx=8, pady=6)
        tk.Label(sz_form, text='מידה:', font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4)
        tk.Entry(sz_form, textvariable=self.attr_size_var, width=18).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(sz_form, text='➕ הוסף', command=self._add_product_size, bg='#27ae60', fg='white').grid(row=0, column=2, padx=6)
        tk.Button(sz_form, text='🗑️ מחק נבחר', command=self._delete_selected_product_size, bg='#e67e22', fg='white').grid(row=0, column=3, padx=4)
        sz_tree_frame = ttk.LabelFrame(sizes_tab, text='מידות', padding=4)
        sz_tree_frame.pack(fill='both', expand=True, padx=8, pady=4)
        self.sizes_tree = ttk.Treeview(sz_tree_frame, columns=('id','name','created_at'), show='headings', height=10)
        for c,t,w in [('id','ID',60),('name','מידה',140),('created_at','נוצר',140)]:
            self.sizes_tree.heading(c, text=t); self.sizes_tree.column(c, width=w, anchor='center')
        sz_vs = ttk.Scrollbar(sz_tree_frame, orient='vertical', command=self.sizes_tree.yview)
        self.sizes_tree.configure(yscroll=sz_vs.set)
        self.sizes_tree.pack(side='left', fill='both', expand=True); sz_vs.pack(side='right', fill='y')
        self._load_sizes_into_tree()

        # Fabric Types
        ft_form = ttk.LabelFrame(ftypes_tab, text='הוספת סוג בד', padding=8)
        ft_form.pack(fill='x', padx=8, pady=6)
        tk.Label(ft_form, text='סוג בד:', font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4)
        tk.Entry(ft_form, textvariable=self.attr_fabric_type_var, width=18).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(ft_form, text='➕ הוסף', command=self._add_fabric_type_item, bg='#27ae60', fg='white').grid(row=0, column=2, padx=6)
        tk.Button(ft_form, text='🗑️ מחק נבחר', command=self._delete_selected_fabric_type_item, bg='#e67e22', fg='white').grid(row=0, column=3, padx=4)
        ft_tree_frame = ttk.LabelFrame(ftypes_tab, text='סוגי בד', padding=4)
        ft_tree_frame.pack(fill='both', expand=True, padx=8, pady=4)
        self.fabric_types_tree = ttk.Treeview(ft_tree_frame, columns=('id','name','created_at'), show='headings', height=10)
        for c,t,w in [('id','ID',60),('name','סוג בד',160),('created_at','נוצר',140)]:
            self.fabric_types_tree.heading(c, text=t); self.fabric_types_tree.column(c, width=w, anchor='center')
        ft_vs = ttk.Scrollbar(ft_tree_frame, orient='vertical', command=self.fabric_types_tree.yview)
        self.fabric_types_tree.configure(yscroll=ft_vs.set)
        self.fabric_types_tree.pack(side='left', fill='both', expand=True); ft_vs.pack(side='right', fill='y')
        self._load_fabric_types_into_tree()

        # Fabric Colors
        fc_form = ttk.LabelFrame(fcolors_tab, text='הוספת צבע בד', padding=8)
        fc_form.pack(fill='x', padx=8, pady=6)
        tk.Label(fc_form, text='צבע בד:', font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4)
        tk.Entry(fc_form, textvariable=self.attr_fabric_color_var, width=18).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(fc_form, text='➕ הוסף', command=self._add_fabric_color_item, bg='#27ae60', fg='white').grid(row=0, column=2, padx=6)
        tk.Button(fc_form, text='🗑️ מחק נבחר', command=self._delete_selected_fabric_color_item, bg='#e67e22', fg='white').grid(row=0, column=3, padx=4)
        fc_tree_frame = ttk.LabelFrame(fcolors_tab, text='צבעי בד', padding=4)
        fc_tree_frame.pack(fill='both', expand=True, padx=8, pady=4)
        self.fabric_colors_tree = ttk.Treeview(fc_tree_frame, columns=('id','name','created_at'), show='headings', height=10)
        for c,t,w in [('id','ID',60),('name','צבע בד',160),('created_at','נוצר',140)]:
            self.fabric_colors_tree.heading(c, text=t); self.fabric_colors_tree.column(c, width=w, anchor='center')
        fc_vs = ttk.Scrollbar(fc_tree_frame, orient='vertical', command=self.fabric_colors_tree.yview)
        self.fabric_colors_tree.configure(yscroll=fc_vs.set)
        self.fabric_colors_tree.pack(side='left', fill='both', expand=True); fc_vs.pack(side='right', fill='y')
        self._load_fabric_colors_into_tree()

        # Print Names
        pn_form = ttk.LabelFrame(prints_tab, text='הוספת שם פרינט', padding=8)
        pn_form.pack(fill='x', padx=8, pady=6)
        tk.Label(pn_form, text='שם פרינט:', font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4)
        tk.Entry(pn_form, textvariable=self.attr_print_name_var, width=18).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(pn_form, text='➕ הוסף', command=self._add_print_name_item, bg='#27ae60', fg='white').grid(row=0, column=2, padx=6)
        tk.Button(pn_form, text='🗑️ מחק נבחר', command=self._delete_selected_print_name_item, bg='#e67e22', fg='white').grid(row=0, column=3, padx=4)
        pn_tree_frame = ttk.LabelFrame(prints_tab, text='שמות פרינט', padding=4)
        pn_tree_frame.pack(fill='both', expand=True, padx=8, pady=4)
        self.print_names_tree = ttk.Treeview(pn_tree_frame, columns=('id','name','created_at'), show='headings', height=10)
        for c,t,w in [('id','ID',60),('name','שם פרינט',160),('created_at','נוצר',140)]:
            self.print_names_tree.heading(c, text=t); self.print_names_tree.column(c, width=w, anchor='center')
        pn_vs = ttk.Scrollbar(pn_tree_frame, orient='vertical', command=self.print_names_tree.yview)
        self.print_names_tree.configure(yscroll=pn_vs.set)
        self.print_names_tree.pack(side='left', fill='both', expand=True); pn_vs.pack(side='right', fill='y')
        self._load_print_names_into_tree()

        # Fabric Categories
        fcg_form = ttk.LabelFrame(fcats_tab, text='הוספת קטגוריית בד', padding=8)
        fcg_form.pack(fill='x', padx=8, pady=6)
        tk.Label(fcg_form, text='קטגוריית בד:', font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4)
        tk.Entry(fcg_form, textvariable=self.attr_fabric_category_var, width=18).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(fcg_form, text='➕ הוסף', command=self._add_fabric_category_item, bg='#27ae60', fg='white').grid(row=0, column=2, padx=6)
        tk.Button(fcg_form, text='🗑️ מחק נבחר', command=self._delete_selected_fabric_category_item, bg='#e67e22', fg='white').grid(row=0, column=3, padx=4)
        fcg_tree_frame = ttk.LabelFrame(fcats_tab, text='קטגוריות בדים', padding=4)
        fcg_tree_frame.pack(fill='both', expand=True, padx=8, pady=4)
        self.fabric_categories_tree = ttk.Treeview(fcg_tree_frame, columns=('id','name','created_at'), show='headings', height=10)
        for c,t,w in [('id','ID',60),('name','קטגוריה',160),('created_at','נוצר',140)]:
            self.fabric_categories_tree.heading(c, text=t); self.fabric_categories_tree.column(c, width=w, anchor='center')
        fcg_vs = ttk.Scrollbar(fcg_tree_frame, orient='vertical', command=self.fabric_categories_tree.yview)
        self.fabric_categories_tree.configure(yscroll=fcg_vs.set)
        self.fabric_categories_tree.pack(side='left', fill='both', expand=True); fcg_vs.pack(side='right', fill='y')
        self._load_fabric_categories_into_tree()

        # Model Names (שם הדגם)
        mn_form = ttk.LabelFrame(modelnames_tab, text='הוסף שם דגם', padding=8)
        mn_form.pack(fill='x', padx=8, pady=6)
        tk.Label(mn_form, text='שם דגם:', font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4)
        tk.Entry(mn_form, textvariable=self.attr_model_name_var, width=18).grid(row=0, column=1, padx=4, pady=4)
        tk.Label(mn_form, text='מחיר תפירה (₪):', font=('Arial',10,'bold')).grid(row=0, column=2, padx=4, pady=4)
        self.attr_model_sewing_price_var = tk.StringVar(value='0')
        tk.Entry(mn_form, textvariable=self.attr_model_sewing_price_var, width=10).grid(row=0, column=3, padx=4, pady=4)
        tk.Button(mn_form, text='➕ הוסף', command=self._add_model_name_item, bg='#27ae60', fg='white').grid(row=0, column=4, padx=6)
        tk.Button(mn_form, text='🗑️ מחק נבחר', command=self._delete_selected_model_name_item, bg='#e67e22', fg='white').grid(row=0, column=5, padx=4)
        tk.Button(mn_form, text='💾 עדכן מחיר', command=self._update_selected_model_sewing_price, bg='#3498db', fg='white').grid(row=0, column=6, padx=4)
        mn_tree_frame = ttk.LabelFrame(modelnames_tab, text='שמות דגם', padding=4)
        mn_tree_frame.pack(fill='both', expand=True, padx=8, pady=4)
        self.model_names_tree = ttk.Treeview(mn_tree_frame, columns=('id','name','sewing_price','created_at'), show='headings', height=10)
        for c,t,w in [('id','ID',60),('name','שם דגם',160),('sewing_price','מחיר תפירה',100),('created_at','נוצר',140)]:
            self.model_names_tree.heading(c, text=t); self.model_names_tree.column(c, width=w, anchor='center')
        mn_vs = ttk.Scrollbar(mn_tree_frame, orient='vertical', command=self.model_names_tree.yview)
        self.model_names_tree.configure(yscroll=mn_vs.set)
        self.model_names_tree.pack(side='left', fill='both', expand=True); mn_vs.pack(side='right', fill='y')
        # לחיצה על שורה מעדכנת את שדה המחיר
        self.model_names_tree.bind('<<TreeviewSelect>>', self._on_model_name_select)
        self._load_model_names_into_tree()

        # Fabric Prices (מחירי בדים)
        self._build_fabric_prices_section(fabric_prices_tab)
        
        # Cost Settings (הגדרות עלויות)
        self._build_cost_settings_section(cost_settings_tab)

    # ===== Products Tree View Toggle =====
    def _create_products_tree(self, cols, headers, widths):
        """יצירת טבלת פריטים עם עמודות מותאמות"""
        # מחיקת טבלה קיימת אם יש
        if hasattr(self, 'products_tree') and self.products_tree:
            self.products_tree.destroy()
        if hasattr(self, 'products_tree_scrollbar') and self.products_tree_scrollbar:
            self.products_tree_scrollbar.destroy()
        
        # יצירת טבלה חדשה
        self.products_tree = ttk.Treeview(self.products_tree_frame, columns=cols, show='headings', height=12)
        for c in cols:
            self.products_tree.heading(c, text=headers[c])
            self.products_tree.column(c, width=widths[c], anchor='center')
        
        self.products_tree_scrollbar = ttk.Scrollbar(self.products_tree_frame, orient='vertical', command=self.products_tree.yview)
        self.products_tree.configure(yscroll=self.products_tree_scrollbar.set)
        self.products_tree.pack(side='left', fill='both', expand=True)
        self.products_tree_scrollbar.pack(side='right', fill='y')

    def _toggle_cost_view(self):
        """החלפה בין תצוגה רגילה לתצוגת עלויות"""
        current_mode = self.cost_view_mode.get()
        new_mode = not current_mode
        self.cost_view_mode.set(new_mode)
        
        if new_mode:
            # עבור לתצוגת עלויות
            self.toggle_view_btn.config(text='📋 תצוגה רגילה', bg='#27ae60')
            self.products_tree_frame.config(text='פריטים - תצוגת עלויות')
            self._create_products_tree(self.cost_cols, self.cost_headers, self.cost_widths)
        else:
            # עבור לתצוגה רגילה
            self.toggle_view_btn.config(text='💰 תצוגת עלויות', bg='#3498db')
            self.products_tree_frame.config(text='פריטים')
            self._create_products_tree(self.regular_cols, self.regular_headers, self.regular_widths)
        
        # טעינת הנתונים מחדש
        self._filter_products_tree()

    def _toggle_fabric_cost_method(self):
        """החלפה בין שיטת חישוב עלות בד לפי מ"ר או לפי משקל"""
        current_method = self.fabric_cost_method.get()
        
        if current_method == 'sqm':
            # עבור לחישוב לפי משקל
            self.fabric_cost_method.set('weight')
            self.toggle_fabric_method_btn.config(text='⚖️ עלות בד: לפי משקל', bg='#e67e22')
        else:
            # עבור לחישוב לפי מ"ר
            self.fabric_cost_method.set('sqm')
            self.toggle_fabric_method_btn.config(text='📐 עלות בד: לפי מ"ר', bg='#9b59b6')
        
        # אם בתצוגת עלויות, טען מחדש
        if self.cost_view_mode.get():
            self._filter_products_tree()

    # ===== Fabric Prices Section =====
    def _build_fabric_prices_section(self, parent):
        """בניית טאב מחירי בדים"""
        # משתנים לטופס
        self.fp_fabric_category_var = tk.StringVar()
        self.fp_fabric_color_var = tk.StringVar()
        self.fp_print_name_var = tk.StringVar()
        self.fp_price_per_sqm_var = tk.StringVar()
        
        # טופס הוספה
        fp_form = ttk.LabelFrame(parent, text='הוספת מחיר בד', padding=10)
        fp_form.pack(fill='x', padx=8, pady=6)
        
        # שורה 1 - קטגוריה, צבע, פרינט
        tk.Label(fp_form, text='קטגוריית בד:', font=('Arial',10,'bold')).grid(row=0, column=0, padx=4, pady=4, sticky='e')
        fabric_cat_names = [r.get('name') for r in getattr(self.data_processor, 'product_fabric_categories', [])]
        self.fp_fabric_category_combo = ttk.Combobox(fp_form, textvariable=self.fp_fabric_category_var, values=fabric_cat_names, width=16, justify='right')
        self.fp_fabric_category_combo.grid(row=0, column=1, padx=4, pady=4, sticky='w')
        
        tk.Label(fp_form, text='צבע בד:', font=('Arial',10,'bold')).grid(row=0, column=2, padx=4, pady=4, sticky='e')
        fabric_color_names = [r.get('name') for r in getattr(self.data_processor, 'product_fabric_colors', [])]
        self.fp_fabric_color_combo = ttk.Combobox(fp_form, textvariable=self.fp_fabric_color_var, values=fabric_color_names, width=16, justify='right')
        self.fp_fabric_color_combo.grid(row=0, column=3, padx=4, pady=4, sticky='w')
        
        tk.Label(fp_form, text='פרינט:', font=('Arial',10,'bold')).grid(row=0, column=4, padx=4, pady=4, sticky='e')
        print_names = [r.get('name') for r in getattr(self.data_processor, 'product_print_names', [])]
        self.fp_print_name_combo = ttk.Combobox(fp_form, textvariable=self.fp_print_name_var, values=print_names, width=16, justify='right')
        self.fp_print_name_combo.grid(row=0, column=5, padx=4, pady=4, sticky='w')
        
        # שורה 2 - מחיר למ"ר
        tk.Label(fp_form, text='מחיר למ"ר (₪):', font=('Arial',10,'bold')).grid(row=1, column=0, padx=4, pady=4, sticky='e')
        tk.Entry(fp_form, textvariable=self.fp_price_per_sqm_var, width=12).grid(row=1, column=1, padx=4, pady=4, sticky='w')
        
        # כפתורים
        tk.Button(fp_form, text='➕ הוסף', command=self._add_fabric_price, bg='#27ae60', fg='white').grid(row=1, column=2, padx=6, pady=4)
        tk.Button(fp_form, text='🗑️ מחק נבחר', command=self._delete_selected_fabric_price, bg='#e67e22', fg='white').grid(row=1, column=3, padx=4, pady=4)
        
        # טבלה
        fp_tree_frame = ttk.LabelFrame(parent, text='טבלת מחירי בדים', padding=6)
        fp_tree_frame.pack(fill='both', expand=True, padx=8, pady=6)
        
        fp_cols = ('id', 'fabric_category', 'fabric_color', 'print_name', 'price_per_sqm', 'created_at')
        self.fabric_prices_tree = ttk.Treeview(fp_tree_frame, columns=fp_cols, show='headings', height=12)
        fp_headers = {
            'id': 'ID',
            'fabric_category': 'קטגוריית בד',
            'fabric_color': 'צבע בד',
            'print_name': 'פרינט',
            'price_per_sqm': 'מחיר למ"ר',
            'created_at': 'נוצר'
        }
        fp_widths = {'id': 50, 'fabric_category': 140, 'fabric_color': 120, 'print_name': 120, 'price_per_sqm': 100, 'created_at': 140}
        for c in fp_cols:
            self.fabric_prices_tree.heading(c, text=fp_headers[c])
            self.fabric_prices_tree.column(c, width=fp_widths[c], anchor='center')
        
        fp_vs = ttk.Scrollbar(fp_tree_frame, orient='vertical', command=self.fabric_prices_tree.yview)
        self.fabric_prices_tree.configure(yscroll=fp_vs.set)
        self.fabric_prices_tree.pack(side='left', fill='both', expand=True)
        fp_vs.pack(side='right', fill='y')
        
        self._load_fabric_prices_into_tree()

    def _build_cost_settings_section(self, parent):
        """בניית טאב הגדרות עלויות גלובליות"""
        # טעינת הגדרות קיימות
        settings = self.data_processor.load_item_cost_settings()
        
        # משתנים
        self.cs_tick_price_var = tk.StringVar(value=str(settings.get('tick_price', 0)))
        self.cs_elastic_price_var = tk.StringVar(value=str(settings.get('elastic_price', 0)))
        self.cs_ribbon_price_var = tk.StringVar(value=str(settings.get('ribbon_price', 0)))
        self.cs_sewing_price_var = tk.StringVar(value=str(settings.get('sewing_price', 0)))
        
        # כותרת
        tk.Label(parent, text='הגדרות מחירי עלות גלובליים', font=('Arial', 14, 'bold'), bg='#f7f9fa', fg='#2c3e50').pack(pady=(10, 5))
        tk.Label(parent, text='מחירים אלו ישמשו לחישוב עלות כל הפריטים בקטלוג', font=('Arial', 10), bg='#f7f9fa', fg='#7f8c8d').pack(pady=(0, 10))
        
        # טופס
        cs_form = ttk.LabelFrame(parent, text='מחירי אביזרים ותפירה', padding=20)
        cs_form.pack(fill='x', padx=20, pady=10)
        
        # שורה 1
        tk.Label(cs_form, text='מחיר טיקטק ליחידה (₪):', font=('Arial',11,'bold')).grid(row=0, column=0, padx=10, pady=10, sticky='e')
        tk.Entry(cs_form, textvariable=self.cs_tick_price_var, width=15, font=('Arial', 11)).grid(row=0, column=1, padx=10, pady=10, sticky='w')
        
        tk.Label(cs_form, text='מחיר גומי ליחידה (₪):', font=('Arial',11,'bold')).grid(row=0, column=2, padx=10, pady=10, sticky='e')
        tk.Entry(cs_form, textvariable=self.cs_elastic_price_var, width=15, font=('Arial', 11)).grid(row=0, column=3, padx=10, pady=10, sticky='w')
        
        # שורה 2
        tk.Label(cs_form, text='מחיר סרט ליחידה (₪):', font=('Arial',11,'bold')).grid(row=1, column=0, padx=10, pady=10, sticky='e')
        tk.Entry(cs_form, textvariable=self.cs_ribbon_price_var, width=15, font=('Arial', 11)).grid(row=1, column=1, padx=10, pady=10, sticky='w')
        
        tk.Label(cs_form, text='מחיר תפירה ברירת מחדל (₪):', font=('Arial',11,'bold')).grid(row=1, column=2, padx=10, pady=10, sticky='e')
        tk.Entry(cs_form, textvariable=self.cs_sewing_price_var, width=15, font=('Arial', 11)).grid(row=1, column=3, padx=10, pady=10, sticky='w')
        
        # כפתור שמירה
        btn_frame = tk.Frame(parent, bg='#f7f9fa')
        btn_frame.pack(pady=20)
        tk.Button(btn_frame, text='💾 שמור הגדרות', command=self._save_cost_settings, bg='#27ae60', fg='white', font=('Arial', 12, 'bold'), padx=30, pady=10).pack()
        
        # הסבר נוסחה
        formula_frame = ttk.LabelFrame(parent, text='נוסחת חישוב עלות פריט', padding=15)
        formula_frame.pack(fill='x', padx=20, pady=10)
        
        formula_text = """
עלות בד = שטח רבוע × מחיר למ"ר
עלות טיקטקים = כמות טיקטקים × מחיר טיקטק
עלות גומי = כמות גומי × מחיר גומי
עלות סרט = כמות סרט × מחיר סרט
עלות תפירה = מחיר תפירה לפי שם הדגם (או מחיר ברירת מחדל)

סה"כ עלות = עלות בד + עלות טיקטקים + עלות גומי + עלות סרט + עלות תפירה

* מחיר תפירה מוגדר בטאב "שם הדגם" לכל דגם בנפרד
* אם לא מוגדר לדגם - ישתמש במחיר ברירת מחדל למעלה
        """
        tk.Label(formula_frame, text=formula_text, font=('Courier New', 10), bg='#f7f9fa', fg='#34495e', justify='right').pack()

    def _load_fabric_prices_into_tree(self):
        """טעינת מחירי בדים לטבלה"""
        if not hasattr(self, 'fabric_prices_tree'):
            return
        for item in self.fabric_prices_tree.get_children():
            self.fabric_prices_tree.delete(item)
        try:
            prices = self.data_processor.load_fabric_prices()
            for rec in prices:
                self.fabric_prices_tree.insert('', 'end', values=(
                    rec.get('id'),
                    rec.get('fabric_category', ''),
                    rec.get('fabric_color', ''),
                    rec.get('print_name', ''),
                    rec.get('price_per_sqm', 0),
                    rec.get('created_at', '')
                ))
        except Exception as e:
            print(f"שגיאה בטעינת מחירי בדים: {e}")

    def _add_fabric_price(self):
        """הוספת מחיר בד חדש"""
        fabric_category = self.fp_fabric_category_var.get().strip()
        fabric_color = self.fp_fabric_color_var.get().strip()
        print_name = self.fp_print_name_var.get().strip()
        price_per_sqm_str = self.fp_price_per_sqm_var.get().strip()
        
        if not fabric_category:
            messagebox.showerror('שגיאה', 'חובה לבחור קטגוריית בד')
            return
        
        try:
            price_per_sqm = float(price_per_sqm_str) if price_per_sqm_str else 0
        except ValueError:
            messagebox.showerror('שגיאה', 'מחיר חייב להיות מספר')
            return
        
        try:
            new_id = self.data_processor.add_fabric_price(fabric_category, fabric_color, print_name, price_per_sqm)
            self._load_fabric_prices_into_tree()
            # ניקוי טופס
            self.fp_fabric_category_var.set('')
            self.fp_fabric_color_var.set('')
            self.fp_print_name_var.set('')
            self.fp_price_per_sqm_var.set('')
            messagebox.showinfo('הצלחה', 'מחיר הבד נוסף בהצלחה')
        except Exception as e:
            messagebox.showerror('שגיאה', str(e))

    def _delete_selected_fabric_price(self):
        """מחיקת מחיר בד נבחר"""
        if not hasattr(self, 'fabric_prices_tree'):
            return
        sel = self.fabric_prices_tree.selection()
        if not sel:
            messagebox.showwarning('אזהרה', 'בחר שורה למחיקה')
            return
        
        deleted = False
        for item in sel:
            vals = self.fabric_prices_tree.item(item, 'values')
            if vals:
                price_id = int(vals[0])
                if self.data_processor.delete_fabric_price(price_id):
                    deleted = True
        
        if deleted:
            self._load_fabric_prices_into_tree()
            messagebox.showinfo('הצלחה', 'מחיר הבד נמחק בהצלחה')

    def _save_cost_settings(self):
        """שמירת הגדרות עלויות גלובליות"""
        try:
            tick_price = float(self.cs_tick_price_var.get() or 0)
            elastic_price = float(self.cs_elastic_price_var.get() or 0)
            ribbon_price = float(self.cs_ribbon_price_var.get() or 0)
            sewing_price = float(self.cs_sewing_price_var.get() or 0)
        except ValueError:
            messagebox.showerror('שגיאה', 'כל המחירים חייבים להיות מספרים')
            return
        
        settings = {
            'tick_price': tick_price,
            'elastic_price': elastic_price,
            'ribbon_price': ribbon_price,
            'sewing_price': sewing_price
        }
        
        if self.data_processor.save_item_cost_settings(settings):
            messagebox.showinfo('הצלחה', 'ההגדרות נשמרו בהצלחה')
        else:
            messagebox.showerror('שגיאה', 'שגיאה בשמירת ההגדרות')

    # ===== LOADERS =====
    def _load_products_catalog_into_tree(self):
        """טעינת קטלוג המוצרים לטבלה (משתמש במערכת הסינון אם קיימת)"""
        if not hasattr(self, 'products_tree'): return
        
        # אם קיים סינון פעיל, השתמש בו
        if hasattr(self, 'product_filter_var'):
            self._filter_products_tree()
            return
            
        # אחרת, טען הכל
        for item in self.products_tree.get_children(): self.products_tree.delete(item)
        try:
            for rec in getattr(self.data_processor, 'products_catalog', []):
                fabric_category_value = rec.get('fabric_category') or 'בלי קטגוריה'
                main_category_value = rec.get('main_category') or 'בגדים'
                self.products_tree.insert('', 'end', values=(
                    rec.get('barcode', ''), rec.get('id'), rec.get('name'), main_category_value, rec.get('category',''), rec.get('size'), rec.get('fabric_type'),
                    rec.get('fabric_color'), rec.get('print_name'), fabric_category_value, rec.get('square_area', 0.0), rec.get('ticks_qty'), rec.get('elastic_qty'),
                    rec.get('ribbon_qty'), rec.get('created_at')
                ))
        except Exception:
            pass

    def _clear_product_filters(self):
        """איפוס כל הסינונים"""
        self.product_filter_var.set('')
        if hasattr(self, 'filter_fabric_category_var'):
            self.filter_fabric_category_var.set('הכל')
        self._filter_products_tree()

    def _filter_products_tree(self):
        """סינון טבלת הפריטים לפי שם הדגם וקטגוריית בד בזמן אמת - תומך בשתי תצוגות"""
        if not hasattr(self, 'products_tree'): return
        
        # קבלת ערך החיפוש לפי שם
        filter_text = self.product_filter_var.get().strip().lower()
        
        # קבלת סינון קטגוריית בד
        fabric_cat_filter = getattr(self, 'filter_fabric_category_var', None)
        fabric_cat_filter = fabric_cat_filter.get() if fabric_cat_filter else "הכל"
        
        # בדיקת מצב תצוגה
        is_cost_view = getattr(self, 'cost_view_mode', None) and self.cost_view_mode.get()
        
        # קבלת שיטת חישוב עלות בד
        fabric_cost_method = getattr(self, 'fabric_cost_method', None)
        fabric_cost_method = fabric_cost_method.get() if fabric_cost_method else 'sqm'
        
        # ניקוי הטבלה
        for item in self.products_tree.get_children():
            self.products_tree.delete(item)
        
        try:
            # טעינת כל הפריטים ממאגר הנתונים
            for rec in getattr(self.data_processor, 'products_catalog', []):
                product_name = rec.get('name', '').lower()
                product_fabric_cat = rec.get('fabric_category', '') or 'בלי קטגוריה'
                
                # בדיקת סינון שם
                name_match = not filter_text or filter_text in product_name
                # בדיקת סינון קטגוריית בד
                fabric_cat_match = fabric_cat_filter == "הכל" or product_fabric_cat == fabric_cat_filter
                
                # אם עובר את כל הסינונים
                if name_match and fabric_cat_match:
                    if is_cost_view:
                        # תצוגת עלויות - חישוב עלויות לפי השיטה הנבחרת
                        costs = self.data_processor.calculate_item_cost(rec, fabric_cost_method)
                        # בדיקה אם יש עלות בד ישירה (מסומן ב-* במצב משקל)
                        has_direct_fabric_cost = rec.get('fabric_cost') is not None and rec.get('fabric_cost') != ''
                        # הצגת סימון * רק במצב משקל ואם יש עלות ישירה
                        if fabric_cost_method == 'weight' and has_direct_fabric_cost:
                            fabric_cost_display = f"₪{costs['fabric_cost']:.2f} *"
                        elif fabric_cost_method == 'weight' and not has_direct_fabric_cost:
                            fabric_cost_display = "₪0.00 (לא הוגדר)"
                        else:
                            fabric_cost_display = f"₪{costs['fabric_cost']:.2f}"
                        self.products_tree.insert('', 'end', values=(
                            rec.get('barcode', ''),
                            rec.get('name', ''),
                            rec.get('size', ''),
                            rec.get('fabric_category', ''),
                            rec.get('fabric_color', ''),
                            rec.get('print_name', ''),
                            fabric_cost_display,
                            f"₪{costs['ticks_cost']:.2f}",
                            f"₪{costs['elastic_cost']:.2f}",
                            f"₪{costs['ribbon_cost']:.2f}",
                            f"₪{costs['sewing_cost']:.2f}",
                            f"₪{costs['total_cost']:.2f}"
                        ))
                    else:
                        # תצוגה רגילה
                        fabric_category_value = rec.get('fabric_category') or 'בלי קטגוריה'
                        main_category_value = rec.get('main_category') or 'בגדים'
                        self.products_tree.insert('', 'end', values=(
                            rec.get('barcode', ''), rec.get('id'), rec.get('name'), main_category_value, rec.get('category',''), 
                            rec.get('size'), rec.get('fabric_type'), rec.get('fabric_color'), 
                            rec.get('print_name'), fabric_category_value, rec.get('square_area', 0.0), 
                            rec.get('ticks_qty'), rec.get('elastic_qty'), rec.get('ribbon_qty'), 
                            rec.get('created_at')
                        ))
        except Exception as e:
            print(f"שגיאה בטעינת פריטים: {e}")

    def _load_accessories_into_tree(self):
        print("🔄 DEBUG: טוען אביזרים לטבלה...")
        if not hasattr(self, 'sewing_accessories_tree'):
            print("❌ DEBUG: sewing_accessories_tree לא קיים")
            return
        for item in self.sewing_accessories_tree.get_children():
            self.sewing_accessories_tree.delete(item)
        try:
            accessories_data = getattr(self.data_processor, 'sewing_accessories', [])
            print(f"📊 DEBUG: נמצאו {len(accessories_data)} אביזרים")
            for rec in accessories_data:
                print(f"➕ DEBUG: מוסיף אביזר: {rec.get('name')} (ID: {rec.get('id')})")
                self.sewing_accessories_tree.insert('', 'end', values=(
                    rec.get('id'), rec.get('name'), rec.get('unit'), rec.get('created_at')
                ))
            print(f"✅ DEBUG: נטענו {len(accessories_data)} אביזרים לטבלה")
        except Exception as e:
            print(f"❌ DEBUG: שגיאה בטעינת אביזרים: {e}")
            pass

    def _add_sewing_accessory(self):
        name = self.acc_name_var.get().strip()
        unit = self.acc_unit_var.get().strip()
        if not name:
            messagebox.showerror("שגיאה", "חובה להזין שם אביזר")
            return
        try:
            new_id = self.data_processor.add_sewing_accessory(name, unit)
            # Clear the form fields
            self.acc_name_var.set('')
            self.acc_unit_var.set('')
            # Reload the entire accessories tree to show the new item
            self._load_accessories_into_tree()
            messagebox.showinfo("הצלחה", f"אביזר '{name}' נוסף בהצלחה")
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _delete_selected_sewing_accessory(self):
        print("=" * 60)
        print("🔴🔴🔴 DEBUG: כפתור 'מחק נבחר' נלחץ בטאב אביזרי תפירה! 🔴🔴🔴")
        print("📍 מיקום: קטלוג מוצרים ופריטים > אביזרי תפירה > כפתור מחק נבחר")
        print("⏰ זמן: " + str(datetime.now()))
        print("=" * 60)
        
        # הודעת דיבאגינג נוספת
        import sys
        print(f"🐍 Python version: {sys.version}")
        print(f"🔍 Self type: {type(self)}")
        print(f"🔍 Has sewing_accessories_tree: {hasattr(self, 'sewing_accessories_tree')}")
        
        if not hasattr(self, 'sewing_accessories_tree'): 
            print("⚠️  טבלת אביזרים לא קיימת")
            return
            
        # בדיקה נוספת - איזה טבלה זה?
        print(f"🔍 DEBUG: sewing_accessories_tree ID: {id(self.sewing_accessories_tree)}")
        print(f"🔍 DEBUG: sewing_accessories_tree type: {type(self.sewing_accessories_tree)}")
        print(f"🔍 DEBUG: sewing_accessories_tree widget name: {self.sewing_accessories_tree.winfo_name()}")
        print(f"🔍 DEBUG: sewing_accessories_tree parent: {self.sewing_accessories_tree.master}")
        sel = self.sewing_accessories_tree.selection()
        print(f"🔍 DEBUG: sel = {sel}")
        print(f"🔍 DEBUG: type(sel) = {type(sel)}")
        print(f"🔍 DEBUG: len(sel) = {len(sel) if sel else 'None'}")
        
        # בדיקה נוספת - כל הפריטים בטבלה
        all_items = self.sewing_accessories_tree.get_children()
        print(f"🔍 DEBUG: כל הפריטים בטבלה: {all_items}")
        print(f"🔍 DEBUG: מספר פריטים בטבלה: {len(all_items)}")
        
        # בדיקה נוספת - נסה לטעון מחדש (רק אם הטבלה ריקה)
        if len(all_items) == 0:
            print("🔄 DEBUG: הטבלה ריקה, מנסה לטעון מחדש...")
            self._load_accessories_into_tree()
            all_items_after = self.sewing_accessories_tree.get_children()
            print(f"🔍 DEBUG: אחרי טעינה מחדש - מספר פריטים: {len(all_items_after)}")
            sel_after = self.sewing_accessories_tree.selection()
            print(f"🔍 DEBUG: אחרי טעינה מחדש - sel: {sel_after}")
        else:
            print("✅ DEBUG: הטבלה לא ריקה, לא טוען מחדש")
            print(f"🔍 DEBUG: הטבלה מכילה {len(all_items)} פריטים")
            # בדיקה נוספת - נסה לקבל את הבחירה שוב
            sel_again = self.sewing_accessories_tree.selection()
            print(f"🔍 DEBUG: בדיקה נוספת - sel: {sel_again}")
            if sel_again:
                print("✅ DEBUG: נמצאה בחירה!")
            else:
                print("❌ DEBUG: עדיין אין בחירה")
        
        if not sel: 
            print("⚠️  לא נבחרו אביזרים למחיקה")
            print("💡 טיפ: לחץ על שורה בטבלה כדי לבחור אותה, ואז לחץ על 'מחק נבחר'")
            messagebox.showwarning("אזהרה", "אנא בחר אביזר למחיקה\nלחץ על שורה בטבלה כדי לבחור אותה")
            return
        
        print(f"📋 נבחרו {len(sel)} אביזר(ים) למחיקה")
        print(f"🔍 אביזרים נבחרים: {[self.sewing_accessories_tree.item(item, 'values')[1] for item in sel]}")
        deleted = False
        deleted_items = []
        
        for item in sel:
            vals = self.sewing_accessories_tree.item(item, 'values')
            if vals:
                accessory_name = vals[1] if len(vals) > 1 else "לא ידוע"
                accessory_id = vals[0]
                print(f"🗑️  מוחק אביזר: {accessory_name} (ID: {accessory_id})")
                
                if self.data_processor.delete_sewing_accessory(int(vals[0])):
                    deleted = True
                    deleted_items.append(accessory_name)
                    print(f"✅ אביזר '{accessory_name}' נמחק בהצלחה")
                else:
                    print(f"❌ שגיאה במחיקת אביזר '{accessory_name}'")
        
        if deleted:
            print(f"🎉 סה\"כ נמחקו {len(deleted_items)} אביזר(ים): {', '.join(deleted_items)}")
            self._load_accessories_into_tree()
            messagebox.showinfo("הצלחה", f"נמחקו {len(deleted_items)} אביזר(ים) בהצלחה")
        else:
            print("❌ לא נמחקו אביזרים")
            messagebox.showerror("שגיאה", "לא ניתן למחוק את האביזרים הנבחרים")

    def _load_categories_into_tree(self):
        if not hasattr(self, 'categories_tree'): return
        for item in self.categories_tree.get_children():
            self.categories_tree.delete(item)
        try:
            for rec in getattr(self.data_processor, 'categories', []):
                self.categories_tree.insert('', 'end', values=(rec.get('id'), rec.get('name'), rec.get('created_at')))
        except Exception:
            pass

    def _load_main_categories_into_tree(self):
        if not hasattr(self, 'main_categories_tree'): return
        for item in self.main_categories_tree.get_children():
            self.main_categories_tree.delete(item)
        try:
            for rec in getattr(self.data_processor, 'main_categories', []):
                self.main_categories_tree.insert('', 'end', values=(rec.get('id'), rec.get('name'), rec.get('created_at')))
        except Exception:
            pass

    def _refresh_main_categories_for_products(self):
        names = [c.get('name','') for c in getattr(self.data_processor, 'main_categories', [])]
        if hasattr(self, 'prod_main_category_combobox'):
            self.prod_main_category_combobox['values'] = names
            cur = self.prod_main_category_var.get()
            if cur not in names:
                self.prod_main_category_var.set('')

    def _apply_main_category_field_visibility(self):
        # When no main category is selected: hide all product fields.
        # When selected: show only fields configured for that category.
        name = self.prod_main_category_var.get().strip() if hasattr(self, 'prod_main_category_var') else ''
        fields = []
        if name:
            rec = self._find_main_category_by_name(name) if hasattr(self, '_find_main_category_by_name') else None
            if rec:
                fields = rec.get('fields') or []
                if not fields and hasattr(self, 'data_processor'):
                    try:
                        fields = self.data_processor.get_main_category_fields(rec.get('id'))
                    except Exception:
                        fields = []
        # If no name chosen -> hide everything
        for key, widgets in getattr(self, '_products_field_widgets', {}).items():
            visible = (bool(name) and (not fields or key in fields))
            for w in widgets:
                try:
                    if visible:
                        w.grid()
                    else:
                        w.grid_remove()
                except Exception:
                    pass

    def _add_category(self):
        name = self.cat_name_var.get().strip()
        if not name:
            messagebox.showerror("שגיאה", "חובה להזין שם קטגוריה")
            return
        try:
            new_id = self.data_processor.add_category(name)
            self.categories_tree.insert('', 'end', values=(new_id, name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            self.cat_name_var.set('')
            self._refresh_categories_for_products()
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _delete_selected_category(self):
        if not hasattr(self, 'categories_tree'): return
        sel = self.categories_tree.selection()
        if not sel: return
        deleted = False
        for item in sel:
            vals = self.categories_tree.item(item, 'values')
            if vals:
                if self.data_processor.delete_category(int(vals[0])):
                    deleted = True
        if deleted:
            self._load_categories_into_tree()
            self._refresh_categories_for_products()

    def _add_main_category(self):
        name = self.main_cat_name_var.get().strip()
        if not name:
            messagebox.showerror("שגיאה", "חובה להזין שם קטגוריה ראשית")
            return
        try:
            new_id = self.data_processor.add_main_category(name)
            self.main_categories_tree.insert('', 'end', values=(new_id, name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            self.main_cat_name_var.set('')
            self._load_main_categories_for_fields()
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _delete_selected_main_category(self):
        if not hasattr(self, 'main_categories_tree'): return
        sel = self.main_categories_tree.selection()
        if not sel: return
        deleted = False
        for item in sel:
            vals = self.main_categories_tree.item(item, 'values')
            if vals:
                if self.data_processor.delete_main_category(int(vals[0])):
                    deleted = True
        if deleted:
            self._load_main_categories_into_tree()
            self._load_main_categories_for_fields()

    def _load_sizes_into_tree(self):
        if not hasattr(self, 'sizes_tree'): return
        for item in self.sizes_tree.get_children(): self.sizes_tree.delete(item)
        for rec in getattr(self.data_processor, 'product_sizes', []):
            self.sizes_tree.insert('', 'end', values=(rec.get('id'), rec.get('name'), rec.get('created_at')))

    def _load_fabric_types_into_tree(self):
        if not hasattr(self, 'fabric_types_tree'): return
        for item in self.fabric_types_tree.get_children(): self.fabric_types_tree.delete(item)
        for rec in getattr(self.data_processor, 'product_fabric_types', []):
            self.fabric_types_tree.insert('', 'end', values=(rec.get('id'), rec.get('name'), rec.get('created_at')))

    def _load_fabric_colors_into_tree(self):
        if not hasattr(self, 'fabric_colors_tree'): return
        for item in self.fabric_colors_tree.get_children(): self.fabric_colors_tree.delete(item)
        for rec in getattr(self.data_processor, 'product_fabric_colors', []):
            self.fabric_colors_tree.insert('', 'end', values=(rec.get('id'), rec.get('name'), rec.get('created_at')))

    def _load_print_names_into_tree(self):
        if not hasattr(self, 'print_names_tree'): return
        for item in self.print_names_tree.get_children(): self.print_names_tree.delete(item)
        for rec in getattr(self.data_processor, 'product_print_names', []):
            self.print_names_tree.insert('', 'end', values=(rec.get('id'), rec.get('name'), rec.get('created_at')))

    def _load_fabric_categories_into_tree(self):
        if not hasattr(self, 'fabric_categories_tree'): return
        for item in self.fabric_categories_tree.get_children(): self.fabric_categories_tree.delete(item)
        for rec in getattr(self.data_processor, 'product_fabric_categories', []):
            self.fabric_categories_tree.insert('', 'end', values=(rec.get('id'), rec.get('name'), rec.get('created_at')))

    def _load_model_names_into_tree(self):
        if not hasattr(self, 'model_names_tree'): return
        for item in self.model_names_tree.get_children(): self.model_names_tree.delete(item)
        for rec in getattr(self.data_processor, 'product_model_names', []):
            sewing_price = rec.get('sewing_price', 0) or 0
            self.model_names_tree.insert('', 'end', values=(
                rec.get('id'), 
                rec.get('name'), 
                f"₪{sewing_price:.2f}" if sewing_price else "₪0.00",
                rec.get('created_at')
            ))

    def _on_model_name_select(self, event):
        """כאשר נבחר שם דגם - הצג את מחיר התפירה שלו"""
        if not hasattr(self, 'model_names_tree'): return
        sel = self.model_names_tree.selection()
        if sel:
            vals = self.model_names_tree.item(sel[0], 'values')
            if vals and len(vals) >= 3:
                # קבל את המחיר מהערך (הסר את הסימן ₪)
                price_str = str(vals[2]).replace('₪', '').strip()
                try:
                    self.attr_model_sewing_price_var.set(price_str)
                except:
                    self.attr_model_sewing_price_var.set('0')

    def _update_selected_model_sewing_price(self):
        """עדכון מחיר תפירה לשם דגם נבחר"""
        if not hasattr(self, 'model_names_tree'): return
        sel = self.model_names_tree.selection()
        if not sel:
            messagebox.showwarning('אזהרה', 'בחר שם דגם לעדכון')
            return
        
        vals = self.model_names_tree.item(sel[0], 'values')
        if not vals:
            return
        
        rec_id = int(vals[0])
        try:
            sewing_price = float(self.attr_model_sewing_price_var.get() or 0)
        except ValueError:
            messagebox.showerror('שגיאה', 'מחיר תפירה חייב להיות מספר')
            return
        
        if self.data_processor.update_model_name_sewing_price(rec_id, sewing_price):
            self._load_model_names_into_tree()
            messagebox.showinfo('הצלחה', 'מחיר התפירה עודכן בהצלחה')
        else:
            messagebox.showerror('שגיאה', 'שגיאה בעדכון מחיר התפירה')

    # ===== ADD =====
    def _add_product_catalog_entry(self):
        name = self.prod_name_var.get().strip()
        if not name:
            messagebox.showerror("שגיאה", "חובה להזין שם מוצר")
            return
        category_raw = self.prod_category_var.get().strip()
        valid_categories = [c.get('name','') for c in getattr(self.data_processor, 'categories', [])]
        # category_raw may be comma-separated multi-select; validate only if provided
        category_tokens = [s.strip() for s in category_raw.split(',') if s.strip()] if category_raw else []
        if category_tokens:
            for ct in category_tokens:
                if ct not in valid_categories:
                    messagebox.showerror("שגיאה", f"תת קטגוריה '{ct}' לא קיימת. הוסף בטאב 'תת קטגוריות' ובחר שוב")
                    return
        # Normalize the category field as a single comma-joined string; allow empty (optional)
        category_value = ",".join(category_tokens) if category_tokens else ''
        sizes_raw = self.prod_size_var.get().strip()
        ftypes_raw = self.prod_fabric_type_var.get().strip()
        fcolors_raw = self.prod_fabric_color_var.get().strip()
        prints_raw = self.prod_print_name_var.get().strip()
        fabric_category_raw = self.prod_fabric_category_var.get().strip()
        square_area_raw = self.prod_square_area_var.get().strip()
        ticks_raw = self.prod_ticks_var.get().strip()
        elastic_raw = self.prod_elastic_var.get().strip()
        ribbon_raw = self.prod_ribbon_var.get().strip()

        def _split(raw):
            if not raw:
                return ['']
            return [s.strip() for s in re.split(r'[;,.\s]+', raw) if s.strip()]

        size_tokens = _split(sizes_raw)
        ft_tokens = _split(ftypes_raw)
        fc_tokens = _split(fcolors_raw)
        pn_tokens = _split(prints_raw)

        def _normalize_size(tok: str) -> str:
            t = tok.strip()
            if not t:
                return t
            if t.startswith('-'):
                t = t[1:]
            if '-' not in t and t.isdigit() and 3 <= len(t) <= 4:
                mid = len(t)//2
                a, b = t[:mid], t[mid:]
                try:
                    ai = int(a); bi = int(b)
                    if 0 <= ai <= 60 and 0 <= bi <= 60 and ai < bi:
                        return f"{ai}-{bi}"
                except Exception:
                    pass
            return t

        size_tokens = [_normalize_size(s) for s in size_tokens]

        from itertools import product
        # Build combinations only across size / fabric type / color / print.
        # Category stays as the full comma-joined list (category_value).
        combos = list(product(size_tokens, ft_tokens, fc_tokens, pn_tokens)) or [( '', '', '', '' )]

        existing = set()
        try:
            for rec in getattr(self.data_processor, 'products_catalog', []):
                existing.add((rec.get('name','').strip(), rec.get('category','').strip(), rec.get('size','').strip(), rec.get('fabric_type','').strip(), rec.get('fabric_color','').strip(), rec.get('print_name','').strip()))
        except Exception:
            existing = set()

        if len(combos) == 1:
            only_sz, only_ft, only_fc, only_pn = combos[0]
            single_key = (name, category_value, only_sz, only_ft, only_fc, only_pn)
            if single_key in existing:
                messagebox.showinfo(
                    "כפילות",
            "המוצר עם הנתונים הללו כבר קיים במערכת:\n"
            f"שם: {name}\nתת קטגוריה: {category_value or '-'}\nמידה: {only_sz or '-'}\nסוג בד: {only_ft or '-'}\nצבע בד: {only_fc or '-'}\nשם פרינט: {only_pn or '-'}"
                )
                return

        added = 0
        try:
            for sz, ft, fc, pn in combos:
                key = (name, category_value, sz, ft, fc, pn)
                if key in existing:
                    continue
                new_id, new_barcode = self.data_processor.add_product_catalog_entry(
                    name, sz, ft, fc, pn, category_value, ticks_raw, elastic_raw, ribbon_raw, fabric_category_raw, square_area_raw
                )
                existing.add(key)
                added += 1
                fabric_category_value = fabric_category_raw or 'בלי קטגוריה'
                main_category_value = self.prod_main_category_var.get().strip() or 'בגדים'
                self.products_tree.insert('', 'end', values=(
                    new_barcode, new_id, name, main_category_value, category_value, sz, ft, fc, pn, fabric_category_value,
                    square_area_raw or 0.0, ticks_raw or 0, elastic_raw or 0, ribbon_raw or 0,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ))
            self.prod_name_var.set(''); self.prod_category_var.set(''); self.prod_fabric_category_var.set(''); self.prod_size_var.set(''); self.prod_fabric_type_var.set(''); self.prod_fabric_color_var.set(''); self.prod_print_name_var.set(''); self.prod_square_area_var.set(''); self.prod_ticks_var.set(''); self.prod_elastic_var.set(''); self.prod_ribbon_var.set('')
            self.selected_sizes.clear(); self.selected_fabric_types.clear(); self.selected_fabric_colors.clear(); self.selected_print_names.clear()
            if hasattr(self, 'size_picker'): self.size_picker.set('')
            if hasattr(self, 'ftype_picker'): self.ftype_picker.set('')
            if hasattr(self, 'fcolor_picker'): self.fcolor_picker.set('')
            if hasattr(self, 'pname_picker'): self.pname_picker.set('')
            if hasattr(self, 'fabric_category_combobox'): self.fabric_category_combobox.set('')
            if hasattr(self, 'model_name_combobox'): self.model_name_combobox.set('')
            if added > 1:
                messagebox.showinfo("הצלחה", f"נוספו {added} וריאנטים למוצר '{name}'")
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _delete_selected_product_entry(self):
        sel = self.products_tree.selection()
        if not sel: return
        ids = []
        for item in sel:
            vals = self.products_tree.item(item, 'values')
            if vals:
                # vals[1] is now the ID (vals[0] is barcode)
                ids.append(int(vals[1]))
        if not ids: return
        deleted_any = False
        for _id in ids:
            if self.data_processor.delete_product_catalog_entry(_id):
                deleted_any = True
        if deleted_any:
            self._load_products_catalog_into_tree()

    def _edit_selected_product(self):
        """פתיחת חלון עריכה לפריט נבחר - עריכת כמויות טיקטקים, גומי וסרט"""
        # בדיקה שנבחר פריט
        sel = self.products_tree.selection()
        if not sel:
            messagebox.showwarning("אזהרה", "אנא בחר פריט לעריכה")
            return
        
        # קבלת הערכים מהשורה הנבחרת (תצוגה רגילה בלבד)
        is_cost_view = getattr(self, 'cost_view_mode', None) and self.cost_view_mode.get()
        if is_cost_view:
            messagebox.showwarning("אזהרה", "עריכה זמינה רק בתצוגה רגילה.\nעבור לתצוגה רגילה תחילה.")
            return
        
        vals = self.products_tree.item(sel[0], 'values')
        if not vals:
            return
        
        # vals בתצוגה רגילה: barcode, id, name, main_category, category, size, fabric_type, fabric_color, print_name, fabric_category, square_area, ticks_qty, elastic_qty, ribbon_qty, created_at
        try:
            product_id = int(vals[1])
            product_name = vals[2]
            product_size = vals[5]
            current_ticks = int(vals[11]) if vals[11] else 0
            current_elastic = int(vals[12]) if vals[12] else 0
            current_ribbon = int(vals[13]) if vals[13] else 0
        except (ValueError, IndexError) as e:
            messagebox.showerror("שגיאה", f"שגיאה בקריאת נתוני הפריט: {e}")
            return
        
        # יצירת חלון עריכה
        top = tk.Toplevel(self.notebook)
        top.title("עריכת כמויות פריט")
        top.grab_set()
        top.resizable(False, False)
        
        frm = ttk.Frame(top, padding=15)
        frm.pack(fill='both', expand=True)
        
        # כותרת עם שם הפריט
        ttk.Label(frm, text=f"פריט: {product_name}", font=('Arial', 11, 'bold')).grid(row=0, column=0, columnspan=2, sticky='w', pady=(0, 5))
        ttk.Label(frm, text=f"מידה: {product_size}", font=('Arial', 10)).grid(row=1, column=0, columnspan=2, sticky='w', pady=(0, 15))
        
        # משתנים לשדות
        ticks_var = tk.StringVar(value=str(current_ticks))
        elastic_var = tk.StringVar(value=str(current_elastic))
        ribbon_var = tk.StringVar(value=str(current_ribbon))
        
        # שדות עריכה
        ttk.Label(frm, text="כמות טיקטקים:", font=('Arial', 10)).grid(row=2, column=0, sticky='e', padx=5, pady=5)
        ttk.Entry(frm, textvariable=ticks_var, width=10).grid(row=2, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frm, text="כמות גומי:", font=('Arial', 10)).grid(row=3, column=0, sticky='e', padx=5, pady=5)
        ttk.Entry(frm, textvariable=elastic_var, width=10).grid(row=3, column=1, sticky='w', padx=5, pady=5)
        
        ttk.Label(frm, text="כמות סרט:", font=('Arial', 10)).grid(row=4, column=0, sticky='e', padx=5, pady=5)
        ttk.Entry(frm, textvariable=ribbon_var, width=10).grid(row=4, column=1, sticky='w', padx=5, pady=5)
        
        # כפתורים
        btn_frame = ttk.Frame(frm)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(15, 0))
        
        def _save():
            try:
                new_ticks = int(ticks_var.get()) if ticks_var.get() else 0
                new_elastic = int(elastic_var.get()) if elastic_var.get() else 0
                new_ribbon = int(ribbon_var.get()) if ribbon_var.get() else 0
                
                if new_ticks < 0 or new_elastic < 0 or new_ribbon < 0:
                    messagebox.showerror("שגיאה", "הכמויות חייבות להיות מספרים חיוביים")
                    return
                
                success = self.data_processor.update_product_quantities(
                    product_id=product_id,
                    ticks_qty=new_ticks,
                    elastic_qty=new_elastic,
                    ribbon_qty=new_ribbon
                )
                
                if success:
                    self._load_products_catalog_into_tree()
                    messagebox.showinfo("הצלחה", "הפריט עודכן בהצלחה")
                    top.destroy()
                else:
                    messagebox.showerror("שגיאה", "לא נמצא פריט לעדכון")
            except ValueError:
                messagebox.showerror("שגיאה", "אנא הזן מספרים שלמים בלבד")
        
        ttk.Button(btn_frame, text="שמור", command=_save).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="ביטול", command=top.destroy).pack(side='left', padx=5)

    def _export_products_catalog(self):
        if not getattr(self.data_processor, 'products_catalog', []):
            messagebox.showerror("שגיאה", "אין מוצרים לייצוא")
            return
        file_path = filedialog.asksaveasfilename(title="ייצוא קטלוג מוצרים", defaultextension='.xlsx', filetypes=[('Excel','*.xlsx')])
        if not file_path: return
        try:
            self.data_processor.export_products_catalog_to_excel(file_path)
            messagebox.showinfo("הצלחה", "הקטלוג יוצא בהצלחה")
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _import_products_catalog_dialog(self):
        top = tk.Toplevel(self.notebook)
        top.title("יבוא קטלוג מוצרים מ-Excel")
        top.grab_set(); top.resizable(False, False)
        frm = ttk.Frame(top, padding=10)
        frm.pack(fill='both', expand=True)

        path_var = tk.StringVar(); overwrite_var = tk.BooleanVar(value=False)

        ttk.Label(frm, text="בחר קובץ Excel בפורמט הייצוא:").grid(row=0, column=0, columnspan=3, sticky='w', pady=(0,6))
        ttk.Entry(frm, textvariable=path_var, width=50).grid(row=1, column=0, columnspan=2, sticky='we', padx=(0,6))
        def _browse():
            p = filedialog.askopenfilename(title="בחר קובץ קטלוג", filetypes=[('Excel','*.xlsx')])
            if p: path_var.set(p)
        ttk.Button(frm, text="עיון...", command=_browse).grid(row=1, column=2, sticky='w')

        ttk.Checkbutton(frm, text="דרוס את הטבלה (במקום להוסיף)", variable=overwrite_var).grid(row=2, column=0, columnspan=3, sticky='w', pady=8)

        btns = ttk.Frame(frm); btns.grid(row=3, column=0, columnspan=3, sticky='e')
        def _do_import():
            path = path_var.get().strip()
            if not path:
                messagebox.showerror("שגיאה", "בחר קובץ לייבוא")
                return
            mode = 'overwrite' if overwrite_var.get() else 'append'
            try:
                res = self.data_processor.import_products_catalog_from_excel(path, mode=mode)
                self._load_products_catalog_into_tree()
                msg = f"יובאו {res.get('imported',0)} רשומות."
                skipped = res.get('skipped_duplicates',0)
                if skipped: msg += f"\nדולגו {skipped} כפילויות."
                if res.get('overwritten'): msg += "\nבוצעה דריסה מלאה של הטבלה."
                messagebox.showinfo("הצלחה", msg); top.destroy()
            except Exception as e:
                messagebox.showerror("שגיאה", str(e))
        ttk.Button(btns, text="ייבוא", command=_do_import).pack(side='right', padx=4)
        ttk.Button(btns, text="ביטול", command=top.destroy).pack(side='right', padx=4)

    # ===== ATTR helpers =====
    def _on_attr_select(self, kind: str):
        picker_map = {
            'sub_category': (self.category_combobox, self.selected_sub_categories, self.prod_category_var),
            'size': (self.size_picker, self.selected_sizes, self.prod_size_var),
            'fabric_type': (self.ftype_picker, self.selected_fabric_types, self.prod_fabric_type_var),
            'fabric_color': (self.fcolor_picker, self.selected_fabric_colors, self.prod_fabric_color_var),
            'print_name': (self.pname_picker, self.selected_print_names, self.prod_print_name_var)
        }
        picker, lst, var = picker_map.get(kind)
        val = picker.get().strip()
        if val and val not in lst:
            lst.append(val)
            var.set(','.join(lst))
        picker.set('')

    def _clear_attr(self, kind: str):
        if kind == 'sub_category':
            self.selected_sub_categories.clear(); self.prod_category_var.set('')
        elif kind == 'size':
            self.selected_sizes.clear(); self.prod_size_var.set('')
        elif kind == 'fabric_type':
            self.selected_fabric_types.clear(); self.prod_fabric_type_var.set('')
        elif kind == 'fabric_color':
            self.selected_fabric_colors.clear(); self.prod_fabric_color_var.set('')
        elif kind == 'print_name':
            self.selected_print_names.clear(); self.prod_print_name_var.set('')

    def _refresh_attribute_pickers(self):
        if hasattr(self, 'size_picker'):
            self.size_picker['values'] = [r.get('name') for r in getattr(self.data_processor,'product_sizes',[])]
        if hasattr(self, 'ftype_picker'):
            self.ftype_picker['values'] = [r.get('name') for r in getattr(self.data_processor,'product_fabric_types',[])]
        if hasattr(self, 'fcolor_picker'):
            self.fcolor_picker['values'] = [r.get('name') for r in getattr(self.data_processor,'product_fabric_colors',[])]
        if hasattr(self, 'pname_picker'):
            self.pname_picker['values'] = [r.get('name') for r in getattr(self.data_processor,'product_print_names',[])]
        if hasattr(self, 'fabric_category_combobox'):
            names = [r.get('name') for r in getattr(self.data_processor, 'product_fabric_categories', [])]
            self.fabric_category_combobox['values'] = names
            if self.prod_fabric_category_var.get() not in names:
                self.prod_fabric_category_var.set('')
        if hasattr(self, 'model_name_combobox'):
            model_names = [r.get('name') for r in getattr(self.data_processor, 'product_model_names', [])]
            self.model_name_combobox['values'] = model_names
            if self.prod_name_var.get() not in model_names:
                self.prod_name_var.set('')

    def _refresh_categories_for_products(self):
        if hasattr(self, 'category_combobox'):
            names = [c.get('name','') for c in getattr(self.data_processor, 'categories', [])]
            self.category_combobox['values'] = names
            # Do not clear self.prod_category_var because it can contain a comma-separated multi-select
            # Just ensure the picker (combobox) has no stale selection
            if self.category_combobox.get() not in names:
                self.category_combobox.set('')
        self._refresh_attribute_pickers()

    # ===== ATTR add/delete =====
    def _add_product_size(self):
        name = self.attr_size_var.get().strip()
        if not name:
            messagebox.showerror('שגיאה', 'חובה להזין מידה')
            return
        try:
            new_id = self.data_processor.add_product_size(name)
            self.sizes_tree.insert('', 'end', values=(new_id, name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            self.attr_size_var.set(''); self._refresh_attribute_pickers()
        except Exception as e:
            messagebox.showerror('שגיאה', str(e))

    def _add_fabric_type_item(self):
        name = self.attr_fabric_type_var.get().strip()
        if not name:
            messagebox.showerror('שגיאה', 'חובה להזין סוג בד')
            return
        try:
            new_id = self.data_processor.add_fabric_type_item(name)
            self.fabric_types_tree.insert('', 'end', values=(new_id, name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            self.attr_fabric_type_var.set(''); self._refresh_attribute_pickers()
        except Exception as e:
            messagebox.showerror('שגיאה', str(e))

    def _add_fabric_color_item(self):
        name = self.attr_fabric_color_var.get().strip()
        if not name:
            messagebox.showerror('שגיאה', 'חובה להזין צבע בד')
            return
        try:
            new_id = self.data_processor.add_fabric_color_item(name)
            self.fabric_colors_tree.insert('', 'end', values=(new_id, name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            self.attr_fabric_color_var.set(''); self._refresh_attribute_pickers()
        except Exception as e:
            messagebox.showerror('שגיאה', str(e))

    def _add_print_name_item(self):
        name = self.attr_print_name_var.get().strip()
        if not name:
            messagebox.showerror('שגיאה', 'חובה להזין שם פרינט')
            return
        try:
            new_id = self.data_processor.add_print_name_item(name)
            self.print_names_tree.insert('', 'end', values=(new_id, name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            self.attr_print_name_var.set(''); self._refresh_attribute_pickers()
        except Exception as e:
            messagebox.showerror('שגיאה', str(e))

    def _add_fabric_category_item(self):
        name = self.attr_fabric_category_var.get().strip()
        if not name:
            messagebox.showerror('שגיאה', 'חובה להזין קטגוריית בד')
            return
        try:
            new_id = self.data_processor.add_fabric_category_item(name)
            self.fabric_categories_tree.insert('', 'end', values=(new_id, name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            self.attr_fabric_category_var.set(''); self._refresh_attribute_pickers()
        except Exception as e:
            messagebox.showerror('שגיאה', str(e))

    def _add_model_name_item(self):
        name = self.attr_model_name_var.get().strip()
        if not name:
            messagebox.showerror('שגיאה', 'חובה להזין שם דגם')
            return
        try:
            sewing_price = float(self.attr_model_sewing_price_var.get() or 0)
        except ValueError:
            sewing_price = 0.0
        try:
            new_id = self.data_processor.add_model_name_item(name, sewing_price)
            self._load_model_names_into_tree()
            self.attr_model_name_var.set('')
            self.attr_model_sewing_price_var.set('0')
        except Exception as e:
            messagebox.showerror('שגיאה', str(e))

    def _delete_selected_product_size(self):
        if not hasattr(self, 'sizes_tree'): return
        sel = self.sizes_tree.selection(); deleted = False
        for item in sel:
            vals = self.sizes_tree.item(item, 'values')
            if vals and self.data_processor.delete_product_size(int(vals[0])):
                deleted = True
        if deleted:
            self._load_sizes_into_tree(); self._refresh_attribute_pickers()

    def _delete_selected_fabric_type_item(self):
        if not hasattr(self, 'fabric_types_tree'): return
        sel = self.fabric_types_tree.selection(); deleted = False
        for item in sel:
            vals = self.fabric_types_tree.item(item, 'values')
            if vals and self.data_processor.delete_fabric_type_item(int(vals[0])):
                deleted = True
        if deleted:
            self._load_fabric_types_into_tree(); self._refresh_attribute_pickers()

    def _delete_selected_fabric_color_item(self):
        if not hasattr(self, 'fabric_colors_tree'): return
        sel = self.fabric_colors_tree.selection(); deleted = False
        for item in sel:
            vals = self.fabric_colors_tree.item(item, 'values')
            if vals and self.data_processor.delete_fabric_color_item(int(vals[0])):
                deleted = True
        if deleted:
            self._load_fabric_colors_into_tree(); self._refresh_attribute_pickers()

    def _delete_selected_print_name_item(self):
        if not hasattr(self, 'print_names_tree'): return
        sel = self.print_names_tree.selection(); deleted = False
        for item in sel:
            vals = self.print_names_tree.item(item, 'values')
            if vals and self.data_processor.delete_print_name_item(int(vals[0])):
                deleted = True
        if deleted:
            self._load_print_names_into_tree(); self._refresh_attribute_pickers()

    def _delete_selected_fabric_category_item(self):
        if not hasattr(self, 'fabric_categories_tree'): return
        sel = self.fabric_categories_tree.selection(); deleted = False
        for item in sel:
            vals = self.fabric_categories_tree.item(item, 'values')
            if vals and self.data_processor.delete_fabric_category_item(int(vals[0])):
                deleted = True
        if deleted:
            self._load_fabric_categories_into_tree(); self._refresh_attribute_pickers()

    def _delete_selected_model_name_item(self):
        if not hasattr(self, 'model_names_tree'): return
        sel = self.model_names_tree.selection(); deleted = False
        for item in sel:
            vals = self.model_names_tree.item(item, 'values')
            if vals and self.data_processor.delete_model_name_item(int(vals[0])):
                deleted = True
        if deleted:
            self._load_model_names_into_tree()

    # ===== barcodes section =====
    def _build_barcodes_section(self, parent):
        """Build the barcodes generation section."""
        # Main frame
        main_frame = tk.Frame(parent, bg='#f7f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(
            main_frame,
            text="ייצור ברקודים EAN-13",
            font=('Arial', 18, 'bold'),
            bg='#f7f9fa',
            fg='#2c3e50'
        )
        title_label.pack(pady=(0, 20))
        
        # Last barcode display frame
        last_barcode_frame = ttk.LabelFrame(main_frame, text="ברקוד אחרון במערכת", padding=20)
        last_barcode_frame.pack(fill='x', pady=(0, 20))
        
        self.last_barcode_var = tk.StringVar()
        self._load_last_barcode()
        
        last_barcode_display = tk.Label(
            last_barcode_frame,
            textvariable=self.last_barcode_var,
            font=('Arial', 28, 'bold'),
            bg='#ecf0f1',
            fg='#2c3e50',
            relief='sunken',
            padx=20,
            pady=15
        )
        last_barcode_display.pack()
        
        # Generation frame
        generation_frame = ttk.LabelFrame(main_frame, text="ייצור ברקודים חדשים", padding=20)
        generation_frame.pack(fill='x', pady=(0, 20))
        
        # Quantity input
        input_frame = tk.Frame(generation_frame, bg='#f7f9fa')
        input_frame.pack(pady=10)
        
        tk.Label(
            input_frame,
            text="כמות ברקודים לייצור:",
            font=('Arial', 12, 'bold'),
            bg='#f7f9fa'
        ).pack(side='right', padx=10)
        
        self.barcode_quantity_var = tk.StringVar(value="10")
        quantity_entry = tk.Entry(
            input_frame,
            textvariable=self.barcode_quantity_var,
            font=('Arial', 12),
            width=10,
            justify='center'
        )
        quantity_entry.pack(side='right', padx=10)
        
        # Generate button
        generate_btn = tk.Button(
            generation_frame,
            text="🔢 ייצר ברקודים",
            command=self._generate_barcodes,
            bg='#27ae60',
            fg='white',
            font=('Arial', 14, 'bold'),
            padx=30,
            pady=10,
            cursor='hand2'
        )
        generate_btn.pack(pady=10)
        
        # Instructions
        instructions_frame = ttk.LabelFrame(main_frame, text="הוראות שימוש", padding=15)
        instructions_frame.pack(fill='x')
        
        instructions_text = """
        1. הברקוד האחרון במערכת מוצג למעלה
        2. הזן את מספר הברקודים שברצונך לייצר
        3. לחץ על "ייצר ברקודים"
        4. הברקודים ייוצרו בקובץ Excel ויישמרו בתיקיית exports/barcodes
        5. המערכת תעדכן אוטומטית את הברקוד האחרון
        """
        
        instructions_label = tk.Label(
            instructions_frame,
            text=instructions_text,
            font=('Arial', 10),
            bg='#f7f9fa',
            fg='#34495e',
            justify='right'
        )
        instructions_label.pack()

    def _load_last_barcode(self):
        """Load the last barcode from JSON file."""
        try:
            barcode_file = 'barcodes_data.json'
            if os.path.exists(barcode_file):
                with open(barcode_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    last_barcode = data.get('last_barcode', '7297555019592')
            else:
                last_barcode = '7297555019592'
            
            self.last_barcode_var.set(last_barcode)
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בטעינת הברקוד האחרון:\n{str(e)}")
            self.last_barcode_var.set('7297555019592')

    def _save_last_barcode(self, barcode):
        """Save the last barcode to JSON file."""
        try:
            barcode_file = 'barcodes_data.json'
            data = {
                'last_barcode': barcode,
                'last_updated': datetime.now().isoformat()
            }
            with open(barcode_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            self.last_barcode_var.set(barcode)
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בשמירת הברקוד האחרון:\n{str(e)}")

    def _calculate_ean13_checksum(self, base_12_digits):
        """
        Calculate EAN-13 checksum digit.
        
        Args:
            base_12_digits: String of 12 digits
            
        Returns:
            Single digit checksum (0-9)
        """
        if len(base_12_digits) != 12:
            raise ValueError("Base must be exactly 12 digits")
        
        # Sum odd positions (1st, 3rd, 5th, etc. - index 0, 2, 4...)
        odd_sum = sum(int(base_12_digits[i]) for i in range(0, 12, 2))
        
        # Sum even positions (2nd, 4th, 6th, etc. - index 1, 3, 5...) and multiply by 3
        even_sum = sum(int(base_12_digits[i]) for i in range(1, 12, 2)) * 3
        
        # Total sum
        total = odd_sum + even_sum
        
        # Checksum is (10 - (total mod 10)) mod 10
        checksum = (10 - (total % 10)) % 10
        
        return str(checksum)

    def _generate_next_barcode(self, current_barcode):
        """
        Generate the next barcode in sequence with proper EAN-13 checksum.
        
        Args:
            current_barcode: Current 13-digit EAN-13 barcode
            
        Returns:
            Next 13-digit EAN-13 barcode
        """
        if len(current_barcode) != 13:
            raise ValueError("Barcode must be exactly 13 digits")
        
        # Remove checksum digit (last digit)
        base_12 = current_barcode[:12]
        
        # Increment the base
        base_number = int(base_12) + 1
        
        # Pad back to 12 digits
        new_base_12 = str(base_number).zfill(12)
        
        # Calculate new checksum
        checksum = self._calculate_ean13_checksum(new_base_12)
        
        # Return full 13-digit barcode
        return new_base_12 + checksum

    def _generate_barcodes(self):
        """Generate barcodes and export to Excel."""
        try:
            # Get quantity
            quantity_str = self.barcode_quantity_var.get().strip()
            if not quantity_str:
                messagebox.showwarning("אזהרה", "אנא הזן כמות ברקודים")
                return
            
            try:
                quantity = int(quantity_str)
            except ValueError:
                messagebox.showwarning("אזהרה", "אנא הזן מספר שלם תקין")
                return
            
            if quantity <= 0:
                messagebox.showwarning("אזהרה", "הכמות חייבת להיות גדולה מאפס")
                return
            
            if quantity > 10000:
                messagebox.showwarning("אזהרה", "הכמות המקסימלית היא 10,000 ברקודים")
                return
            
            # Get current last barcode
            current_barcode = self.last_barcode_var.get()
            
            # Generate list of barcodes
            barcodes = []
            next_barcode = current_barcode
            
            for i in range(quantity):
                next_barcode = self._generate_next_barcode(next_barcode)
                barcodes.append(next_barcode)
            
            # Export to Excel
            self._export_barcodes_to_excel(barcodes)
            
            # Update last barcode in system
            self._save_last_barcode(barcodes[-1])
            
            messagebox.showinfo(
                "הצלחה",
                f"נוצרו {quantity} ברקודים בהצלחה!\n\n"
                f"הברקוד האחרון: {barcodes[-1]}\n"
                f"הקובץ נשמר בתיקיית exports/barcodes"
            )
            
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בייצור ברקודים:\n{str(e)}")

    def _export_barcodes_to_excel(self, barcodes):
        """
        Export barcodes list to Excel file.
        
        Args:
            barcodes: List of barcode strings
        """
        try:
            # Try to use openpyxl if available, otherwise use xlsxwriter
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                use_openpyxl = True
            except ImportError:
                try:
                    import xlsxwriter
                    use_openpyxl = False
                except ImportError:
                    messagebox.showerror(
                        "שגיאה",
                        "לא נמצאה ספריית openpyxl או xlsxwriter.\n"
                        "אנא התקן אחת מהן עם: pip install openpyxl"
                    )
                    return
            
            # Create exports directory if it doesn't exist
            export_dir = os.path.join('exports', 'barcodes')
            os.makedirs(export_dir, exist_ok=True)
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(export_dir, f"ברקודים_{timestamp}.xlsx")
            
            if use_openpyxl:
                # Create workbook with openpyxl
                wb = Workbook()
                ws = wb.active
                ws.title = "ברקודים"
                
                # Header
                ws['A1'] = "מספר ברקוד"
                header_cell = ws['A1']
                header_cell.font = Font(bold=True, size=14)
                header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_cell.font = Font(bold=True, size=14, color="FFFFFF")
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Add barcodes
                for idx, barcode in enumerate(barcodes, start=2):
                    cell = ws[f'A{idx}']
                    cell.value = barcode
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=12)
                
                # Set column width
                ws.column_dimensions['A'].width = 25
                
                # Save workbook
                wb.save(filename)
                
            else:
                # Create workbook with xlsxwriter
                workbook = xlsxwriter.Workbook(filename)
                worksheet = workbook.add_worksheet("ברקודים")
                
                # Define formats
                header_format = workbook.add_format({
                    'bold': True,
                    'font_size': 14,
                    'bg_color': '#366092',
                    'font_color': 'white',
                    'align': 'center',
                    'valign': 'vcenter'
                })
                
                cell_format = workbook.add_format({
                    'font_size': 12,
                    'align': 'center',
                    'valign': 'vcenter'
                })
                
                # Set column width
                worksheet.set_column('A:A', 25)
                
                # Write header
                worksheet.write('A1', "מספר ברקוד", header_format)
                
                # Write barcodes
                for idx, barcode in enumerate(barcodes, start=1):
                    worksheet.write(f'A{idx+1}', barcode, cell_format)
                
                # Close workbook
                workbook.close()
            
            # Ask if user wants to open the file
            if messagebox.askyesno("פתיחת קובץ", f"הקובץ נשמר בהצלחה.\nלפתוח את הקובץ?"):
                os.startfile(filename)
                
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בייצוא לאקסל:\n{str(e)}")

    # ===== Cuts (גזרות) section =====
    def _build_cuts_section(self, parent):
        """Build the cuts catalog section: form + table + actions."""
        form = ttk.LabelFrame(parent, text="הוספת גזרה", padding=10)
        form.pack(fill='x', padx=10, pady=6)

        self.cut_file_name_var = tk.StringVar()
        self.cut_file_path_var = tk.StringVar()
        self.cut_product_name_var = tk.StringVar()
        self.cut_sizes_var = tk.StringVar()
        self.cut_category_var = tk.StringVar()
        self.cut_image_path_var = tk.StringVar()

        row = 0
        tk.Label(form, text="שם קובץ אופטיטקס:", font=('Arial', 10, 'bold')).grid(row=row, column=0, padx=4, pady=4, sticky='e')
        tk.Entry(form, textvariable=self.cut_file_name_var, width=28).grid(row=row, column=1, padx=4, pady=4, sticky='w')
        tk.Label(form, text="נתיב קובץ:", font=('Arial', 10, 'bold')).grid(row=row, column=2, padx=(12,4), pady=4, sticky='e')
        tk.Entry(form, textvariable=self.cut_file_path_var, width=36).grid(row=row, column=3, padx=4, pady=4, sticky='w')
        tk.Button(form, text="בחר קובץ...", command=self._browse_cut_file).grid(row=row, column=4, padx=4, pady=4)
        row += 1

        tk.Label(form, text="שם הפריט:", font=('Arial', 10, 'bold')).grid(row=row, column=0, padx=4, pady=4, sticky='e')
        model_names = [r.get('name', '') for r in getattr(self.data_processor, 'product_model_names', []) if r.get('name')]
        self.cut_product_combo = ttk.Combobox(form, textvariable=self.cut_product_name_var, values=sorted(set(model_names)), width=26)
        self.cut_product_combo.grid(row=row, column=1, padx=4, pady=4, sticky='w')
        tk.Label(form, text="מידות:", font=('Arial', 10, 'bold')).grid(row=row, column=2, padx=(12,4), pady=4, sticky='e')
        tk.Entry(form, textvariable=self.cut_sizes_var, width=36).grid(row=row, column=3, columnspan=2, padx=4, pady=4, sticky='w')
        row += 1

        tk.Label(form, text="קטגוריה:", font=('Arial', 10, 'bold')).grid(row=row, column=0, padx=4, pady=4, sticky='e')
        self.cut_category_combo = ttk.Combobox(form, textvariable=self.cut_category_var, width=26)
        self.cut_category_combo.grid(row=row, column=1, padx=4, pady=4, sticky='w')
        tk.Label(form, text="תמונה:", font=('Arial', 10, 'bold')).grid(row=row, column=2, padx=(12,4), pady=4, sticky='e')
        tk.Entry(form, textvariable=self.cut_image_path_var, width=32).grid(row=row, column=3, padx=4, pady=4, sticky='w')
        tk.Button(form, text="בחר תמונה...", command=self._browse_cut_image).grid(row=row, column=4, padx=4, pady=4)
        row += 1

        tk.Button(form, text="➕ הוסף גזרה", command=self._add_cut, bg='#27ae60', fg='white').grid(row=row, column=1, padx=8, pady=6)

        tree_frame = ttk.LabelFrame(parent, text="טבלת גזרות", padding=6)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=6)
        cols = ('file_name', 'file_path', 'product_name', 'sizes', 'category', 'image_path')
        self.cuts_tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=14)
        headers = {'file_name': 'שם קובץ', 'file_path': 'נתיב קובץ', 'product_name': 'שם פריט', 'sizes': 'מידות', 'category': 'קטגוריה', 'image_path': 'תמונה'}
        widths = {'file_name': 180, 'file_path': 280, 'product_name': 120, 'sizes': 120, 'category': 120, 'image_path': 200}
        for c in cols:
            self.cuts_tree.heading(c, text=headers[c])
            self.cuts_tree.column(c, width=widths[c], anchor='center')
        vs = ttk.Scrollbar(tree_frame, orient='vertical', command=self.cuts_tree.yview)
        self.cuts_tree.configure(yscroll=vs.set)
        self.cuts_tree.pack(side='left', fill='both', expand=True)
        vs.pack(side='right', fill='y')
        self.cuts_tree.bind('<Double-1>', self._on_cuts_tree_double_click)

        actions = tk.Frame(parent, bg='#f7f9fa')
        actions.pack(fill='x', padx=10, pady=(0, 8))
        tk.Button(actions, text="🗑 מחק נבחר", command=self._delete_cut, bg='#e67e22', fg='white').pack(side='left', padx=5)
        tk.Button(actions, text="💾 שמור", command=self._save_cuts_catalog, bg='#3498db', fg='white').pack(side='left', padx=5)

        self._refresh_cuts_tree()
        self._update_cuts_category_combo()

    def _update_cuts_category_combo(self):
        """Update category combobox values from existing cuts."""
        cuts = getattr(self.data_processor, 'cuts_catalog', [])
        cats = sorted(set(c.get('category', '') for c in cuts if c.get('category')))
        if hasattr(self, 'cut_category_combo'):
            self.cut_category_combo['values'] = cats

    def _load_cuts_catalog(self):
        """Load cuts from data processor (already in memory)."""
        getattr(self.data_processor, 'cuts_catalog', [])

    def _save_cuts_catalog(self):
        """Save cuts to JSON file."""
        try:
            dp = self.data_processor
            dp._save_json_list(dp.cuts_catalog_file, dp.cuts_catalog)
            messagebox.showinfo("שמירה", "גזרות נשמרו בהצלחה")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בשמירת גזרות: {e}")

    def _refresh_cuts_tree(self):
        """Refresh the cuts tree from data."""
        if not hasattr(self, 'cuts_tree'):
            return
        for iid in self.cuts_tree.get_children():
            self.cuts_tree.delete(iid)
        cuts = getattr(self.data_processor, 'cuts_catalog', [])
        for c in cuts:
            self.cuts_tree.insert('', 'end', values=(
                c.get('file_name', ''),
                c.get('file_path', ''),
                c.get('product_name', ''),
                c.get('sizes', ''),
                c.get('category', ''),
                c.get('image_path', '')
            ))
        self._update_cuts_category_combo()

    def _browse_cut_file(self):
        """Browse for Optitex .pds file; set file_path and file_name."""
        path = filedialog.askopenfilename(
            title="בחר קובץ אופטיטקס",
            filetypes=[("Optitex PDS", "*.pds"), ("All files", "*.*")]
        )
        if path:
            self.cut_file_path_var.set(path)
            self.cut_file_name_var.set(os.path.basename(path))

    def _browse_cut_image(self):
        """Browse for image file."""
        path = filedialog.askopenfilename(
            title="בחר תמונה",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.gif *.bmp"), ("All files", "*.*")]
        )
        if path:
            self.cut_image_path_var.set(path)

    def _add_cut(self):
        """Add a new cut record."""
        file_name = (self.cut_file_name_var.get() or '').strip()
        file_path = (self.cut_file_path_var.get() or '').strip()
        product_name = (self.cut_product_name_var.get() or '').strip()
        sizes = (self.cut_sizes_var.get() or '').strip()
        category = (self.cut_category_var.get() or '').strip()
        image_path = (self.cut_image_path_var.get() or '').strip()

        if not file_name:
            messagebox.showwarning("חסר", "אנא הזן שם קובץ אופטיטקס")
            return

        dp = self.data_processor
        cuts = getattr(dp, 'cuts_catalog', [])
        new_id = max([c.get('id', 0) for c in cuts], default=0) + 1
        record = {
            'id': new_id,
            'file_name': file_name,
            'file_path': file_path,
            'image_path': image_path,
            'product_name': product_name,
            'sizes': sizes,
            'category': category,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        dp.cuts_catalog = cuts + [record]
        self._refresh_cuts_tree()
        self._save_cuts_catalog()

        self.cut_file_name_var.set('')
        self.cut_file_path_var.set('')
        self.cut_product_name_var.set('')
        self.cut_sizes_var.set('')
        self.cut_category_var.set('')
        self.cut_image_path_var.set('')

    def _delete_cut(self):
        """Delete selected cut."""
        sel = self.cuts_tree.selection()
        if not sel:
            messagebox.showwarning("אזהרה", "בחר שורה למחיקה")
            return
        if not messagebox.askyesno("אישור", "למחוק את הגזרה הנבחרת?"):
            return
        vals = self.cuts_tree.item(sel[0], 'values')
        if not vals:
            return
        file_name, file_path = vals[0], vals[1]
        dp = self.data_processor
        dp.cuts_catalog = [c for c in dp.cuts_catalog if c.get('file_name') != file_name or c.get('file_path') != file_path]
        self._refresh_cuts_tree()
        self._save_cuts_catalog()

    def _on_cuts_tree_double_click(self, event):
        """On double-click: show image popup if image path set."""
        sel = self.cuts_tree.selection()
        if not sel:
            return
        vals = self.cuts_tree.item(sel[0], 'values')
        if len(vals) < 6:
            return
        image_path = (vals[5] or '').strip()
        if not image_path or not os.path.exists(image_path):
            messagebox.showinfo("תמונה", "אין תמונה או שהנתיב לא קיים")
            return
        try:
            from PIL import Image, ImageTk
            win = tk.Toplevel(self.root)
            win.title("תמונת גזרה")
            im = Image.open(image_path)
            im.thumbnail((600, 600))
            photo = ImageTk.PhotoImage(im)
            lbl = tk.Label(win, image=photo)
            lbl.image = photo
            lbl.pack(padx=10, pady=10)
        except Exception as e:
            messagebox.showerror("שגיאה", f"לא ניתן להציג תמונה: {e}")
