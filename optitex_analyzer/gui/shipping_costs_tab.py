"""Shipping costs and fabrics tab for managing shipping expenses and fabric costs."""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import json
import os
import subprocess
from .shipping_companies_tab import ShippingCompaniesTabMixin

class ShippingCostsTabMixin(ShippingCompaniesTabMixin):
    """Mixin for shipping costs and fabrics management tab."""
    
    def _create_shipping_costs_tab(self):
        """Create the shipping costs and fabrics tab."""
        tab = tk.Frame(self.notebook, bg='#f7f9fa')
        self.notebook.add(tab, text="עלויות משלוחים ובדים")
        
        # Create inner notebook for sub-tabs
        inner_nb = ttk.Notebook(tab)
        inner_nb.pack(fill='both', expand=True, padx=4, pady=4)
        
        # Shipping costs sub-tab
        shipping_costs_page = tk.Frame(inner_nb, bg='#f7f9fa')
        inner_nb.add(shipping_costs_page, text="עלויות משלוחים")
        self._build_shipping_costs_content(shipping_costs_page)
        
        # Shipping companies sub-tab
        shipping_companies_page = tk.Frame(inner_nb, bg='#f7f9fa')
        inner_nb.add(shipping_companies_page, text="חברות עמילות/שילוח")
        self._build_shipping_companies_content(shipping_companies_page)
    
    def _build_shipping_costs_content(self, container):
        """Build the shipping costs and fabrics content."""
        # Title
        title_label = tk.Label(
            container, 
            text="ניהול עלויות משלוחים ובדים", 
            font=('Arial', 16, 'bold'), 
            bg='#f7f9fa', 
            fg='#2c3e50'
        )
        title_label.pack(pady=(10, 20))
        
        # Input form frame
        input_frame = ttk.LabelFrame(container, text="הוספת משלוח חדש", padding=20)
        input_frame.pack(fill='x', padx=20, pady=10)
        
        # Row 1 - Name and Date
        tk.Label(input_frame, text="שם:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.shipping_name_var = tk.StringVar()
        self.shipping_name_combo = ttk.Combobox(input_frame, textvariable=self.shipping_name_var, width=18, font=('Arial', 10), state='readonly')
        self.shipping_name_combo.grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(input_frame, text="תאריך משלוח:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.shipping_date_var = tk.StringVar(value=datetime.now().strftime('%d/%m/%Y'))
        tk.Entry(input_frame, textvariable=self.shipping_date_var, width=15, font=('Arial', 10)).grid(row=0, column=3, padx=5, pady=5)
        
        # Row 2 - Cub and Total Weight
        tk.Label(input_frame, text="Cub:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.cub_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.cub_var, width=20, font=('Arial', 10)).grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(input_frame, text="משקל כולל:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky='w', padx=5, pady=5)
        self.total_weight_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.total_weight_var, width=15, font=('Arial', 10)).grid(row=1, column=3, padx=5, pady=5)
        
        # Row 3 - Quantity and Product Price in USD
        tk.Label(input_frame, text="כמות גלילים:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        self.rolls_quantity_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.rolls_quantity_var, width=15, font=('Arial', 10)).grid(row=2, column=1, padx=5, pady=5)
        
        tk.Label(input_frame, text="מחיר סחורה בדולר:", font=('Arial', 10, 'bold')).grid(row=2, column=2, sticky='w', padx=5, pady=5)
        self.product_price_usd_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.product_price_usd_var, width=15, font=('Arial', 10)).grid(row=2, column=3, padx=5, pady=5)
        
        # Row 3 - USD Exchange Rate
        tk.Label(input_frame, text="שער הדולר:", font=('Arial', 10, 'bold')).grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.usd_rate_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.usd_rate_var, width=15, font=('Arial', 10)).grid(row=3, column=1, padx=5, pady=5)
        
        # Row 4 - Shipping costs
        tk.Label(input_frame, text="עלות משלוח סופית (ללא מע״מ):", font=('Arial', 10, 'bold')).grid(row=4, column=0, sticky='w', padx=5, pady=5)
        self.final_shipping_cost_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.final_shipping_cost_var, width=20, font=('Arial', 10)).grid(row=4, column=1, padx=5, pady=5)
        
        tk.Label(input_frame, text="משלוח פנימי:", font=('Arial', 10, 'bold')).grid(row=4, column=2, sticky='w', padx=5, pady=5)
        self.domestic_shipping_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.domestic_shipping_var, width=15, font=('Arial', 10)).grid(row=4, column=3, padx=5, pady=5)
        
        
        # Row 5 - Packing List Upload
        tk.Label(input_frame, text="PACKING LIST:", font=('Arial', 10, 'bold')).grid(row=5, column=0, sticky='w', padx=5, pady=5)
        self.packing_list_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.packing_list_var, width=30, font=('Arial', 10), state='readonly').grid(row=5, column=1, padx=5, pady=5)
        tk.Button(input_frame, text="📁 בחר קובץ", command=self._select_packing_list_file, bg='#3498db', fg='white', font=('Arial', 9)).grid(row=5, column=2, padx=5, pady=5)
        tk.Button(input_frame, text="🗑 נקה", command=self._clear_packing_list, bg='#e74c3c', fg='white', font=('Arial', 9)).grid(row=5, column=3, padx=5, pady=5)
        
        # Row 6 - Payment Request Upload
        tk.Label(input_frame, text="דרישת תשלום:", font=('Arial', 10, 'bold')).grid(row=6, column=0, sticky='w', padx=5, pady=5)
        self.payment_request_var = tk.StringVar()
        tk.Entry(input_frame, textvariable=self.payment_request_var, width=30, font=('Arial', 10), state='readonly').grid(row=6, column=1, padx=5, pady=5)
        tk.Button(input_frame, text="📁 בחר קובץ", command=self._select_payment_request_file, bg='#3498db', fg='white', font=('Arial', 9)).grid(row=6, column=2, padx=5, pady=5)
        tk.Button(input_frame, text="🗑 נקה", command=self._clear_payment_request, bg='#e74c3c', fg='white', font=('Arial', 9)).grid(row=6, column=3, padx=5, pady=5)
        
        # Buttons
        buttons_frame = tk.Frame(input_frame)
        buttons_frame.grid(row=7, column=0, columnspan=4, pady=10)
        
        tk.Button(
            buttons_frame,
            text="הוסף משלוח",
            command=self._add_shipping_record,
            bg='#27ae60',
            fg='white',
            font=('Arial', 10, 'bold'),
            width=15
        ).pack(side='left', padx=5)
        
        tk.Button(
            buttons_frame,
            text="נקה שדות",
            command=self._clear_shipping_inputs,
            bg='#e74c3c',
            fg='white',
            font=('Arial', 10, 'bold'),
            width=15
        ).pack(side='left', padx=5)
        
        tk.Button(
            buttons_frame,
            text="ייצא לאקסל",
            command=self._export_to_excel,
            bg='#3498db',
            fg='white',
            font=('Arial', 10, 'bold'),
            width=15
        ).pack(side='left', padx=5)
        
        # Data table frame
        table_frame = ttk.LabelFrame(container, text="רשימת משלוחים", padding=10)
        table_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Add title above table
        title_label = tk.Label(
            table_frame, 
            text="מחירים ללא המע״מ של ישראל", 
            font=('Arial', 12, 'bold'), 
            bg='#f7f9fa', 
            fg='#2c3e50'
        )
        title_label.pack(pady=(0, 10))
        
        # Add sorting button
        sort_frame = tk.Frame(table_frame)
        sort_frame.pack(fill='x', pady=(0, 10))
        tk.Button(
            sort_frame,
            text="📅 מיין לפי תאריך (חדש למעלה)",
            command=self._sort_shipping_table_by_date,
            bg='#3498db',
            fg='white',
            font=('Arial', 9, 'bold')
        ).pack(side='right')
        
        # Treeview for data display
        columns = ('name', 'date', 'cub', 'total_weight', 'rolls_quantity', 'product_price_usd', 'usd_rate',
                  'final_shipping_cost', 'domestic_shipping', 'final_cost_incl_domestic', 
                  'total_price_per_kg', 'total_price_per_cubic', 'fabric_shipping_cost_percent')
        
        headers = {
            'name': 'שם',
            'date': 'תאריך משלוח',
            'cub': 'Cub',
            'total_weight': 'משקל כולל',
            'rolls_quantity': 'כמות גלילים',
            'product_price_usd': 'מחיר סחורה בדולר',
            'usd_rate': 'שער הדולר',
            'final_shipping_cost': 'עלות משלוח סופית',
            'domestic_shipping': 'משלוח פנימי',
            'final_cost_incl_domestic': 'עלות סופית כולל משלוח פנימי',
            'total_price_per_kg': 'עלות משלוח לק״ג',
            'total_price_per_cubic': 'מחיר כולל למטר מעוקב',
            'fabric_shipping_cost_percent': 'אחוז עלות משלוח בד'
        }
        
        self.shipping_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        # Configure columns
        for col in columns:
            self.shipping_tree.heading(col, text=headers[col])
            self.shipping_tree.column(col, width=100, anchor='center')
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.shipping_tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient='horizontal', command=self.shipping_tree.xview)
        self.shipping_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack treeview and scrollbars
        self.shipping_tree.pack(side='left', fill='both', expand=True)
        v_scrollbar.pack(side='right', fill='y')
        h_scrollbar.pack(side='bottom', fill='x')
        
        # Bind double-click to open files
        self.shipping_tree.bind('<Double-1>', self._on_row_double_click)
        
        # Load shipping companies for combobox
        self._load_shipping_companies_for_combobox()
        
        # Load existing data
        self._load_shipping_data()
        
        # Load CSV data if exists
        self._load_csv_data()
        
        # Sort table by date after loading all data
        self._sort_shipping_table_by_date()
    
    def _add_shipping_record(self):
        """Add a new shipping record."""
        try:
            # Get input values
            name = self.shipping_name_var.get().strip()
            date = self.shipping_date_var.get().strip()
            cub = self.cub_var.get().strip()
            total_weight = self.total_weight_var.get().strip()
            rolls_quantity = self.rolls_quantity_var.get().strip()
            product_price_usd = self.product_price_usd_var.get().strip()
            usd_rate = self.usd_rate_var.get().strip()
            final_shipping_cost = self.final_shipping_cost_var.get().strip()
            domestic_shipping = self.domestic_shipping_var.get().strip()
            packing_list_file = self.packing_list_var.get().strip()
            payment_request_file = self.payment_request_var.get().strip()
            
            # Validate required fields
            if not name or not date:
                messagebox.showerror("שגיאה", "שם ותאריך הם שדות חובה")
                return
            
            # Convert to numbers for calculations
            cub_val = float(cub) if cub else 0
            total_weight_val = float(total_weight) if total_weight else 0
            rolls_quantity_val = int(rolls_quantity) if rolls_quantity else 0
            product_price_usd_val = float(product_price_usd) if product_price_usd else 0
            usd_rate_val = float(usd_rate) if usd_rate else 0
            final_shipping_cost_val = float(final_shipping_cost) if final_shipping_cost else 0
            domestic_shipping_val = float(domestic_shipping) if domestic_shipping else 0
            
            # Calculate derived values
            final_cost_incl_domestic = final_shipping_cost_val + domestic_shipping_val
            
            if total_weight_val > 0:
                total_price_per_kg = final_cost_incl_domestic / total_weight_val
            else:
                total_price_per_kg = 0
            
            if cub_val > 0:
                total_price_per_cubic = final_cost_incl_domestic / cub_val
            else:
                total_price_per_cubic = 0
            
            # Calculate fabric shipping cost percentage (simplified without fabric cost per kg)
            if total_price_per_kg > 0:
                fabric_shipping_cost_percent = 100  # Default to 100% since we don't have fabric cost
            else:
                fabric_shipping_cost_percent = 0
            
            # Create record
            record = {
                'name': name,
                'date': date,
                'cub': cub_val,
                'total_weight': total_weight_val,
                'rolls_quantity': rolls_quantity_val,
                'product_price_usd': product_price_usd_val,
                'usd_rate': usd_rate_val,
                'final_shipping_cost': final_shipping_cost_val,
                'domestic_shipping': domestic_shipping_val,
                'final_cost_incl_domestic': final_cost_incl_domestic,
                'total_price_per_kg': total_price_per_kg,
                'total_price_per_cubic': total_price_per_cubic,
                'fabric_shipping_cost_percent': fabric_shipping_cost_percent,
                'packing_list': packing_list_file,
                'payment_request': payment_request_file
            }
            
            # Add to treeview
            values = (
                name, date, f"{cub_val:.2f}", f"{total_weight_val:.1f}",
                str(rolls_quantity_val), f"{product_price_usd_val:.2f}", f"{usd_rate_val:.2f}", f"{final_shipping_cost_val:.0f}", f"{domestic_shipping_val:.0f}",
                f"{final_cost_incl_domestic:.0f}", f"{total_price_per_kg:.2f}", f"{total_price_per_cubic:.0f}",
                f"{fabric_shipping_cost_percent:.0f}%"
            )
            
            self.shipping_tree.insert('', 'end', values=values)
            
            # Save the new record directly to file
            self._save_new_record_to_file(record)
            
            # Sort table by date
            self._sort_shipping_table_by_date()
            
            # Clear inputs
            self._clear_shipping_inputs()
            
            messagebox.showinfo("הצלחה", "המשלוח נוסף בהצלחה")
            
        except ValueError as e:
            messagebox.showerror("שגיאה", f"אנא הזן מספרים תקינים: {str(e)}")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בהוספת המשלוח: {str(e)}")
    
    def _clear_shipping_inputs(self):
        """Clear all input fields."""
        # Don't clear the shipping name combobox, just reset to first option
        if hasattr(self, 'shipping_name_combo') and self.shipping_name_combo['values']:
            self.shipping_name_combo.set(self.shipping_name_combo['values'][0])
        else:
            self.shipping_name_var.set("")
        self.shipping_date_var.set(datetime.now().strftime('%d/%m/%Y'))
        self.cub_var.set("")
        self.total_weight_var.set("")
        self.rolls_quantity_var.set("")
        self.product_price_usd_var.set("")
        self.usd_rate_var.set("")
        self.final_shipping_cost_var.set("")
        self.domestic_shipping_var.set("")
        self.packing_list_var.set("")
        self.payment_request_var.set("")
    

    def _load_shipping_companies_for_combobox(self):
        """Load shipping companies data for the combobox."""
        try:
            data_file = "shipping_companies.json"
            if os.path.exists(data_file):
                with open(data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # Extract company names
                company_names = [record.get('company_name', '') for record in data if record.get('company_name')]
                
                # Update combobox values
                self.shipping_name_combo['values'] = company_names
                
                # Set default value if there are companies
                if company_names:
                    self.shipping_name_combo.set(company_names[0])
        except Exception as e:
            print(f"Error loading shipping companies for combobox: {e}")
            # Set empty values if loading fails
            self.shipping_name_combo['values'] = []
    
    def refresh_shipping_companies_combobox(self):
        """Refresh the shipping companies combobox with updated data."""
        self._load_shipping_companies_for_combobox()
    
    def _sort_shipping_table_by_date(self):
        """Sort the shipping table by date (newest first)."""
        try:
            # Get all items from the tree
            items = []
            for item in self.shipping_tree.get_children():
                values = self.shipping_tree.item(item)['values']
                if values and len(values) > 1:  # Make sure we have at least name and date
                    date_str = values[1]  # Date is in column 1
                    items.append((item, values, date_str))
            
            # Sort by date (newest first)
            def parse_date(date_str):
                try:
                    # Handle different date formats
                    if '/' in date_str:
                        # Format: DD/MM/YYYY
                        day, month, year = date_str.split('/')
                        return datetime(int(year), int(month), int(day))
                    elif '-' in date_str:
                        # Format: YYYY-MM-DD
                        year, month, day = date_str.split('-')
                        return datetime(int(year), int(month), int(day))
                    else:
                        # If we can't parse, put it at the end
                        return datetime.min
                except:
                    return datetime.min
            
            items.sort(key=lambda x: parse_date(x[2]), reverse=True)
            
            # Clear the tree and re-insert sorted items
            for item in self.shipping_tree.get_children():
                self.shipping_tree.delete(item)
            
            for item, values, date_str in items:
                self.shipping_tree.insert('', 'end', values=values)
                
        except Exception as e:
            print(f"Error sorting shipping table: {e}")
    
    def _on_payment_request_double_click(self, event):
        """Handle double-click on payment request column to open file."""
        # Get the item that was clicked
        item = self.shipping_tree.selection()[0] if self.shipping_tree.selection() else None
        if not item:
            return
        
        # Get the values of the selected row
        values = self.shipping_tree.item(item, 'values')
        if len(values) < 13:  # Need at least 13 columns
            return
        
        # Get payment request from the record data (not from display)
        payment_request_file = ""
        try:
            # Get the record ID or name to find the full record
            record_name = values[0] if values else ""
            if record_name:
                # Load the full record from file to get the hidden fields
                data_file = "shipping_costs.json"
                if os.path.exists(data_file):
                    with open(data_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    # Find the matching record
                    for record in data:
                        if record.get('name') == record_name:
                            payment_request_file = record.get('payment_request', '')
                            break
        except Exception:
            pass
        if not payment_request_file or payment_request_file.strip() == "":
            messagebox.showinfo("מידע", "אין קובץ דרישת תשלום לשורה זו")
            return
        
        # Try to open the file
        try:
            
            file_path = os.path.join("payment_requests", payment_request_file)
            if os.path.exists(file_path):
                # Try to open with default application
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS and Linux
                    subprocess.call(['open', file_path])
                else:
                    subprocess.call(['xdg-open', file_path])
            else:
                messagebox.showerror("שגיאה", f"קובץ לא נמצא: {file_path}")
                
        except Exception as e:
            messagebox.showerror("שגיאה", f"לא ניתן לפתוח את הקובץ: {str(e)}")
    
    def _save_shipping_data(self):
        """Save shipping data to JSON file."""
        try:
            # Load existing data to preserve packing_list and payment_request
            existing_data = []
            try:
                if os.path.exists("shipping_costs.json"):
                    with open("shipping_costs.json", 'r', encoding='utf-8') as f:
                        existing_data = json.load(f)
            except Exception:
                pass
            
            # Get all data from treeview
            data = []
            for item in self.shipping_tree.get_children():
                values = self.shipping_tree.item(item)['values']
                if values and len(values) >= 13:
                    # Convert values back to proper format
                    record = {
                        'name': values[0],
                        'date': values[1],
                        'cub': float(values[2]) if values[2] else 0,
                        'total_weight': float(values[3]) if values[3] else 0,
                        'rolls_quantity': int(values[4]) if values[4] else 0,
                        'product_price_usd': float(values[5]) if values[5] else 0,
                        'usd_rate': float(values[6]) if values[6] else 0,
                        'final_shipping_cost': float(values[7]) if values[7] else 0,
                        'domestic_shipping': float(values[8]) if values[8] else 0,
                        'final_cost_incl_domestic': float(values[9]) if values[9] else 0,
                        'total_price_per_kg': float(values[10]) if values[10] else 0,
                        'total_price_per_cubic': float(values[11]) if values[11] else 0,
                        'fabric_shipping_cost_percent': float(values[12].replace('%', '')) if values[12] and values[12] != '' else 0,
                        'packing_list': '',
                        'payment_request': ''
                    }
                    
                    # Find matching record in existing data to preserve packing_list and payment_request
                    for existing_record in existing_data:
                        if (existing_record.get('name') == record['name'] and 
                            existing_record.get('date') == record['date']):
                            record['packing_list'] = existing_record.get('packing_list', '')
                            record['payment_request'] = existing_record.get('payment_request', '')
                            break
                    
                    data.append(record)
            
            # Save to file
            with open("shipping_costs.json", 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"Error saving shipping data: {e}")
            messagebox.showerror("שגיאה", f"שגיאה בשמירת הנתונים: {str(e)}")
    
    def _load_shipping_data(self):
        """Load shipping data from JSON file."""
        try:
            data_file = "shipping_costs.json"
            if os.path.exists(data_file):
                with open(data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # Clear existing data
                for item in self.shipping_tree.get_children():
                    self.shipping_tree.delete(item)
                
                # Load data into treeview
                for record in data:
                    values = (
                        record.get('name', ''),
                        record.get('date', ''),
                        f"{record.get('cub', 0):.2f}",
                        f"{record.get('total_weight', 0):.1f}",
                        str(record.get('rolls_quantity', 0)),
                        f"{record.get('product_price_usd', 0):.2f}",
                        f"{record.get('usd_rate', 0):.2f}",
                        f"{record.get('final_shipping_cost', 0):.0f}",
                        f"{record.get('domestic_shipping', 0):.0f}",
                        f"{record.get('final_cost_incl_domestic', 0):.0f}",
                        f"{record.get('total_price_per_kg', 0):.2f}",
                        f"{record.get('total_price_per_cubic', 0):.0f}",
                        f"{record.get('fabric_shipping_cost_percent', 0):.0f}%"
                    )
                    self.shipping_tree.insert('', 'end', values=values)
                    
        except Exception as e:
            print(f"Error loading shipping data: {e}")
            messagebox.showerror("שגיאה", f"שגיאה בטעינת הנתונים: {str(e)}")
    
    def _select_packing_list_file(self):
        """Select packing list file."""
        file_path = filedialog.askopenfilename(
            title="בחר קובץ Packing List",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            # Copy file to packing_lists directory
            try:
                import shutil
                filename = os.path.basename(file_path)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                name, ext = os.path.splitext(filename)
                new_filename = f"{name}_{timestamp}{ext}"
                
                # Create directory if it doesn't exist
                os.makedirs("packing_lists", exist_ok=True)
                
                # Copy file
                dest_path = os.path.join("packing_lists", new_filename)
                shutil.copy2(file_path, dest_path)
                
                # Update the variable
                self.packing_list_var.set(new_filename)
                
            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה בהעתקת הקובץ: {str(e)}")
    
    def _clear_packing_list(self):
        """Clear packing list selection."""
        self.packing_list_var.set("")
    
    def _select_payment_request_file(self):
        """Select payment request file."""
        file_path = filedialog.askopenfilename(
            title="בחר קובץ דרישת תשלום",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_path:
            # Copy file to payment_requests directory
            try:
                import shutil
                filename = os.path.basename(file_path)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                name, ext = os.path.splitext(filename)
                new_filename = f"{name}_{timestamp}{ext}"
                
                # Create directory if it doesn't exist
                os.makedirs("payment_requests", exist_ok=True)
                
                # Copy file
                dest_path = os.path.join("payment_requests", new_filename)
                shutil.copy2(file_path, dest_path)
                
                # Update the variable
                self.payment_request_var.set(new_filename)
                
            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה בהעתקת הקובץ: {str(e)}")
    
    def _clear_payment_request(self):
        """Clear payment request selection."""
        self.payment_request_var.set("")
    
    def _on_row_double_click(self, event):
        """Handle double-click on table row to show file options."""
        item = self.shipping_tree.selection()[0] if self.shipping_tree.selection() else None
        if not item:
            return
        
        values = self.shipping_tree.item(item, 'values')
        if len(values) < 13:
            return
        
        # Get the record name and date to find the full record
        record_name = values[0]
        record_date = values[1]
        
        try:
            # Load the full record from file
            packing_list = ""
            payment_request = ""
            
            if os.path.exists("shipping_costs.json"):
                with open("shipping_costs.json", 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # Find the matching record
                for record in data:
                    if (record.get('name') == record_name and 
                        record.get('date') == record_date):
                        packing_list = record.get('packing_list', '')
                        payment_request = record.get('payment_request', '')
                        break
            
            # Show file options dialog
            self._show_file_options_dialog(record_name, record_date, packing_list, payment_request)
                        
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בטעינת הנתונים: {str(e)}")
    
    def _show_file_options_dialog(self, record_name, record_date, packing_list, payment_request):
        """Show dialog with file options."""
        # Create dialog window
        dialog = tk.Toplevel()
        dialog.title(f"קבצים עבור {record_name} - {record_date}")
        dialog.geometry("400x300")
        dialog.resizable(False, False)
        
        # Center the dialog
        dialog.transient(self.root if hasattr(self, 'root') else None)
        dialog.grab_set()
        
        # Main frame
        main_frame = tk.Frame(dialog, bg='#f7f9fa')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(
            main_frame,
            text=f"קבצים עבור {record_name}",
            font=('Arial', 14, 'bold'),
            bg='#f7f9fa',
            fg='#2c3e50'
        )
        title_label.pack(pady=(0, 20))
        
        # Packing list section
        packing_frame = tk.Frame(main_frame, bg='#f7f9fa')
        packing_frame.pack(fill='x', pady=10)
        
        tk.Label(
            packing_frame,
            text="Packing List:",
            font=('Arial', 12, 'bold'),
            bg='#f7f9fa',
            fg='#2c3e50'
        ).pack(anchor='w')
        
        if packing_list and os.path.exists(os.path.join("packing_lists", packing_list)):
            tk.Label(
                packing_frame,
                text=f"📄 {packing_list}",
                font=('Arial', 10),
                bg='#f7f9fa',
                fg='#27ae60'
            ).pack(anchor='w', pady=(5, 0))
            
            tk.Button(
                packing_frame,
                text="פתח Packing List",
                command=lambda: self._open_file(os.path.join("packing_lists", packing_list)),
                bg='#3498db',
                fg='white',
                font=('Arial', 10, 'bold'),
                width=20
            ).pack(anchor='w', pady=(5, 0))
        else:
            tk.Label(
                packing_frame,
                text="אין קובץ Packing List",
                font=('Arial', 10),
                bg='#f7f9fa',
                fg='#e74c3c'
            ).pack(anchor='w', pady=(5, 0))
        
        # Payment request section
        payment_frame = tk.Frame(main_frame, bg='#f7f9fa')
        payment_frame.pack(fill='x', pady=10)
        
        tk.Label(
            payment_frame,
            text="דרישת תשלום:",
            font=('Arial', 12, 'bold'),
            bg='#f7f9fa',
            fg='#2c3e50'
        ).pack(anchor='w')
        
        if payment_request and os.path.exists(os.path.join("payment_requests", payment_request)):
            tk.Label(
                payment_frame,
                text=f"📄 {payment_request}",
                font=('Arial', 10),
                bg='#f7f9fa',
                fg='#27ae60'
            ).pack(anchor='w', pady=(5, 0))
            
            tk.Button(
                payment_frame,
                text="פתח דרישת תשלום",
                command=lambda: self._open_file(os.path.join("payment_requests", payment_request)),
                bg='#e74c3c',
                fg='white',
                font=('Arial', 10, 'bold'),
                width=20
            ).pack(anchor='w', pady=(5, 0))
        else:
            tk.Label(
                payment_frame,
                text="אין קובץ דרישת תשלום",
                font=('Arial', 10),
                bg='#f7f9fa',
                fg='#e74c3c'
            ).pack(anchor='w', pady=(5, 0))
        
        # Close button
        close_button = tk.Button(
            main_frame,
            text="סגור",
            command=dialog.destroy,
            bg='#95a5a6',
            fg='white',
            font=('Arial', 10, 'bold'),
            width=15
        )
        close_button.pack(pady=(20, 0))
    
    def _open_file(self, file_path):
        """Open a file with default application."""
        try:
            if os.path.exists(file_path):
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS and Linux
                    subprocess.call(['open', file_path])
                else:
                    subprocess.call(['xdg-open', file_path])
            else:
                messagebox.showerror("שגיאה", f"קובץ לא נמצא: {file_path}")
        except Exception as e:
            messagebox.showerror("שגיאה", f"לא ניתן לפתוח את הקובץ: {str(e)}")
    
    def _save_new_record_to_file(self, new_record):
        """Save a new record directly to the JSON file."""
        try:
            # Load existing data
            data = []
            if os.path.exists("shipping_costs.json"):
                with open("shipping_costs.json", 'r', encoding='utf-8') as f:
                    data = json.load(f)
            
            # Add the new record
            data.append(new_record)
            
            # Save back to file
            with open("shipping_costs.json", 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            print(f"Error saving new record: {e}")
            messagebox.showerror("שגיאה", f"שגיאה בשמירת הרשומה החדשה: {str(e)}")
    
    def _load_csv_data(self):
        """Load data from CSV file if exists."""
        # This function can be implemented if needed
        pass
    
    def _export_to_excel(self):
        """Export shipping data to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "עלויות משלוחים"
            
            # Headers
            headers = [
                'שם', 'תאריך משלוח', 'Cub', 'משקל כולל', 'כמות גלילים',
                'מחיר סחורה בדולר', 'שער הדולר', 'עלות משלוח סופית',
                'משלוח פנימי', 'עלות סופית כולל משלוח פנימי',
                'עלות משלוח לק״ג', 'מחיר כולל למטר מעוקב', 'אחוז עלות משלוח בד'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Write data
            for row, item in enumerate(self.shipping_tree.get_children(), 2):
                values = self.shipping_tree.item(item)['values']
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"shipping_costs_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("exports", filename)
            os.makedirs("exports", exist_ok=True)
            wb.save(filepath)
            
            messagebox.showinfo("הצלחה", f"הנתונים יוצאו בהצלחה לקובץ: {filename}")
            
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בייצוא לאקסל: {str(e)}")
