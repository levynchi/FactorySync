import tkinter as tk
from tkinter import ttk, messagebox
import os
from datetime import datetime

class DeliveryNoteMethodsMixin:
    def _refresh_driver_names_for_delivery(self):
        """טעינת שמות המובילים לקומבובוקס תעודת משלוח."""
        try:
            import json, os
            path = os.path.join(os.getcwd(), 'drivers.json')
            names = []
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        for d in data:
                            name = (d.get('name') or '').strip()
                            if name:
                                names.append(name)
            names = sorted({n for n in names})
            # Update driver comboboxes across tabs if present
            if hasattr(self, 'pkg_driver_combo'):
                self.pkg_driver_combo['values'] = names
            if hasattr(self, 'fs_pkg_driver_combo'):
                self.fs_pkg_driver_combo['values'] = names
        except Exception:
            pass

        # Internal state lists
        self._delivery_lines = []
        self._packages = []
        self._accessories = []

    # ---- Helpers (products / variants) ----
    def _refresh_delivery_products_allowed(self, initial: bool = False):
        try:
            catalog = getattr(self.data_processor, 'products_catalog', []) or []
            names = sorted({ (rec.get('name') or '').strip() for rec in catalog if rec.get('name') })
            self._delivery_products_allowed = names
            self._delivery_products_allowed_full = list(names)
            if hasattr(self, 'dn_product_combo'):
                try:
                    cur = self.dn_product_var.get()
                    self.dn_product_combo['values'] = self._delivery_products_allowed_full
                    if cur and cur not in self._delivery_products_allowed_full:
                        self.dn_product_var.set('')
                except Exception: pass
        except Exception:
            self._delivery_products_allowed = []
            self._delivery_products_allowed_full = []

    def _update_delivery_size_options(self):
        product = (self.dn_product_var.get() or '').strip(); sizes = []
        if product:
            try:
                catalog = getattr(self.data_processor, 'products_catalog', []) or []
                for rec in catalog:
                    if (rec.get('name') or '').strip() == product:
                        sz = (rec.get('size') or '').strip()
                        if sz: sizes.append(sz)
                import re
                def _size_key(s):
                    m = re.match(r"(\d+)", s); base = int(m.group(1)) if m else 0
                    return (base, s)
                sizes = sorted({s for s in sizes}, key=_size_key)
            except Exception: sizes = []
        if hasattr(self, 'dn_size_combo'):
            try:
                self.dn_size_combo['values'] = sizes
                if sizes:
                    self.dn_size_combo.state(['!disabled','readonly'])
                    if self.dn_size_var.get() not in sizes: self.dn_size_var.set('')
                else:
                    self.dn_size_var.set(''); self.dn_size_combo.set('')
                    try: self.dn_size_combo.state(['disabled'])
                    except Exception: pass
            except Exception: pass

    def _update_delivery_fabric_type_options(self):
        product = (self.dn_product_var.get() or '').strip(); fabric_types = []
        if product:
            try:
                catalog = getattr(self.data_processor, 'products_catalog', []) or []
                for rec in catalog:
                    if (rec.get('name') or '').strip() == product:
                        ft = (rec.get('fabric_type') or '').strip()
                        if ft: fabric_types.append(ft)
                fabric_types = sorted({f for f in fabric_types})
            except Exception: fabric_types = []
        if hasattr(self, 'dn_fabric_type_combo'):
            try:
                self.dn_fabric_type_combo['values'] = fabric_types
                if fabric_types:
                    self.dn_fabric_type_combo.state(['!disabled','readonly'])
                    if self.dn_fabric_type_var.get() not in fabric_types: self.dn_fabric_type_var.set('')
                else:
                    self.dn_fabric_type_var.set(''); self.dn_fabric_type_combo.set('')
                    try: self.dn_fabric_type_combo.state(['disabled'])
                    except Exception: pass
            except Exception: pass

    def _update_delivery_fabric_color_options(self):
        product = (self.dn_product_var.get() or '').strip(); chosen_ft = (self.dn_fabric_type_var.get() or '').strip(); colors = []
        if product:
            try:
                catalog = getattr(self.data_processor, 'products_catalog', []) or []
                for rec in catalog:
                    if (rec.get('name') or '').strip() == product:
                        if chosen_ft and (rec.get('fabric_type') or '').strip() != chosen_ft:
                            continue
                        col = (rec.get('fabric_color') or '').strip()
                        if col: colors.append(col)
                colors = sorted({c for c in colors})
            except Exception: colors = []
        if hasattr(self, 'dn_fabric_color_combo'):
            try:
                self.dn_fabric_color_combo['values'] = colors
                if colors:
                    self.dn_fabric_color_combo.state(['!disabled','readonly'])
                    if self.dn_fabric_color_var.get() not in colors:
                        self.dn_fabric_color_var.set('לבן' if 'לבן' in colors else '')
                else:
                    self.dn_fabric_color_var.set(''); self.dn_fabric_color_combo.set('')
                    try: self.dn_fabric_color_combo.state(['disabled'])
                    except Exception: pass
            except Exception: pass

    def _update_delivery_fabric_category_auto(self):
        """Auto-fill fabric category based on a matching product variant from products_catalog.
        Match keys: name, size, fabric_type, fabric_color, print_name (when available)."""
        try:
            product = (self.dn_product_var.get() or '').strip()
            if not product:
                self.dn_fabric_category_var.set(''); return
            size = (self.dn_size_var.get() or '').strip()
            ft = (self.dn_fabric_type_var.get() or '').strip()
            col = (self.dn_fabric_color_var.get() or '').strip()
            pn = (self.dn_print_name_var.get() or '').strip()
            catalog = getattr(self.data_processor, 'products_catalog', []) or []
            best = None
            best_score = -1
            for rec in catalog:
                if (rec.get('name') or '').strip() != product:
                    continue
                score = 0
                # optional exact matches add score
                if size and (rec.get('size') or '').strip() == size: score += 1
                if ft and (rec.get('fabric_type') or '').strip() == ft: score += 1
                if col and (rec.get('fabric_color') or '').strip() == col: score += 1
                if pn and (rec.get('print_name') or '').strip() == pn: score += 1
                if score > best_score:
                    best_score = score; best = rec
            if best and best.get('fabric_category'):
                self.dn_fabric_category_var.set((best.get('fabric_category') or '').strip())
            else:
                # fallback: blank
                self.dn_fabric_category_var.set('')
        except Exception:
            try:
                self.dn_fabric_category_var.set('')
            except Exception:
                pass

    # ---- Lines ops ----
    def _add_delivery_line(self):
        product = self.dn_product_var.get().strip(); size = self.dn_size_var.get().strip(); qty_raw = self.dn_qty_var.get().strip(); note = self.dn_note_var.get().strip()
        fabric_type = self.dn_fabric_type_var.get().strip(); fabric_color = self.dn_fabric_color_var.get().strip(); print_name = self.dn_print_name_var.get().strip() or 'חלק'
        fabric_category = getattr(self, 'dn_fabric_category_var', None)
        fabric_category = fabric_category.get().strip() if fabric_category else ''
        if not product or not qty_raw:
            messagebox.showerror("שגיאה", "חובה לבחור מוצר ולהזין כמות"); return
        if self._delivery_products_allowed and product not in self._delivery_products_allowed:
            messagebox.showerror("שגיאה", "יש לבחור מוצר מהרשימה בלבד"); return
        try:
            qty = int(qty_raw); assert qty > 0
        except Exception:
            messagebox.showerror("שגיאה", "כמות חייבת להיות מספר חיובי"); return
        category = getattr(self, 'dn_category_var', None)
        category = category.get().strip() if category else ''
        line = {'product': product, 'size': size, 'fabric_type': fabric_type, 'fabric_color': fabric_color, 'fabric_category': fabric_category, 'print_name': print_name, 'category': category, 'quantity': qty, 'note': note}
        self._delivery_lines.append(line)
        # columns: product,size,fabric_type,fabric_color,fabric_category,print_name,quantity,note
        # columns: product,size,fabric_type,fabric_color,fabric_category,print_name,category,quantity,note
        self.delivery_tree.insert('', 'end', values=(product,size,fabric_type,fabric_color,fabric_category,print_name,category,qty,note))
        self.dn_size_var.set(''); self.dn_qty_var.set(''); self.dn_note_var.set('')
        try: self.dn_product_combo['values'] = self._delivery_products_allowed_full
        except Exception: pass
        self._update_delivery_summary()

    def _delete_delivery_selected(self):
        sel = self.delivery_tree.selection()
        if not sel: return
        all_items = self.delivery_tree.get_children(); indices = [all_items.index(i) for i in sel]
        for item in sel: self.delivery_tree.delete(item)
        for idx in sorted(indices, reverse=True):
            if 0 <= idx < len(self._delivery_lines): del self._delivery_lines[idx]
        self._update_delivery_summary()

    def _clear_delivery_lines(self):
        self._delivery_lines = []
        for item in self.delivery_tree.get_children(): self.delivery_tree.delete(item)
        self._update_delivery_summary()

    def _update_delivery_summary(self):
        total_rows = len(self._delivery_lines); total_qty = sum(l.get('quantity',0) for l in self._delivery_lines)
        self.delivery_summary_var.set(f"{total_rows} שורות | {total_qty} כמות")

    def _save_delivery_note(self):
        supplier = self.dn_supplier_name_var.get().strip(); date_str = self.dn_date_var.get().strip()
        valid_names = set(self._get_supplier_names()) if hasattr(self,'_get_supplier_names') else set()
        if not supplier or (valid_names and supplier not in valid_names):
            messagebox.showerror("שגיאה", "יש לבחור שם ספק מהרשימה"); return
        if not self._delivery_lines and not self._accessories:
            messagebox.showerror("שגיאה", "אין שורות מוצרים או אביזרי תפירה לשמירה"); return
        # אם רשימת ה-packages בזיכרון ריקה אך יש שורות בטבלה, נשחזר מהרשימה המוצגת
        try:
            if (not self._packages) and hasattr(self, 'packages_tree'):
                items = self.packages_tree.get_children()
                rebuilt = []
                for it in items:
                    vals = self.packages_tree.item(it, 'values') or ()
                    if vals:
                        try:
                            rebuilt.append({'package_type': vals[0], 'quantity': int(vals[1]), 'driver': vals[2] if len(vals) > 2 else ''})
                        except Exception:
                            # אם הכמות אינה מספר – נשמור כמחרוזת כדי לא לאבד נתון
                            rebuilt.append({'package_type': vals[0], 'quantity': vals[1], 'driver': vals[2] if len(vals) > 2 else ''})
                if rebuilt:
                    self._packages = rebuilt
        except Exception:
            pass
        # Confirm saving without any packages defined (parallel to supplier intake tab)
        try:
            if not self._packages:
                proceed = messagebox.askyesno(
                    "אישור",
                    "לא הוזנו צורות אריזה (שקיות / שקים / בדים).\nהאם לשמור את התעודה ללא כמות שקים?"
                )
                if not proceed:
                    return
        except Exception:
            pass
        try:
            # שימוש בשיטה החדשה לאחר פיצול הקבצים (עם נפילה אחורה)
            if hasattr(self.data_processor, 'add_delivery_note'):
                new_id = self.data_processor.add_delivery_note(supplier, date_str, self._delivery_lines, packages=self._packages, accessories=self._accessories)
            else:
                new_id = self.data_processor.add_supplier_receipt(supplier, date_str, self._delivery_lines, packages=self._packages, accessories=self._accessories, receipt_kind='delivery_note')
            messagebox.showinfo("הצלחה", f"תעודה נשמרה (ID: {new_id})")
            self._clear_delivery_lines()
            self._clear_packages()
            self._clear_accessories()
            try:
                self._refresh_delivery_notes_list()
            except Exception:
                pass
            # עדכון טאב הובלות אם קיים
            try:
                self._notify_new_receipt_saved()
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    # ---- Sewing Accessories ops ----
    def _load_sewing_accessories_for_delivery(self):
        """Load sewing accessories for the delivery note combobox."""
        try:
            accessories = getattr(self.data_processor, 'sewing_accessories', []) or []
            names = []
            for acc in accessories:
                name = (acc.get('name') or '').strip()
                if name:
                    names.append(name)
            names = sorted({n for n in names})
            if hasattr(self, 'dn_accessory_combo'):
                self.dn_accessory_combo['values'] = names
                # Store all names for autocomplete
                self._all_sewing_accessories = names
        except Exception:
            if hasattr(self, 'dn_accessory_combo'):
                self.dn_accessory_combo['values'] = []
                self._all_sewing_accessories = []

    def _filter_sewing_accessories(self, event=None):
        """Filter sewing accessories based on user input for autocomplete."""
        try:
            if not hasattr(self, '_all_sewing_accessories'):
                return
            
            current_text = self.dn_accessory_var.get().strip()
            if not current_text:
                # Show all accessories if no text, sorted intelligently
                sorted_accessories = self._sort_accessories_intelligently(self._all_sewing_accessories)
                self.dn_accessory_combo['values'] = sorted_accessories
                return
            
            # Filter accessories that start with the current text
            filtered = [name for name in self._all_sewing_accessories 
                       if name.lower().startswith(current_text.lower())]
            
            # Sort filtered results intelligently
            sorted_filtered = self._sort_accessories_intelligently(filtered)
            
            # Update combobox values
            self.dn_accessory_combo['values'] = sorted_filtered
            
            # Show dropdown if there are matches - use a safer method
            if sorted_filtered and event and event.type == '2':  # KeyPress event
                # Only open dropdown on key press, not on other events
                try:
                    self.dn_accessory_combo.event_generate('<Button-1>')
                except:
                    pass
        except Exception:
            pass

    def _sort_accessories_intelligently(self, accessories):
        """Sort accessories intelligently - numbers first, then alphabetically."""
        def sort_key(name):
            # Extract numbers from the name for sorting
            import re
            numbers = re.findall(r'\d+', name)
            if numbers:
                # Use the first number found for sorting
                return (0, int(numbers[0]), name)  # 0 = numbers first
            else:
                # No numbers, sort alphabetically
                return (1, name)  # 1 = letters after numbers
        
        return sorted(accessories, key=sort_key)

    def _on_accessory_click(self, event=None):
        """Handle accessory combobox click to show filtered options."""
        try:
            # Refresh the filtered list when clicking
            self._filter_sewing_accessories()
            
            # Open dropdown after a short delay to ensure it's ready
            self.after(50, lambda: self.dn_accessory_combo.event_generate('<Button-1>'))
        except Exception:
            pass

    def _on_accessory_selected(self, event=None):
        """Handle accessory selection to update unit field."""
        try:
            selected = self.dn_accessory_var.get().strip()
            if not selected:
                self.dn_accessory_unit_var.set('')
                return
            
            accessories = getattr(self.data_processor, 'sewing_accessories', []) or []
            for acc in accessories:
                if (acc.get('name') or '').strip() == selected:
                    unit = (acc.get('unit') or '').strip()
                    self.dn_accessory_unit_var.set(unit)
                    return
            self.dn_accessory_unit_var.set('')
        except Exception:
            pass

    def _add_accessory_line(self):
        """Add a sewing accessory line to the delivery note."""
        accessory = self.dn_accessory_var.get().strip()
        qty_raw = self.dn_accessory_qty_var.get().strip()
        unit = self.dn_accessory_unit_var.get().strip()
        
        if not accessory or not qty_raw:
            messagebox.showerror("שגיאה", "חובה לבחור אביזר תפירה ולהזין כמות")
            return
        
        try:
            qty = int(qty_raw)
            assert qty > 0
        except Exception:
            messagebox.showerror("שגיאה", "כמות חייבת להיות מספר חיובי")
            return
        
        # Check if accessory already exists
        for acc in self._accessories:
            if acc.get('accessory') == accessory:
                messagebox.showwarning("אזהרה", "האביזר כבר קיים ברשימה")
                return
        
        record = {'accessory': accessory, 'unit': unit, 'quantity': qty}
        self._accessories.append(record)
        
        # Debug message
        print(f"DEBUG: Added accessory: {accessory}, Total accessories: {len(self._accessories)}")
        
        # Refresh the accessories tree to show all items
        self._refresh_accessories_tree()
        
        # Clear the input fields
        self.dn_accessory_qty_var.set('')
        
        # Update summary
        self._update_delivery_summary()

    def _refresh_accessories_tree(self):
        """Refresh the accessories tree to show all current accessories."""
        print(f"DEBUG: Refreshing accessories tree, count: {len(self._accessories)}")
        
        # Clear existing items
        for item in self.delivery_accessories_tree.get_children():
            self.delivery_accessories_tree.delete(item)
        
        # Insert all accessories from the list
        for i, acc in enumerate(self._accessories):
            values = (
                acc.get('accessory', ''),
                acc.get('unit', ''),
                acc.get('quantity', 0)
            )
            print(f"DEBUG: Inserting accessory {i+1}: {values}")
            self.delivery_accessories_tree.insert('', 'end', values=values)
        
        # Force update of the tree display
        self.delivery_accessories_tree.update_idletasks()
        print(f"DEBUG: Tree refresh complete, items in tree: {len(self.delivery_accessories_tree.get_children())}")

    def _delete_selected_accessory(self):
        """Delete selected accessory from the list."""
        sel = self.delivery_accessories_tree.selection()
        if not sel:
            return
        all_items = self.delivery_accessories_tree.get_children()
        indices = [all_items.index(i) for i in sel]
        for idx in sorted(indices, reverse=True):
            if 0 <= idx < len(self._accessories):
                del self._accessories[idx]
        
        # Refresh the tree to show updated list
        self._refresh_accessories_tree()
        self._update_delivery_summary()

    def _clear_accessories(self):
        """Clear all accessories from the list."""
        self._accessories = []
        # Refresh the tree to show empty list
        self._refresh_accessories_tree()
        self._update_delivery_summary()

    # ---- Packages ops ----
    def _add_package_line(self):
        pkg_type = (self.pkg_type_var.get() or '').strip()
        qty_raw = (self.pkg_qty_var.get() or '').strip()
        if not pkg_type or not qty_raw:
            messagebox.showerror("שגיאה", "חובה לבחור פריט הובלה ולהזין כמות")
            return
        try:
            qty = int(qty_raw); assert qty > 0
        except Exception:
            messagebox.showerror("שגיאה", "כמות חייבת להיות מספר חיובי")
            return
        driver = (getattr(self, 'pkg_driver_var', tk.StringVar()).get() or '').strip()
        record = {'package_type': pkg_type, 'quantity': qty, 'driver': driver}
        self._packages.append(record)
        self.packages_tree.insert('', 'end', values=(pkg_type, qty, driver))
        self.pkg_qty_var.set('')

    def _delete_selected_package(self):
        sel = self.packages_tree.selection()
        if not sel: return
        all_items = self.packages_tree.get_children(); indices = [all_items.index(i) for i in sel]
        for item in sel: self.packages_tree.delete(item)
        for idx in sorted(indices, reverse=True):
            if 0 <= idx < len(self._packages): del self._packages[idx]

    def _clear_packages(self):
        self._packages = []
        for item in self.packages_tree.get_children():
            self.packages_tree.delete(item)

    # ---- Delivery notes list ----
    def _refresh_delivery_notes_list(self):
        try:
            # Always refresh from disk to include external additions
            self.data_processor.refresh_supplier_receipts()
            notes = self.data_processor.get_delivery_notes()
        except Exception:
            notes = []
        # Clear
        if hasattr(self, 'delivery_notes_tree'):
            for iid in self.delivery_notes_tree.get_children():
                self.delivery_notes_tree.delete(iid)
            for rec in notes:
                pkg_summary = ', '.join(f"{p.get('package_type')}:{p.get('quantity')}" for p in rec.get('packages', [])[:4])
                if len(rec.get('packages', [])) > 4:
                    pkg_summary += ' ...'
                # add delete icon in the last column to match ('id','date','supplier','total','packages','delete')
                self.delivery_notes_tree.insert('', 'end', values=(rec.get('id'), rec.get('date'), rec.get('supplier'), rec.get('total_quantity'), pkg_summary, '🗑'))

    # ---- View single delivery note ----
    def _open_selected_delivery_note_view(self, event=None):
        try:
            if not hasattr(self, 'delivery_notes_tree'): return
            sel = self.delivery_notes_tree.selection()
            if not sel: return
            vals = self.delivery_notes_tree.item(sel[0], 'values')
            if not vals: return
            note_id = vals[0]
            # איתור הרשומה המלאה
            try:
                self.data_processor.refresh_supplier_receipts()
            except Exception:
                pass
            full_list = getattr(self.data_processor, 'delivery_notes', [])
            rec = None
            for r in full_list:
                if str(r.get('id')) == str(note_id):
                    rec = r; break
            if not rec:
                messagebox.showerror("שגיאה", "תעודה לא נמצאה")
                return
            win = tk.Toplevel(self.notebook)
            win.title(f"תעודת משלוח #{rec.get('id')}")
            win.geometry('760x520')
            win.transient(self.notebook.winfo_toplevel())
            header = tk.Frame(win, pady=6)
            header.pack(fill='x')
            def _lbl(text):
                return tk.Label(header, text=text, font=('Arial',10,'bold'))
            _lbl(f"ID: {rec.get('id')}").grid(row=0,column=0,padx=6,sticky='w')
            _lbl(f"תאריך: {rec.get('date')}").grid(row=0,column=1,padx=6,sticky='w')
            _lbl(f"ספק: {rec.get('supplier')}").grid(row=0,column=2,padx=6,sticky='w')
            # תיקון מחרוזת f לא נכונה שגרמה להצגת הטקסט {rec.get('total_quantity')} במקום הערך
            _lbl(f"סה\"כ כמות: {rec.get('total_quantity')}").grid(row=0,column=3,padx=6,sticky='w')
            # קונטיינר לפאנלים
            body = tk.PanedWindow(win, orient='vertical')
            body.pack(fill='both', expand=True, padx=8, pady=4)
            # שורות מוצרים
            lines_frame = tk.LabelFrame(body, text='שורות מוצרים')
            body.add(lines_frame, stretch='always')
            lines_cols = ('product','size','fabric_type','fabric_color','fabric_category','print_name','quantity','note')
            lines_tree = ttk.Treeview(lines_frame, columns=lines_cols, show='headings', height=8)
            headers_map = {'product':'מוצר','size':'מידה','fabric_type':'סוג בד','fabric_color':'צבע בד','fabric_category':'קטגורית בד','print_name':'פרינט','quantity':'כמות','note':'הערה'}
            widths_map = {'product':140,'size':70,'fabric_type':110,'fabric_color':110,'fabric_category':120,'print_name':110,'quantity':60,'note':160}
            for c in lines_cols:
                lines_tree.heading(c, text=headers_map[c])
                lines_tree.column(c, width=widths_map[c], anchor='center')
            for line in rec.get('lines', []) or []:
                lines_tree.insert('', 'end', values=(line.get('product'), line.get('size'), line.get('fabric_type'), line.get('fabric_color'), line.get('fabric_category',''), line.get('print_name'), line.get('quantity'), line.get('note')))
            lines_tree.pack(fill='both', expand=True, padx=4, pady=4)
            
            # אביזרי תפירה
            if rec.get('accessories'):
                acc_frame = tk.LabelFrame(body, text='אביזרי תפירה')
                body.add(acc_frame, stretch='always')
                acc_cols = ('accessory','unit','quantity')
                acc_tree = ttk.Treeview(acc_frame, columns=acc_cols, show='headings', height=6)
                acc_headers = {'accessory':'אביזר תפירה','unit':'יחידה','quantity':'כמות'}
                acc_widths = {'accessory':200,'unit':100,'quantity':80}
                for c in acc_cols:
                    acc_tree.heading(c, text=acc_headers[c])
                    acc_tree.column(c, width=acc_widths[c], anchor='center')
                for acc in rec.get('accessories', []) or []:
                    acc_tree.insert('', 'end', values=(acc.get('accessory'), acc.get('unit'), acc.get('quantity')))
                acc_tree.pack(fill='both', expand=True, padx=4, pady=4)
            
            # חבילות הובלה
            pkg_frame = tk.LabelFrame(body, text='פרטי הובלה / חבילות')
            body.add(pkg_frame, stretch='always')
            pkg_cols = ('package_type','quantity','driver')
            pkg_tree = ttk.Treeview(pkg_frame, columns=pkg_cols, show='headings', height=6)
            pkg_headers = {'package_type':'פריט הובלה','quantity':'כמות','driver':'מוביל'}
            pkg_widths = {'package_type':140,'quantity':70,'driver':120}
            for c in pkg_cols:
                pkg_tree.heading(c, text=pkg_headers[c])
                pkg_tree.column(c, width=pkg_widths[c], anchor='center')
            for p in rec.get('packages', []) or []:
                pkg_tree.insert('', 'end', values=(p.get('package_type'), p.get('quantity'), p.get('driver')))
            pkg_tree.pack(fill='both', expand=True, padx=4, pady=4)
            # Actions: Open in Excel + Close
            btns = tk.Frame(win)
            btns.pack(fill='x', pady=6, padx=4)

            def _export_dn_to_excel_and_open():
                try:
                    # Lazy import to avoid hard dependency during normal UI flow
                    from openpyxl import Workbook
                    from openpyxl.styles import Border, Side, Alignment, Font
                    try:
                        from openpyxl.worksheet.page import PageMargins  # type: ignore
                    except Exception:
                        PageMargins = None  # type: ignore
                except Exception as e:
                    try:
                        messagebox.showerror("שגיאה", f"נדרש openpyxl לצורך יצוא לאקסל:\n{e}")
                    except Exception:
                        pass
                    return
                try:
                    # Prepare workbook
                    wb = Workbook(); ws = wb.active; ws.title = "תעודת משלוח"
                    # Right-to-Left sheet for Hebrew
                    try:
                        ws.sheet_view.rightToLeft = True
                    except Exception:
                        pass
                    # Page setup: A4 portrait, fit to width (center on page for print)
                    try:
                        ws.page_setup.orientation = 'portrait'
                    except Exception:
                        pass
                    try:
                        # A4 paper size code is 9 in Excel
                        ws.page_setup.paperSize = 9
                    except Exception:
                        pass
                    try:
                        ws.page_setup.fitToWidth = 1
                        ws.page_setup.fitToHeight = 0
                    except Exception:
                        pass
                    try:
                        ws.print_options.horizontalCentered = True
                        ws.print_options.verticalCentered = False
                    except Exception:
                        pass
                    try:
                        if PageMargins:
                            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
                    except Exception:
                        pass
                    # Business header (from settings): Prefer Logo image; fallback to Business Name + VAT/Type line
                    try:
                        s = getattr(self, 'settings', None)
                        def _sget(k, default=""):
                            try:
                                return (s.get(k, default) if s else default) or ""
                            except Exception:
                                return ""
                        biz_name = _sget('business.name')
                        biz_type = _sget('business.type')
                        biz_vat  = _sget('business.vat_id')
                        logo_path = _sget('business.logo_path')
                    except Exception:
                        biz_name = biz_type = biz_vat = logo_path = ""
                    # Insert styled business header if available
                    try:
                        last_col = 4  # we use 4 columns in the table below
                        inserted_logo = False
                        # Try to place logo image if path configured
                        try:
                            from openpyxl.drawing.image import Image as XLImage  # type: ignore
                            if logo_path and os.path.exists(logo_path):
                                img = XLImage(logo_path)
                                # Set specific dimensions as requested: height 4.66 cm, width 15.73 cm
                                try:
                                    # Convert cm to pixels: 1 cm = 37.8 pixels (96 DPI / 2.54 cm/inch)
                                    img.height = 4.66 * 37.8  # 4.66 cm in pixels
                                    img.width = 15.73 * 37.8  # 15.73 cm in pixels
                                except Exception:
                                    pass
                                ws.add_image(img, 'A1')
                                # Set row height to accommodate logo (Excel row height is in points, 1 cm = 28.35 points)
                                try:
                                    ws.row_dimensions[1].height = 4.66 * 28.35  # Convert cm to Excel row height units
                                except Exception:
                                    ws.row_dimensions[1].height = 120
                                # Add spacing rows under the image
                                ws.append([]); ws.append([])
                                inserted_logo = True
                        except Exception:
                            inserted_logo = False
                        # If no logo, fallback to biz name line
                        if not inserted_logo and biz_name:
                            ws.append([biz_name])
                            r = ws.max_row
                            try:
                                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
                            except Exception:
                                pass
                            try:
                                ws.cell(row=r, column=1).font = Font(size=16, bold=True)
                                ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
                            except Exception:
                                pass
                        # Regardless, include type/VAT line under header if present
                        if (biz_type or biz_vat):
                            line = f"{(biz_type or '').strip()} {(biz_vat or '').strip()}".strip()
                            ws.append([line])
                            r = ws.max_row
                            try:
                                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
                            except Exception:
                                pass
                            try:
                                ws.cell(row=r, column=1).font = Font(size=16)
                                ws.cell(row=r, column=1).alignment = Alignment(horizontal='right')
                            except Exception:
                                pass
                        if inserted_logo or biz_name or biz_vat or biz_type:
                            ws.append([])  # spacing
                    except Exception:
                        pass

                    # Document info header (keys and values)
                    doc_info_start = ws.max_row + 1
                    ws.append(["מסמך", f"תעודת משלוח #{rec.get('id')}"])
                    # תאריך בפורמט D/M/Y
                    try:
                        raw_date = (rec.get('date') or '').strip()
                        def _parse_dt(s: str):
                            for fmt in ("%Y-%m-%d","%d/%m/%Y","%d-%m-%Y","%Y/%m/%d","%m/%d/%Y","%d.%m.%Y"):
                                try:
                                    # Return date object to avoid time component in Excel
                                    return datetime.strptime(s, fmt).date()
                                except Exception:
                                    pass
                            return None
                        d_obj = _parse_dt(raw_date)
                        if d_obj is not None:
                            ws.append(["תאריך", d_obj])
                            try:
                                # Apply Excel date number format on the just-added cell
                                r = ws.max_row
                                dc = ws.cell(row=r, column=2)
                                dc.number_format = 'd/m/yyyy'
                                dc.alignment = Alignment(horizontal='right')
                            except Exception:
                                pass
                        else:
                            # fallback quick reformat (YYYY-MM-DD -> D/M/Y)
                            parts = raw_date.replace('.', '-').replace('/', '-').split('-')
                            if len(parts) == 3 and len(parts[0]) == 4:
                                try:
                                    y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                                    ws.append(["תאריך", f"{d}/{m}/{y}"])
                                except Exception:
                                    ws.append(["תאריך", raw_date])
                            else:
                                ws.append(["תאריך", raw_date])
                    except Exception:
                        ws.append(["תאריך", rec.get('date')])
                    ws.append(["ספק", rec.get('supplier')])
                    doc_info_end = ws.max_row
                    try:
                        # Style: right-align, bold keys, slightly larger font
                        for row in ws.iter_rows(min_row=doc_info_start, max_row=doc_info_end, min_col=1, max_col=2):
                            try:
                                row[0].font = Font(bold=True, size=16)
                                row[0].alignment = Alignment(horizontal='right')
                            except Exception:
                                pass
                            try:
                                row[1].font = Font(size=16)
                                row[1].alignment = Alignment(horizontal='right')
                            except Exception:
                                pass
                        # אם הוזנה ישירות כתאריך - מספר פורמט ד/מ/שנה (אימות נוסף)
                        try:
                            dcell = ws.cell(row=doc_info_start+1, column=2)
                            dcell.number_format = 'd/m/yyyy'
                        except Exception:
                            pass
                    except Exception:
                        pass
                    ws.append([])
                    # Lines table (A4-ready): Model | Size | Description (fabric type + print) | Quantity
                    ws.append(["שורות מוצרים"])
                    try:
                        # Merge and style the section title
                        r = ws.max_row
                        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                        tcell = ws.cell(row=r, column=1)
                        tcell.font = Font(bold=True, size=16)
                        tcell.alignment = Alignment(horizontal='right')
                    except Exception:
                        pass
                    header_row = ["שם הדגם", "מידה", "תיאור", "כמות"]
                    ws.append(header_row)
                    
                    # Add gray header row (row 10) with bold headers
                    try:
                        from openpyxl.styles import PatternFill
                        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        header_row_num = ws.max_row
                        for col in range(1, 5):  # Columns A to D
                            cell = ws.cell(row=header_row_num, column=col)
                            cell.fill = gray_fill
                            cell.font = Font(bold=True, size=16)
                            cell.alignment = Alignment(horizontal='center')
                    except Exception:
                        pass
                    
                    start_data_row = ws.max_row + 1
                    for line in (rec.get('lines', []) or []):
                        model = line.get('product')
                        size = (line.get('size') or '')
                        fabric_type = (line.get('fabric_type') or '')
                        color = (line.get('fabric_color') or '')
                        print_name = (line.get('print_name') or '')
                        # Compose description: fabric_type + color + print_name (skip empties)
                        parts = [p for p in [fabric_type, color, print_name] if p]
                        desc = " | ".join(parts)
                        qty = line.get('quantity')
                        ws.append([model, size, desc, qty])
                    # שורת סה"כ כמות מתחת לטבלה (תגית בעמודה C והמספר בעמודה D)
                    try:
                        try:
                            total_qty = sum(int((l or {}).get('quantity', 0) or 0) for l in (rec.get('lines', []) or []))
                        except Exception:
                            total_qty = (rec.get('total_quantity') or 0)
                        # A, B empty; C label; D numeric total
                        ws.append([None, None, "סה\"כ כמות", total_qty])
                        r_tot = ws.max_row
                        try:
                            ws.cell(row=r_tot, column=3).font = Font(bold=True, size=16)
                            ws.cell(row=r_tot, column=3).alignment = Alignment(horizontal='right')
                            ws.cell(row=r_tot, column=4).alignment = Alignment(horizontal='center')
                        except Exception:
                            pass
                    except Exception:
                        pass
                    
                    # אביזרי תפירה
                    if rec.get('accessories'):
                        ws.append([])  # רווח
                        ws.append(["אביזרי תפירה"])
                        try:
                            # Merge and style the section title
                            r = ws.max_row
                            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                            tcell = ws.cell(row=r, column=1)
                            tcell.font = Font(bold=True, size=16)
                            tcell.alignment = Alignment(horizontal='right')
                        except Exception:
                            pass
                        acc_header_row = ["אביזר תפירה", "יחידה", "כמות", ""]
                        ws.append(acc_header_row)
                        acc_start_row = ws.max_row + 1
                        for acc in (rec.get('accessories', []) or []):
                            accessory = acc.get('accessory', '')
                            unit = acc.get('unit', '')
                            qty = acc.get('quantity', '')
                            ws.append([accessory, unit, qty, ""])
                        acc_end_row = ws.max_row
                        # Style accessories table
                        try:
                            thin = Side(style='thin', color='000000')
                            border = Border(left=thin, right=thin, top=thin, bottom=thin)
                            # Bold header row
                            for cell in ws[ws.cell(row=acc_start_row-1, column=1).row]:
                                if cell.row == acc_start_row-1:
                                    try:
                                        cell.font = Font(bold=True, size=16)
                                    except Exception:
                                        pass
                            for row in ws.iter_rows(min_row=acc_start_row-1, max_row=acc_end_row, min_col=1, max_col=4):
                                for cell in row:
                                    try:
                                        cell.border = border
                                    except Exception:
                                        pass
                                    try:
                                        if cell.row >= acc_start_row:
                                            cell.font = Font(size=16)
                                    except Exception:
                                        pass
                                # Align Hebrew text to the right
                                try:
                                    row[0].alignment = Alignment(horizontal='right')
                                    row[1].alignment = Alignment(horizontal='right')
                                    row[2].alignment = Alignment(horizontal='center')
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    
                    end_data_row = ws.max_row
                    # Style: black grid on lines area + fonts and alignments
                    try:
                        thin = Side(style='thin', color='000000')
                        border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        # Bold header row
                        for cell in ws[ws.cell(row=start_data_row-1, column=1).row]:
                            if cell.row == start_data_row-1:
                                try:
                                    cell.font = Font(bold=True, size=16)
                                except Exception:
                                    pass
                        for row in ws.iter_rows(min_row=start_data_row-1, max_row=end_data_row, min_col=1, max_col=4):
                            for cell in row:
                                try:
                                    cell.border = border
                                except Exception:
                                    pass
                                try:
                                    # Data font size a bit larger for readability
                                    if cell.row >= start_data_row:
                                        cell.font = Font(size=16)
                                except Exception:
                                    pass
                            # Align quantity center/right
                            try:
                                row[3].alignment = Alignment(horizontal='center')
                            except Exception:
                                pass
                            # Align Hebrew text to the right for Model/Size/Description
                            try:
                                row[0].alignment = Alignment(horizontal='right')
                                row[1].alignment = Alignment(horizontal='right')
                                row[2].alignment = Alignment(horizontal='right')
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # Autosize columns A, B, C, D with optimal width for all content (like Excel AutoFit)
                    try:
                        for col_letter in ['A', 'B', 'C', 'D']:
                            max_width = 0
                            col_num = ord(col_letter) - ord('A') + 1
                            
                            # Check all cells in this column
                            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                                for cell in row:
                                    try:
                                        val = str(cell.value) if cell.value is not None else ""
                                        if val.strip():  # Only count non-empty cells
                                            # Calculate character width more accurately
                                            char_width = 0
                                            for char in val:
                                                # Hebrew characters are wider
                                                if '\u0590' <= char <= '\u05FF':
                                                    char_width += 1.2
                                                # Numbers and English are narrower
                                                elif char.isdigit() or char.isalpha():
                                                    char_width += 0.6
                                                # Special characters
                                                else:
                                                    char_width += 0.8
                                            
                                            # Add some padding for cell borders
                                            char_width += 2
                                            max_width = max(max_width, char_width)
                                    except Exception:
                                        pass
                            
                            # Set width with reasonable limits (Excel-like behavior)
                            if max_width > 0:
                                # Minimum 8, maximum 50 (Excel standard)
                                ws.column_dimensions[col_letter].width = min(max(8, max_width), 50)
                            else:
                                ws.column_dimensions[col_letter].width = 12  # Default width if no content
                    except Exception:
                        pass
                    # Save to exports folder
                    base_dir = os.path.join(os.getcwd(), 'exports', 'delivery_notes')
                    try:
                        os.makedirs(base_dir, exist_ok=True)
                    except Exception:
                        pass
                    safe_id = rec.get('id')
                    safe_date = (rec.get('date') or '').replace('/', '-').replace(':', '-')
                    fname = f"delivery_note_{safe_id}_{safe_date}.xlsx" if safe_id is not None else f"delivery_note_{safe_date}.xlsx"
                    out_path = os.path.join(base_dir, fname)
                    try:
                        wb.save(out_path)
                    except PermissionError as e:
                        # Fallback: add timestamp suffix if file is locked/open
                        try:
                            ts = datetime.now().strftime('%H%M%S')
                            alt_path = out_path.replace('.xlsx', f'_{ts}.xlsx')
                            wb.save(alt_path)
                            out_path = alt_path
                        except Exception:
                            try:
                                messagebox.showerror("שגיאה", f"שמירת קובץ נכשלה (קובץ פתוח?):\n{e}")
                            except Exception:
                                pass
                            return
                    except Exception as e:
                        try:
                            messagebox.showerror("שגיאה", f"שמירת קובץ נכשלה:\n{e}")
                        except Exception:
                            pass
                        return
                    # Open in Excel (Windows association)
                    try:
                        os.startfile(out_path)  # type: ignore[attr-defined]
                    except Exception as e:
                        try:
                            messagebox.showinfo("נשמר", f"הקובץ נשמר ב:\n{out_path}\n(לא הצלחתי לפתוח אוטומטית)")
                        except Exception:
                            pass

                except Exception as e:
                    try:
                        messagebox.showerror("שגיאה", str(e))
                    except Exception:
                        pass

            tk.Button(btns, text='🖨 פתח באקסל', command=_export_dn_to_excel_and_open, bg='#27ae60', fg='white').pack(side='right', padx=4)
            tk.Button(btns, text='סגור', command=win.destroy).pack(side='right')
        except Exception as e:
            try:
                messagebox.showerror("שגיאה", str(e))
            except Exception:
                pass

    # ===== Fabrics Send (שליחת בדים) =====
    def _fs_import_barcodes_from_file(self):
        """Import barcodes from a CSV or Excel file and add valid ones to the send list.
        Supported:
        - CSV: one barcode per line, or in a column named 'barcode' (case-insensitive).
        - Excel (.xlsx): first sheet, one column with barcodes (any header), or a column named 'barcode'.
        Duplicates (already in list) are skipped. Invalid/not-found barcodes are reported at the end.
        """
        try:
            from tkinter import filedialog
            path = filedialog.askopenfilename(title='בחר קובץ ברקודים', filetypes=[('Excel/CSV','*.xlsx;*.csv'),('Excel','*.xlsx'),('CSV','*.csv'),('All files','*.*')])
        except Exception:
            path = ''
        if not path:
            return
        import os
        ext = os.path.splitext(path)[1].lower()
        barcodes = []
        # Read file
        try:
            if ext == '.csv':
                import csv
                with open(path, 'r', encoding='utf-8-sig') as f:
                    rdr = csv.reader(f)
                    rows = list(rdr)
                if not rows:
                    rows = []
                # Try header-based
                if rows and any('barcode' in (c or '').strip().lower() for c in rows[0]):
                    header = [ (c or '').strip().lower() for c in rows[0] ]
                    try:
                        idx = header.index('barcode')
                        for r in rows[1:]:
                            if idx < len(r):
                                bc = (r[idx] or '').strip()
                                if bc:
                                    barcodes.append(bc)
                    except Exception:
                        pass
                # Fallback: single column or first cell per row
                if not barcodes:
                    for r in rows:
                        if not r: continue
                        bc = (r[0] or '').strip()
                        if bc:
                            barcodes.append(bc)
            elif ext == '.xlsx':
                try:
                    from openpyxl import load_workbook  # type: ignore
                except Exception as e:
                    try: messagebox.showerror('שגיאה', f"נדרש openpyxl לקריאת אקסל: {e}")
                    except Exception: pass
                    return
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                # Try to detect header row
                header = None
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    values = [ (str(c) if c is not None else '').strip() for c in row ]
                    if not any(values):
                        continue
                    if header is None and any('barcode' == v.lower() for v in values):
                        header = [ v.lower() for v in values ]
                        try:
                            idx = header.index('barcode')
                        except Exception:
                            idx = None
                        continue
                    if header is not None:
                        # header already detected, take value at 'barcode' column
                        try:
                            if idx is not None and idx < len(values):
                                bc = values[idx]
                                if bc:
                                    barcodes.append(bc)
                                continue
                        except Exception:
                            pass
                    # Fallback: take first non-empty cell
                    first = next((v for v in values if v), '')
                    if first:
                        barcodes.append(first)
            else:
                try: messagebox.showerror('שגיאה', 'סוג קובץ לא נתמך. נא לבחור CSV או XLSX')
                except Exception: pass
                return
        except Exception as e:
            try: messagebox.showerror('שגיאה', f'קריאת הקובץ נכשלה: {e}')
            except Exception: pass
            return
        # Normalize barcodes
        barcodes = [ str(b).strip() for b in barcodes if str(b).strip() ]
        if not barcodes:
            try: messagebox.showwarning('מידע', 'לא נמצאו ברקודים בקובץ')
            except Exception: pass
            return
        
        # Remove duplicates from the imported file itself
        original_count = len(barcodes)
        barcodes = list(dict.fromkeys(barcodes))  # Remove duplicates while preserving order
        duplicates_in_file = original_count - len(barcodes)
        
        # Build fast lookup from inventory
        try:
            records = getattr(self.data_processor, 'fabrics_inventory', []) or []
        except Exception:
            records = []
        index = { str(r.get('barcode')).strip(): r for r in records if str(r.get('barcode') or '').strip() }
        
        # Current list to prevent duplicates
        current = set()
        try:
            for iid in self.fs_tree.get_children():
                vals = self.fs_tree.item(iid, 'values') or []
                if vals:
                    current.add(str(vals[0]))
        except Exception:
            pass
        
        added = 0
        invalid = []
        skipped_existing = 0
        
        for bc in barcodes:
            if bc in current:
                skipped_existing += 1
                continue
            rec = index.get(bc)
            if not rec:
                invalid.append(bc)
                continue
            values = (
                rec.get('barcode',''), rec.get('fabric_type',''), rec.get('color_name',''), rec.get('color_no',''), rec.get('design_code',''),
                rec.get('width',''), f"{rec.get('net_kg',0):.2f}", f"{rec.get('meters',0):.2f}", f"{rec.get('price',0):.2f}", rec.get('location',''), rec.get('status','במלאי')
            )
            try:
                self.fs_tree.insert('', 'end', values=values)
                added += 1
            except Exception:
                pass
        self._fs_update_summary()
        # Report result
        try:
            msg = f"נוספו {added} ברקודים"
            
            # Report duplicates removed from file
            if duplicates_in_file > 0:
                msg += f"\nהוסרו {duplicates_in_file} כפילויות מהקובץ"
            
            # Report skipped existing barcodes
            if skipped_existing > 0:
                msg += f"\nדולגו {skipped_existing} ברקודים שכבר קיימים ברשימה"
            
            # Report invalid barcodes
            if invalid:
                # Limit preview to first 20 for usability
                preview = ', '.join(invalid[:20]) + (" ..." if len(invalid) > 20 else '')
                msg += f"\nלא נמצאו במלאי: {len(invalid)}\n{preview}"
            
            messagebox.showinfo('ייבוא ברקודים', msg)
        except Exception:
            pass
    def _fs_add_fabric_by_barcode(self):
        """Add fabric record by barcode into the shipment list (from fabrics inventory)."""
        try:
            barcode = (getattr(self, 'fs_barcode_var', None).get() or '').strip()
        except Exception:
            barcode = ''
        if not barcode:
            return
        # find in inventory
        try:
            records = getattr(self.data_processor, 'fabrics_inventory', []) or []
        except Exception:
            records = []
        rec = next((r for r in records if str(r.get('barcode') or '').strip() == barcode), None)
        if not rec:
            try: messagebox.showwarning('לא נמצא', 'ברקוד זה לא נמצא במלאי הבדים')
            except Exception: pass
            return
        # prevent duplicates in current table
        try:
            for iid in self.fs_tree.get_children():
                vals = self.fs_tree.item(iid, 'values') or []
                if vals and str(vals[0]) == barcode:
                    try: messagebox.showinfo('מידע', 'ברקוד זה כבר ברשימה')
                    except Exception: pass
                    return
        except Exception:
            pass
        values = (
            rec.get('barcode',''), rec.get('fabric_type',''), rec.get('color_name',''), rec.get('color_no',''), rec.get('design_code',''),
            rec.get('width',''), f"{rec.get('net_kg',0):.2f}", f"{rec.get('meters',0):.2f}", f"{rec.get('price',0):.2f}", rec.get('location',''), rec.get('status','במלאי')
        )
        try:
            self.fs_tree.insert('', 'end', values=values)
            self.fs_barcode_var.set('')
        except Exception:
            pass
        self._fs_update_summary()

    def _fs_add_manual_fabric(self):
        """Add fabric manually without barcode (for fabrics with missing labels)."""
        # Create manual fabric entry dialog
        win = tk.Toplevel(self.notebook)
        win.title("הוספת בד ידנית")
        win.geometry('500x400')
        win.transient(self.notebook.winfo_toplevel())
        win.grab_set()  # Make modal
        
        # Center the window
        win.update_idletasks()
        x = (win.winfo_screenwidth() // 2) - (500 // 2)
        y = (win.winfo_screenheight() // 2) - (400 // 2)
        win.geometry(f'500x400+{x}+{y}')
        
        # Main frame
        main_frame = tk.Frame(win, padx=20, pady=20)
        main_frame.pack(fill='both', expand=True)
        
        # Title
        tk.Label(main_frame, text="הוספת בד ידנית (ללא בר קוד)", font=('Arial', 14, 'bold')).pack(pady=(0, 20))
        
        # Form fields
        fields_frame = tk.Frame(main_frame)
        fields_frame.pack(fill='x', pady=10)
        
        # Fabric type
        tk.Label(fields_frame, text="סוג בד:", width=15, anchor='e').grid(row=0, column=0, sticky='e', padx=(0, 10), pady=5)
        fabric_type_var = tk.StringVar()
        fabric_type_combo = ttk.Combobox(fields_frame, textvariable=fabric_type_var, width=20, state='readonly')
        fabric_type_combo['values'] = ['INTERLOCK', 'RIB', 'COTTON', 'POLYESTER', 'OTHER']
        fabric_type_combo.grid(row=0, column=1, sticky='w', pady=5)
        
        # Color name
        tk.Label(fields_frame, text="צבע:", width=15, anchor='e').grid(row=1, column=0, sticky='e', padx=(0, 10), pady=5)
        color_name_var = tk.StringVar()
        color_name_combo = ttk.Combobox(fields_frame, textvariable=color_name_var, width=20)
        color_name_combo['values'] = ['לבן', 'שחור', 'אפור', 'כחול', 'אדום', 'ירוק', 'צהוב', 'ורוד', 'חום', 'אחר']
        color_name_combo.grid(row=1, column=1, sticky='w', pady=5)
        
        # Color number
        tk.Label(fields_frame, text="מס' צבע:", width=15, anchor='e').grid(row=2, column=0, sticky='e', padx=(0, 10), pady=5)
        color_no_var = tk.StringVar()
        tk.Entry(fields_frame, textvariable=color_no_var, width=22).grid(row=2, column=1, sticky='w', pady=5)
        
        # Design code
        tk.Label(fields_frame, text="Desen Kodu:", width=15, anchor='e').grid(row=3, column=0, sticky='e', padx=(0, 10), pady=5)
        design_code_var = tk.StringVar(value='חלק')
        tk.Entry(fields_frame, textvariable=design_code_var, width=22).grid(row=3, column=1, sticky='w', pady=5)
        
        # Width
        tk.Label(fields_frame, text="רוחב (ס\"מ):", width=15, anchor='e').grid(row=4, column=0, sticky='e', padx=(0, 10), pady=5)
        width_var = tk.StringVar()
        tk.Entry(fields_frame, textvariable=width_var, width=22).grid(row=4, column=1, sticky='w', pady=5)
        
        # Net weight
        tk.Label(fields_frame, text="ק\"ג נטו:", width=15, anchor='e').grid(row=5, column=0, sticky='e', padx=(0, 10), pady=5)
        net_kg_var = tk.StringVar()
        tk.Entry(fields_frame, textvariable=net_kg_var, width=22).grid(row=5, column=1, sticky='w', pady=5)
        
        # Meters
        tk.Label(fields_frame, text="מטרים:", width=15, anchor='e').grid(row=6, column=0, sticky='e', padx=(0, 10), pady=5)
        meters_var = tk.StringVar()
        tk.Entry(fields_frame, textvariable=meters_var, width=22).grid(row=6, column=1, sticky='w', pady=5)
        
        # Price
        tk.Label(fields_frame, text="מחיר:", width=15, anchor='e').grid(row=7, column=0, sticky='e', padx=(0, 10), pady=5)
        price_var = tk.StringVar()
        tk.Entry(fields_frame, textvariable=price_var, width=22).grid(row=7, column=1, sticky='w', pady=5)
        
        # Buttons
        buttons_frame = tk.Frame(main_frame)
        buttons_frame.pack(fill='x', pady=(20, 0))
        
        def add_fabric():
            # Validate required fields
            if not fabric_type_var.get().strip():
                messagebox.showerror("שגיאה", "חובה לבחור סוג בד")
                return
            if not color_name_var.get().strip():
                messagebox.showerror("שגיאה", "חובה להזין צבע")
                return
            if not net_kg_var.get().strip():
                messagebox.showerror("שגיאה", "חובה להזין ק\"ג נטו")
                return
            if not meters_var.get().strip():
                messagebox.showerror("שגיאה", "חובה להזין מטרים")
                return
            
            # Validate numeric fields
            try:
                net_kg = float(net_kg_var.get().replace(',', '.'))
                meters = float(meters_var.get().replace(',', '.'))
                price = float(price_var.get().replace(',', '.')) if price_var.get().strip() else 0.0
                width = int(width_var.get()) if width_var.get().strip() else 0
            except ValueError:
                messagebox.showerror("שגיאה", "ערכים מספריים לא תקינים")
                return
            
            # Save to unbarcoded fabrics table instead of regular inventory
            try:
                # Add to unbarcoded fabrics
                new_id = self.data_processor.add_unbarcoded_fabric(
                    fabric_type=fabric_type_var.get().strip(),
                    manufacturer='',  # Not used in manual entry
                    color=color_name_var.get().strip(),
                    shade=color_no_var.get().strip(),
                    notes=f"Desen: {design_code_var.get().strip()}, רוחב: {width}cm, ק\"ג: {net_kg:.2f}, מטרים: {meters:.2f}, מחיר: {price:.2f}"
                )
                
                # Add to shipment tree for display (with barcode "0")
                values = (
                    '0',  # barcode
                    fabric_type_var.get().strip(),
                    color_name_var.get().strip(),
                    color_no_var.get().strip(),
                    design_code_var.get().strip(),
                    str(width),
                    f"{net_kg:.2f}",
                    f"{meters:.2f}",
                    f"{price:.2f}",
                    'ידני',
                    'במלאי'
                )
                
                self.fs_tree.insert('', 'end', values=values)
                self._fs_update_summary()
                win.destroy()
                messagebox.showinfo("הצלחה", f"הבד נוסף בהצלחה (ID: {new_id}) ונשמר בטאב 'בדים בלי ברקוד'")
            except Exception as e:
                messagebox.showerror("שגיאה", f"שגיאה בהוספת הבד: {str(e)}")
        
        tk.Button(buttons_frame, text="הוסף בד", command=add_fabric, bg='#27ae60', fg='white', font=('Arial', 10, 'bold')).pack(side='right', padx=(10, 0))
        tk.Button(buttons_frame, text="ביטול", command=win.destroy, bg='#95a5a6', fg='white').pack(side='right')

    def _fs_remove_selected(self):
        try:
            sel = self.fs_tree.selection()
            for item in sel:
                self.fs_tree.delete(item)
        except Exception:
            pass
        self._fs_update_summary()

    def _fs_clear_all(self):
        try:
            for item in self.fs_tree.get_children():
                self.fs_tree.delete(item)
        except Exception:
            pass
        self._fs_update_summary()

    def _fs_update_summary(self):
        try:
            count = len(self.fs_tree.get_children())
        except Exception:
            count = 0
        try:
            self.fs_summary_var.set(f"{count} בדים")
        except Exception:
            pass

    # Packages helpers for fabrics send
    def _fs_add_package_line(self):
        ptype = (getattr(self, 'fs_pkg_type_var', None).get() or '').strip()
        qty_raw = (getattr(self, 'fs_pkg_qty_var', None).get() or '').strip()
        driver = (getattr(self, 'fs_pkg_driver_var', None).get() or '').strip()
        if not ptype or not qty_raw:
            try: messagebox.showerror('שגיאה', 'יש לבחור פריט הובלה ולהזין כמות')
            except Exception: pass
            return
        try:
            qty = int(qty_raw); assert qty > 0
        except Exception:
            try: messagebox.showerror('שגיאה', 'כמות חייבת להיות מספר חיובי')
            except Exception: pass
            return
        try:
            self._fs_packages.append({'package_type': ptype, 'quantity': qty, 'driver': driver})
        except Exception:
            self._fs_packages = [{'package_type': ptype, 'quantity': qty, 'driver': driver}]
        try:
            self.fs_packages_tree.insert('', 'end', values=(ptype, qty, driver))
            self.fs_pkg_qty_var.set('')
        except Exception:
            pass

    def _fs_delete_selected_package(self):
        try:
            sel = self.fs_packages_tree.selection()
            all_items = self.fs_packages_tree.get_children()
            indices = [all_items.index(i) for i in sel]
            for i in sel:
                self.fs_packages_tree.delete(i)
            for idx in sorted(indices, reverse=True):
                if 0 <= idx < len(self._fs_packages):
                    del self._fs_packages[idx]
        except Exception:
            pass

    def _fs_clear_packages(self):
        try:
            self._fs_packages = []
        except Exception:
            pass
        try:
            for i in self.fs_packages_tree.get_children():
                self.fs_packages_tree.delete(i)
        except Exception:
            pass

    def _fs_save_shipment(self):
        """Save fabrics shipment: set status to 'נשלח' for barcodes and persist a shipment doc."""
        # collect barcodes from table
        barcodes = []
        try:
            for iid in self.fs_tree.get_children():
                vals = self.fs_tree.item(iid, 'values') or []
                if vals:
                    barcodes.append(str(vals[0]))
        except Exception:
            pass
        if not barcodes:
            try: messagebox.showwarning('שגיאה', 'אין בדים לשמירה')
            except Exception: pass
            return
        # Require supplier selection
        try:
            supplier = (getattr(self, 'fs_supplier_var', None).get() or '').strip()
        except Exception:
            supplier = ''
        valid_suppliers = []
        try:
            if hasattr(self, '_get_supplier_names'):
                valid_suppliers = self._get_supplier_names() or []
        except Exception:
            valid_suppliers = []
        if not supplier or (valid_suppliers and supplier not in valid_suppliers):
            try: messagebox.showerror('שגיאה', 'יש לבחור ספק יעד מהרשימה לפני שמירה')
            except Exception: pass
            return
        # packages
        packages = []
        try:
            for iid in self.fs_packages_tree.get_children():
                vals = self.fs_packages_tree.item(iid, 'values') or []
                if vals:
                    try:
                        packages.append({'package_type': vals[0], 'quantity': int(vals[1]), 'driver': vals[2] if len(vals) > 2 else ''})
                    except Exception:
                        packages.append({'package_type': vals[0], 'quantity': vals[1], 'driver': vals[2] if len(vals) > 2 else ''})
        except Exception:
            packages = []
        # If no driver is provided in any package, warn and confirm proceed
        try:
            has_driver = any((p.get('driver') or '').strip() for p in packages)
            if not has_driver:
                proceed = messagebox.askyesno('אישור', 'לא הוזן שם מוביל. האם לשמור בלי מוביל?')
                if not proceed:
                    return
        except Exception:
            pass
        # status+location update for each barcode (bulk for efficiency)
        # Skip manual fabrics (barcode "0") as they are stored in unbarcoded table
        updated = 0
        manual_fabrics = 0
        try:
            # Separate real barcodes from manual fabrics
            real_barcodes = [bc for bc in barcodes if bc != '0']
            manual_fabrics = len([bc for bc in barcodes if bc == '0'])
            
            if real_barcodes:
                updates = [{'barcode': bc, 'status': 'נשלח', 'location': supplier} for bc in real_barcodes]
                if hasattr(self.data_processor, 'bulk_update_fabrics'):
                    updated = self.data_processor.bulk_update_fabrics(updates)
                else:
                    # fallback per item
                    for bc in real_barcodes:
                        try:
                            if hasattr(self.data_processor, 'update_fabric_status') and self.data_processor.update_fabric_status(bc, 'נשלח'):
                                # also try to set location field directly
                                try:
                                    inv = getattr(self.data_processor, 'fabrics_inventory', []) or []
                                    for r in inv:
                                        if str(r.get('barcode')) == str(bc):
                                            r['location'] = supplier
                                            break
                                    self.data_processor.save_fabrics_inventory()
                                except Exception:
                                    pass
                                updated += 1
                        except Exception:
                            pass
            
            # For manual fabrics, we don't need to update inventory since they're in unbarcoded table
            # They are already saved there when added manually
        except Exception:
            pass
        # persist shipment in data processor
        new_id = None
        try:
            if hasattr(self.data_processor, 'add_fabrics_shipment'):
                # We can compute simple summary from the first row if available
                ft = cn = cno = ''
                net_kg = meters = 0
                try:
                    first_vals = self.fs_tree.item(self.fs_tree.get_children()[0], 'values') or []
                    ft = first_vals[1] if len(first_vals) > 1 else ''
                    cn = first_vals[2] if len(first_vals) > 2 else ''
                    cno = first_vals[3] if len(first_vals) > 3 else ''
                except Exception:
                    pass
                # sum net_kg/meters
                try:
                    for iid in self.fs_tree.get_children():
                        vals = self.fs_tree.item(iid, 'values') or []
                        if len(vals) >= 9:
                            try:
                                net_kg += float(str(vals[6]).replace(',', '.'))
                            except Exception:
                                pass
                            try:
                                meters += float(str(vals[7]).replace(',', '.'))
                            except Exception:
                                pass
                except Exception:
                    pass
                # date: reuse delivery note date if present
                date_str = ''
                try:
                    date_str = (getattr(self, 'dn_date_var', None).get() or '').strip()
                except Exception:
                    date_str = ''
                new_id = self.data_processor.add_fabrics_shipment(barcodes, packages, date_str=date_str, fabric_type=ft, color_name=cn, color_no=cno, net_kg=net_kg, meters=meters, supplier=supplier)
        except Exception:
            new_id = None
        try:
            if new_id:
                msg = f"נשמרה שליחת בדים (ID: {new_id}) לספק '{supplier}'.\n"
                if updated > 0:
                    msg += f"עודכנו {updated} בדים לסטטוס 'נשלח' והמיקום עודכן ל-{supplier}.\n"
                if manual_fabrics > 0:
                    msg += f"כללו {manual_fabrics} בדים ידניים שנשמרו בטאב 'בדים בלי ברקוד'."
                messagebox.showinfo('הצלחה', msg)
            else:
                msg = f"עודכנו {updated} בדים לסטטוס 'נשלח' והמיקום עודכן ל-{supplier}."
                if manual_fabrics > 0:
                    msg += f"\nכללו {manual_fabrics} בדים ידניים שנשמרו בטאב 'בדים בלי ברקוד'."
                messagebox.showinfo('הצלחה', msg)
        except Exception:
            pass
        # clear UI and refresh inventory tab if available
        self._fs_clear_all(); self._fs_clear_packages()
        try:
            if hasattr(self, 'fs_supplier_var'):
                self.fs_supplier_var.set('')
                if hasattr(self, 'fs_supplier_summary_var'):
                    self.fs_supplier_summary_var.set('ספק: -')
        except Exception:
            pass
        try:
            if hasattr(self, '_refresh_fabrics_table'):
                self._refresh_fabrics_table()
        except Exception:
            pass
        try:
            self._fs_refresh_shipments_list()
        except Exception:
            pass

    # Saved shipments list helpers (browse/delete/open)
    def _fs_refresh_shipments_list(self):
        try:
            if hasattr(self.data_processor, 'refresh_fabrics_shipments'):
                self.data_processor.refresh_fabrics_shipments()
            records = getattr(self.data_processor, 'fabrics_shipments', []) or []
        except Exception:
            records = []
        tree = getattr(self, 'fabrics_shipments_tree', None)
        if not tree:
            return
        try:
            for iid in tree.get_children():
                tree.delete(iid)
        except Exception:
            pass
        for rec in records:
            tree.insert('', 'end', values=(
                rec.get('id',''), rec.get('date',''), rec.get('count_barcodes',0), '🗑'
            ))

    def _fs_on_click_shipments(self, event):
        tree = getattr(self, 'fabrics_shipments_tree', None)
        if not tree:
            return
        region = tree.identify('region', event.x, event.y)
        if region != 'cell':
            return
        col = tree.identify_column(event.x)
        # delete column is now #4 after simplifying columns
        if col != '#4':
            return
        row_id = tree.identify_row(event.y)
        if not row_id:
            return
        values = tree.item(row_id, 'values') or []
        if not values:
            return
        try:
            rec_id = int(values[0])
        except Exception:
            return
        try:
            if not messagebox.askyesno('אישור', f"למחוק שליחת בדים ID {rec_id}?"):
                return
        except Exception:
            pass
        try:
            if hasattr(self.data_processor, 'delete_fabrics_shipment') and self.data_processor.delete_fabrics_shipment(rec_id):
                tree.delete(row_id)
                try:
                    messagebox.showinfo('נמחק', f"שליחת בדים {rec_id} נמחקה")
                except Exception:
                    pass
        except Exception:
            pass

    def _open_selected_fabrics_shipment_view(self, event=None):
        tree = getattr(self, 'fabrics_shipments_tree', None)
        if not tree:
            return
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], 'values') or []
        if not vals:
            return
        try:
            ship_id = int(vals[0])
        except Exception:
            return
        # find full record
        try:
            rec = next((r for r in getattr(self.data_processor, 'fabrics_shipments', []) if int(r.get('id',-1)) == ship_id), None)
        except Exception:
            rec = None
        if not rec:
            try: messagebox.showerror('שגיאה', 'רשומה לא נמצאה')
            except Exception: pass
            return
        # Open a view window with meta + per-barcode details from inventory
        win = tk.Toplevel(self.notebook)
        win.title(f"שליחת בדים #{ship_id}")
        win.geometry('720x520')
        header = tk.Frame(win, pady=6)
        header.pack(fill='x')
        def _lbl(t):
            return tk.Label(header, text=t, font=('Arial',10,'bold'))
        _lbl(f"ID: {rec.get('id')}").grid(row=0,column=0,padx=6,sticky='w')
        _lbl(f"תאריך: {rec.get('date')}").grid(row=0,column=1,padx=6,sticky='w')
        _lbl(f"ספק: {rec.get('supplier','')}").grid(row=0,column=2,padx=6,sticky='w')
        # table of barcode details
        body = tk.Frame(win)
        body.pack(fill='both', expand=True, padx=8, pady=6)
        cols = ('barcode','fabric_type','color','net_kg','meters')
        tree2 = ttk.Treeview(body, columns=cols, show='headings', height=16)
        tree2.heading('barcode', text='ברקוד')
        tree2.heading('fabric_type', text='סוג בד')
        tree2.heading('color', text='צבע')
        tree2.heading('net_kg', text='ק"ג נטו')
        tree2.heading('meters', text='מטרים')
        tree2.column('barcode', width=160, anchor='center')
        tree2.column('fabric_type', width=140, anchor='center')
        tree2.column('color', width=120, anchor='center')
        tree2.column('net_kg', width=90, anchor='center')
        tree2.column('meters', width=90, anchor='center')
        vs = ttk.Scrollbar(body, orient='vertical', command=tree2.yview)
        tree2.configure(yscroll=vs.set)
        tree2.grid(row=0, column=0, sticky='nsew'); vs.grid(row=0, column=1, sticky='ns')
        body.grid_columnconfigure(0, weight=1); body.grid_rowconfigure(0, weight=1)
        # map barcode -> inventory record for details
        inventory = []
        try:
            inventory = getattr(self.data_processor, 'fabrics_inventory', []) or []
        except Exception:
            inventory = []
        inv_map = {}
        try:
            inv_map = { str(r.get('barcode','')).strip(): r for r in inventory }
        except Exception:
            inv_map = {}
        for bc in rec.get('barcodes', []) or []:
            r = inv_map.get(str(bc).strip()) or {}
            color = (r.get('color_name','') or '')
            if r.get('color_no'):
                color = f"{color} {r.get('color_no')}".strip()
            try:
                nk = f"{float(r.get('net_kg',0)):.2f}"
            except Exception:
                nk = r.get('net_kg',0)
            try:
                mt = f"{float(r.get('meters',0)):.2f}"
            except Exception:
                mt = r.get('meters',0)
            tree2.insert('', 'end', values=(
                bc,
                r.get('fabric_type',''),
                color,
                nk,
                mt
            ))
        btns = tk.Frame(win)
        btns.pack(fill='x', pady=6)
        # Export to Excel (similar to delivery notes view)
        def _export_fs_to_excel_and_open():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Border, Side, Alignment, Font
                try:
                    from openpyxl.worksheet.page import PageMargins  # type: ignore
                except Exception:
                    PageMargins = None  # type: ignore
                try:
                    from openpyxl.utils import get_column_letter  # type: ignore
                except Exception:
                    def get_column_letter(c):
                        return 'A'
                try:
                    from openpyxl.drawing.image import Image as XLImage  # type: ignore
                except Exception:
                    XLImage = None  # type: ignore
            except Exception as e:
                try:
                    messagebox.showerror('שגיאה', f"נדרש openpyxl לצורך יצוא לאקסל:\n{e}")
                except Exception:
                    pass
                return
            try:
                wb = Workbook(); ws = wb.active; ws.title = 'שליחת בדים'
                try:
                    ws.sheet_view.rightToLeft = True
                except Exception:
                    pass
                # Page setup
                try: ws.page_setup.orientation = 'portrait'
                except Exception: pass
                try: ws.page_setup.paperSize = 9
                except Exception: pass
                try: ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
                except Exception: pass
                try:
                    ws.print_options.horizontalCentered = True
                    ws.print_options.verticalCentered = False
                except Exception:
                    pass
                try:
                    if PageMargins:
                        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
                except Exception:
                    pass
                # Header (logo from Business Details). We'll place it later after autosize to fit full width.
                logo_path = ''
                has_logo = False
                try:
                    s = getattr(self, 'settings', None)
                    if s:
                        lp = (s.get('business.logo_path', '') or '').strip()
                        if lp and os.path.exists(lp):
                            logo_path = lp
                            has_logo = True
                except Exception:
                    logo_path = ''
                    has_logo = False
                # Pre-add spacing so details start below the future logo area
                if has_logo:
                    ws.append([])
                    ws.append([])
                # Doc info
                ws.append(["מסמך", f"שליחת בדים #{rec.get('id')}"])
                # Date as text
                ws.append(["תאריך", rec.get('date','')])
                ws.append(["ספק", rec.get('supplier','')])
                try:
                    for row in ws.iter_rows(min_row=ws.max_row-2, max_row=ws.max_row, min_col=1, max_col=2):
                        row[0].font = Font(bold=True, size=16)
                        row[0].alignment = Alignment(horizontal='right')
                        row[1].alignment = Alignment(horizontal='right')
                except Exception:
                    pass
                ws.append([])
                # Table header
                headers = ["ברקוד","סוג בד","צבע","ק\"ג נטו","מטרים"]
                ws.append(headers)
                header_row = ws.max_row
                # Inventory map
                try:
                    inventory = getattr(self.data_processor, 'fabrics_inventory', []) or []
                    inv_map = { str(r.get('barcode','')).strip(): r for r in inventory }
                except Exception:
                    inv_map = {}
                total_kg = 0.0; total_m = 0.0
                for bc in rec.get('barcodes', []) or []:
                    r = inv_map.get(str(bc).strip()) or {}
                    color = (r.get('color_name','') or '')
                    if r.get('color_no'):
                        color = f"{color} {r.get('color_no')}".strip()
                    try:
                        nk = float(str(r.get('net_kg',0)).replace(',', '.'))
                    except Exception:
                        nk = 0.0
                    try:
                        mt = float(str(r.get('meters',0)).replace(',', '.'))
                    except Exception:
                        mt = 0.0
                    total_kg += nk; total_m += mt
                    ws.append([bc, r.get('fabric_type',''), color, nk, mt])
                # Totals row
                ws.append([None, None, 'סה"כ', total_kg, total_m])
                total_row = ws.max_row
                # Style
                try:
                    thin = Side(style='thin', color='000000')
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for cell in ws[header_row]:
                        cell.font = Font(bold=True, size=16)
                        cell.alignment = Alignment(horizontal='center')
                    for row in ws.iter_rows(min_row=header_row, max_row=total_row, min_col=1, max_col=5):
                        for cell in row:
                            cell.border = border
                            if cell.column <= 3:
                                cell.alignment = Alignment(horizontal='right')
                            else:
                                cell.alignment = Alignment(horizontal='center')
                    # Bold totals label and numbers
                    ws.cell(row=total_row, column=3).font = Font(bold=True)
                    ws.cell(row=total_row, column=4).font = Font(bold=True)
                    ws.cell(row=total_row, column=5).font = Font(bold=True)
                except Exception:
                    pass
                # Autosize
                try:
                    for col in ws.columns:
                        max_len = 0; col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                val = str(cell.value) if cell.value is not None else ""
                                if len(val) > max_len: max_len = len(val)
                            except Exception:
                                pass
                        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)
                except Exception:
                    pass
                # Now place the logo sized to the full sheet width (A4 fit width)
                if XLImage and has_logo and logo_path:
                    try:
                        img = XLImage(logo_path)
                        # Approximate total pixel width across used columns (1..5)
                        def col_width_to_px(w: float) -> int:
                            try:
                                return int((float(w) + 0.75) * 7)
                            except Exception:
                                return 64
                        total_px = 0
                        for c in range(1, 6):
                            w = ws.column_dimensions[get_column_letter(c)].width
                            if w is None:
                                w = 10
                            total_px += col_width_to_px(w)
                        # Scale image to full width
                        if getattr(img, 'width', None):
                            ratio = float(total_px) / float(max(1, img.width))
                            img.width = int(img.width * ratio)
                            if getattr(img, 'height', None):
                                img.height = int(img.height * ratio)
                        ws.add_image(img, 'A1')
                        # Set row 1 height roughly to image height (convert px->points ~ *0.75)
                        try:
                            h_px = getattr(img, 'height', None)
                            if h_px:
                                ws.row_dimensions[1].height = float(h_px) * 0.75
                        except Exception:
                            pass
                    except Exception:
                        pass
                # Save
                base_dir = os.path.join(os.getcwd(), 'exports', 'fabrics_shipments')
                try: os.makedirs(base_dir, exist_ok=True)
                except Exception: pass
                safe_id = rec.get('id'); safe_date = (rec.get('date') or '').replace('/', '-').replace(':', '-')
                fname = f"fabrics_shipment_{safe_id}_{safe_date}.xlsx" if safe_id is not None else f"fabrics_shipment_{safe_date}.xlsx"
                out_path = os.path.join(base_dir, fname)
                try:
                    wb.save(out_path)
                except PermissionError as e:
                    try:
                        ts = datetime.now().strftime('%H%M%S')
                        alt_path = out_path.replace('.xlsx', f'_{ts}.xlsx')
                        wb.save(alt_path)
                        out_path = alt_path
                    except Exception:
                        try: messagebox.showerror('שגיאה', f"שמירת קובץ נכשלה (קובץ פתוח?):\n{e}")
                        except Exception: pass
                        return
                except Exception as e:
                    try: messagebox.showerror('שגיאה', f"שמירת קובץ נכשלה:\n{e}")
                    except Exception: pass
                    return
                try:
                    os.startfile(out_path)  # type: ignore[attr-defined]
                except Exception:
                    try: messagebox.showinfo('נשמר', f"הקובץ נשמר ב:\n{out_path}\n(לא הצלחתי לפתוח אוטומטית)")
                    except Exception: pass
            except Exception as e:
                try: messagebox.showerror('שגיאה', str(e))
                except Exception: pass

        tk.Button(btns, text='🖨 פתח באקסל', command=_export_fs_to_excel_and_open, bg='#27ae60', fg='white').pack(side='right', padx=4)
        tk.Button(btns, text='סגור', command=win.destroy).pack(side='right')
