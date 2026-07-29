import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import os
from datetime import datetime
from .. import theme

# סימן כיווניות RTL לטקסט עברי (Right-To-Left Mark)
RLM = '\u200f'


class DrawingsManagerTabMixin:
    """Mixin עבור טאב מנהל ציורים.

    This is the same implementation previously in gui/drawings_manager_tab.py,
    moved here to organize the feature under its own package.
    """

    def _format_products_details(self, products):
        """עיצוב פירוט דגמים ומידות בפורמט המבוקש"""
        if not products:
            return ""
        
        formatted_parts = []
        for product in products:
            product_name = product.get('שם המוצר', '')
            if not product_name:
                continue
                
            # איסוף כל המידות למוצר זה
            size_parts = []
            for size_info in product.get('מידות', []):
                size = size_info.get('מידה', '')
                quantity = size_info.get('כמות', 0)
                if size and quantity > 0:
                    # הוספת LTR mark לפני המספרים כדי להבטיח הצגה נכונה
                    size_parts.append(f"[\u202D{size}X{int(quantity)}\u202C]")
            
            if size_parts:
                formatted_parts.append(f"{''.join(size_parts)}{product_name}")
        
        # הוספת RLM בתחילת המחרוזת כולה כדי להבטיח קריאה נכונה מימין לשמאל
        return RLM + "".join(formatted_parts)

    # === Helpers: context menu for text inputs ===
    def _attach_paste_menu(self, widget: tk.Widget):
        """Attach a simple right-click menu with a Paste action to a text-capable widget (e.g., Entry).

        Windows users expect right-click paste. This binds Button-3 to show a menu with 'הדבק'.
        """
        def _show_menu(event, w=widget):
            try:
                menu = tk.Menu(w, tearoff=0)
                menu.add_command(label="הדבק", command=lambda: w.event_generate("<<Paste>>"))
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                try:
                    menu.grab_release()
                except Exception:
                    pass
            return "break"
        try:
            widget.bind("<Button-3>", _show_menu)
        except Exception:
            pass

    def _create_drawings_manager_tab(self):
        tab = tk.Frame(self.notebook, bg=theme.PAGE_BG)
        self.notebook.add(tab, text="מנהל ציורים")
        self._drawings_tab = tab
        tk.Label(tab, text="מנהל ציורים - טבלה מקומית", font=(theme.FONT_FAMILY, 16, 'bold'), bg=theme.PAGE_BG, fg=theme.DARK).pack(pady=10)
        # Inner notebook to host drawings table and embedded converter
        inner_nb = ttk.Notebook(tab)
        inner_nb.pack(fill='both', expand=True, padx=6, pady=(0, 6))
        table_page = tk.Frame(inner_nb, bg=theme.PAGE_BG)
        converter_page = tk.Frame(inner_nb, bg=theme.PAGE_BG)
        cut_drawings_page = tk.Frame(inner_nb, bg=theme.PAGE_BG)
        product_map_page = tk.Frame(inner_nb, bg=theme.PAGE_BG)
        area_calc_page = tk.Frame(inner_nb, bg=theme.PAGE_BG)
        inner_nb.add(table_page, text="טבלת ציורים")
        inner_nb.add(converter_page, text="ממיר קבצים")
        inner_nb.add(product_map_page, text="מיפוי מוצרים")
        inner_nb.add(area_calc_page, text="שטח רבוע לדגם מידה")
        # Embed cut drawings (returned drawings) tab if builder exists
        try:
            if hasattr(self, '_build_returned_drawings_content'):
                inner_nb.add(cut_drawings_page, text="קליטת ציור שנחתך")
                self._build_returned_drawings_content(cut_drawings_page)
        except Exception:
            pass

        actions = tk.Frame(table_page, bg=theme.PAGE_BG)
        actions.pack(fill='x', padx=12, pady=(0, 8))
        left = tk.Frame(actions, bg=theme.PAGE_BG); left.pack(side='left')
        tk.Button(left, text="🔄 רענן", command=self._refresh_drawings_tree, bg=theme.PRIMARY, fg='white', font=(theme.FONT_FAMILY, 10, 'bold'), width=10).pack(side='left', padx=4)
        tk.Button(left, text="📊 ייצא לאקסל", command=self._export_drawings_to_excel_tab, bg=theme.SUCCESS, fg='white', font=(theme.FONT_FAMILY, 10, 'bold'), width=12).pack(side='left', padx=4)
        right = tk.Frame(actions, bg=theme.PAGE_BG); right.pack(side='right')
        tk.Button(right, text="❌ מחק נבחר", command=self._delete_selected_drawing_tab, bg=theme.WARNING, fg='white', font=(theme.FONT_FAMILY, 10, 'bold'), width=10).pack(side='right', padx=4)

        # מערכת סינון
        filter_frame = tk.Frame(table_page, bg=theme.PAGE_BG)
        filter_frame.pack(fill='x', padx=12, pady=(0, 8))
        
        tk.Label(filter_frame, text="סינון:", font=(theme.FONT_FAMILY, 10, 'bold'), bg=theme.PAGE_BG).pack(side='left', padx=(0, 8))
        
        # סינון לפי ספק
        tk.Label(filter_frame, text="ספק:", font=(theme.FONT_FAMILY, 9), bg=theme.PAGE_BG).pack(side='left', padx=(0, 4))
        self.drawings_supplier_filter_var = tk.StringVar()
        self.drawings_supplier_filter_cb = ttk.Combobox(filter_frame, textvariable=self.drawings_supplier_filter_var, width=20, state='readonly')
        self.drawings_supplier_filter_cb.pack(side='left', padx=(0, 8))
        self.drawings_supplier_filter_cb.bind('<<ComboboxSelected>>', self._apply_drawings_filters)
        
        # סינון לפי סטטוס
        tk.Label(filter_frame, text="סטטוס:", font=(theme.FONT_FAMILY, 9), bg=theme.PAGE_BG).pack(side='left', padx=(0, 4))
        self.drawings_status_filter_var = tk.StringVar()
        self.drawings_status_filter_cb = ttk.Combobox(filter_frame, textvariable=self.drawings_status_filter_var, width=15, state='readonly')
        self.drawings_status_filter_cb.pack(side='left', padx=(0, 8))
        self.drawings_status_filter_cb.bind('<<ComboboxSelected>>', self._apply_drawings_filters)
        
        # כפתור נקה סינון
        tk.Button(filter_frame, text="🗑️ נקה סינון", command=self._clear_drawings_filters, bg=theme.MUTED, fg='white', font=(theme.FONT_FAMILY, 9)).pack(side='left', padx=8)

        table_frame = tk.Frame(table_page, bg=theme.CARD_BG)
        table_frame.pack(fill='both', expand=True, padx=12, pady=8)
        cols = ("id", "file_name", "created_at", "products", "total_quantity", "estimated_layers", "actual_layers", "products_details", "sent_to_supplier", "status", "excel")
        self.drawings_tree = ttk.Treeview(table_frame, columns=cols, show='headings')
        headers = {"id": "ID", "file_name": "שם הקובץ", "created_at": "תאריך יצירה", "products": "מוצרים", "total_quantity": "סך כמויות", "estimated_layers": "שכבות משוערת", "actual_layers": "שכבות בפועל", "products_details": "פירוט דגמים ומידות", "sent_to_supplier": "נשלח לספק", "status": "סטטוס", "excel": "Excel"}
        widths = {"id": 70, "file_name": 260, "created_at": 140, "products": 80, "total_quantity": 90, "estimated_layers": 100, "actual_layers": 100, "products_details": 300, "sent_to_supplier": 100, "status": 90, "excel": 60}
        for c in cols:
            self.drawings_tree.heading(c, text=headers[c])
            self.drawings_tree.column(c, width=widths[c], anchor='center')
        vs = ttk.Scrollbar(table_frame, orient='vertical', command=self.drawings_tree.yview)
        self.drawings_tree.configure(yscroll=vs.set)
        self.drawings_tree.grid(row=0, column=0, sticky='nsew')
        vs.grid(row=0, column=1, sticky='ns')
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)
        self.drawings_tree.bind('<Double-1>', self._on_drawings_double_click)
        self.drawings_tree.bind('<Button-3>', self._on_drawings_right_click)
        self.drawings_tree.bind('<Button-1>', self._on_drawings_click)
        self._drawing_status_menu = tk.Menu(self.drawings_tree, tearoff=0)
        for st in ("טרם נשלח", "נשלח", "הוחזר", "נחתך"):
            self._drawing_status_menu.add_command(label=st, command=lambda s=st: self._change_selected_drawing_status(s))
        self.drawings_stats_var = tk.StringVar(value="אין נתונים")
        tk.Label(table_page, textvariable=self.drawings_stats_var, bg=theme.DARK_2, fg='white', anchor='w', padx=10, font=(theme.FONT_FAMILY, 10)).pack(fill='x', side='bottom')
        
        # אתחול רשימות הסינון
        self._refresh_drawings_filter_options()

        # Build converter content inside second inner tab if available
        try:
            if hasattr(self, '_build_converter_tab_content'):
                self._build_converter_tab_content(converter_page)
        except Exception:
            pass
        self._populate_drawings_tree()
        self._update_drawings_stats()
        # Build product mapping tab
        try:
            self._build_product_mapping_tab(product_map_page)
        except Exception:
            pass
        
        # Build area calculation tab
        try:
            self._build_area_calculation_tab(area_calc_page)
        except Exception:
            pass

    # === Product Mapping Tab ===
    def _build_product_mapping_tab(self, container: tk.Widget):
        wrapper = tk.Frame(container, bg=theme.PAGE_BG)
        wrapper.pack(fill='both', expand=True, padx=10, pady=8)
        tk.Label(wrapper, text="ניהול מיפוי מוצרים (קובץ מוצרים.xlsx)", font=(theme.FONT_FAMILY, 14, 'bold'), bg=theme.PAGE_BG).pack(anchor='e', pady=(0, 8))

        actions = tk.Frame(wrapper, bg=theme.PAGE_BG); actions.pack(fill='x', pady=(0, 6))
        tk.Button(actions, text="🔄 רענן", command=self._refresh_product_mapping_table, bg=theme.PRIMARY, fg='white').pack(side='right', padx=4)
        tk.Button(actions, text="💾 שמור לקובץ", command=self._save_product_mapping, bg=theme.DARK, fg='white').pack(side='right', padx=4)

        form = tk.Frame(wrapper, bg=theme.PANEL_BG); form.pack(fill='x', pady=(0, 6))
        tk.Label(form, text="file name:", bg=theme.PANEL_BG).grid(row=0, column=0, sticky='w', padx=6, pady=4)
        self.pm_file_name_var = tk.StringVar()
        file_name_entry = tk.Entry(form, textvariable=self.pm_file_name_var, width=26)
        file_name_entry.grid(row=0, column=1, sticky='w', padx=4, pady=4)
        self._attach_paste_menu(file_name_entry)
        tk.Label(form, text="product name:", bg=theme.PANEL_BG).grid(row=0, column=2, sticky='w', padx=10, pady=4)
        self.pm_product_name_var = tk.StringVar()
        # Load model names list and create a combobox for selection
        try:
            self._load_model_names_options()
        except Exception:
            self._model_names_list = []
        self.pm_product_name_combo = ttk.Combobox(
            form,
            textvariable=self.pm_product_name_var,
            width=24,
            values=getattr(self, '_model_names_list', []),
            state='normal'  # allow typing if list is empty or custom value needed
        )
        self.pm_product_name_combo.grid(row=0, column=3, sticky='w', padx=4, pady=4)
        tk.Label(form, text="unit quantity:", bg=theme.PANEL_BG).grid(row=0, column=4, sticky='w', padx=10, pady=4)
        self.pm_unit_qty_var = tk.StringVar(value='1')
        self.pm_unit_qty_spin = tk.Spinbox(form, from_=1, to=999, textvariable=self.pm_unit_qty_var, width=5)
        self.pm_unit_qty_spin.grid(row=0, column=5, sticky='w', padx=4, pady=4)
        tk.Button(form, text="➕ הוסף/עדכן", command=self._add_product_mapping_row, bg=theme.SUCCESS, fg='white').grid(row=0, column=6, padx=8)
        tk.Button(form, text="🗑️ מחק נבחר", command=self._delete_selected_product_mapping, bg=theme.WARNING, fg='white').grid(row=0, column=7, padx=4)
        tk.Button(form, text="❌ נקה שדות", command=lambda: (self.pm_file_name_var.set(''), self.pm_product_name_var.set(''), self.pm_unit_qty_var.set('1')), bg=theme.DANGER, fg='white').grid(row=0, column=8, padx=4)

        cols = ('file_name', 'product_name', 'unit_qty')
        self.product_mapping_tree = ttk.Treeview(wrapper, columns=cols, show='headings', height=12)
        self.product_mapping_tree.heading('file_name', text='file name')
        self.product_mapping_tree.heading('product_name', text='product name')
        self.product_mapping_tree.heading('unit_qty', text='unit quantity')
        self.product_mapping_tree.column('file_name', width=210, anchor='w')
        self.product_mapping_tree.column('product_name', width=220, anchor='w')
        self.product_mapping_tree.column('unit_qty', width=110, anchor='center')
        vs = ttk.Scrollbar(wrapper, orient='vertical', command=self.product_mapping_tree.yview)
        self.product_mapping_tree.configure(yscroll=vs.set)
        self.product_mapping_tree.pack(side='left', fill='both', expand=True)
        vs.pack(side='right', fill='y')
        self.product_mapping_tree.bind('<<TreeviewSelect>>', self._on_product_mapping_select)

        self._product_mapping_rows = []  # each row includes optional 'unit quantity'
        self._load_product_mapping_initial()

    # === Model names support ===
    def _get_model_names_path(self):
        import os
        return os.path.join(os.getcwd(), 'model_names.json')

    def _load_model_names_options(self):
        """Load model names (product names) from model_names.json and update combobox if exists."""
        import os, json
        path = self._get_model_names_path()
        names = []
        try:
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # Expecting list of { id, name, created_at }
                for item in data or []:
                    name = (item.get('name') or '').strip()
                    if name:
                        names.append(name)
        except Exception:
            names = []
        # unique + sort by name
        try:
            names = sorted({n for n in names})
        except Exception:
            names = names
        self._model_names_list = names
        # Update combobox values if created
        if hasattr(self, 'pm_product_name_combo'):
            try:
                self.pm_product_name_combo['values'] = self._model_names_list
            except Exception:
                pass

    def _get_products_excel_path(self):
        import os
        return os.path.join(os.getcwd(), 'קובץ מוצרים.xlsx')

    def _load_product_mapping_initial(self):
        self._product_mapping_rows = []
        path = self._get_products_excel_path()
        try:
            import pandas as pd, os
            if os.path.exists(path):
                df = pd.read_excel(path)
                for _, r in df.iterrows():
                    fn = str(r.get('file name') or '').strip(); pn = str(r.get('product name') or '').strip()
                    uq_raw = r.get('unit quantity', 1)
                    try:
                        uq = int(uq_raw)
                        if uq <= 0: uq = 1
                    except Exception:
                        uq = 1
                    if fn and pn:
                        self._product_mapping_rows.append({'file name': fn, 'product name': pn, 'unit quantity': uq})
        except Exception:
            pass
        self._populate_product_mapping_tree()

    def _populate_product_mapping_tree(self):
        if not hasattr(self, 'product_mapping_tree'): return
        for iid in self.product_mapping_tree.get_children(): self.product_mapping_tree.delete(iid)
        for row in self._product_mapping_rows:
            self.product_mapping_tree.insert('', 'end', values=(row['file name'], row['product name'], row.get('unit quantity', 1)))

    def _refresh_product_mapping_table(self):
        # Refresh available model names and table content
        try:
            self._load_model_names_options()
        except Exception:
            pass
        self._load_product_mapping_initial()

    def _add_product_mapping_row(self):
        fn = (self.pm_file_name_var.get() or '').strip(); pn = (self.pm_product_name_var.get() or '').strip(); uq_txt = (self.pm_unit_qty_var.get() or '').strip()
        if not fn or not pn:
            messagebox.showerror("שגיאה", "יש למלא file name + product name"); return
        try:
            uq = int(uq_txt)
            if uq <= 0: raise ValueError
        except Exception:
            messagebox.showerror("שגיאה", "unit quantity חייב להיות מספר >= 1")
            return
        replaced = False
        for row in self._product_mapping_rows:
            if row['file name'].lower() == fn.lower():
                row['file name'] = fn; row['product name'] = pn; row['unit quantity'] = uq; replaced = True; break
        if not replaced:
            self._product_mapping_rows.append({'file name': fn, 'product name': pn, 'unit quantity': uq})
        self._populate_product_mapping_tree(); self.pm_file_name_var.set(''); self.pm_product_name_var.set(''); self.pm_unit_qty_var.set('1')
        # Auto-save silently so changes persist across restarts
        try:
            self._save_product_mapping(False)
        except Exception:
            pass

    def _delete_selected_product_mapping(self):
        if not hasattr(self, 'product_mapping_tree'): return
        sel = self.product_mapping_tree.selection();
        if not sel: return
        values = self.product_mapping_tree.item(sel[0], 'values')
        if not values: return
        fn = values[0]
        self._product_mapping_rows = [r for r in self._product_mapping_rows if r['file name'] != fn]
        self._populate_product_mapping_tree()
        # Auto-save silently after delete
        try:
            self._save_product_mapping(False)
        except Exception:
            pass

    def _on_product_mapping_select(self, event=None):
        sel = self.product_mapping_tree.selection();
        if not sel: return
        vals = self.product_mapping_tree.item(sel[0], 'values');
        if not vals: return
        self.pm_file_name_var.set(vals[0]); self.pm_product_name_var.set(vals[1])
        # ensure combobox shows the selected value
        try:
            if hasattr(self, 'pm_product_name_combo'):
                self.pm_product_name_combo.set(vals[1])
        except Exception:
            pass
        if len(vals) > 2:
            try:
                self.pm_unit_qty_var.set(str(vals[2]))
            except Exception:
                self.pm_unit_qty_var.set('1')

    def _save_product_mapping(self, show_message: bool = True):
        path = self._get_products_excel_path()
        try:
            import pandas as pd
            if not self._product_mapping_rows:
                messagebox.showerror("שגיאה", "אין שורות לשמירה"); return
            normalized = []
            for r in self._product_mapping_rows:
                uq = r.get('unit quantity', 1)
                try:
                    uq = int(uq)
                    if uq <= 0: uq = 1
                except Exception:
                    uq = 1
                normalized.append({'file name': r.get('file name',''), 'product name': r.get('product name',''), 'unit quantity': uq})
            df = pd.DataFrame(normalized)[['file name','product name','unit quantity']]
            df.to_excel(path, index=False)
            if show_message:
                messagebox.showinfo("הצלחה", f"נשמר {path}")
            try:
                if hasattr(self, 'file_analyzer') and path:
                    self.file_analyzer.load_products_mapping(path)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _show_drawings_manager_tab(self):
        for i in range(len(self.notebook.tabs())):
            if self.notebook.tab(i, 'text') == "מנהל ציורים": self.notebook.select(i); break

    def _populate_drawings_tree(self):
        for item in self.drawings_tree.get_children(): self.drawings_tree.delete(item)
        for record in self.data_processor.drawings_data:
            products_count = len(record.get('מוצרים', [])); total_quantity = record.get('סך כמויות', 0)
            # הצגת תאריך ללא שעת יצירה (רק חלק התאריך)
            created_raw = record.get('תאריך יצירה','')
            created_date_only = created_raw.split()[0] if isinstance(created_raw, str) and created_raw else created_raw
            sent_flag = record.get('נשלח לספק')
            # אם יש שם ספק – נציג אותו; אחרת נשמור לוגיקה קיימת (כן/לא/ריק)
            supplier_name = (record.get('נמען') or '').strip()
            if supplier_name:
                sent_display = supplier_name
            else:
                sent_display = 'כן' if sent_flag is True else ('לא' if sent_flag is False else '')
            estimated_layers = record.get('כמות שכבות משוערת', '—')
            # שכבות בפועל - מוצג רק לציורים שנחתכו
            status = record.get('status', 'נשלח')
            actual_layers = record.get('שכבות', '') if status == 'נחתך' else '—'
            products_details = self._format_products_details(record.get('מוצרים', []))
            self.drawings_tree.insert('', 'end', values=(
                record.get('id',''),
                record.get('שם הקובץ',''),
                created_date_only,
                products_count,
                f"{total_quantity:.1f}" if isinstance(total_quantity,(int,float)) else total_quantity,
                estimated_layers,
                actual_layers,
                products_details,
                sent_display,
                record.get('status','נשלח'),
                "📄"  # excel icon
            ))

    def _update_drawings_stats(self):
        total_drawings = len(self.data_processor.drawings_data); total_quantity = sum(r.get('סך כמויות', 0) for r in self.data_processor.drawings_data)
        self.drawings_stats_var.set(f"סך הכל: {total_drawings} ציורים | סך כמויות: {total_quantity:.1f}")

    def _refresh_drawings_tree(self):
        if hasattr(self.data_processor, 'refresh_drawings_data'):
            try: self.data_processor.refresh_drawings_data()
            except Exception: pass
        # עדכון רשימות הסינון
        self._refresh_drawings_filter_options()
        # החלת הסינון הנוכחי או הצגת הכל
        self._apply_drawings_filters()

    def _on_drawings_double_click(self, event):
        item_id = self.drawings_tree.focus();
        if not item_id: return
        vals = self.drawings_tree.item(item_id, 'values');
        if not vals: return
        try: rec_id = int(vals[0])
        except Exception: return
        record = self.data_processor.get_drawing_by_id(rec_id) if hasattr(self.data_processor, 'get_drawing_by_id') else None
        if not record: return
        self._show_drawing_details(record)

    def _on_drawings_right_click(self, event):
        row_id = self.drawings_tree.identify_row(event.y)
        if row_id: self.drawings_tree.selection_set(row_id)
        sel = self.drawings_tree.selection()
        if not sel: return
        menu = tk.Menu(self.drawings_tree, tearoff=0)
        menu.add_command(label="📋 הצג פרטים", command=lambda: self._on_drawings_double_click(None))
        menu.add_cascade(label="סטטוס", menu=self._drawing_status_menu)
        menu.add_separator(); menu.add_command(label="🗑️ מחק", command=self._delete_selected_drawing_tab)
        try: menu.tk_popup(event.x_root, event.y_root)
        finally: menu.grab_release()

    def _change_selected_drawing_status(self, new_status):
        sel = self.drawings_tree.selection();
        if not sel: return
        vals = self.drawings_tree.item(sel[0], 'values');
        if not vals: return
        try: rec_id = int(vals[0])
        except Exception: return
        if hasattr(self.data_processor, 'update_drawing_status') and self.data_processor.update_drawing_status(rec_id, new_status):
            # עדכון עמודת הסטטוס לפי שם העמודה כדי לא להיות תלוי אינדקסים
            new_vals = list(vals)
            cols = list(self.drawings_tree['columns'])
            try:
                status_idx = cols.index('status')
                if 0 <= status_idx < len(new_vals):
                    new_vals[status_idx] = new_status
            except Exception:
                pass
            self.drawings_tree.item(sel[0], values=new_vals)

    # === Click handling for per-row Excel export column ===
    def _on_drawings_click(self, event):
        try:
            col_id = self.drawings_tree.identify_column(event.x)  # e.g. '#1'
            row_id = self.drawings_tree.identify_row(event.y)
            if not row_id:
                return
            # Determine column name
            columns = self.drawings_tree['columns']
            idx = int(col_id.replace('#','')) - 1
            if 0 <= idx < len(columns) and columns[idx] == 'excel':
                vals = self.drawings_tree.item(row_id, 'values')
                if not vals: return
                try:
                    rec_id = int(vals[0])
                except Exception:
                    return
                record = self.data_processor.get_drawing_by_id(rec_id) if hasattr(self.data_processor, 'get_drawing_by_id') else None
                if record:
                    self._export_single_drawing_to_excel(record)
                return 'break'  # prevent selection change flicker
        except Exception:
            pass

    def _export_single_drawing_to_excel(self, record):
        """Create a temporary Excel file (דגם/מידה/כמות) with RTL, bold + size16 headers, size16 centered cells."""
        try:
            import pandas as pd, tempfile, os
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            rows = []
            for product in record.get('מוצרים', []):
                for size_info in product.get('מידות', []):
                    rows.append({
                        'דגם': product.get('שם המוצר',''),
                        'מידה': size_info.get('מידה',''),
                        'כמות': size_info.get('כמות',0)
                    })
            if not rows:
                messagebox.showwarning("אין נתונים", "אין שורות לייצוא לציור זה")
                return
            # DataFrame without 'הערה' column per user request
            df = pd.DataFrame(rows, columns=['דגם','מידה','כמות'])
            # Create temp file
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f"_drawing_{record.get('id')}.xlsx")
            tmp_path = tmp.name
            tmp.close()
            df.to_excel(tmp_path, index=False)
            # Style header
            wb = load_workbook(tmp_path)
            ws = wb.active
            # גליון RTL
            try: ws.sheet_view.rightToLeft = True
            except Exception: pass
            # הוספת שורות מידע מעל הטבלה: נזיז 8 שורות כדי לפנות מקום ללוגו ומידע עסקי
            ws.insert_rows(1, amount=8)  # headers יעברו לשורה 9
            max_col = ws.max_column
            raw_dt = record.get('תאריך יצירה','')
            # שמירת תאריך בלבד ללא זמן; אם כולל זמן נחתוך ברווח הראשון
            formatted_date = ''
            if isinstance(raw_dt, str):
                if ' ' in raw_dt:
                    formatted_date = raw_dt.split()[0]
                else:
                    # אם כבר רק תאריך באורך 10 (YYYY-MM-DD) נשאיר
                    formatted_date = raw_dt
            # שם העסק מתוך טאב "פרטי עסק"
            business_name = ''
            logo_path = ''
            try:
                s = getattr(self, 'settings', None)
                if s:
                    business_name = s.get('business.name', '') or ''
                    logo_path = s.get('business.logo_path', '') or ''
            except Exception:
                business_name = ''
                logo_path = ''
            
            # A1: לוגו העסק בלבד (ללא שם עסק) - מוגבל עד שורה 5
            try:
                from openpyxl.drawing.image import Image as XLImage
                if logo_path and os.path.exists(logo_path):
                    # הוספת לוגו בלבד
                    img = XLImage(logo_path)
                    # התאמת גודל הלוגו - מוגבל בדיוק לשורות 1-5
                    img.width = 564  # רוחב מופחת ב-6% (600 * 0.94)
                    img.height = 200  # גובה חזר למקור
                    # הוספת הלוגו לתא A1
                    ws.add_image(img, 'A1')
                    # הגדלת גובה השורות 1-5 כדי להכיל את הלוגו בדיוק
                    for row in range(1, 6):
                        ws.row_dimensions[row].height = 40  # 40 פיקסלים לכל שורה = 200 סה"כ
            except Exception:
                pass  # אם יש בעיה עם הלוגו, לא נוסיף כלום לתא A1
            
            # מזג בין A1 ועד העמודה האחרונה בהדר (לפחות C) - מיזוג שורות 1-5
            try:
                last_col_letter = ws.cell(row=1, column=max_col).column_letter
                ws.merge_cells(start_row=1, start_column=1, end_row=5, end_column=max_col)
            except Exception:
                last_col_letter = 'C'

            # העברת כל המידע העסקי לשורה 6
            # A6: ציור מספר, B6: סוג בד
            ws.cell(row=6, column=1, value=f"ציור מספר: {record.get('id','')}")
            ws.cell(row=6, column=2, value=f"סוג בד: {record.get('סוג בד','')}")

            # B7: תאריך יצירה עם הטקסט המלא
            if formatted_date:
                ws.cell(row=7, column=2, value=f"תאריך יצירה: {formatted_date}")

            # A7: כמות השכבות המשוערת
            layers_count = record.get('שכבות', '') or record.get('כמות שכבות משוערת', '')
            if layers_count:
                ws.cell(row=7, column=1, value=f"כמות שכבות: {layers_count}")

            # שורה 8: אורך ורוחב הציור
            drawing_length = record.get('אורך ציור', '')
            drawing_width = record.get('רוחב ציור', '')
            if drawing_length:
                # המרת סנטימטרים למטרים
                try:
                    length_meters = float(drawing_length) / 100
                    ws.cell(row=8, column=1, value=f"אורך ציור: {length_meters:.2f} מטרים")
                except (ValueError, TypeError):
                    ws.cell(row=8, column=1, value=f"אורך ציור: {drawing_length} ס\"מ")
            if drawing_width:
                ws.cell(row=8, column=2, value=f"רוחב ציור: {drawing_width} ס\"מ")

            # עיצוב גודל כמו הכותרות (16) לשורות 6-8
            meta_font = Font(size=16)
            for (r, c) in [(6,1), (6,2), (7,1), (7,2), (8,1), (8,2)]:
                try:
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None:
                        cell.font = meta_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                except Exception:
                    pass

            header_row_index = 9
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_font = Font(bold=True, size=16)
            base_font = Font(size=16)
            
            # יישור וגבולות לכל התאים + רוחב עמודות אוטומטי
            thin = Side(border_style='thin', color='000000')
            
            # יישור וגבולות לכל התאים
            for r in range(header_row_index, ws.max_row+1):
                for c in range(1, ws.max_column+1):
                    cell = ws.cell(row=r, column=c)
                    if r == header_row_index:
                        # כותרות
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = header_fill
                    else:
                        # נתונים
                        cell.font = base_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # רוחב עמודות אוטומטי - חישוב ידני
            for c in range(1, ws.max_column+1):
                col_letter = ws.cell(row=header_row_index, column=c).column_letter
                max_length = 0
                
                # בדיקת אורך התוכן בכל התאים בעמודה
                for r in range(1, ws.max_row+1):
                    cell_value = ws.cell(row=r, column=c).value
                    if cell_value:
                        # חישוב אורך התוכן (תווים עבריים נחשבים כפול)
                        cell_length = len(str(cell_value))
                        # הוספת מרווח נוסף לתווים עבריים
                        if any('\u0590' <= char <= '\u05FF' for char in str(cell_value)):
                            cell_length = int(cell_length * 1.5)
                        max_length = max(max_length, cell_length)
                
                # הגדרת רוחב העמודה (מינימום 10, מקסימום 50)
                column_width = max(10, min(50, max_length + 2))
                ws.column_dimensions[col_letter].width = column_width
            # Optional metadata sheet
            meta = wb.create_sheet('פרטי ציור')
            try: meta.sheet_view.rightToLeft = True
            except Exception: pass
            meta.append(['שם קובץ', record.get('שם הקובץ','')])
            # התאמה לשם החדש 'ציור מספר'
            meta.append(['ציור מספר', record.get('id','')])
            # תאריך ביצירת גליון המטא - גם כאן רק תאריך
            meta_date = raw_dt.split()[0] if isinstance(raw_dt, str) and raw_dt else raw_dt
            meta.append(['תאריך יצירה', meta_date])
            meta.append(['סוג בד', record.get('סוג בד','')])
            meta.append(['כמות שכבות משוערת', record.get('כמות שכבות משוערת','—')])
            meta.append(['סטטוס', record.get('status','')])
            for row in meta.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2):
                for cell in row:
                    if cell.column == 1:
                        cell.font = Font(bold=True, size=16)
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.font = base_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # הוספת טבלת ברקודים מתחת לטבלה הראשית
            last_data_row = ws.max_row
            barcode_start_row = last_data_row + 3  # רווח של שורתיים
            
            # המשפט מעל הטבלה
            ws.cell(row=barcode_start_row, column=1, value="רשום את מספרי הברקודים ושדך לדף")
            barcode_title_cell = ws.cell(row=barcode_start_row, column=1)
            barcode_title_cell.font = Font(bold=True, size=16)
            barcode_title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # מיזוג התא לרוחב שתי עמודות
            ws.merge_cells(start_row=barcode_start_row, start_column=1, end_row=barcode_start_row, end_column=2)
            
            # יצירת 10 שורות עם גבולות שחורים בעמודות A ו-B
            thick_border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
            
            for row_num in range(barcode_start_row + 1, barcode_start_row + 11):  # 10 שורות
                for col_num in [1, 2]:  # עמודות A ו-B
                    cell = ws.cell(row=row_num, column=col_num, value="")
                    cell.border = thick_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            wb.save(tmp_path)
            # Open with Excel
            try:
                os.startfile(tmp_path)  # type: ignore[attr-defined]
            except Exception:
                messagebox.showinfo("קובץ נוצר", f"הקובץ נוצר ב:\n{tmp_path}")
        except Exception as e:
            messagebox.showerror("שגיאה", f"כשל ביצירת קובץ Excel: {e}")

    def _print_drawing_record(self, record):
        """Open a printable-style window and offer system print if possible."""
        # Header info first (RTL lines)
        header_lines = []
        header_lines.append(f"ציור: {record.get('שם הקובץ','')}")
        header_lines.append(f"ID: {record.get('id','')}")
        header_lines.append(f"תאריך יצירה: {record.get('תאריך יצירה','')}")
        if 'סוג בד' in record:
            header_lines.append(f"סוג בד: {record.get('סוג בד')}")
        header_lines.append(f"סטטוס: {record.get('status','')}")

        # Flatten products into single table rows: product(model), size, quantity
        rows = []
        import re
        def _size_key(size_str: str):
            """מפתח מיון לגדלים (תמיכה בפורמטים שונים כמו 0-3, 12m-18m, 3, 24-30)."""
            if not isinstance(size_str, str):
                return 0, size_str
            # החלפת אות m (months) כדי להשוות מספרים בלבד
            cleaned = size_str.lower().replace('m','')
            # קח את הספרות הראשונות
            m = re.match(r"(\d+)", cleaned.strip())
            base = int(m.group(1)) if m else 0
            # עדיפות: טווח לפני מספר בודד? נשתמש באורך לקביעת סדר יציב
            return base, cleaned
        for product in record.get('מוצרים', []):
            prod_name = product.get('שם המוצר','')
            # מיון לוגי של המידות לפני יצירת השורות
            sorted_sizes = sorted(product.get('מידות', []), key=lambda si: _size_key(si.get('מידה','')))
            for size_info in sorted_sizes:
                size = size_info.get('מידה','')
                qty = size_info.get('כמות',0)
                rows.append((prod_name, size, qty))

        # Determine column widths (in characters)
        prod_w = max([len(str(r[0])) for r in rows] + [4])
        size_w = max([len(str(r[1])) for r in rows] + [4])
        qty_w  = max([len(str(r[2])) for r in rows] + [4])

        # Build table header (remember RTL: we want Product on right, then Size, then Quantity on left)
        # Using monospaced font; we compose as Product | Size | Quantity in logical order, RLM will render RTL
        table_lines = []
        header_row = f"{ 'דגם'.ljust(prod_w) }  { 'מידה'.ljust(size_w) }  { 'כמות'.rjust(qty_w) }"
        sep_row = '-' * len(header_row)
        table_lines.append(header_row)
        table_lines.append(sep_row)
        for prod, size, qty in rows:
            line = f"{ str(prod).ljust(prod_w) }  { str(size).ljust(size_w) }  { str(qty).rjust(qty_w) }"
            table_lines.append(line)

        content = "\n".join([RLM + l for l in header_lines]) + "\n\n" + "\n".join([RLM + l for l in table_lines])

        top = tk.Toplevel(self.root); top.title("תצוגת הדפסה"); top.geometry('600x700'); top.configure(bg=theme.PAGE_BG)
        txt = scrolledtext.ScrolledText(top, font=('Courier New', 10), wrap='word')
        txt.pack(fill='both', expand=True, padx=8, pady=8)
        txt.tag_configure('rtl', justify='right')
        txt.insert(tk.END, content, 'rtl'); txt.config(state='disabled')
        btns = tk.Frame(top, bg=theme.PAGE_BG); btns.pack(fill='x', pady=6)
        tk.Button(btns, text="הדפס", command=lambda: self._attempt_system_print(content), bg=theme.DARK, fg='white', width=12).pack(side='left', padx=10)
        tk.Button(btns, text="סגור", command=top.destroy, bg=theme.MUTED, fg='white', width=12).pack(side='right', padx=10)

    def _attempt_system_print(self, text_content: str):
        """Try to print using Windows default printer. Fallback: copy to clipboard."""
        try:
            import tempfile, os, sys
            # ודא שהקובץ מתחיל ב-RLM כדי לשמר יישור RTL גם בהדפסה טקסטואלית
            if not text_content.startswith(RLM):
                text_content = RLM + text_content
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.txt', mode='w', encoding='utf-8')
            tmp.write(text_content)
            tmp.close()
            # Windows specific printing via shell
            if sys.platform.startswith('win'):
                try:
                    os.startfile(tmp.name, 'print')  # type: ignore[attr-defined]
                    messagebox.showinfo("הודעה", "נשלח להדפסה (ייתכן עיכוב שניות ספורות)")
                    return
                except Exception:
                    pass
            # Fallback copy to clipboard
            self.root.clipboard_clear(); self.root.clipboard_append(text_content)
            messagebox.showinfo("הודעה", "לא ניתן להדפיס ישירות. התוכן הועתק ללוח – הדבק לקובץ ונסה להדפיס.")
        except Exception as e:
            messagebox.showerror("שגיאה", f"כשל בהדפסה: {e}")

    def _show_drawing_details(self, record):
        top = tk.Toplevel(self.root); top.title(f"פרטי ציור - {record.get('שם הקובץ','')}"); top.geometry('900x700'); top.configure(bg=theme.PAGE_BG)
        tk.Label(top, text=f"פרטי ציור: {record.get('שם הקובץ','')}", font=(theme.FONT_FAMILY, 14, 'bold'), bg=theme.PAGE_BG, anchor='e', justify='right').pack(pady=10, fill='x')
        info = tk.LabelFrame(top, text="מידע כללי", bg=theme.PAGE_BG); info.pack(fill='x', padx=12, pady=6)
        base_txt = (
            f"ID: {record.get('id','')}\n"
            f"תאריך יצירה: {record.get('תאריך יצירה','')}\n"
            f"מספר מוצרים: {len(record.get('מוצרים', []))}\n"
            f"סך הכמויות: {record.get('סך כמויות',0)}"
        )
        if 'סוג בד' in record:
            base_txt += f"\nסוג בד: {record.get('סוג בד')}"
        if 'נמען' in record:
            base_txt += f"\nנמען (ספק): {record.get('נמען')}"
        if 'כמות שכבות משוערת' in record:
            base_txt += f"\nכמות שכבות משוערת: {record.get('כמות שכבות משוערת')}"
        # הוספת תצוגת מידות הציור אם קיימות
        if 'רוחב ציור' in record:
            marker_width = record.get('רוחב ציור')
            base_txt += f"\nרוחב ציור: {marker_width:.2f} ס״ם"
        if 'אורך ציור' in record:
            marker_length = record.get('אורך ציור')
            base_txt += f"\nאורך ציור: {marker_length:.2f} ס״ם"
        # הוספת תצוגת שכבות בפועל אם הציור נחתך
        if 'שכבות' in record and record.get('status') == 'נחתך':
            actual_layers = record.get('שכבות')
            base_txt += f"\nשכבות בפועל (נחתך): {actual_layers}"
        # הוספת תצוגת משקל ומטרים אם הציור נחתך
        if record.get('status') == 'נחתך':
            if 'משקל כולל' in record:
                total_weight = record.get('משקל כולל')
                base_txt += f"\nמשקל כולל נגזר: {total_weight:.2f} ק״ג"
            if 'מטרים כוללים' in record:
                total_meters = record.get('מטרים כוללים')
                base_txt += f"\nמטרים כוללים נגזרו: {total_meters:.2f}"
            # חישוב משקל לכל שכבה
            if 'משקל כולל' in record and 'שכבות' in record:
                total_weight = record.get('משקל כולל')
                layers = record.get('שכבות')
                if layers is not None and layers > 0:
                    weight_per_layer = total_weight / layers
                    base_txt += f"\nמשקל לכל שכבה: {weight_per_layer:.2f} ק״ג"
        status_val = record.get('status','')
        base_txt += f"\nסטטוס: {status_val}"
        tk.Label(info, text=base_txt, bg=theme.PAGE_BG, justify='right', anchor='e').pack(fill='x', padx=8, pady=6)
        tk.Label(top, text="פירוט מוצרים ומידות:", font=(theme.FONT_FAMILY, 12, 'bold'), bg=theme.PAGE_BG, anchor='e', justify='right').pack(anchor='e', padx=12, pady=(6, 2), fill='x')
        st = scrolledtext.ScrolledText(top, height=20, font=('Courier New', 10), wrap='word')
        st.pack(fill='both', expand=True, padx=12, pady=4)
        st.tag_configure('rtl', justify='right')
        layers_used = None  # (הוסר חישוב שכבות מתוך "ציורים חוזרים")
        overall_expected = 0
        for product in record.get('מוצרים', []):
            st.insert(tk.END, RLM + f"\n📦 {product.get('שם המוצר','')}\n", 'rtl')
            st.insert(tk.END, RLM + "="*60 + "\n", 'rtl')
            total_prod_q = 0; total_expected_product = 0
            # מיון לוגי של המידות להצגה עקבית
            import re
            def _size_key(size_str: str):
                if not isinstance(size_str, str):
                    return 0, size_str
                cleaned = size_str.lower().replace('m','')
                m = re.match(r"(\d+)", cleaned.strip())
                base = int(m.group(1)) if m else 0
                return base, cleaned
            product_sizes_sorted = sorted(product.get('מידות', []), key=lambda si: _size_key(si.get('מידה','')))
            for size_info in product_sizes_sorted:
                size = size_info.get('מידה',''); quantity = size_info.get('כמות',0); note = size_info.get('הערה',''); total_prod_q += quantity
                line = f"   מידה {size:>8}: {quantity:>8}"
                if layers_used and isinstance(layers_used, int) and layers_used > 0:
                    expected_qty = quantity * layers_used; total_expected_product += expected_qty; overall_expected += expected_qty
                    line += f"  | לאחר גזירה (שכבות {layers_used}): {expected_qty}"
                if note: line += f"  - {note}"
                st.insert(tk.END, RLM + line + "\n", 'rtl')
            st.insert(tk.END, RLM + f"\nסך עבור מוצר זה: {total_prod_q}", 'rtl')
            if total_expected_product:
                st.insert(tk.END, RLM + f" | סך צפוי לאחר גזירה: {total_expected_product}", 'rtl')
            st.insert(tk.END, RLM + "\n" + "-"*60 + "\n", 'rtl')
        if layers_used and overall_expected:
            st.insert(tk.END, RLM + f"\n➡ סך כמות צפויה לאחר גזירה לכל הציור: {overall_expected}\n", 'rtl')
        st.config(state='disabled')
        tk.Button(top, text="סגור", command=top.destroy, bg=theme.MUTED, fg='white', font=(theme.FONT_FAMILY, 11, 'bold'), width=12).pack(pady=10)

    def _delete_selected_drawing_tab(self):
        sel = self.drawings_tree.selection();
        if not sel: return
        vals = self.drawings_tree.item(sel[0], 'values');
        if not vals: return
        rec_id = vals[0]; file_name = vals[1]
        if not messagebox.askyesno("אישור מחיקה", f"למחוק את הציור:\n{file_name}? פעולה זו אינה הפיכה"): return
        deleted = False
        if hasattr(self.data_processor, 'delete_drawing'):
            try: deleted = self.data_processor.delete_drawing(rec_id)
            except Exception: deleted = False
        if not deleted:
            before = len(self.data_processor.drawings_data)
            self.data_processor.drawings_data = [r for r in self.data_processor.drawings_data if str(r.get('id')) != str(rec_id)]
            if len(self.data_processor.drawings_data) < before and hasattr(self.data_processor, 'save_drawings_data'):
                try: self.data_processor.save_drawings_data(); deleted = True
                except Exception: pass
        if deleted:
            self._refresh_drawings_tree(); messagebox.showinfo("הצלחה", "נמחק בהצלחה")
        else:
            messagebox.showerror("שגיאה", "המחיקה נכשלה")


    def _export_drawings_to_excel_tab(self):
        if not self.data_processor.drawings_data:
            messagebox.showwarning("אזהרה", "אין ציורים לייצוא"); return
        file_path = filedialog.asksaveasfilename(title="ייצא ציורים לאקסל", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not file_path: return
        try:
            self.data_processor.export_drawings_to_excel(file_path)
            messagebox.showinfo("הצלחה", f"הציורים יוצאו אל:\n{file_path}")
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _refresh_drawings_filter_options(self):
        """רענון רשימות הסינון"""
        try:
            # רשימת ספקים
            suppliers = set()
            for record in self.data_processor.drawings_data:
                supplier = record.get('נמען', '').strip()
                if supplier:
                    suppliers.add(supplier)
            
            supplier_list = ['הכל'] + sorted(list(suppliers))
            self.drawings_supplier_filter_cb['values'] = supplier_list
            self.drawings_supplier_filter_var.set('הכל')
            
            # רשימת סטטוסים
            status_list = ['הכל', 'טרם נשלח', 'נשלח', 'הוחזר', 'נחתך']
            self.drawings_status_filter_cb['values'] = status_list
            self.drawings_status_filter_var.set('הכל')
        except Exception:
            pass

    def _apply_drawings_filters(self, event=None):
        """החלת סינון על טבלת הציורים"""
        try:
            supplier_filter = self.drawings_supplier_filter_var.get()
            status_filter = self.drawings_status_filter_var.get()
            
            # ניקוי הטבלה
            for item in self.drawings_tree.get_children():
                self.drawings_tree.delete(item)
            
            # הוספת רשומות מסוננות
            filtered_count = 0
            for record in self.data_processor.drawings_data:
                # סינון לפי ספק
                if supplier_filter != 'הכל':
                    record_supplier = record.get('נמען', '').strip()
                    if record_supplier != supplier_filter:
                        continue
                
                # סינון לפי סטטוס
                if status_filter != 'הכל':
                    record_status = record.get('status', '').strip()
                    if record_status != status_filter:
                        continue
                
                # הוספה לטבלה
                products_count = len(record.get('מוצרים', []))
                total_quantity = record.get('סך כמויות', 0)
                
                # הצגת תאריך ללא שעת יצירה
                created_raw = record.get('תאריך יצירה','')
                created_date_only = created_raw.split()[0] if isinstance(created_raw, str) and created_raw else created_raw
                
                sent_flag = record.get('נשלח לספק')
                supplier_name = (record.get('נמען') or '').strip()
                if supplier_name:
                    sent_display = supplier_name
                else:
                    sent_display = 'כן' if sent_flag is True else ('לא' if sent_flag is False else '')
                
                estimated_layers = record.get('כמות שכבות משוערת', '—')
                # שכבות בפועל - מוצג רק לציורים שנחתכו
                status = record.get('status', 'נשלח')
                actual_layers = record.get('שכבות', '') if status == 'נחתך' else '—'
                products_details = self._format_products_details(record.get('מוצרים', []))
                
                self.drawings_tree.insert('', 'end', values=(
                    record.get('id',''),
                    record.get('שם הקובץ',''),
                    created_date_only,
                    products_count,
                    f"{total_quantity:.1f}" if isinstance(total_quantity,(int,float)) else total_quantity,
                    estimated_layers,
                    actual_layers,
                    products_details,
                    sent_display,
                    record.get('status','נשלח'),
                    "📄"
                ))
                filtered_count += 1
            
            # עדכון סטטיסטיקות
            self._update_drawings_stats_filtered(filtered_count, supplier_filter, status_filter)
        except Exception:
            pass

    def _clear_drawings_filters(self):
        """ניקוי כל הסינונים"""
        try:
            self.drawings_supplier_filter_var.set('הכל')
            self.drawings_status_filter_var.set('הכל')
            self._refresh_drawings_tree()
        except Exception:
            pass

    def _update_drawings_stats_filtered(self, filtered_count, supplier_filter, status_filter):
        """עדכון סטטיסטיקות עם מידע על הסינון"""
        try:
            total_drawings = len(self.data_processor.drawings_data)
            total_quantity = sum(r.get('סך כמויות', 0) for r in self.data_processor.drawings_data)
            
            filter_info = ""
            if supplier_filter != 'הכל' or status_filter != 'הכל':
                filter_info = f" | מסונן: {filtered_count} ציורים"
                if supplier_filter != 'הכל':
                    filter_info += f" (ספק: {supplier_filter})"
                if status_filter != 'הכל':
                    filter_info += f" (סטטוס: {status_filter})"
            
            self.drawings_stats_var.set(f"סך הכל: {total_drawings} ציורים | סך כמויות: {total_quantity:.1f}{filter_info}")
        except Exception:
            pass

    # === Area Calculation Tab ===
    def _build_area_calculation_tab(self, container: tk.Widget):
        """בניית טאב חישוב שטח כולל לדגם מידה"""
        wrapper = tk.Frame(container, bg=theme.PAGE_BG)
        wrapper.pack(fill='both', expand=True, padx=10, pady=8)
        
        # כותרת
        tk.Label(wrapper, text="שטח רבוע לדגם מידה", font=(theme.FONT_FAMILY, 14, 'bold'), bg=theme.PAGE_BG, fg=theme.DARK).pack(pady=(0, 15))
        
        # מסגרת הזנת נתונים
        input_frame = ttk.LabelFrame(wrapper, text="הזנת נתונים", padding=15)
        input_frame.pack(fill='x', pady=(0, 15))
        
        # אתחול משתנים וטעינת נתונים
        self._load_area_data_from_file()
        
        # שם דגם - נבחר מקטלוג המוצרים
        tk.Label(input_frame, text="שם הדגם:", font=(theme.FONT_FAMILY, 10, 'bold')).grid(row=0, column=0, sticky='w', padx=5, pady=8)
        self.area_product_name_var = tk.StringVar()
        
        # קבלת רשימת שמות הדגמים מקטלוג המוצרים
        product_names = self._get_unique_product_names_from_catalog()
        self.area_product_combo = ttk.Combobox(input_frame, textvariable=self.area_product_name_var, 
                                             values=product_names, state='readonly', width=25, justify='right')
        self.area_product_combo.grid(row=0, column=1, sticky='w', padx=5, pady=8)
        self.area_product_combo.bind('<<ComboboxSelected>>', self._on_area_product_selected)
        
        # מידה - תתעדכן בהתאם לדגם הנבחר
        tk.Label(input_frame, text="מידה:", font=(theme.FONT_FAMILY, 10, 'bold')).grid(row=0, column=2, sticky='w', padx=5, pady=8)
        self.area_size_var = tk.StringVar()
        self.area_size_combo = ttk.Combobox(input_frame, textvariable=self.area_size_var, 
                                          state='readonly', width=15, justify='right')
        self.area_size_combo.grid(row=0, column=3, sticky='w', padx=5, pady=8)
        
        # מ"ר של הגיזרה
        tk.Label(input_frame, text="מ\"ר של הגיזרה:", font=(theme.FONT_FAMILY, 10, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=8)
        self.area_sqm_var = tk.StringVar()
        area_entry = tk.Entry(input_frame, textvariable=self.area_sqm_var, width=20, justify='center')
        area_entry.grid(row=1, column=1, sticky='w', padx=5, pady=8)
        self._attach_paste_menu(area_entry)
        
        # כפתורי פעולה
        buttons_frame = tk.Frame(input_frame)
        buttons_frame.grid(row=1, column=2, columnspan=2, sticky='w', padx=15, pady=8)
        
        tk.Button(buttons_frame, text="➕ הוסף", command=self._add_area_calculation_entry,
                 bg=theme.SUCCESS, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='left', padx=5)
        tk.Button(buttons_frame, text="🧮 חשב שטח רבוע", command=self._calculate_total_area,
                 bg=theme.PRIMARY, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='left', padx=5)
        tk.Button(buttons_frame, text="🗑️ נקה הכל", command=self._clear_area_calculations,
                 bg=theme.DANGER, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='left', padx=5)
        
        # טבלת נתונים
        table_frame = ttk.LabelFrame(wrapper, text="רשימת מ\"ר", padding=10)
        table_frame.pack(fill='both', expand=True, pady=(0, 15))
        
        # יצירת הטבלה
        cols = ('product_name', 'size', 'total_area')
        self.area_calc_tree = ttk.Treeview(table_frame, columns=cols, show='headings', height=12)
        
        headers = {
            'product_name': 'שם דגם',
            'size': 'מידה', 
            'total_area': 'שטח רבוע'
        }
        
        widths = {
            'product_name': 250,
            'size': 150,
            'total_area': 200
        }
        
        for col in cols:
            self.area_calc_tree.heading(col, text=headers[col])
            self.area_calc_tree.column(col, width=widths[col], anchor='center')
        
        # סרגל גלילה
        area_scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.area_calc_tree.yview)
        self.area_calc_tree.configure(yscroll=area_scrollbar.set)
        self.area_calc_tree.grid(row=0, column=0, sticky='nsew')
        area_scrollbar.grid(row=0, column=1, sticky='ns')
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)
        
        # כפתור מחיקת שורה נבחרת
        tk.Button(table_frame, text="❌ מחק שורה נבחרת", command=self._delete_selected_area_row,
                 bg=theme.WARNING, fg='white', font=(theme.FONT_FAMILY, 9)).grid(row=1, column=0, pady=5, sticky='w')
        
        # טעינת הנתונים השמורים לטבלה
        self._populate_area_data_from_storage()

    def _get_unique_product_names_from_catalog(self):
        """קבלת רשימת שמות דגמים ייחודיים מקטלוג המוצרים"""
        try:
            catalog = getattr(self.data_processor, 'products_catalog', []) or []
            unique_names = list(set(item.get('name', '') for item in catalog if item.get('name', '').strip()))
            return sorted([name for name in unique_names if name])
        except Exception:
            return []

    def _on_area_product_selected(self, event=None):
        """עדכון רשימת המידות בהתאם לדגם הנבחר"""
        try:
            selected_product = self.area_product_name_var.get()
            if not selected_product:
                self.area_size_combo['values'] = []
                return
            
            # חיפוש כל המידות הזמינות לדגם הנבחר
            catalog = getattr(self.data_processor, 'products_catalog', []) or []
            available_sizes = []
            for item in catalog:
                if item.get('name', '') == selected_product:
                    size = item.get('size', '').strip()
                    if size and size not in available_sizes:
                        available_sizes.append(size)
            
            # מיון המידות
            self.area_size_combo['values'] = sorted(available_sizes)
            self.area_size_var.set('')  # איפוס הבחירה הנוכחית
        except Exception:
            self.area_size_combo['values'] = []

    def _add_area_calculation_entry(self):
        """הוספת רשומה חדשה לחישוב שטח"""
        try:
            product_name = self.area_product_name_var.get().strip()
            size = self.area_size_var.get().strip()
            sqm_text = self.area_sqm_var.get().strip()
            
            if not product_name:
                messagebox.showerror("שגיאה", "יש לבחור שם דגם")
                return
            
            if not size:
                messagebox.showerror("שגיאה", "יש לבחור מידה")
                return
            
            if not sqm_text:
                messagebox.showerror("שגיאה", "יש להזין מ\"ר של הגיזרה")
                return
            
            try:
                sqm_per_piece = float(sqm_text.replace(',', '.'))
                if sqm_per_piece <= 0:
                    raise ValueError("ערך חייב להיות חיובי")
            except ValueError:
                messagebox.showerror("שגיאה", "מ\"ר של הגיזרה חייב להיות מספר חיובי")
                return
            
            # בדיקה אם כבר קיימת רשומה עם אותו דגם ומידה
            for item in self.area_calc_tree.get_children():
                values = self.area_calc_tree.item(item, 'values')
                if values[0] == product_name and values[1] == size:
                    messagebox.showwarning("אזהרה", f"כבר קיימת רשומה עבור {product_name} מידה {size}")
                    return
            
            # הוספה לטבלה
            self.area_calc_tree.insert('', 'end', values=(
                product_name,
                size,
                f"{sqm_per_piece:.6f}"
            ))
            
            # שמירת הנתונים
            self._save_area_data_entry(product_name, size, sqm_per_piece)
            
            # איפוס השדות
            self.area_product_name_var.set('')
            self.area_size_var.set('')
            self.area_sqm_var.set('')
            self.area_size_combo['values'] = []
            
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בהוספת רשומה: {str(e)}")

    def _calculate_total_area(self):
        """חישוב השטח הכולל בהתבסס על נתוני הציורים"""
        try:
            if not self.area_calc_tree.get_children():
                messagebox.showwarning("אזהרה", "אין רשומות לחישוב")
                return
            
            total_area = 0.0
            drawings_data = getattr(self.data_processor, 'drawings_data', []) or []
            
            # עבור כל שורה בטבלה
            for item in self.area_calc_tree.get_children():
                values = list(self.area_calc_tree.item(item, 'values'))
                product_name = values[0]
                size = values[1]
                
                # קבלת מ"ר לחתיכה מהמאגר
                sqm_per_piece = self._get_area_data_entry(product_name, size)
                if sqm_per_piece is None:
                    continue
                
                # חיפוש כמות בכל הציורים
                total_quantity = 0
                for drawing in drawings_data:
                    for product in drawing.get('מוצרים', []):
                        if product.get('שם המוצר', '') == product_name:
                            for size_info in product.get('מידות', []):
                                if size_info.get('מידה', '') == size:
                                    total_quantity += size_info.get('כמות', 0)
                
                # חישוב שטח עבור המוצר הזה - רק להצגת הודעת מידע
                product_area = total_quantity * sqm_per_piece
                
                # כעת לא משנים את הערך בטבלה, רק שומרים את מ"ר המקורי
            
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בחישוב: {str(e)}")

    def _delete_selected_area_row(self):
        """מחיקת השורה הנבחרת מהטבלה ומהמאגר"""
        try:
            selected = self.area_calc_tree.selection()
            if not selected:
                messagebox.showwarning("אזהרה", "יש לבחור שורה למחיקה")
                return
            
            # קבלת הנתונים של השורה הנבחרת
            values = self.area_calc_tree.item(selected[0], 'values')
            product_name = values[0]
            size = values[1]
            
            # מחיקה מהמאגר
            if hasattr(self.data_processor, 'area_data'):
                self.data_processor.area_data = [
                    entry for entry in self.data_processor.area_data 
                    if not (entry.get('product_name') == product_name and entry.get('size') == size)
                ]
                self._save_area_data_to_file()
            
            # מחיקה מהטבלה
            self.area_calc_tree.delete(selected[0])
            
            # עדכון התוצאה
            self._calculate_total_area()
            
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה במחיקת שורה: {str(e)}")

    def _clear_area_calculations(self):
        """ניקוי כל החישובים מהטבלה ומהמאגר"""
        try:
            if messagebox.askyesno("אישור", "האם אתה בטוח שברצונך למחוק את כל החישובים?"):
                # ניקוי המאגר
                if hasattr(self.data_processor, 'area_data'):
                    self.data_processor.area_data = []
                    self._save_area_data_to_file()
                
                # ניקוי הטבלה
                for item in self.area_calc_tree.get_children():
                    self.area_calc_tree.delete(item)
                
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בניקוי: {str(e)}")

    def _save_area_data_entry(self, product_name: str, size: str, sqm_per_piece: float):
        """שמירת רשומת מ"ר לדגם/מידה במאגר הנתונים"""
        try:
            # קבלת רשימת הנתונים הקיימת או יצירת חדשה
            if not hasattr(self.data_processor, 'area_data'):
                self.data_processor.area_data = []
            
            # בדיקה אם כבר קיימת רשומה
            for entry in self.data_processor.area_data:
                if entry.get('product_name') == product_name and entry.get('size') == size:
                    entry['sqm_per_piece'] = sqm_per_piece
                    entry['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    self._save_area_data_to_file()
                    return
            
            # הוספת רשומה חדשה
            new_entry = {
                'product_name': product_name,
                'size': size,
                'sqm_per_piece': sqm_per_piece,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            self.data_processor.area_data.append(new_entry)
            self._save_area_data_to_file()
            
        except Exception as e:
            print(f"שגיאה בשמירת נתוני שטח: {e}")

    def _get_area_data_entry(self, product_name: str, size: str) -> float | None:
        """קבלת מ"ר לחתיכה עבור דגם/מידה מהמאגר"""
        try:
            if not hasattr(self.data_processor, 'area_data'):
                return None
            
            for entry in self.data_processor.area_data:
                if entry.get('product_name') == product_name and entry.get('size') == size:
                    return float(entry.get('sqm_per_piece', 0))
            return None
        except Exception:
            return None

    def _save_area_data_to_file(self):
        """שמירת נתוני השטח לקובץ"""
        try:
            area_data_file = os.path.join(os.getcwd(), 'area_data.json')
            with open(area_data_file, 'w', encoding='utf-8') as f:
                json.dump(self.data_processor.area_data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"שגיאה בשמירת קובץ נתוני שטח: {e}")

    def _load_area_data_from_file(self):
        """טעינת נתוני השטח מהקובץ"""
        try:
            area_data_file = os.path.join(os.getcwd(), 'area_data.json')
            if os.path.exists(area_data_file):
                with open(area_data_file, 'r', encoding='utf-8') as f:
                    self.data_processor.area_data = json.load(f)
            else:
                self.data_processor.area_data = []
        except Exception as e:
            print(f"שגיאה בטעינת נתוני שטח: {e}")
            self.data_processor.area_data = []

    def _populate_area_data_from_storage(self):
        """מילוי הטבלה מהנתונים השמורים"""
        try:
            if not hasattr(self.data_processor, 'area_data'):
                return
            
            # ניקוי הטבלה הנוכחית
            for item in self.area_calc_tree.get_children():
                self.area_calc_tree.delete(item)
            
            # הוספת הנתונים השמורים
            for entry in self.data_processor.area_data:
                product_name = entry.get('product_name', '')
                size = entry.get('size', '')
                sqm_per_piece = entry.get('sqm_per_piece', 0)
                if product_name and size:
                    self.area_calc_tree.insert('', 'end', values=(
                        product_name,
                        size,
                        f"{sqm_per_piece:.6f}"
                    ))
        except Exception as e:
            print(f"שגיאה במילוי טבלת שטח: {e}")
