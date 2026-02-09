import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import json
import os
import calendar as _cal

class ShipmentsTabMixin:
    """Mixin לטאב 'הובלות' המציג שורות אריזה מכל הקליטות והתעודות.

    כל רשומת קליטה / תעודת משלוח יכולה להכיל רשימת packages (package_type, quantity).
    הטאב מציג פירוק שטוח של כל החבילות עם מקורן.
    """

    def _create_shipments_tab(self):
        tab = tk.Frame(self.notebook, bg='#f7f9fa')
        self.notebook.add(tab, text="הובלות")

        # פנימי: Notebook לתת-טאבים (סיכום / מובילים)
        inner_nb = ttk.Notebook(tab)
        inner_nb.pack(fill='both', expand=True, padx=5, pady=5)

        # --- עמוד סיכום הובלות ---
        shipments_page = tk.Frame(inner_nb, bg='#f7f9fa')
        inner_nb.add(shipments_page, text='סיכום הובלות')
        tk.Label(shipments_page, text="הובלות - סיכום הובלה מקליטות ותעודות", font=('Arial',14,'bold'), bg='#f7f9fa', fg='#2c3e50').pack(pady=6)

        # פריים סינון
        filter_frame = tk.LabelFrame(shipments_page, text='סינון', bg='#f7f9fa', font=('Arial',9,'bold'))
        filter_frame.pack(fill='x', padx=8, pady=(0,4))
        
        # שורה ראשונה: מוביל ומצב תשלום
        filter_row1 = tk.Frame(filter_frame, bg='#f7f9fa')
        filter_row1.pack(fill='x', padx=4, pady=4)
        
        tk.Label(filter_row1, text='מוביל:', bg='#f7f9fa', font=('Arial',9)).pack(side='right', padx=(4,2))
        self.shipments_filter_driver_var = tk.StringVar(value='הכל')
        self.shipments_filter_driver_combo = ttk.Combobox(filter_row1, textvariable=self.shipments_filter_driver_var, width=15, state='readonly')
        self.shipments_filter_driver_combo['values'] = ['הכל']
        self.shipments_filter_driver_combo.pack(side='right', padx=2)
        self.shipments_filter_driver_combo.bind('<<ComboboxSelected>>', lambda e: self._refresh_shipments_table())
        
        tk.Label(filter_row1, text='מצב תשלום:', bg='#f7f9fa', font=('Arial',9)).pack(side='right', padx=(10,2))
        self.shipments_filter_paid_var = tk.StringVar(value='הכל')
        filter_paid_combo = ttk.Combobox(filter_row1, textvariable=self.shipments_filter_paid_var, width=12, state='readonly')
        filter_paid_combo['values'] = ['הכל', 'רק שולם', 'רק לא שולם']
        filter_paid_combo.pack(side='right', padx=2)
        filter_paid_combo.bind('<<ComboboxSelected>>', lambda e: self._refresh_shipments_table())

        tk.Label(filter_row1, text='סוג הובלה:', bg='#f7f9fa', font=('Arial',9)).pack(side='right', padx=(10,2))
        self.shipments_filter_kind_var = tk.StringVar(value='הכל')
        self.shipments_filter_kind_combo = ttk.Combobox(filter_row1, textvariable=self.shipments_filter_kind_var, width=14, state='readonly')
        self.shipments_filter_kind_combo['values'] = ['הכל', 'קליטה', 'הובלה', 'קליטת בדים', 'שליחת בדים']
        self.shipments_filter_kind_combo.pack(side='right', padx=2)
        self.shipments_filter_kind_combo.bind('<<ComboboxSelected>>', lambda e: self._refresh_shipments_table())

        tk.Label(filter_row1, text='פריט הובלה:', bg='#f7f9fa', font=('Arial',9)).pack(side='right', padx=(10,2))
        self.shipments_filter_package_type_var = tk.StringVar(value='הכל')
        self.shipments_filter_package_type_combo = ttk.Combobox(filter_row1, textvariable=self.shipments_filter_package_type_var, width=14, state='readonly')
        self.shipments_filter_package_type_combo['values'] = ['הכל']
        self.shipments_filter_package_type_combo.pack(side='right', padx=2)
        self.shipments_filter_package_type_combo.bind('<<ComboboxSelected>>', lambda e: self._refresh_shipments_table())
        
        # שורה שנייה: תאריכים
        filter_row2 = tk.Frame(filter_frame, bg='#f7f9fa')
        filter_row2.pack(fill='x', padx=4, pady=4)
        
        tk.Label(filter_row2, text='תאריך מ:', bg='#f7f9fa', font=('Arial',9)).pack(side='right', padx=(4,2))
        self.shipments_filter_date_from_var = tk.StringVar()
        date_from_entry = tk.Entry(filter_row2, textvariable=self.shipments_filter_date_from_var, width=12, font=('Arial',9))
        date_from_entry.pack(side='right', padx=2)
        date_from_entry.bind('<KeyRelease>', lambda e: self._refresh_shipments_table())
        tk.Button(filter_row2, text='📅', width=2, command=lambda: self._open_date_picker(date_from_entry, self.shipments_filter_date_from_var)).pack(side='right', padx=2)
        
        tk.Label(filter_row2, text='עד:', bg='#f7f9fa', font=('Arial',9)).pack(side='right', padx=(10,2))
        self.shipments_filter_date_to_var = tk.StringVar()
        date_to_entry = tk.Entry(filter_row2, textvariable=self.shipments_filter_date_to_var, width=12, font=('Arial',9))
        date_to_entry.pack(side='right', padx=2)
        date_to_entry.bind('<KeyRelease>', lambda e: self._refresh_shipments_table())
        tk.Button(filter_row2, text='📅', width=2, command=lambda: self._open_date_picker(date_to_entry, self.shipments_filter_date_to_var)).pack(side='right', padx=2)
        
        tk.Button(filter_row2, text='🔄 נקה סינון', command=self._clear_shipments_filter, bg='#95a5a6', fg='white', font=('Arial',8)).pack(side='right', padx=(10,2))

        toolbar = tk.Frame(shipments_page, bg='#f7f9fa')
        toolbar.pack(fill='x', padx=8, pady=(0,4))
        
        # כפתורי פעולה בצד ימין
        tk.Button(toolbar, text="🔄 רענן", command=self._refresh_shipments_table, bg='#3498db', fg='white').pack(side='right', padx=4)
        tk.Button(toolbar, text='🗑 מחק שורה נבחרת', command=self._delete_selected_shipment_row, bg='#c0392b', fg='white').pack(side='right', padx=4)
        tk.Button(toolbar, text='✓ סמן כשולם', command=self._mark_shipment_as_paid, bg='#27ae60', fg='white').pack(side='right', padx=4)
        tk.Button(toolbar, text='✗ בטל שולם', command=self._mark_shipment_as_unpaid, bg='#e67e22', fg='white').pack(side='right', padx=4)
        
        # בקרי סידור בצד שמאל
        tk.Label(toolbar, text='סידור לפי:', bg='#f7f9fa', font=('Arial',9)).pack(side='left', padx=(4,2))
        self.shipments_sort_var = tk.StringVar(value='date_desc')
        sort_options = [
            ('תאריך (חדש לישן)', 'date_desc'),
            ('תאריך (ישן לחדש)', 'date_asc'),
            ('פריט הובלה', 'package_type'),
            ('סוג', 'kind'),
            ('מוביל', 'driver')
        ]
        sort_combo = ttk.Combobox(toolbar, textvariable=self.shipments_sort_var, width=18, state='readonly')
        sort_combo['values'] = [opt[0] for opt in sort_options]
        sort_combo.pack(side='left', padx=2)
        # מיפוי בין תצוגה לערך
        self._sort_display_to_value = {opt[0]: opt[1] for opt in sort_options}
        self._sort_value_to_display = {opt[1]: opt[0] for opt in sort_options}
        sort_combo.current(0)
        sort_combo.bind('<<ComboboxSelected>>', lambda e: self._refresh_shipments_table())

        # Container for tree + scrollbar
        tree_container = tk.Frame(shipments_page, bg='#f7f9fa')
        tree_container.pack(fill='both', expand=True, padx=10, pady=6)

        columns = ('id','kind','date','package_type','quantity','driver','paid')
        self.shipments_tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=17)
        headers = {
            'id': 'מספר תעודה',
            'kind': 'סוג',
            'date': 'תאריך',
            'package_type': 'פריט הובלה',
            'quantity': 'כמות',
            'driver': 'שם המוביל',
            'paid': 'האם שולם'
        }
        widths = {'id':110,'kind':90,'date':110,'package_type':140,'quantity':80,'driver':110,'paid':80}
        for c in columns:
            self.shipments_tree.heading(c, text=headers[c])
            self.shipments_tree.column(c, width=widths[c], anchor='center')
        vs = ttk.Scrollbar(tree_container, orient='vertical', command=self.shipments_tree.yview)
        self.shipments_tree.configure(yscroll=vs.set)
        self.shipments_tree.pack(side='left', fill='both', expand=True)
        vs.pack(side='left', fill='y')

        # Summary label below tree
        self.shipments_summary_var = tk.StringVar(value='')
        self.shipments_summary_label = tk.Label(shipments_page, textvariable=self.shipments_summary_var,
            font=('Arial', 11, 'bold'), bg='#f7f9fa', fg='#2c3e50', anchor='e')
        self.shipments_summary_label.pack(fill='x', padx=15, pady=(2, 6))

        self._refresh_shipments_table()

        # --- עמוד מובילים ---
        drivers_page = tk.Frame(inner_nb, bg='#f7f9fa')
        inner_nb.add(drivers_page, text='מובילים')

        form = tk.LabelFrame(drivers_page, text='הוספת / עדכון מוביל', bg='#f7f9fa')
        form.pack(fill='x', padx=10, pady=8)
        tk.Label(form, text='שם:', bg='#f7f9fa').grid(row=0, column=0, padx=4, pady=4, sticky='e')
        self.driver_name_var = tk.StringVar()
        tk.Entry(form, textvariable=self.driver_name_var, width=30).grid(row=0, column=1, padx=4, pady=4)
        tk.Label(form, text='טלפון:', bg='#f7f9fa').grid(row=0, column=2, padx=4, pady=4, sticky='e')
        self.driver_phone_var = tk.StringVar()
        tk.Entry(form, textvariable=self.driver_phone_var, width=20).grid(row=0, column=3, padx=4, pady=4)
        tk.Button(form, text='➕ שמור/עדכן', command=self._add_or_update_driver, bg='#27ae60', fg='white').grid(row=0, column=4, padx=6, pady=4)
        tk.Button(form, text='🗑 מחק נבחר', command=self._delete_driver, bg='#c0392b', fg='white').grid(row=0, column=5, padx=6, pady=4)

        drivers_columns = ('name','phone')
        self.drivers_tree = ttk.Treeview(drivers_page, columns=drivers_columns, show='headings', height=15)
        self.drivers_tree.heading('name', text='שם מוביל')
        self.drivers_tree.heading('phone', text='טלפון')
        self.drivers_tree.column('name', width=180, anchor='center')
        self.drivers_tree.column('phone', width=140, anchor='center')
        dscroll = ttk.Scrollbar(drivers_page, orient='vertical', command=self.drivers_tree.yview)
        self.drivers_tree.configure(yscroll=dscroll.set)
        self.drivers_tree.pack(side='left', fill='both', expand=True, padx=(10,0), pady=6)
        dscroll.pack(side='left', fill='y', pady=6)

        self._load_drivers()
        self._refresh_drivers_table()
        self._update_shipments_filter_drivers()

        # --- עמוד דו"ח חישוב הובלות לתשלום ---
        payment_report_page = tk.Frame(inner_nb, bg='#f7f9fa')
        inner_nb.add(payment_report_page, text='דו"ח חישוב הובלות לתשלום')
        
        tk.Label(payment_report_page, text="דו\"ח חישוב הובלות לתשלום", font=('Arial',14,'bold'), bg='#f7f9fa', fg='#2c3e50').pack(pady=10)
        
        # פריים לבחירת פרמטרים
        params_frame = tk.LabelFrame(payment_report_page, text='בחירת פרמטרים', bg='#f7f9fa', font=('Arial',10,'bold'))
        params_frame.pack(fill='x', padx=20, pady=10)
        
        # שורה 1: בחירת מוביל
        row1 = tk.Frame(params_frame, bg='#f7f9fa')
        row1.pack(fill='x', padx=10, pady=8)
        tk.Label(row1, text='מוביל:', bg='#f7f9fa', font=('Arial',10)).pack(side='right', padx=5)
        self.payment_driver_var = tk.StringVar()
        self.payment_driver_combo = ttk.Combobox(row1, textvariable=self.payment_driver_var, width=25, state='readonly')
        self.payment_driver_combo.pack(side='right', padx=5)
        
        # שורה 2: תאריך התחלה
        row2 = tk.Frame(params_frame, bg='#f7f9fa')
        row2.pack(fill='x', padx=10, pady=8)
        tk.Label(row2, text='תאריך התחלה:', bg='#f7f9fa', font=('Arial',10)).pack(side='right', padx=5)
        self.payment_start_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        payment_start_entry = tk.Entry(row2, textvariable=self.payment_start_date_var, width=15, font=('Arial',10))
        payment_start_entry.pack(side='right', padx=5)
        tk.Button(row2, text='📅', width=3, command=lambda: self._open_date_picker(payment_start_entry, self.payment_start_date_var)).pack(side='right', padx=2)
        
        # שורה 3: תאריך סוף
        row3 = tk.Frame(params_frame, bg='#f7f9fa')
        row3.pack(fill='x', padx=10, pady=8)
        tk.Label(row3, text='תאריך סוף:', bg='#f7f9fa', font=('Arial',10)).pack(side='right', padx=5)
        self.payment_end_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        payment_end_entry = tk.Entry(row3, textvariable=self.payment_end_date_var, width=15, font=('Arial',10))
        payment_end_entry.pack(side='right', padx=5)
        tk.Button(row3, text='📅', width=3, command=lambda: self._open_date_picker(payment_end_entry, self.payment_end_date_var)).pack(side='right', padx=2)
        
        # שורה 4: סינון לפי מצב תשלום
        row4 = tk.Frame(params_frame, bg='#f7f9fa')
        row4.pack(fill='x', padx=10, pady=8)
        tk.Label(row4, text='סינון לפי מצב תשלום:', bg='#f7f9fa', font=('Arial',10)).pack(side='right', padx=5)
        self.payment_filter_var = tk.StringVar(value='לא שולם')
        payment_filter_combo = ttk.Combobox(row4, textvariable=self.payment_filter_var, width=22, state='readonly')
        payment_filter_combo['values'] = ['הכל', 'רק לא שולם', 'רק שולם']
        payment_filter_combo.current(1)  # ברירת מחדל: רק לא שולם
        payment_filter_combo.pack(side='right', padx=5)
        
        # כפתורי חישוב ושמירה
        btn_frame = tk.Frame(params_frame, bg='#f7f9fa')
        btn_frame.pack(fill='x', padx=10, pady=10)
        tk.Button(btn_frame, text='📊 חשב דו"ח', command=self._calculate_payment_report, bg='#2ecc71', fg='white', font=('Arial',11,'bold'), width=20).pack(side='right', padx=5)
        self.save_report_btn = tk.Button(btn_frame, text='💾 שמור דוח', command=self._save_shipment_report, bg='#3498db', fg='white', font=('Arial',11,'bold'), width=20, state='disabled')
        self.save_report_btn.pack(side='right', padx=5)
        
        # פריים לתוצאות
        results_frame = tk.LabelFrame(payment_report_page, text='תוצאות', bg='#f7f9fa', font=('Arial',10,'bold'))
        results_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # תצוגת תוצאות
        self.payment_results_text = tk.Text(results_frame, height=15, width=60, font=('Arial',12), bg='white', fg='#2c3e50', state='disabled')
        self.payment_results_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        # עדכון רשימת מובילים בטאב החדש
        self._update_payment_drivers_list()

        # --- עמוד מחירון הובלות ---
        pricing_page = tk.Frame(inner_nb, bg='#f7f9fa')
        inner_nb.add(pricing_page, text='מחירון')
        
        tk.Label(pricing_page, text="מחירון הובלות", font=('Arial',14,'bold'), bg='#f7f9fa', fg='#2c3e50').pack(pady=10)
        
        # פריים לעדכון מחירים
        update_frame = tk.LabelFrame(pricing_page, text='עדכון מחירים', bg='#f7f9fa', font=('Arial',10,'bold'))
        update_frame.pack(fill='x', padx=20, pady=10)
        
        # שורה 1: בחירת מוביל
        pricing_row1 = tk.Frame(update_frame, bg='#f7f9fa')
        pricing_row1.pack(fill='x', padx=10, pady=8)
        tk.Label(pricing_row1, text='בחר מוביל:', bg='#f7f9fa', font=('Arial',10)).pack(side='right', padx=5)
        self.pricing_driver_var = tk.StringVar()
        self.pricing_driver_combo = ttk.Combobox(pricing_row1, textvariable=self.pricing_driver_var, width=25, state='readonly')
        self.pricing_driver_combo.pack(side='right', padx=5)
        self.pricing_driver_combo.bind('<<ComboboxSelected>>', self._load_driver_pricing)
        
        # שורה 2: מחיר לבד
        pricing_row2 = tk.Frame(update_frame, bg='#f7f9fa')
        pricing_row2.pack(fill='x', padx=10, pady=8)
        tk.Label(pricing_row2, text='מחיר לבד (₪):', bg='#f7f9fa', font=('Arial',10)).pack(side='right', padx=5)
        self.pricing_fabric_var = tk.StringVar(value='0')
        tk.Entry(pricing_row2, textvariable=self.pricing_fabric_var, width=15, font=('Arial',10)).pack(side='right', padx=5)
        
        # שורה 3: מחיר לשק
        pricing_row3 = tk.Frame(update_frame, bg='#f7f9fa')
        pricing_row3.pack(fill='x', padx=10, pady=8)
        tk.Label(pricing_row3, text='מחיר לשק (₪):', bg='#f7f9fa', font=('Arial',10)).pack(side='right', padx=5)
        self.pricing_bag_var = tk.StringVar(value='0')
        tk.Entry(pricing_row3, textvariable=self.pricing_bag_var, width=15, font=('Arial',10)).pack(side='right', padx=5)
        
        # שורה 4: מחיר לשקית קטנה
        pricing_row4 = tk.Frame(update_frame, bg='#f7f9fa')
        pricing_row4.pack(fill='x', padx=10, pady=8)
        tk.Label(pricing_row4, text='מחיר לשקית קטנה (₪):', bg='#f7f9fa', font=('Arial',10)).pack(side='right', padx=5)
        self.pricing_small_bag_var = tk.StringVar(value='0')
        tk.Entry(pricing_row4, textvariable=self.pricing_small_bag_var, width=15, font=('Arial',10)).pack(side='right', padx=5)
        
        # כפתור שמירה
        pricing_btn_frame = tk.Frame(update_frame, bg='#f7f9fa')
        pricing_btn_frame.pack(fill='x', padx=10, pady=10)
        tk.Button(pricing_btn_frame, text='💾 שמור מחירים', command=self._save_driver_pricing, bg='#3498db', fg='white', font=('Arial',11,'bold'), width=20).pack()
        
        # טבלת מחירים קיימת
        pricing_list_frame = tk.LabelFrame(pricing_page, text='מחירון נוכחי', bg='#f7f9fa', font=('Arial',10,'bold'))
        pricing_list_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        pricing_columns = ('driver', 'fabric', 'bag', 'small_bag')
        self.pricing_tree = ttk.Treeview(pricing_list_frame, columns=pricing_columns, show='headings', height=12)
        self.pricing_tree.heading('driver', text='שם המוביל')
        self.pricing_tree.heading('fabric', text='מחיר לבד (₪)')
        self.pricing_tree.heading('bag', text='מחיר לשק (₪)')
        self.pricing_tree.heading('small_bag', text='מחיר לשקית קטנה (₪)')
        self.pricing_tree.column('driver', width=180, anchor='center')
        self.pricing_tree.column('fabric', width=120, anchor='center')
        self.pricing_tree.column('bag', width=120, anchor='center')
        self.pricing_tree.column('small_bag', width=150, anchor='center')
        
        pricing_scroll = ttk.Scrollbar(pricing_list_frame, orient='vertical', command=self.pricing_tree.yview)
        self.pricing_tree.configure(yscroll=pricing_scroll.set)
        self.pricing_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        pricing_scroll.pack(side='left', fill='y', pady=10)
        
        # עדכון רשימת מובילים ומחירון
        self._update_pricing_drivers_list()
        self._refresh_pricing_table()

        # --- עמוד היסטוריית דוחות ---
        reports_history_page = tk.Frame(inner_nb, bg='#f7f9fa')
        inner_nb.add(reports_history_page, text='היסטוריית דוחות')
        
        tk.Label(reports_history_page, text="היסטוריית דוחות הובלות", font=('Arial',14,'bold'), bg='#f7f9fa', fg='#2c3e50').pack(pady=10)
        
        # כפתור רענון
        toolbar_reports = tk.Frame(reports_history_page, bg='#f7f9fa')
        toolbar_reports.pack(fill='x', padx=8, pady=(0,4))
        tk.Button(toolbar_reports, text="🔄 רענן", command=self._refresh_reports_history, bg='#3498db', fg='white').pack(side='right', padx=4)
        
        # טבלת דוחות
        reports_columns = ('report_id', 'created_at', 'driver', 'period', 'total_cost', 'open_excel')
        self.reports_history_tree = ttk.Treeview(reports_history_page, columns=reports_columns, show='headings', height=15)
        self.reports_history_tree.heading('report_id', text='מספר דוח')
        self.reports_history_tree.heading('created_at', text='תאריך יצירה')
        self.reports_history_tree.heading('driver', text='מוביל')
        self.reports_history_tree.heading('period', text='תקופה')
        self.reports_history_tree.heading('total_cost', text='סכום לתשלום (₪)')
        self.reports_history_tree.heading('open_excel', text='פתיחה')
        self.reports_history_tree.column('report_id', width=80, anchor='center')
        self.reports_history_tree.column('created_at', width=140, anchor='center')
        self.reports_history_tree.column('driver', width=120, anchor='center')
        self.reports_history_tree.column('period', width=180, anchor='center')
        self.reports_history_tree.column('total_cost', width=120, anchor='center')
        self.reports_history_tree.column('open_excel', width=100, anchor='center')
        
        # דאבל קליק לפתיחת אקסל
        self.reports_history_tree.bind('<Double-Button-1>', self._on_report_double_click)
        
        reports_scroll = ttk.Scrollbar(reports_history_page, orient='vertical', command=self.reports_history_tree.yview)
        self.reports_history_tree.configure(yscroll=reports_scroll.set)
        self.reports_history_tree.pack(side='left', fill='both', expand=True, padx=(10,0), pady=6)
        reports_scroll.pack(side='left', fill='y', pady=6)
        
        # טעינת היסטוריה
        self._refresh_reports_history()
        
        # אתחול משתנה לדוח נוכחי
        self._current_report_data = None

    # ---- Drivers management ----
    def _drivers_file_path(self):
        return os.path.join(os.getcwd(), 'drivers.json')

    def _load_drivers(self):
        self._drivers = []
        try:
            path = self._drivers_file_path()
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        self._drivers = [d for d in data if isinstance(d, dict) and 'name' in d]
        except Exception:
            self._drivers = []

    def _save_drivers(self):
        try:
            path = self._drivers_file_path()
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(self._drivers, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror('שגיאה', f'שגיאה בשמירת מובילים: {e}')

    def _refresh_drivers_table(self):
        if not hasattr(self, 'drivers_tree'):
            return
        for iid in self.drivers_tree.get_children():
            self.drivers_tree.delete(iid)
        for d in self._drivers:
            self.drivers_tree.insert('', 'end', values=(d.get('name',''), d.get('phone','')))
    
    def _update_shipments_filter_drivers(self):
        """עדכון רשימת המובילים בסינון."""
        try:
            if not hasattr(self, 'shipments_filter_driver_combo'):
                return
            driver_names = ['הכל'] + [d.get('name', '') for d in self._drivers if d.get('name')]
            current = self.shipments_filter_driver_var.get()
            self.shipments_filter_driver_combo['values'] = driver_names
            if current in driver_names:
                self.shipments_filter_driver_var.set(current)
            else:
                self.shipments_filter_driver_var.set('הכל')
        except Exception:
            pass
    
    def _clear_shipments_filter(self):
        """ניקוי כל הסינונים."""
        try:
            self.shipments_filter_driver_var.set('הכל')
            self.shipments_filter_paid_var.set('הכל')
            self.shipments_filter_kind_var.set('הכל')
            self.shipments_filter_package_type_var.set('הכל')
            self.shipments_filter_date_from_var.set('')
            self.shipments_filter_date_to_var.set('')
            self._refresh_shipments_table()
        except Exception:
            pass

    def _add_or_update_driver(self):
        name = (self.driver_name_var.get() or '').strip()
        phone = (self.driver_phone_var.get() or '').strip()
        if not name:
            messagebox.showwarning('חסר שם', 'נא להזין שם מוביל')
            return
        # עדכון אם קיים
        for d in self._drivers:
            if d.get('name') == name:
                d['phone'] = phone
                break
        else:
            self._drivers.append({'name': name, 'phone': phone})
        self._save_drivers()
        self._refresh_drivers_table()
        self._update_payment_drivers_list()  # עדכון גם בטאב דו"ח תשלום
        self._update_pricing_drivers_list()  # עדכון גם בטאב מחירון
        self._refresh_pricing_table()  # עדכון טבלת מחירון
        self._update_shipments_filter_drivers()  # עדכון גם בסינון הובלות
        self.driver_name_var.set('')
        self.driver_phone_var.set('')

    def _delete_driver(self):
        sel = self.drivers_tree.selection()
        if not sel:
            return
        values = self.drivers_tree.item(sel[0], 'values')
        if not values:
            return
        name = values[0]
        self._drivers = [d for d in self._drivers if d.get('name') != name]
        self._save_drivers()
        self._refresh_drivers_table()
        self._update_payment_drivers_list()  # עדכון גם בטאב דו"ח תשלום
        self._update_pricing_drivers_list()  # עדכון גם בטאב מחירון
        self._refresh_pricing_table()  # עדכון טבלת מחירון
        self._update_shipments_filter_drivers()  # עדכון גם בסינון הובלות

    # ---- Data build ----
    def _refresh_shipments_table(self):
        """רענון טבלת ההובלות ע"י איסוף כל ה-packages מכל הרשומות."""
        try:
            # לוודא טעינה עדכנית מהדיסק
            if hasattr(self.data_processor, 'refresh_supplier_receipts'):
                self.data_processor.refresh_supplier_receipts()
            # טען גם קליטות בדים ושליחות בדים
            try:
                if hasattr(self.data_processor, 'refresh_fabrics_intakes'):
                    self.data_processor.refresh_fabrics_intakes()
            except Exception:
                pass
            try:
                if hasattr(self.data_processor, 'refresh_fabrics_shipments'):
                    self.data_processor.refresh_fabrics_shipments()
            except Exception:
                pass
            supplier_intakes = getattr(self.data_processor, 'supplier_intakes', [])
            delivery_notes = getattr(self.data_processor, 'delivery_notes', [])
            fabrics_intakes = getattr(self.data_processor, 'fabrics_intakes', [])
            fabrics_shipments = getattr(self.data_processor, 'fabrics_shipments', [])
            # קריאה לקובץ מורשת ישן במידת הצורך (תאימות לאחור)
            legacy = []
            try:
                legacy_path = getattr(self.data_processor, 'supplier_receipts_file', None)
                if legacy_path and os.path.exists(legacy_path):
                    # נשתמש בפונקציה הפנימית אם קיימת
                    if hasattr(self.data_processor, '_load_json_list'):
                        legacy = self.data_processor._load_json_list(legacy_path) or []
                    else:
                        with open(legacy_path, 'r', encoding='utf-8') as f:
                            import json as _json
                            legacy = _json.load(f) or []
            except Exception:
                legacy = []
            rows = []
            def collect(source_list, receipt_kind):
                for rec in source_list:
                    rec_id = rec.get('id')
                    date_str = rec.get('date') or ''
                    # validate / normalize date for sorting
                    try:
                        sort_dt = datetime.strptime(date_str, '%Y-%m-%d')
                    except Exception:
                        try:
                            sort_dt = datetime.strptime(date_str[:10], '%Y-%m-%d')
                        except Exception:
                            sort_dt = datetime.min
                    for idx, pkg in enumerate(rec.get('packages', []) or []):
                        rows.append({
                            'rec_id': rec_id,
                            'receipt_kind': receipt_kind,
                            'kind': 'קליטה' if receipt_kind == 'supplier_intake' else ('הובלה' if receipt_kind == 'delivery_note' else ('קליטת בדים' if receipt_kind == 'fabrics_intake' else ('שליחת בדים' if receipt_kind == 'fabrics_shipment' else receipt_kind))),
                            'date': date_str,
                            'sort_dt': sort_dt,
                            'pkg_index': idx,
                            'package_type': pkg.get('package_type',''),
                            'quantity': pkg.get('quantity',''),
                            'driver': pkg.get('driver',''),
                            'paid': pkg.get('paid', False)
                        })
            collect(supplier_intakes, 'supplier_intake')
            collect(delivery_notes, 'delivery_note')
            collect(fabrics_intakes, 'fabrics_intake')
            collect(fabrics_shipments, 'fabrics_shipment')
            # הוספת נתוני מורשת שאינם קיימים כבר ברשימות החדשות
            try:
                existing_keys = {( 'supplier_intake', r.get('id') ) for r in supplier_intakes}
                existing_keys |= {( 'delivery_note', r.get('id') ) for r in delivery_notes}
                for rec in legacy or []:
                    rk = rec.get('receipt_kind') or 'supplier_intake'
                    key = (rk, rec.get('id'))
                    if key in existing_keys:
                        continue
                    collect([rec], rk)
            except Exception:
                pass
            # עדכון רשימת פריטי הובלה לסינון
            if hasattr(self, 'shipments_filter_package_type_combo'):
                pkg_types = sorted(set(r.get('package_type', '') for r in rows if r.get('package_type')))
                current_pt = getattr(self, 'shipments_filter_package_type_var', None)
                if current_pt:
                    cur_val = current_pt.get()
                    self.shipments_filter_package_type_combo['values'] = ['הכל'] + pkg_types
                    if cur_val in ['הכל'] + pkg_types:
                        current_pt.set(cur_val)
                    else:
                        current_pt.set('הכל')
            # מיון לפי בחירת המשתמש
            sort_mode = getattr(self, 'shipments_sort_var', None)
            if sort_mode:
                sort_value = self._sort_display_to_value.get(sort_mode.get(), 'date_desc')
            else:
                sort_value = 'date_desc'
            
            if sort_value == 'date_desc':
                rows.sort(key=lambda r: (r['sort_dt'], r['rec_id']), reverse=True)
            elif sort_value == 'date_asc':
                rows.sort(key=lambda r: (r['sort_dt'], r['rec_id']), reverse=False)
            elif sort_value == 'package_type':
                rows.sort(key=lambda r: (r['package_type'], r['sort_dt']), reverse=False)
            elif sort_value == 'kind':
                rows.sort(key=lambda r: (r['kind'], r['sort_dt']), reverse=False)
            elif sort_value == 'driver':
                rows.sort(key=lambda r: (r['driver'], r['sort_dt']), reverse=False)
            else:
                rows.sort(key=lambda r: (r['sort_dt'], r['rec_id']), reverse=True)
            
            # סינון לפי קריטריונים
            filtered_rows = []
            filter_driver = getattr(self, 'shipments_filter_driver_var', None)
            filter_paid = getattr(self, 'shipments_filter_paid_var', None)
            filter_kind = getattr(self, 'shipments_filter_kind_var', None)
            filter_package_type = getattr(self, 'shipments_filter_package_type_var', None)
            filter_date_from = getattr(self, 'shipments_filter_date_from_var', None)
            filter_date_to = getattr(self, 'shipments_filter_date_to_var', None)
            
            for r in rows:
                # סינון מוביל
                if filter_driver and filter_driver.get() != 'הכל':
                    if r.get('driver', '') != filter_driver.get():
                        continue
                
                # סינון סוג הובלה
                if filter_kind and filter_kind.get() != 'הכל':
                    if r.get('kind', '') != filter_kind.get():
                        continue
                
                # סינון פריט הובלה
                if filter_package_type and filter_package_type.get() != 'הכל':
                    if r.get('package_type', '') != filter_package_type.get():
                        continue
                
                # סינון מצב תשלום
                if filter_paid:
                    paid_filter_val = filter_paid.get()
                    if paid_filter_val == 'רק שולם' and not r.get('paid', False):
                        continue
                    elif paid_filter_val == 'רק לא שולם' and r.get('paid', False):
                        continue
                
                # סינון תאריך מ-
                if filter_date_from and filter_date_from.get():
                    try:
                        date_from = datetime.strptime(filter_date_from.get(), '%Y-%m-%d')
                        if r['sort_dt'] < date_from:
                            continue
                    except Exception:
                        pass
                
                # סינון תאריך עד
                if filter_date_to and filter_date_to.get():
                    try:
                        date_to = datetime.strptime(filter_date_to.get(), '%Y-%m-%d')
                        if r['sort_dt'] > date_to:
                            continue
                    except Exception:
                        pass
                
                filtered_rows.append(r)
            
            rows = filtered_rows
        except Exception:
            rows = []
        if hasattr(self, 'shipments_tree'):
            # איפוס מפת-מטא של שורות
            self._shipments_row_meta = {}
            for iid in self.shipments_tree.get_children():
                self.shipments_tree.delete(iid)
            for r in rows:
                paid_display = '✓' if r.get('paid') else ''
                iid = self.shipments_tree.insert('', 'end', values=(r['rec_id'], r['kind'], r['date'], r['package_type'], r['quantity'], r.get('driver',''), paid_display))
                # שמירת מטא כדי לאפשר מחיקה מדויקת של פריט הובלה
                self._shipments_row_meta[iid] = {
                    'rec_id': r['rec_id'],
                    'receipt_kind': r.get('receipt_kind'),
                    'pkg_index': r.get('pkg_index'),
                    'package_type': r.get('package_type'),
                    'quantity': r.get('quantity'),
                    'driver': r.get('driver'),
                    'paid': r.get('paid', False)
                }

            # סיכום כמות פריטים
            total_qty = 0
            for r in rows:
                q = r.get('quantity', 0)
                try:
                    total_qty += int(q) if q not in (None, '') else 0
                except (TypeError, ValueError):
                    pass
            if hasattr(self, 'shipments_summary_var'):
                self.shipments_summary_var.set(f'סה"כ כמות פריטים: {total_qty}')

    def _delete_selected_shipment_row(self):
        """מחק את שורת ההובלה המסומנת מהמקור (קליטה/תעודת משלוח) ושמור."""
        if not hasattr(self, 'shipments_tree'):
            return
        sel = self.shipments_tree.selection()
        if not sel:
            messagebox.showinfo('אין בחירה', 'נא לבחור שורת הובלה למחיקה')
            return
        iid = sel[0]
        # נסה לקבל מטא; אם חסר ננסה לשחזר מהערכים
        meta = getattr(self, '_shipments_row_meta', {}).get(iid) if hasattr(self, '_shipments_row_meta') else None
        values = self.shipments_tree.item(iid, 'values') or ()
        # ערכי תצוגה
        try:
            rec_id_val = values[0]
            kind_display = values[1]
            # date_str not used; keep values[2] for index consistency only
            package_type = values[3]
            quantity_display = values[4]
            driver_display = values[5] if len(values) > 5 else ''
        except Exception:
            messagebox.showerror('שגיאה', 'אין נתונים תקפים בשורה שנבחרה')
            return
        # קביעת receipt_kind
        receipt_kind = (meta or {}).get('receipt_kind')
        if not receipt_kind:
            receipt_kind = 'supplier_intake' if kind_display == 'קליטה' else ('delivery_note' if kind_display == 'הובלה' else ('fabrics_intake' if kind_display == 'קליטת בדים' else ('fabrics_shipment' if kind_display == 'שליחת בדים' else '')))
        if receipt_kind not in ('supplier_intake', 'delivery_note', 'fabrics_intake', 'fabrics_shipment'):
            messagebox.showerror('שגיאה', 'לא ניתן לזהות את סוג הרשומה של שורת ההובלה')
            return
        # המרה ל-int בטוח ל-id והכמות
        try:
            rec_id = int(rec_id_val)
        except Exception:
            # נסה להתייחס כמחרוזת להשוואה
            rec_id = rec_id_val
        try:
            qty_val = int(quantity_display)
        except Exception:
            try:
                qty_val = int(str(quantity_display).strip())
            except Exception:
                qty_val = None
        # מציאת הרשומה במקור
        if receipt_kind == 'supplier_intake':
            records = getattr(self.data_processor, 'supplier_intakes', [])
        elif receipt_kind == 'delivery_note':
            records = getattr(self.data_processor, 'delivery_notes', [])
        elif receipt_kind == 'fabrics_intake':
            records = getattr(self.data_processor, 'fabrics_intakes', [])
        else:  # fabrics_shipment
            records = getattr(self.data_processor, 'fabrics_shipments', [])
        target_rec = None
        for i, r in enumerate(records):
            if str(r.get('id')) == str(rec_id):
                target_rec = r
                break
        if target_rec is None:
            messagebox.showerror('שגיאה', f"לא נמצאה רשומת מקור ID {rec_id}")
            return
        # מציאת אינדקס הפריט למחיקה
        pkg_index = (meta or {}).get('pkg_index')
        packages = (target_rec.get('packages') or [])
        if pkg_index is None or not (0 <= int(pkg_index) < len(packages)):
            # נסה לאתר לפי התאמת שדות
            pkg_index = None
            for idx, pkg in enumerate(packages):
                try:
                    if (str(pkg.get('package_type','')) == str(package_type) and
                        (int(pkg.get('quantity',0)) == qty_val if qty_val is not None else True) and
                        str(pkg.get('driver','')) == str(driver_display)):
                        pkg_index = idx
                        break
                except Exception:
                    continue
        if pkg_index is None:
            messagebox.showerror('שגיאה', 'לא נמצא פריט הובלה מתאים למחיקה')
            return
        # אישור משתמש
        if not messagebox.askyesno('אישור מחיקה', f"למחוק את פריט ההובלה '{package_type}' (כמות {quantity_display}) מתעודה {rec_id}?"):
            return
        # מחיקה ושמירה
        try:
            del packages[int(pkg_index)]
            target_rec['packages'] = packages
            # שמירה לקובץ המתאים
            if receipt_kind == 'supplier_intake':
                save_ok = self.data_processor._save_json_list(self.data_processor.supplier_intakes_file, records)
            elif receipt_kind == 'delivery_note':
                save_ok = self.data_processor._save_json_list(self.data_processor.delivery_notes_file, records)
            elif receipt_kind == 'fabrics_intake':
                save_ok = self.data_processor._save_json_list(self.data_processor.fabrics_intakes_file, records)
            else:  # fabrics_shipment
                save_ok = self.data_processor._save_json_list(self.data_processor.fabrics_shipments_file, records)
            if save_ok and hasattr(self.data_processor, '_rebuild_combined_receipts'):
                self.data_processor._rebuild_combined_receipts()
            # רענון קליטות בדים או שליחות בדים במידת הצורך
            try:
                if receipt_kind == 'fabrics_intake' and hasattr(self.data_processor, 'refresh_fabrics_intakes'):
                    self.data_processor.refresh_fabrics_intakes()
                elif receipt_kind == 'fabrics_shipment' and hasattr(self.data_processor, 'refresh_fabrics_shipments'):
                    self.data_processor.refresh_fabrics_shipments()
            except Exception:
                pass
            # ריענון טבלה
            self._refresh_shipments_table()
        except Exception as e:
            messagebox.showerror('שגיאה', f'כשל במחיקת פריט הובלה: {e}')

    def _mark_shipment_as_paid(self):
        """סמן את ההובלה הנבחרת כשולמה."""
        if not hasattr(self, 'shipments_tree'):
            return
        sel = self.shipments_tree.selection()
        if not sel:
            messagebox.showinfo('אין בחירה', 'נא לבחור שורת הובלה לסימון כשולמה')
            return
        self._update_shipment_paid_status(sel[0], True)

    def _mark_shipment_as_unpaid(self):
        """בטל סימון שולם עבור ההובלה הנבחרת."""
        if not hasattr(self, 'shipments_tree'):
            return
        sel = self.shipments_tree.selection()
        if not sel:
            messagebox.showinfo('אין בחירה', 'נא לבחור שורת הובלה לביטול סימון שולם')
            return
        self._update_shipment_paid_status(sel[0], False)

    def _update_shipment_paid_status(self, tree_iid, paid_status: bool):
        """עדכן את מצב השולם של הובלה ושמור למסד נתונים."""
        try:
            # קבל מטא נתונים
            meta = getattr(self, '_shipments_row_meta', {}).get(tree_iid) if hasattr(self, '_shipments_row_meta') else None
            if not meta:
                messagebox.showerror('שגיאה', 'לא נמצאו מטא נתונים לשורה זו')
                return
            
            rec_id = meta.get('rec_id')
            receipt_kind = meta.get('receipt_kind')
            pkg_index = meta.get('pkg_index')
            
            # מציאת הרשומה במקור
            if receipt_kind == 'supplier_intake':
                records = getattr(self.data_processor, 'supplier_intakes', [])
            elif receipt_kind == 'delivery_note':
                records = getattr(self.data_processor, 'delivery_notes', [])
            elif receipt_kind == 'fabrics_intake':
                records = getattr(self.data_processor, 'fabrics_intakes', [])
            else:  # fabrics_shipment
                records = getattr(self.data_processor, 'fabrics_shipments', [])
            
            target_rec = None
            for r in records:
                if str(r.get('id')) == str(rec_id):
                    target_rec = r
                    break
            
            if target_rec is None:
                messagebox.showerror('שגיאה', f"לא נמצאה רשומת מקור ID {rec_id}")
                return
            
            # עדכן את מצב השולם
            packages = target_rec.get('packages') or []
            if pkg_index is not None and 0 <= int(pkg_index) < len(packages):
                packages[int(pkg_index)]['paid'] = paid_status
                
                # שמירה לקובץ המתאים
                if receipt_kind == 'supplier_intake':
                    save_ok = self.data_processor._save_json_list(self.data_processor.supplier_intakes_file, records)
                elif receipt_kind == 'delivery_note':
                    save_ok = self.data_processor._save_json_list(self.data_processor.delivery_notes_file, records)
                elif receipt_kind == 'fabrics_intake':
                    save_ok = self.data_processor._save_json_list(self.data_processor.fabrics_intakes_file, records)
                else:  # fabrics_shipment
                    save_ok = self.data_processor._save_json_list(self.data_processor.fabrics_shipments_file, records)
                
                if save_ok:
                    # ריענון רשימות
                    try:
                        if receipt_kind == 'fabrics_intake' and hasattr(self.data_processor, 'refresh_fabrics_intakes'):
                            self.data_processor.refresh_fabrics_intakes()
                        elif receipt_kind == 'fabrics_shipment' and hasattr(self.data_processor, 'refresh_fabrics_shipments'):
                            self.data_processor.refresh_fabrics_shipments()
                    except Exception:
                        pass
                    
                    # ריענון טבלה
                    self._refresh_shipments_table()
                else:
                    messagebox.showerror('שגיאה', 'כשל בשמירת השינויים')
            else:
                messagebox.showerror('שגיאה', 'אינדקס חבילה לא תקין')
        except Exception as e:
            messagebox.showerror('שגיאה', f'כשל בעדכון מצב שולם: {e}')

    # ---- Hook from save actions ----
    def _notify_new_receipt_saved(self):
        """קריאה מהטאבים של קליטה / תעודת משלוח לאחר שמירה כדי לרענן הובלות."""
        try:
            if hasattr(self, '_refresh_shipments_table'):
                self._refresh_shipments_table()
        except Exception:
            pass

    # ---- Payment Report Functions ----
    def _update_payment_drivers_list(self):
        """עדכון רשימת המובילים בטאב דו\"ח תשלום."""
        try:
            if not hasattr(self, 'payment_driver_combo'):
                return
            driver_names = [d.get('name', '') for d in self._drivers if d.get('name')]
            self.payment_driver_combo['values'] = driver_names
            if driver_names:
                self.payment_driver_combo.current(0)
        except Exception:
            pass

    def _calculate_payment_report(self):
        """חישוב דו\"ח הובלות לתשלום לפי מוביל וטווח תאריכים."""
        try:
            # קבלת פרמטרים
            driver_name = (self.payment_driver_var.get() or '').strip()
            start_date_str = (self.payment_start_date_var.get() or '').strip()
            end_date_str = (self.payment_end_date_var.get() or '').strip()
            payment_filter = (self.payment_filter_var.get() or 'הכל').strip()
            
            # בדיקת תקינות
            if not driver_name:
                messagebox.showwarning('חסר מוביל', 'נא לבחור מוביל')
                return
            
            if not start_date_str or not end_date_str:
                messagebox.showwarning('חסרים תאריכים', 'נא למלא תאריך התחלה וסוף')
                return
            
            # המרת תאריכים
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            except Exception:
                messagebox.showerror('שגיאה', 'פורמט תאריך לא תקין. השתמש ב-YYYY-MM-DD')
                return
            
            if start_date > end_date:
                messagebox.showerror('שגיאה', 'תאריך ההתחלה גדול מתאריך הסוף')
                return
            
            # רענון נתונים
            try:
                if hasattr(self.data_processor, 'refresh_supplier_receipts'):
                    self.data_processor.refresh_supplier_receipts()
                if hasattr(self.data_processor, 'refresh_fabrics_intakes'):
                    self.data_processor.refresh_fabrics_intakes()
                if hasattr(self.data_processor, 'refresh_fabrics_shipments'):
                    self.data_processor.refresh_fabrics_shipments()
            except Exception:
                pass
            
            # איסוף נתונים
            supplier_intakes = getattr(self.data_processor, 'supplier_intakes', [])
            delivery_notes = getattr(self.data_processor, 'delivery_notes', [])
            fabrics_intakes = getattr(self.data_processor, 'fabrics_intakes', [])
            fabrics_shipments = getattr(self.data_processor, 'fabrics_shipments', [])
            
            # סיכום כמויות לפי סוג חבילה
            package_counts = {}
            total_packages = 0
            
            def process_records(records_list):
                """עיבוד רשומות וסיכום כמויות."""
                nonlocal total_packages
                for rec in records_list:
                    # בדיקת תאריך
                    date_str = rec.get('date', '')
                    try:
                        rec_date = datetime.strptime(date_str, '%Y-%m-%d')
                    except Exception:
                        try:
                            rec_date = datetime.strptime(date_str[:10], '%Y-%m-%d')
                        except Exception:
                            continue
                    
                    # האם בטווח התאריכים?
                    if not (start_date <= rec_date <= end_date):
                        continue
                    
                    # עיבוד חבילות
                    for pkg in rec.get('packages', []) or []:
                        pkg_driver = (pkg.get('driver', '') or '').strip()
                        
                        # האם זה המוביל הנבחר?
                        if pkg_driver != driver_name:
                            continue
                        
                        # בדיקת מצב תשלום לפי הפילטר
                        pkg_paid = pkg.get('paid', False)
                        if payment_filter == 'רק לא שולם' and pkg_paid:
                            continue  # דלג על חבילות ששולמו
                        elif payment_filter == 'רק שולם' and not pkg_paid:
                            continue  # דלג על חבילות שלא שולמו
                        # אם payment_filter == 'הכל', אל תדלג
                        
                        pkg_type = pkg.get('package_type', 'לא מוגדר')
                        pkg_qty = pkg.get('quantity', 0)
                        
                        try:
                            pkg_qty = int(pkg_qty)
                        except Exception:
                            pkg_qty = 0
                        
                        # צבירת כמויות
                        if pkg_type not in package_counts:
                            package_counts[pkg_type] = 0
                        package_counts[pkg_type] += pkg_qty
                        total_packages += pkg_qty
            
            # עיבוד כל הרשומות
            process_records(supplier_intakes)
            process_records(delivery_notes)
            process_records(fabrics_intakes)
            process_records(fabrics_shipments)
            
            # קבלת מחירים של המוביל
            driver_pricing = {}
            for d in self._drivers:
                if d.get('name') == driver_name:
                    driver_pricing = d.get('pricing', {})
                    break
            
            # הצגת תוצאות
            self.payment_results_text.config(state='normal')
            self.payment_results_text.delete('1.0', 'end')
            
            # כותרת
            report_title = f"דו\"ח הובלות - {driver_name}\n"
            report_title += f"תקופה: {start_date_str} עד {end_date_str}\n"
            report_title += f"סינון: {payment_filter}\n"
            report_title += "=" * 50 + "\n\n"
            self.payment_results_text.insert('end', report_title, 'title')
            
            if not package_counts:
                no_data_msg = f"לא נמצאו הובלות בתקופה זו למוביל זה"
                if payment_filter == 'רק לא שולם':
                    no_data_msg += " (שלא שולמו)"
                elif payment_filter == 'רק שולם':
                    no_data_msg += " (ששולמו)"
                no_data_msg += ".\n"
                self.payment_results_text.insert('end', no_data_msg, 'no_data')
            else:
                # תצוגת סיכום לפי סוג חבילה עם חישובי מחיר
                self.payment_results_text.insert('end', "סיכום כמויות ועלויות:\n\n", 'header')
                
                total_cost = 0
                for pkg_type, qty in sorted(package_counts.items()):
                    # קבלת מחיר לפי סוג החבילה
                    price = driver_pricing.get(pkg_type, 0)
                    cost = qty * price
                    total_cost += cost
                    
                    # הצגת השורה
                    if price > 0:
                        line = f"  • {pkg_type}: {qty} × {price:.2f} ₪ = {cost:.2f} ₪\n"
                        self.payment_results_text.insert('end', line, 'data')
                    else:
                        line = f"  • {pkg_type}: {qty} (ללא מחיר במחירון)\n"
                        self.payment_results_text.insert('end', line, 'data_warning')
                
                self.payment_results_text.insert('end', "\n" + "-" * 50 + "\n", 'separator')
                total_line = f"סה\"כ חבילות: {total_packages}\n"
                self.payment_results_text.insert('end', total_line, 'total')
                cost_line = f"סה\"כ לתשלום: {total_cost:.2f} ₪\n"
                self.payment_results_text.insert('end', cost_line, 'total_cost')
            
            # עיצוב טקסט
            self.payment_results_text.tag_config('title', font=('Arial', 13, 'bold'), foreground='#2c3e50')
            self.payment_results_text.tag_config('header', font=('Arial', 12, 'bold'), foreground='#34495e')
            self.payment_results_text.tag_config('data', font=('Arial', 11), foreground='#2c3e50')
            self.payment_results_text.tag_config('data_warning', font=('Arial', 11), foreground='#e67e22')
            self.payment_results_text.tag_config('separator', foreground='#95a5a6')
            self.payment_results_text.tag_config('total', font=('Arial', 12, 'bold'), foreground='#27ae60')
            self.payment_results_text.tag_config('total_cost', font=('Arial', 13, 'bold'), foreground='#27ae60')
            self.payment_results_text.tag_config('no_data', font=('Arial', 11), foreground='#e74c3c')
            
            self.payment_results_text.config(state='disabled')
            
            # שמירת נתוני הדוח לשימוש בשמירה
            if package_counts:
                self._current_report_data = {
                    'driver': driver_name,
                    'start_date': start_date_str,
                    'end_date': end_date_str,
                    'payment_filter': payment_filter,
                    'package_counts': package_counts,
                    'total_packages': total_packages,
                    'total_cost': total_cost if 'total_cost' in locals() else 0,
                    'driver_pricing': driver_pricing,
                    'items_data': []  # נמלא בשמירה
                }
                # אפשר כפתור שמירה
                if hasattr(self, 'save_report_btn'):
                    self.save_report_btn.config(state='normal')
            else:
                self._current_report_data = None
                if hasattr(self, 'save_report_btn'):
                    self.save_report_btn.config(state='disabled')
            
        except Exception as e:
            messagebox.showerror('שגיאה', f'שגיאה בחישוב דו\"ח: {e}')
            self._current_report_data = None
            if hasattr(self, 'save_report_btn'):
                self.save_report_btn.config(state='disabled')

    # ---- Pricing Management Functions ----
    def _update_pricing_drivers_list(self):
        """עדכון רשימת המובילים בטאב מחירון."""
        try:
            if not hasattr(self, 'pricing_driver_combo'):
                return
            driver_names = [d.get('name', '') for d in self._drivers if d.get('name')]
            self.pricing_driver_combo['values'] = driver_names
            if driver_names:
                self.pricing_driver_combo.current(0)
                # טען מחירים של המוביל הראשון
                self._load_driver_pricing(None)
        except Exception:
            pass

    def _load_driver_pricing(self, event):
        """טעינת מחירי ההובלה של המוביל הנבחר."""
        try:
            driver_name = (self.pricing_driver_var.get() or '').strip()
            if not driver_name:
                return
            
            # מצא את המוביל
            driver = None
            for d in self._drivers:
                if d.get('name') == driver_name:
                    driver = d
                    break
            
            if driver:
                pricing = driver.get('pricing', {})
                self.pricing_fabric_var.set(str(pricing.get('בד', 0)))
                self.pricing_bag_var.set(str(pricing.get('שק', 0)))
                self.pricing_small_bag_var.set(str(pricing.get('שקית קטנה', 0)))
            else:
                # אם לא נמצא, אפס את השדות
                self.pricing_fabric_var.set('0')
                self.pricing_bag_var.set('0')
                self.pricing_small_bag_var.set('0')
        except Exception:
            pass

    def _save_driver_pricing(self):
        """שמירת מחירי ההובלה של המוביל הנבחר."""
        try:
            driver_name = (self.pricing_driver_var.get() or '').strip()
            if not driver_name:
                messagebox.showwarning('חסר מוביל', 'נא לבחור מוביל')
                return
            
            # קבלת מחירים
            try:
                fabric_price = float(self.pricing_fabric_var.get() or '0')
                bag_price = float(self.pricing_bag_var.get() or '0')
                small_bag_price = float(self.pricing_small_bag_var.get() or '0')
            except ValueError:
                messagebox.showerror('שגיאה', 'נא להזין מחירים תקינים (מספרים בלבד)')
                return
            
            # מצא את המוביל ועדכן מחירים
            driver_found = False
            for d in self._drivers:
                if d.get('name') == driver_name:
                    d['pricing'] = {
                        'בד': fabric_price,
                        'שק': bag_price,
                        'שקית קטנה': small_bag_price
                    }
                    driver_found = True
                    break
            
            if not driver_found:
                messagebox.showerror('שגיאה', 'לא נמצא מוביל בשם זה')
                return
            
            # שמור לקובץ
            self._save_drivers()
            self._refresh_pricing_table()
            messagebox.showinfo('הצלחה', f'המחירים של {driver_name} נשמרו בהצלחה')
            
        except Exception as e:
            messagebox.showerror('שגיאה', f'שגיאה בשמירת מחירים: {e}')

    def _refresh_pricing_table(self):
        """רענון טבלת המחירון."""
        try:
            if not hasattr(self, 'pricing_tree'):
                return
            
            # נקה טבלה
            for iid in self.pricing_tree.get_children():
                self.pricing_tree.delete(iid)
            
            # הוסף שורות
            for driver in self._drivers:
                name = driver.get('name', '')
                pricing = driver.get('pricing', {})
                fabric_price = pricing.get('בד', 0)
                bag_price = pricing.get('שק', 0)
                small_bag_price = pricing.get('שקית קטנה', 0)
                
                self.pricing_tree.insert('', 'end', values=(
                    name,
                    f'{fabric_price:.2f}' if fabric_price else '0.00',
                    f'{bag_price:.2f}' if bag_price else '0.00',
                    f'{small_bag_price:.2f}' if small_bag_price else '0.00'
                ))
        except Exception:
            pass

    # ---- Shipment Reports Functions ----
    def _save_shipment_report(self):
        """שמירת דוח הובלות - יצירת אקסל, עדכון סטטוס שולם ושמירת מטא-דאטה."""
        try:
            if not self._current_report_data:
                messagebox.showwarning('אין דוח', 'נא לחשב דוח לפני השמירה')
                return
            
            # בקש אישור
            if not messagebox.askyesno('אישור שמירה', 
                'שמירת הדוח תסמן את כל ההובלות כשולמו.\nהאם להמשיך?'):
                return
            
            # איסוף פריטי הדוח מחדש (עם מחיר בלבד)
            report_items = self._collect_report_items()
            if not report_items:
                messagebox.showerror('שגיאה', 'לא נמצאו פריטים לדוח')
                return
            
            # יצירת קובץ אקסל
            excel_filename = self._create_shipment_report_excel(report_items)
            if not excel_filename:
                return
            
            # עדכון סטטוס שולם
            self._mark_report_items_as_paid(report_items)
            
            # שמירת מטא-דאטה
            report_id = self._save_report_metadata(excel_filename)
            
            # הודעה למשתמש
            messagebox.showinfo('הצלחה', f'הדוח נשמר בהצלחה!\nמספר דוח: {report_id}\nקובץ: {excel_filename}')
            
            # רענון היסטוריה וסיכום הובלות
            self._refresh_reports_history()
            self._refresh_shipments_table()
            
            # נקה דוח נוכחי ונטרל כפתור
            self._current_report_data = None
            if hasattr(self, 'save_report_btn'):
                self.save_report_btn.config(state='disabled')
            
        except Exception as e:
            messagebox.showerror('שגיאה', f'שגיאה בשמירת דוח: {e}')

    def _collect_report_items(self):
        """איסוף כל פריטי ההובלה שהשתתפו בדוח."""
        if not self._current_report_data:
            return []
        
        items = []
        driver_name = self._current_report_data['driver']
        start_date_str = self._current_report_data['start_date']
        end_date_str = self._current_report_data['end_date']
        payment_filter = self._current_report_data['payment_filter']
        driver_pricing = self._current_report_data['driver_pricing']
        
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        except:
            return []
        
        # איסוף נתונים
        supplier_intakes = getattr(self.data_processor, 'supplier_intakes', [])
        delivery_notes = getattr(self.data_processor, 'delivery_notes', [])
        fabrics_intakes = getattr(self.data_processor, 'fabrics_intakes', [])
        fabrics_shipments = getattr(self.data_processor, 'fabrics_shipments', [])
        
        def collect_from_records(records_list, receipt_kind):
            for rec in records_list:
                rec_id = rec.get('id')
                date_str = rec.get('date', '')
                try:
                    rec_date = datetime.strptime(date_str, '%Y-%m-%d')
                except:
                    try:
                        rec_date = datetime.strptime(date_str[:10], '%Y-%m-%d')
                    except:
                        continue
                
                if not (start_date <= rec_date <= end_date):
                    continue
                
                for pkg_index, pkg in enumerate(rec.get('packages', []) or []):
                    pkg_driver = (pkg.get('driver', '') or '').strip()
                    if pkg_driver != driver_name:
                        continue
                    
                    pkg_paid = pkg.get('paid', False)
                    if payment_filter == 'רק לא שולם' and pkg_paid:
                        continue
                    elif payment_filter == 'רק שולם' and not pkg_paid:
                        continue
                    
                    pkg_type = pkg.get('package_type', 'לא מוגדר')
                    price = driver_pricing.get(pkg_type, 0)
                    
                    # רק פריטים עם מחיר
                    if price <= 0:
                        continue
                    
                    pkg_qty = pkg.get('quantity', 0)
                    try:
                        pkg_qty = int(pkg_qty)
                    except:
                        pkg_qty = 0
                    
                    kind_display = 'קליטה' if receipt_kind == 'supplier_intake' else ('הובלה' if receipt_kind == 'delivery_note' else ('קליטת בדים' if receipt_kind == 'fabrics_intake' else 'שליחת בדים'))
                    
                    items.append({
                        'rec_id': rec_id,
                        'receipt_kind': receipt_kind,
                        'pkg_index': pkg_index,
                        'kind': kind_display,
                        'date': date_str,
                        'package_type': pkg_type,
                        'quantity': pkg_qty,
                        'driver': pkg_driver,
                        'price': price,
                        'cost': pkg_qty * price
                    })
        
        collect_from_records(supplier_intakes, 'supplier_intake')
        collect_from_records(delivery_notes, 'delivery_note')
        collect_from_records(fabrics_intakes, 'fabrics_intake')
        collect_from_records(fabrics_shipments, 'fabrics_shipment')
        
        return items

    def _create_shipment_report_excel(self, report_items):
        """יצירת קובץ אקסל עם הדוח."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messagebox.showerror('שגיאה', 'חסרה ספריית openpyxl. נא להתקין: pip install openpyxl')
            return None
        
        try:
            # יצירת תיקייה
            reports_dir = os.path.join(os.getcwd(), 'exports', 'shipment_reports')
            os.makedirs(reports_dir, exist_ok=True)
            
            # שם קובץ
            driver_name = self._current_report_data['driver']
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # מציאת מספר דוח רץ
            report_num = 1
            while True:
                filename = f"דוח_הובלות_{driver_name}_{timestamp}_{report_num:03d}.xlsx"
                filepath = os.path.join(reports_dir, filename)
                if not os.path.exists(filepath):
                    break
                report_num += 1
            
            # יצירת ספר עבודה
            wb = Workbook()
            ws = wb.active
            ws.title = "דוח הובלות"
            
            # כותרת
            ws['A1'] = f"דוח הובלות - {driver_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"תקופה: {self._current_report_data['start_date']} עד {self._current_report_data['end_date']}"
            ws['A3'] = f"סינון: {self._current_report_data['payment_filter']}"
            
            # סיכום
            ws['A5'] = "סיכום כמויות ועלויות:"
            ws['A5'].font = Font(bold=True, size=12)
            
            row = 6
            package_counts = self._current_report_data['package_counts']
            driver_pricing = self._current_report_data['driver_pricing']
            
            for pkg_type in sorted(package_counts.keys()):
                qty = package_counts[pkg_type]
                price = driver_pricing.get(pkg_type, 0)
                if price > 0:
                    cost = qty * price
                    ws[f'A{row}'] = f"  • {pkg_type}: {qty} × {price:.2f} ₪ = {cost:.2f} ₪"
                    row += 1
            
            row += 1
            ws[f'A{row}'] = f"סה\"כ חבילות: {self._current_report_data['total_packages']}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = f"סה\"כ לתשלום: {self._current_report_data['total_cost']:.2f} ₪"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="006100")
            
            # טבלה
            row += 2
            headers = ['מספר תעודה', 'סוג', 'תאריך', 'פריט הובלה', 'כמות', 'מוביל', 'מחיר ליחידה', 'עלות']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # נתונים
            for item in report_items:
                row += 1
                ws.cell(row=row, column=1, value=item['rec_id'])
                ws.cell(row=row, column=2, value=item['kind'])
                ws.cell(row=row, column=3, value=item['date'])
                ws.cell(row=row, column=4, value=item['package_type'])
                ws.cell(row=row, column=5, value=item['quantity'])
                ws.cell(row=row, column=6, value=item['driver'])
                ws.cell(row=row, column=7, value=f"{item['price']:.2f}")
                ws.cell(row=row, column=8, value=f"{item['cost']:.2f}")
            
            # רוחב עמודות
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12
            
            # שמירה
            wb.save(filepath)
            
            return filename
            
        except Exception as e:
            messagebox.showerror('שגיאה', f'שגיאה ביצירת קובץ אקסל: {e}')
            return None

    def _mark_report_items_as_paid(self, report_items):
        """עדכון סטטוס שולם לכל הפריטים בדוח."""
        try:
            # קיבוץ לפי סוג רשומה
            updates_by_kind = {}
            for item in report_items:
                kind = item['receipt_kind']
                if kind not in updates_by_kind:
                    updates_by_kind[kind] = []
                updates_by_kind[kind].append(item)
            
            # עדכון כל סוג
            for receipt_kind, items in updates_by_kind.items():
                if receipt_kind == 'supplier_intake':
                    records = getattr(self.data_processor, 'supplier_intakes', [])
                    file_path = self.data_processor.supplier_intakes_file
                elif receipt_kind == 'delivery_note':
                    records = getattr(self.data_processor, 'delivery_notes', [])
                    file_path = self.data_processor.delivery_notes_file
                elif receipt_kind == 'fabrics_intake':
                    records = getattr(self.data_processor, 'fabrics_intakes', [])
                    file_path = self.data_processor.fabrics_intakes_file
                else:  # fabrics_shipment
                    records = getattr(self.data_processor, 'fabrics_shipments', [])
                    file_path = self.data_processor.fabrics_shipments_file
                
                # עדכון פריטים
                for item in items:
                    for rec in records:
                        if str(rec.get('id')) == str(item['rec_id']):
                            packages = rec.get('packages', [])
                            pkg_index = item['pkg_index']
                            if 0 <= pkg_index < len(packages):
                                packages[pkg_index]['paid'] = True
                            break
                
                # שמירה
                self.data_processor._save_json_list(file_path, records)
            
        except Exception as e:
            messagebox.showerror('שגיאה', f'שגיאה בעדכון סטטוס שולם: {e}')

    def _save_report_metadata(self, excel_filename):
        """שמירת מטא-דאטה של הדוח."""
        try:
            reports_file = os.path.join(os.getcwd(), 'shipment_reports.json')
            
            # טעינת דוחות קיימים
            if os.path.exists(reports_file):
                try:
                    with open(reports_file, 'r', encoding='utf-8') as f:
                        reports = json.load(f)
                except:
                    reports = []
            else:
                reports = []
            
            # מציאת מספר דוח חדש
            if reports:
                max_id = max(r.get('id', 0) for r in reports)
                new_id = max_id + 1
            else:
                new_id = 1
            
            # יצירת רשומה
            report_record = {
                'id': new_id,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'driver': self._current_report_data['driver'],
                'start_date': self._current_report_data['start_date'],
                'end_date': self._current_report_data['end_date'],
                'payment_filter': self._current_report_data['payment_filter'],
                'total_packages': self._current_report_data['total_packages'],
                'total_cost': self._current_report_data['total_cost'],
                'excel_file': excel_filename
            }
            
            reports.append(report_record)
            
            # שמירה
            with open(reports_file, 'w', encoding='utf-8') as f:
                json.dump(reports, f, ensure_ascii=False, indent=2)
            
            return new_id
            
        except Exception as e:
            messagebox.showerror('שגיאה', f'שגיאה בשמירת מטא-דאטה: {e}')
            return 0

    def _refresh_reports_history(self):
        """רענון טבלת היסטוריית הדוחות."""
        try:
            if not hasattr(self, 'reports_history_tree'):
                return
            
            # ניקוי טבלה
            for iid in self.reports_history_tree.get_children():
                self.reports_history_tree.delete(iid)
            
            # טעינת דוחות
            reports_file = os.path.join(os.getcwd(), 'shipment_reports.json')
            if not os.path.exists(reports_file):
                return
            
            try:
                with open(reports_file, 'r', encoding='utf-8') as f:
                    reports = json.load(f)
            except:
                return
            
            # הוספה לטבלה (הפוך - חדשים ראשונים)
            for report in reversed(reports):
                period = f"{report.get('start_date', '')} - {report.get('end_date', '')}"
                self.reports_history_tree.insert('', 'end', values=(
                    report.get('id', ''),
                    report.get('created_at', ''),
                    report.get('driver', ''),
                    period,
                    f"{report.get('total_cost', 0):.2f}",
                    '📄 פתח'
                ))
        except Exception as e:
            pass

    def _on_report_double_click(self, event):
        """פתיחת קובץ אקסל בדאבל קליק."""
        try:
            selection = self.reports_history_tree.selection()
            if not selection:
                return
            
            values = self.reports_history_tree.item(selection[0], 'values')
            if not values:
                return
            
            report_id = values[0]
            
            # טעינת דוחות
            reports_file = os.path.join(os.getcwd(), 'shipment_reports.json')
            if not os.path.exists(reports_file):
                return
            
            with open(reports_file, 'r', encoding='utf-8') as f:
                reports = json.load(f)
            
            # מציאת דוח
            report = None
            for r in reports:
                if str(r.get('id')) == str(report_id):
                    report = r
                    break
            
            if not report:
                messagebox.showerror('שגיאה', 'לא נמצא דוח')
                return
            
            # פתיחת קובץ
            excel_filename = report.get('excel_file')
            filepath = os.path.join(os.getcwd(), 'exports', 'shipment_reports', excel_filename)
            
            if not os.path.exists(filepath):
                messagebox.showerror('שגיאה', f'קובץ לא נמצא: {excel_filename}')
                return
            
            # פתיחה
            os.startfile(filepath)
            
        except Exception as e:
            messagebox.showerror('שגיאה', f'שגיאה בפתיחת קובץ: {e}')
