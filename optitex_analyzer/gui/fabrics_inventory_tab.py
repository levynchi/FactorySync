import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from . import theme

class FabricsInventoryTabMixin:
    """Mixin לטאב מלאי בדים."""
    def _create_fabrics_inventory_tab(self):
        tab = tk.Frame(self.notebook, bg=theme.PAGE_BG); self.notebook.add(tab, text="מלאי בדים")
        tk.Label(tab, text="מלאי בדים", font=(theme.FONT_FAMILY, 16, 'bold'), bg=theme.PAGE_BG, fg=theme.DARK).pack(pady=8)
        # Action bar
        actions = tk.Frame(tab, bg=theme.PAGE_BG); actions.pack(fill='x', padx=15, pady=5)
        tk.Button(actions, text="⬇️ הורד תבנית אקסל למשלוח", command=self._export_fabrics_template_excel, bg=theme.SUCCESS, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='right', padx=5)
        tk.Button(actions, text="📤 הדפס לאקסל", command=self._export_current_fabrics_to_excel, bg=theme.TEAL, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='right', padx=5)
        tk.Button(actions, text="📥 הכנס משלוח בדים (CSV)", command=self._import_fabrics_csv, bg=theme.PRIMARY_DARK, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='right', padx=5)
        tk.Button(actions, text="🔄 רענן", command=self._refresh_fabrics_table, bg=theme.PRIMARY, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='right', padx=5)

        inner_notebook = ttk.Notebook(tab); inner_notebook.pack(fill='both', expand=True, padx=10, pady=(0,5))
        inventory_tab = tk.Frame(inner_notebook, bg=theme.CARD_BG); inner_notebook.add(inventory_tab, text="נתוני מלאי")
        unbarcoded_tab = tk.Frame(inner_notebook, bg=theme.CARD_BG); inner_notebook.add(unbarcoded_tab, text="בדים בלי ברקוד")
        barcode_search_tab = tk.Frame(inner_notebook, bg=theme.CARD_BG); inner_notebook.add(barcode_search_tab, text="חיפוש לפי ברקוד")

        # Filter bar for inventory
        filter_frame = tk.Frame(inventory_tab, bg=theme.CARD_BG); filter_frame.pack(fill='x', padx=5, pady=(6,0))
        # Variables
        self.fabrics_filter_type_var = tk.StringVar(value='')
        self.fabrics_filter_color_var = tk.StringVar(value='')
        self.fabrics_filter_location_var = tk.StringVar(value='')
        self.fabrics_filter_status_var = tk.StringVar(value='')
        self.fabrics_filter_intake_date_var = tk.StringVar(value='')
        # Widgets (placed right-to-left)
        tk.Label(filter_frame, text=':סטטוס', bg=theme.CARD_BG).pack(side='right', padx=(6,2))
        self.fabrics_filter_status_cb = ttk.Combobox(filter_frame, textvariable=self.fabrics_filter_status_var, state='readonly', width=14, values=('', 'במלאי','נשלח','נגזר'))
        self.fabrics_filter_status_cb.pack(side='right', padx=(0,10))

        tk.Label(filter_frame, text=':תאריך קליטה', bg=theme.CARD_BG).pack(side='right', padx=(6,2))
        self.fabrics_filter_intake_date_cb = ttk.Combobox(filter_frame, textvariable=self.fabrics_filter_intake_date_var, state='readonly', width=16)
        self.fabrics_filter_intake_date_cb.pack(side='right', padx=(0,10))

        tk.Label(filter_frame, text=':מיקום', bg=theme.CARD_BG).pack(side='right', padx=(6,2))
        self.fabrics_filter_location_cb = ttk.Combobox(filter_frame, textvariable=self.fabrics_filter_location_var, width=18)
        self.fabrics_filter_location_cb.pack(side='right', padx=(0,10))

        tk.Label(filter_frame, text=':צבע', bg=theme.CARD_BG).pack(side='right', padx=(6,2))
        self.fabrics_filter_color_cb = ttk.Combobox(filter_frame, textvariable=self.fabrics_filter_color_var, width=18)
        self.fabrics_filter_color_cb.pack(side='right', padx=(0,10))

        tk.Label(filter_frame, text=':סוג בד', bg=theme.CARD_BG).pack(side='right', padx=(6,2))
        self.fabrics_filter_type_cb = ttk.Combobox(filter_frame, textvariable=self.fabrics_filter_type_var, width=20, state='readonly')
        self.fabrics_filter_type_cb.pack(side='right', padx=(0,10))

        # Actions
        tk.Button(filter_frame, text='נקה', command=lambda: self._clear_fabrics_filters()).pack(side='left', padx=(0,6))
        tk.Button(filter_frame, text='החל סינון', command=lambda: self._apply_fabrics_filters()).pack(side='left')
        self.fabrics_filter_info_var = tk.StringVar(value='')
        tk.Label(filter_frame, textvariable=self.fabrics_filter_info_var, bg=theme.CARD_BG, fg=theme.SUBTEXT).pack(side='left', padx=10)

        # Bind quick-apply
        self.fabrics_filter_status_cb.bind('<<ComboboxSelected>>', lambda e: self._apply_fabrics_filters())
        self.fabrics_filter_type_cb.bind('<<ComboboxSelected>>', lambda e: self._apply_fabrics_filters())
        self.fabrics_filter_color_cb.bind('<<ComboboxSelected>>', lambda e: self._apply_fabrics_filters())
        self.fabrics_filter_location_cb.bind('<<ComboboxSelected>>', lambda e: self._apply_fabrics_filters())
        self.fabrics_filter_intake_date_cb.bind('<<ComboboxSelected>>', lambda e: self._apply_fabrics_filters())

        # Inventory table
        table_frame = tk.Frame(inventory_tab, bg=theme.CARD_BG); table_frame.pack(fill='both', expand=True, padx=5, pady=5)
        cols = ('barcode','fabric_type','color_name','color_no','design_code','width','net_kg','meters','price','location','intake_date','status')
        self.fabrics_tree = ttk.Treeview(table_frame, columns=cols, show='headings')
        headers = {'barcode':'ברקוד','fabric_type':'סוג בד','color_name':'צבע','color_no':'מס׳ צבע','design_code':'Desen Kodu','width':'רוחב','net_kg':'ק"ג נטו','meters':'מטרים','price':'מחיר','location':'מיקום','intake_date':'תאריך קליטה','status':'סטטוס'}
        widths = {'barcode':120,'fabric_type':140,'color_name':110,'color_no':80,'design_code':110,'width':60,'net_kg':80,'meters':80,'price':80,'location':90,'intake_date':120,'status':80}
        for c in cols:
            self.fabrics_tree.heading(c, text=headers[c]); self.fabrics_tree.column(c, width=widths[c], anchor='center')
        vsb = ttk.Scrollbar(table_frame, orient='vertical', command=self.fabrics_tree.yview); self.fabrics_tree.configure(yscroll=vsb.set)
        self.fabrics_tree.grid(row=0,column=0,sticky='nsew'); vsb.grid(row=0,column=1,sticky='ns')
        table_frame.grid_columnconfigure(0,weight=1); table_frame.grid_rowconfigure(0,weight=1)
        self._fabric_status_menu = tk.Menu(self.fabrics_tree, tearoff=0)
        for status in ("במלאי","נשלח","נגזר"):
            self._fabric_status_menu.add_command(label=status, command=lambda s=status: self._change_selected_fabric_status(s))
        self.fabrics_tree.bind('<Button-3>', self._on_fabrics_right_click)

        # Logs tab
        logs_tab = tk.Frame(inner_notebook, bg=theme.CARD_BG); inner_notebook.add(logs_tab, text="קבצים שעלו")
        tk.Label(logs_tab, text="דאבל־קליק על קובץ כדי לפתוח באקסל", bg=theme.CARD_BG, fg=theme.SUBTEXT,
                 font=(theme.FONT_FAMILY, 9), anchor='e').pack(fill='x', padx=8, pady=(6, 0))
        logs_frame = tk.Frame(logs_tab, bg=theme.CARD_BG); logs_frame.pack(fill='both', expand=True, padx=5, pady=5)
        log_cols = ('id','file_name','imported_at','records_added','delete')
        self.fabrics_logs_tree = ttk.Treeview(logs_frame, columns=log_cols, show='headings')
        log_headers = {'id':'ID','file_name':'שם קובץ','imported_at':'תאריך העלאה','records_added':'רשומות','delete':'מחיקה'}
        log_widths = {'id':50,'file_name':220,'imported_at':140,'records_added':70,'delete':60}
        for c in log_cols:
            self.fabrics_logs_tree.heading(c, text=log_headers[c]); self.fabrics_logs_tree.column(c, width=log_widths[c], anchor='center')
        lsvb = ttk.Scrollbar(logs_frame, orient='vertical', command=self.fabrics_logs_tree.yview); self.fabrics_logs_tree.configure(yscroll=lsvb.set)
        self.fabrics_logs_tree.grid(row=0,column=0,sticky='nsew'); lsvb.grid(row=0,column=1,sticky='ns')
        logs_frame.grid_columnconfigure(0,weight=1); logs_frame.grid_rowconfigure(0,weight=1)
        self.fabrics_logs_tree.bind('<Button-1>', self._handle_logs_click)
        self.fabrics_logs_tree.bind('<Double-1>', self._open_fabric_import_in_excel)

        # Unbarcoded fabrics UI
        ub_actions = tk.Frame(unbarcoded_tab, bg=theme.CARD_BG); ub_actions.pack(fill='x', padx=6, pady=6)
        tk.Button(ub_actions, text="➕ הוסף", command=self._ub_add_dialog, bg=theme.SUCCESS, fg='white').pack(side='right', padx=4)
        tk.Button(ub_actions, text="🗑️ מחק נבחר", command=self._ub_delete_selected, bg=theme.WARNING, fg='white').pack(side='right')
        ub_frame = tk.Frame(unbarcoded_tab, bg=theme.CARD_BG); ub_frame.pack(fill='both', expand=True, padx=6, pady=(0,6))
        ub_cols = ('id','created_at','fabric_type','manufacturer','color','shade','notes')
        self.ub_tree = ttk.Treeview(ub_frame, columns=ub_cols, show='headings')
        ub_headers = {'id':'', 'created_at':'תאריך','fabric_type':'סוג בד','manufacturer':'יצרן הבד','color':'צבע','shade':'גוון','notes':'הערות'}
        ub_widths = {'id':60,'created_at':140,'fabric_type':160,'manufacturer':160,'color':100,'shade':80,'notes':240}
        for c in ub_cols:
            self.ub_tree.heading(c, text=ub_headers[c])
            if c == 'id':
                self.ub_tree.column(c, width=0, minwidth=0, stretch=False)
            else:
                self.ub_tree.column(c, width=ub_widths[c], anchor='center')
        ub_vsb = ttk.Scrollbar(ub_frame, orient='vertical', command=self.ub_tree.yview); self.ub_tree.configure(yscroll=ub_vsb.set)
        self.ub_tree.grid(row=0,column=0,sticky='nsew'); ub_vsb.grid(row=0,column=1,sticky='ns')
        ub_frame.grid_columnconfigure(0,weight=1); ub_frame.grid_rowconfigure(0,weight=1)
        self._populate_unbarcoded_table()

        # Barcode search tab
        self._build_barcode_search_tab(barcode_search_tab)

        # Footer summary
        self.fabrics_summary_var = tk.StringVar(value="אין נתונים")
        tk.Label(tab, textvariable=self.fabrics_summary_var, bg=theme.DARK, fg='white', anchor='w', padx=12, font=(theme.FONT_FAMILY,10)).pack(fill='x', side='bottom')
        # Initialize filters list values, then populate
        try:
            self._refresh_fabric_filter_values()
        except Exception:
            pass
        self._populate_fabrics_table(); self._populate_fabrics_logs(); self._update_fabrics_summary()

    def _export_fabrics_template_excel(self):
        """יוצר קובץ Excel ריק עם כותרות בסדר שהיבוא (CSV) מצפה לו."""
        # סדר ושמות העמודות כפי שהפונקציה import_fabrics_csv מצפה להם
        headers = [
            'BARCODE NO',
            'סוג בד',
            'COLOR NAME',
            'COLOR NO',
            'Desen Kodu',
            'WIDTH',
            'GR',
            'NET KG',
            'GROSS KG',
            'METER',
            'PRICE',
            'TOTAL',
            'location',
            'Last Modified',
            'מטרה',
        ]
        # בחירת נתיב שמירה
        from tkinter import filedialog, messagebox
        default_name = 'fabrics_shipment_template.xlsx'
        path = filedialog.asksaveasfilename(title='שמירת תבנית משלוח בדים', defaultextension='.xlsx', initialfile=default_name, filetypes=[('Excel','*.xlsx')])
        if not path:
            return
        try:
            # יצירת קובץ Excel עם הכותרות בלבד
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, Alignment  # type: ignore
            from openpyxl.utils import get_column_letter  # type: ignore
            wb = Workbook()
            ws = wb.active
            ws.title = 'Shipment'
            try:
                # תצוגת RTL כדי להקל על הזנה בעברית
                ws.sheet_view.rightToLeft = True
            except Exception:
                pass
            # כתיבת כותרות בשורה הראשונה
            for col_idx, name in enumerate(headers, start=1):
                c = ws.cell(row=1, column=col_idx, value=name)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center')
                # רוחב עמודה אוטומטי בסיסי לפי אורך הטקסט
                try:
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(28, len(name) + 4))
                except Exception:
                    pass
            # שורת עזרה אופציונלית (לא חובה)
            ws.cell(row=2, column=1, value='')
            wb.save(path)
            try:
                messagebox.showinfo('נוצר קובץ', f'הקובץ נשמר בהצלחה:\n{path}\n\nהערה: ליבוא בתוכנה יש לשמור/להמיר את הקובץ ל-CSV עם אותן כותרות.')
            except Exception:
                pass
        except Exception as e:
            try:
                messagebox.showerror('שגיאה', f'כשל ביצירת תבנית: {e}')
            except Exception:
                pass

    def _populate_fabrics_table(self, records=None):
        # Decide records based on filters if not provided
        if records is None:
            base = list(getattr(self.data_processor, 'fabrics_inventory', []) or [])
            if self._has_active_fabrics_filters():
                records = self._filter_fabrics(base)
            else:
                records = base[-1000:]
        # Render
        for item in self.fabrics_tree.get_children(): self.fabrics_tree.delete(item)
        # map import_log_id to imported_at
        try:
            logs = getattr(self.data_processor, 'fabrics_import_logs', None)
            if logs is None:
                logs = self.data_processor.load_fabrics_import_logs()
        except Exception:
            logs = []
        try:
            log_date_map = { int(r.get('id')): (r.get('imported_at') or '') for r in logs if isinstance(r.get('id'), int) }
        except Exception:
            log_date_map = {}
        for rec in records:
            intake_dt = ''
            try:
                ilid = rec.get('import_log_id')
                if ilid is not None and str(ilid).isdigit():
                    intake_dt = log_date_map.get(int(ilid), '') or ''
            except Exception:
                intake_dt = ''
            if not intake_dt:
                intake_dt = rec.get('last_modified','') or rec.get('Last Modified','') or ''
            self.fabrics_tree.insert('', 'end', values=(
                rec.get('barcode',''), rec.get('fabric_type',''), rec.get('color_name',''), rec.get('color_no',''), rec.get('design_code',''), rec.get('width',''),
                f"{rec.get('net_kg',0):.2f}", f"{rec.get('meters',0):.2f}", f"{rec.get('price',0):.2f}", rec.get('location',''), intake_dt, rec.get('status','במלאי')
            ))

    def _export_current_fabrics_to_excel(self):
        """ייצוא נתוני מלאי (כפי שמופיעים בטבלה, כולל תאריך קליטה) לאקסל."""
        base = list(getattr(self.data_processor, 'fabrics_inventory', []) or [])
        
        # בדיקה אם יש פילטרים פעילים והצגת דיאלוג בחירה
        if self._has_active_fabrics_filters():
            choice = messagebox.askyesnocancel(
                "ייצוא לאקסל",
                "יש פילטרים פעילים.\n\nהאם לייצא רק את הנתונים המסוננים?\n\nכן = נתונים מסוננים בלבד\nלא = כל הנתונים"
            )
            if choice is None:  # ביטול
                return
            elif choice:  # כן - נתונים מסוננים
                records = self._filter_fabrics(base)
            else:  # לא - כל הנתונים
                records = base
        else:
            records = base
        try:
            logs = getattr(self.data_processor, 'fabrics_import_logs', None)
            if logs is None:
                logs = self.data_processor.load_fabrics_import_logs()
        except Exception:
            logs = []
        try:
            log_date_map = { int(r.get('id')): (r.get('imported_at') or '') for r in logs if isinstance(r.get('id'), int) }
        except Exception:
            log_date_map = {}
        path = filedialog.asksaveasfilename(title='ייצוא נתוני מלאי', defaultextension='.xlsx', initialfile='fabrics_inventory.xlsx', filetypes=[('Excel','*.xlsx')])
        if not path:
            return
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, Alignment  # type: ignore
            wb = Workbook(); ws = wb.active; ws.title = 'מלאי בדים'
            try:
                ws.sheet_view.rightToLeft = True
            except Exception:
                pass
            headers = ['ברקוד','סוג בד','צבע','מס׳ צבע','Desen Kodu','רוחב','ק"ג נטו','מטרים','מחיר','מיקום','תאריך קליטה','סטטוס']
            for j, h in enumerate(headers, start=1):
                c = ws.cell(row=1, column=j, value=h)
                c.font = Font(bold=True); c.alignment = Alignment(horizontal='center')
            r_index = 2
            for rec in records:
                intake_dt = ''
                try:
                    ilid = rec.get('import_log_id')
                    if ilid is not None and str(ilid).isdigit():
                        intake_dt = log_date_map.get(int(ilid), '') or ''
                except Exception:
                    intake_dt = ''
                if not intake_dt:
                    intake_dt = rec.get('last_modified','') or rec.get('Last Modified','') or ''
                row = [
                    rec.get('barcode',''), rec.get('fabric_type',''), rec.get('color_name',''), rec.get('color_no',''), rec.get('design_code',''), rec.get('width',''),
                    float(rec.get('net_kg',0) or 0), float(rec.get('meters',0) or 0), float(rec.get('price',0) or 0), rec.get('location',''), intake_dt, rec.get('status','במלאי')
                ]
                for j, v in enumerate(row, start=1): ws.cell(row=r_index, column=j, value=v)
                r_index += 1
            wb.save(path)
            # פתיחת הקובץ באקסל
            try:
                os.startfile(path)
            except Exception:
                pass
        except Exception as e:
            try: messagebox.showerror('שגיאה', f'כשל ביצוא המלאי: {e}')
            except Exception: pass
        # Update quick info
        try:
            total = len(records)
            self.fabrics_filter_info_var.set(f"תוצאות: {total}" if self._has_active_fabrics_filters() else '')
        except Exception:
            pass

    def _has_active_fabrics_filters(self):
        return any([
            (self.fabrics_filter_type_var.get() or '').strip(),
            (self.fabrics_filter_color_var.get() or '').strip(),
            (self.fabrics_filter_location_var.get() or '').strip(),
            (self.fabrics_filter_status_var.get() or '').strip(),
            (self.fabrics_filter_intake_date_var.get() or '').strip(),
        ])

    def _filter_fabrics(self, records):
        t = (self.fabrics_filter_type_var.get() or '').strip()
        c = (self.fabrics_filter_color_var.get() or '').strip()
        loc = (self.fabrics_filter_location_var.get() or '').strip()
        st = (self.fabrics_filter_status_var.get() or '').strip()
        intake_date = (self.fabrics_filter_intake_date_var.get() or '').strip()
        
        # Get intake date mapping
        try:
            logs = getattr(self.data_processor, 'fabrics_import_logs', None)
            if logs is None:
                logs = self.data_processor.load_fabrics_import_logs()
        except Exception:
            logs = []
        try:
            log_date_map = { int(r.get('id')): (r.get('imported_at') or '') for r in logs if isinstance(r.get('id'), int) }
        except Exception:
            log_date_map = {}
        
        def match(rec):
            if t and (rec.get('fabric_type','') != t):
                return False
            if st and (rec.get('status','במלאי') != st):
                return False
            if c and (c.lower() not in (rec.get('color_name','') or '').lower()):
                return False
            if loc and (loc.lower() not in (rec.get('location','') or '').lower()):
                return False
            if intake_date:
                # Get the intake date for this record
                rec_intake_date = ''
                try:
                    ilid = rec.get('import_log_id')
                    if ilid is not None and str(ilid).isdigit():
                        rec_intake_date = log_date_map.get(int(ilid), '') or ''
                except Exception:
                    pass
                if not rec_intake_date:
                    rec_intake_date = rec.get('last_modified','') or rec.get('Last Modified','') or ''
                
                # Compare dates (only the date part, not time)
                if rec_intake_date:
                    rec_date = rec_intake_date.split(' ')[0] if ' ' in rec_intake_date else rec_intake_date
                    filter_date = intake_date.split(' ')[0] if ' ' in intake_date else intake_date
                    if rec_date != filter_date:
                        return False
                else:
                    return False
            return True
        return [r for r in records if match(r)]

    def _apply_fabrics_filters(self):
        self._populate_fabrics_table()
        self._update_fabrics_summary()

    def _clear_fabrics_filters(self):
        self.fabrics_filter_type_var.set(''); self.fabrics_filter_color_var.set(''); self.fabrics_filter_location_var.set(''); self.fabrics_filter_status_var.set(''); self.fabrics_filter_intake_date_var.set('')
        self._populate_fabrics_table()
        self._update_fabrics_summary()

    def _refresh_fabric_filter_values(self):
        inv = getattr(self.data_processor, 'fabrics_inventory', []) or []
        # Unique values for comboboxes
        types = sorted({(r.get('fabric_type') or '').strip() for r in inv if (r.get('fabric_type') or '').strip()})
        colors = sorted({(r.get('color_name') or '').strip() for r in inv if (r.get('color_name') or '').strip()})
        locs = sorted({(r.get('location') or '').strip() for r in inv if (r.get('location') or '').strip()})
        
        # Get unique intake dates
        try:
            logs = getattr(self.data_processor, 'fabrics_import_logs', None)
            if logs is None:
                logs = self.data_processor.load_fabrics_import_logs()
        except Exception:
            logs = []
        try:
            log_date_map = { int(r.get('id')): (r.get('imported_at') or '') for r in logs if isinstance(r.get('id'), int) }
        except Exception:
            log_date_map = {}
        
        # Collect all intake dates
        intake_dates = set()
        for rec in inv:
            try:
                ilid = rec.get('import_log_id')
                if ilid is not None and str(ilid).isdigit():
                    intake_date = log_date_map.get(int(ilid), '') or ''
                    if intake_date:
                        # Extract just the date part
                        date_part = intake_date.split(' ')[0] if ' ' in intake_date else intake_date
                        intake_dates.add(date_part)
            except Exception:
                pass
        
        intake_dates = sorted(list(intake_dates))
        
        # Preserve selections if still valid
        cur_t, cur_c, cur_l, cur_d = self.fabrics_filter_type_var.get(), self.fabrics_filter_color_var.get(), self.fabrics_filter_location_var.get(), self.fabrics_filter_intake_date_var.get()
        self.fabrics_filter_type_cb['values'] = [''] + types
        self.fabrics_filter_color_cb['values'] = [''] + colors
        self.fabrics_filter_location_cb['values'] = [''] + locs
        self.fabrics_filter_intake_date_cb['values'] = [''] + intake_dates
        if cur_t not in self.fabrics_filter_type_cb['values']: self.fabrics_filter_type_var.set('')
        if cur_c not in self.fabrics_filter_color_cb['values']: self.fabrics_filter_color_var.set('')
        if cur_l not in self.fabrics_filter_location_cb['values']: self.fabrics_filter_location_var.set('')
        if cur_d not in self.fabrics_filter_intake_date_cb['values']: self.fabrics_filter_intake_date_var.set('')

    def _on_fabrics_right_click(self, event):
        row_id = self.fabrics_tree.identify_row(event.y)
        if row_id:
            self.fabrics_tree.selection_set(row_id)
            try:
                self._fabric_status_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self._fabric_status_menu.grab_release()

    def _change_selected_fabric_status(self, new_status):
        sel = self.fabrics_tree.selection()
        if not sel: return
        values = list(self.fabrics_tree.item(sel[0], 'values'))
        if not values: return
        barcode = values[0]
        if self.data_processor.update_fabric_status(barcode, new_status):
            values[-1] = new_status; self.fabrics_tree.item(sel[0], values=values)

    def _update_fabrics_summary(self):
        # קבלת נתונים לפי סינון אם יש
        base = list(getattr(self.data_processor, 'fabrics_inventory', []) or [])
        if self._has_active_fabrics_filters():
            records = self._filter_fabrics(base)
        else:
            records = base
        
        # חישוב סיכום לפי הנתונים המסוננים
        total_records = len(records)
        total_meters = sum(float(rec.get('meters', 0) or 0) for rec in records)
        total_net_kg = sum(float(rec.get('net_kg', 0) or 0) for rec in records)
        
        # הצגת סיכום עם אינדיקציה אם יש סינון פעיל
        if self._has_active_fabrics_filters():
            self.fabrics_summary_var.set(f"תוצאות מסוננות: {total_records} | מטרים: {total_meters:.2f} | ק\"ג נטו: {total_net_kg:.2f}")
        else:
            self.fabrics_summary_var.set(f"סה\"כ רשומות: {total_records} | מטרים: {total_meters:.2f} | ק\"ג נטו: {total_net_kg:.2f}")

    def _refresh_fabrics_table(self):
        self.data_processor.fabrics_inventory = self.data_processor.load_fabrics_inventory()
        try:
            self._refresh_fabric_filter_values()
        except Exception:
            pass
        self._populate_fabrics_table()
        if hasattr(self.data_processor, 'fabrics_import_logs'):
            self.data_processor.fabrics_import_logs = self.data_processor.load_fabrics_import_logs(); self._populate_fabrics_logs()
        # Refresh unbarcoded list
        try:
            self.data_processor.refresh_fabrics_unbarcoded(); self._populate_unbarcoded_table()
        except Exception:
            pass
        self._update_fabrics_summary()

    def _import_fabrics_csv(self):
        file_path = filedialog.askopenfilename(title="בחר קובץ CSV של משלוח בדים", filetypes=[("CSV files","*.csv"),("All files","*.*")])
        if not file_path: return
        try:
            added = self.data_processor.import_fabrics_csv(file_path); self._refresh_fabrics_table(); messagebox.showinfo("הצלחה", f"נוספו {added} רשומות מהמשלוח")
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _populate_fabrics_logs(self):
        for item in self.fabrics_logs_tree.get_children(): self.fabrics_logs_tree.delete(item)
        logs = getattr(self.data_processor, 'fabrics_import_logs', [])
        for log in sorted(logs, key=lambda x: x.get('id', 0)):
            self.fabrics_logs_tree.insert('', 'end', values=(log.get('id',''), log.get('file_name',''), log.get('imported_at',''), log.get('records_added',''), '🗑'))

    def _handle_logs_click(self, event):
        region = self.fabrics_logs_tree.identify('region', event.x, event.y)
        if region != 'cell': return
        col = self.fabrics_logs_tree.identify_column(event.x)
        if col != '#5': return
        item_id = self.fabrics_logs_tree.identify_row(event.y)
        if not item_id: return
        values = self.fabrics_logs_tree.item(item_id, 'values')
        if not values: return
        try: log_id = int(values[0])
        except Exception: return
        if not messagebox.askyesno("אישור", "למחוק רשומת לוג זו?"): return
        result = self.data_processor.delete_fabric_import_log_and_fabrics(log_id)
        if result.get('logs_deleted'):
            self._populate_fabrics_logs(); self._populate_fabrics_table()

    def _open_fabric_import_in_excel(self, event):
        """דאבל-קליק על קובץ שעלה: פתיחת ה-CSV המקורי, או שחזור לאקסל מהרשומות אם המקור לא קיים."""
        col = self.fabrics_logs_tree.identify_column(event.x)
        if col == '#5':  # עמודת מחיקה - מטופלת בקליק בודד
            return
        item_id = self.fabrics_logs_tree.identify_row(event.y)
        if not item_id:
            return
        values = self.fabrics_logs_tree.item(item_id, 'values')
        if not values:
            return
        try:
            log_id = int(values[0])
        except Exception:
            return
        logs = getattr(self.data_processor, 'fabrics_import_logs', []) or []
        log = next((l for l in logs if l.get('id') == log_id), None)
        if not log:
            return
        # 1) פתיחת הקובץ המקורי אם עדיין קיים במחשב
        full_path = (log.get('full_path') or '').strip()
        if full_path and os.path.exists(full_path):
            try:
                os.startfile(full_path)
                return
            except Exception:
                pass
        # 2) שחזור לאקסל מהרשומות ששויכו להעלאה זו
        records = self.data_processor.get_fabrics_by_import_log(log_id)
        if not records:
            messagebox.showinfo(
                "אין נתונים",
                "הקובץ המקורי לא נמצא במחשב, ולא נותרו רשומות מלאי מהעלאה זו לשחזור.")
            return
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, Alignment  # type: ignore
            wb = Workbook(); ws = wb.active; ws.title = 'Shipment'
            try:
                ws.sheet_view.rightToLeft = True
            except Exception:
                pass
            headers = ['BARCODE NO', 'סוג בד', 'COLOR NAME', 'COLOR NO', 'Desen Kodu', 'WIDTH', 'GR',
                       'NET KG', 'GROSS KG', 'METER', 'PRICE', 'TOTAL', 'location', 'Last Modified', 'מטרה']
            keys = ['barcode', 'fabric_type', 'color_name', 'color_no', 'design_code', 'width', 'gr',
                    'net_kg', 'gross_kg', 'meters', 'price', 'total', 'location', 'last_modified', 'purpose']
            for j, h in enumerate(headers, start=1):
                c = ws.cell(row=1, column=j, value=h)
                c.font = Font(bold=True); c.alignment = Alignment(horizontal='center')
            for i, rec in enumerate(records, start=2):
                for j, k in enumerate(keys, start=1):
                    ws.cell(row=i, column=j, value=rec.get(k, ''))
            out_dir = os.path.join(os.getcwd(), 'exports', 'fabrics_imports')
            os.makedirs(out_dir, exist_ok=True)
            base = os.path.splitext(log.get('file_name') or f'import_{log_id}')[0]
            safe = ''.join(ch if (ch.isalnum() or ch in ' -_.א-ת') else '_' for ch in base).strip() or f'import_{log_id}'
            out_path = os.path.join(out_dir, f'{safe}.xlsx')
            wb.save(out_path)
            try:
                os.startfile(out_path)
            except Exception:
                messagebox.showinfo('נוצר קובץ', f'הקובץ שוחזר ונשמר ב:\n{out_path}')
        except Exception as e:
            messagebox.showerror('שגיאה', f'כשל בשחזור הקובץ לאקסל: {e}')

    # ===== Unbarcoded fabrics helpers =====
    def _populate_unbarcoded_table(self):
        tree = getattr(self, 'ub_tree', None)
        if not tree: return
        for item in tree.get_children(): tree.delete(item)
        rows = getattr(self.data_processor, 'fabrics_unbarcoded', []) or []
        for r in rows:
            tree.insert('', 'end', values=(
                r.get('id',''),
                r.get('created_at',''),
                r.get('fabric_type',''),
                r.get('manufacturer',''),
                r.get('color',''),
                r.get('shade',''),
                r.get('notes','')
            ))

    def _ub_add_dialog(self):
        win = tk.Toplevel(self.root)
        win.title('הוספת בד ללא ברקוד')
        form = tk.Frame(win, padx=10, pady=10)
        form.pack(fill='both', expand=True)
        labels = ['סוג בד','יצרן הבד','צבע','גוון','הערות']
        vars_ = [tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar()]
        for i, lbl in enumerate(labels):
            tk.Label(form, text=lbl).grid(row=i, column=0, sticky='e', padx=4, pady=4)
            tk.Entry(form, textvariable=vars_[i], width=30).grid(row=i, column=1, sticky='w', padx=4, pady=4)
        btns = tk.Frame(form); btns.grid(row=len(labels), column=0, columnspan=2, sticky='e', pady=(8,0))
        def _do_add():
            try:
                new_id = self.data_processor.add_unbarcoded_fabric(vars_[0].get(), vars_[1].get(), vars_[2].get(), vars_[3].get(), vars_[4].get())
                self._populate_unbarcoded_table()
                try: messagebox.showinfo('נשמר', f'נוסף (ID: {new_id})')
                except Exception: pass
                win.destroy()
            except Exception as e:
                messagebox.showerror('שגיאה', str(e))
        tk.Button(btns, text='שמירה', command=_do_add, bg=theme.DARK, fg='white').pack(side='right', padx=4)
        tk.Button(btns, text='ביטול', command=win.destroy).pack(side='right')

    def _ub_delete_selected(self):
        sel = self.ub_tree.selection()
        if not sel: return
        item = sel[0]
        vals = self.ub_tree.item(item, 'values') or []
        if not vals: return
        try:
            rec_id = int(vals[0])
        except Exception:
            return
        try:
            if not messagebox.askyesno('אישור', f"למחוק רשומה {rec_id}?"):
                return
        except Exception:
            pass
        if self.data_processor.delete_unbarcoded_fabric(rec_id):
            self._populate_unbarcoded_table()

    def _build_barcode_search_tab(self, container):
        """בניית טאב חיפוש לפי ברקוד"""
        # כותרת
        tk.Label(container, text="חיפוש לפי ברקוד", font=(theme.FONT_FAMILY, 14, 'bold'), bg=theme.CARD_BG).pack(pady=10)
        
        # שדה סריקת ברקוד
        barcode_frame = tk.Frame(container, bg=theme.CARD_BG)
        barcode_frame.pack(fill='x', padx=20, pady=10)
        
        tk.Label(barcode_frame, text="ברקוד:", font=(theme.FONT_FAMILY, 12, 'bold'), bg=theme.CARD_BG).pack(side='right', padx=(0, 8))
        self.barcode_search_var = tk.StringVar()
        barcode_entry = tk.Entry(barcode_frame, textvariable=self.barcode_search_var, font=('Consolas', 12), width=25)
        barcode_entry.pack(side='right', padx=(0, 8))
        barcode_entry.bind('<Return>', self._add_barcode_to_search)
        
        tk.Button(barcode_frame, text="➕ הוסף", command=self._add_barcode_to_search, bg=theme.SUCCESS, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='right', padx=8)
        tk.Button(barcode_frame, text="🗑️ מחק נבחר", command=self._remove_selected_barcode, bg=theme.WARNING, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='left', padx=4)
        tk.Button(barcode_frame, text="🧹 נקה הכל", command=self._clear_all_barcodes, bg=theme.DANGER, fg='white', font=(theme.FONT_FAMILY, 10, 'bold')).pack(side='left', padx=4)
        
        # טבלת ברקודים שנסרקו
        table_frame = tk.Frame(container, bg=theme.CARD_BG)
        table_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        cols = ('barcode', 'fabric_type', 'color_name', 'color_no', 'design_code', 'width', 'net_kg', 'meters', 'price', 'location', 'status')
        self.barcode_search_tree = ttk.Treeview(table_frame, columns=cols, show='headings', height=15)
        
        headers = {
            'barcode': 'ברקוד', 'fabric_type': 'סוג בד', 'color_name': 'צבע', 'color_no': 'מס׳ צבע',
            'design_code': 'Desen', 'width': 'רוחב', 'net_kg': 'ק"ג נטו', 'meters': 'מטרים',
            'price': 'מחיר', 'location': 'מיקום', 'status': 'סטטוס'
        }
        
        widths = {
            'barcode': 120, 'fabric_type': 130, 'color_name': 100, 'color_no': 70,
            'design_code': 100, 'width': 60, 'net_kg': 70, 'meters': 70,
            'price': 70, 'location': 80, 'status': 80
        }
        
        for c in cols:
            self.barcode_search_tree.heading(c, text=headers[c])
            self.barcode_search_tree.column(c, width=widths[c], anchor='center')
        
        # סרגל גלילה
        search_vsb = ttk.Scrollbar(table_frame, orient='vertical', command=self.barcode_search_tree.yview)
        self.barcode_search_tree.configure(yscroll=search_vsb.set)
        self.barcode_search_tree.grid(row=0, column=0, sticky='nsew')
        search_vsb.grid(row=0, column=1, sticky='ns')
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)
        
        # תחשיב כולל
        summary_frame = tk.Frame(container, bg=theme.DARK, relief='raised', bd=2)
        summary_frame.pack(fill='x', padx=20, pady=(10, 20))
        
        self.barcode_search_summary_var = tk.StringVar(value="אין ברקודים נסרקו")
        tk.Label(summary_frame, textvariable=self.barcode_search_summary_var, bg=theme.DARK, fg='white', 
                font=(theme.FONT_FAMILY, 12, 'bold'), padx=15, pady=8).pack()
        
        # אתחול רשימת ברקודים
        self._scanned_barcodes_list = []

    def _add_barcode_to_search(self, event=None):
        """הוספת ברקוד לחיפוש"""
        barcode = self.barcode_search_var.get().strip()
        if not barcode:
            return
        
        # בדיקה אם הברקוד כבר קיים
        if barcode in self._scanned_barcodes_list:
            messagebox.showinfo("כפילות", f"הברקוד {barcode} כבר נסרק")
            self.barcode_search_var.set("")
            return
        
        # חיפוש הברקוד במלאי
        fabric = None
        for item in self.data_processor.fabrics_inventory:
            if str(item.get('barcode', '')).strip() == barcode:
                fabric = item
                break
        
        if not fabric:
            messagebox.showerror("ברקוד לא נמצא", f"הברקוד {barcode} לא נמצא במלאי")
            self.barcode_search_var.set("")
            return
        
        # הוספה לרשימה ולטבלה
        self._scanned_barcodes_list.append(barcode)
        
        values = (
            fabric.get('barcode', ''),
            fabric.get('fabric_type', ''),
            fabric.get('color_name', ''),
            fabric.get('color_no', ''),
            fabric.get('design_code', ''),
            fabric.get('width', ''),
            f"{fabric.get('net_kg', 0):.2f}",
            f"{fabric.get('meters', 0):.2f}",
            f"{fabric.get('price', 0):.2f}",
            fabric.get('location', ''),
            fabric.get('status', 'במלאי')
        )
        
        self.barcode_search_tree.insert('', 'end', values=values)
        self.barcode_search_var.set("")
        self._update_barcode_search_summary()

    def _remove_selected_barcode(self):
        """מחיקת ברקוד נבחר"""
        selected = self.barcode_search_tree.selection()
        if not selected:
            return
        
        for item in selected:
            values = self.barcode_search_tree.item(item, 'values')
            if values:
                barcode = values[0]
                if barcode in self._scanned_barcodes_list:
                    self._scanned_barcodes_list.remove(barcode)
            self.barcode_search_tree.delete(item)
        
        self._update_barcode_search_summary()

    def _clear_all_barcodes(self):
        """ניקוי כל הברקודים"""
        self._scanned_barcodes_list.clear()
        for item in self.barcode_search_tree.get_children():
            self.barcode_search_tree.delete(item)
        self._update_barcode_search_summary()

    def _update_barcode_search_summary(self):
        """עדכון תחשיב כולל"""
        if not self._scanned_barcodes_list:
            self.barcode_search_summary_var.set("אין ברקודים נסרקו")
            return
        
        total_weight = 0.0
        total_meters = 0.0
        cut_count = 0
        
        for barcode in self._scanned_barcodes_list:
            fabric = next((item for item in self.data_processor.fabrics_inventory 
                          if str(item.get('barcode', '')).strip() == barcode), None)
            if fabric:
                total_weight += float(fabric.get('net_kg', 0))
                total_meters += float(fabric.get('meters', 0))
                if fabric.get('status', '') == 'נגזר':
                    cut_count += 1
        
        summary_text = f"סה\"כ: {len(self._scanned_barcodes_list)} ברקודים | "
        summary_text += f"משקל: {total_weight:.2f} ק\"ג | "
        summary_text += f"מטרים: {total_meters:.2f} | "
        summary_text += f"נגזרו: {cut_count} ברקודים"
        
        self.barcode_search_summary_var.set(summary_text)
