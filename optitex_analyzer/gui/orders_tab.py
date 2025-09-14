import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import json
import os

class OrdersTabMixin:
    """Mixin for orders management functionality."""
    
    def _create_orders_tab(self):
        """Create the orders management tab."""
        tab = tk.Frame(self.notebook, bg='#f7f9fa')
        self.notebook.add(tab, text="הזמנות")
        
        # Title
        tk.Label(
            tab, 
            text="ניהול הזמנות", 
            font=('Arial', 16, 'bold'), 
            bg='#f7f9fa', 
            fg='#2c3e50'
        ).pack(pady=4)
        
        # Create inner notebook for sub-tabs
        inner_nb = ttk.Notebook(tab)
        inner_nb.pack(fill='both', expand=True, padx=6, pady=4)
        
        # Create sub-tabs
        create_order_tab = tk.Frame(inner_nb, bg='#f7f9fa')
        customers_tab = tk.Frame(inner_nb, bg='#f7f9fa')
        orders_history_tab = tk.Frame(inner_nb, bg='#f7f9fa')
        
        inner_nb.add(create_order_tab, text="יצירת הזמנה")
        inner_nb.add(customers_tab, text="ניהול לקוחות")
        inner_nb.add(orders_history_tab, text="היסטוריית הזמנות")
        
        # Build each sub-tab
        self._build_create_order_tab(create_order_tab)
        self._build_customers_tab(customers_tab)
        self._build_orders_history_tab(orders_history_tab)
    
    def _build_create_order_tab(self, parent):
        """Build the order creation tab."""
        # Order header form
        header_frame = ttk.LabelFrame(parent, text="פרטי הזמנה", padding=10)
        header_frame.pack(fill='x', padx=10, pady=6)
        
        # Customer selection
        tk.Label(header_frame, text="לקוח:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=4, pady=4)
        self.order_customer_var = tk.StringVar()
        self.order_customer_combo = ttk.Combobox(
            header_frame, 
            textvariable=self.order_customer_var, 
            state='readonly', 
            width=25, 
            justify='right'
        )
        self.order_customer_combo.grid(row=0, column=1, sticky='w', padx=2, pady=4)
        
        # Order date
        tk.Label(header_frame, text="תאריך הזמנה:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky='w', padx=4, pady=4)
        self.order_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        tk.Entry(header_frame, textvariable=self.order_date_var, width=12, state='readonly').grid(row=0, column=3, sticky='w', padx=2, pady=4)
        
        # Order number
        tk.Label(header_frame, text="מספר הזמנה:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=4, pady=4)
        self.order_number_var = tk.StringVar()
        tk.Entry(header_frame, textvariable=self.order_number_var, width=15).grid(row=1, column=1, sticky='w', padx=2, pady=4)
        
        # Order notes
        tk.Label(header_frame, text="הערות:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky='w', padx=4, pady=4)
        self.order_notes_var = tk.StringVar()
        tk.Entry(header_frame, textvariable=self.order_notes_var, width=30).grid(row=1, column=3, sticky='w', padx=2, pady=4)
        
        # Product selection form
        product_frame = ttk.LabelFrame(parent, text="הוספת מוצר להזמנה", padding=10)
        product_frame.pack(fill='x', padx=10, pady=6)
        
        # Main category selection
        tk.Label(product_frame, text="קטגוריה ראשית:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=4, pady=4)
        self.order_main_category_var = tk.StringVar()
        self.order_main_category_combo = ttk.Combobox(
            product_frame, 
            textvariable=self.order_main_category_var, 
            state='readonly', 
            width=15, 
            justify='right'
        )
        self.order_main_category_combo.grid(row=0, column=1, sticky='w', padx=2, pady=4)
        self.order_main_category_combo.bind('<<ComboboxSelected>>', lambda e: self._on_main_category_select())
        
        # Product fields
        tk.Label(product_frame, text="מוצר:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=4, pady=4)
        self.order_product_var = tk.StringVar()
        self.order_product_combo = ttk.Combobox(
            product_frame, 
            textvariable=self.order_product_var, 
            state='readonly', 
            width=20, 
            justify='right'
        )
        self.order_product_combo.grid(row=1, column=1, sticky='w', padx=2, pady=4)
        self.order_product_combo.bind('<<ComboboxSelected>>', lambda e: self._on_product_select())
        
        tk.Label(product_frame, text="מידות:", font=('Arial', 10, 'bold')).grid(row=1, column=2, sticky='w', padx=4, pady=4)
        # Create frame for size selection
        size_frame = ttk.Frame(product_frame)
        size_frame.grid(row=1, column=3, sticky='w', padx=2, pady=4)
        
        self.order_size_var = tk.StringVar()
        self.order_size_combo = ttk.Combobox(
            size_frame, 
            textvariable=self.order_size_var, 
            state='readonly', 
            width=12, 
            justify='right'
        )
        self.order_size_combo.pack(side='left', padx=(0, 4))
        self.order_size_combo.bind('<<ComboboxSelected>>', lambda e: self._on_size_select())
        
        # Display selected sizes
        self.order_sizes_display_var = tk.StringVar()
        self.order_sizes_display_entry = tk.Entry(size_frame, textvariable=self.order_sizes_display_var, width=20, state='readonly')
        self.order_sizes_display_entry.pack(side='left', padx=(0, 4))
        
        # Clear sizes button
        self.btn_clear_sizes = tk.Button(size_frame, text='נקה', command=lambda: self._clear_sizes(), width=4)
        self.btn_clear_sizes.pack(side='left')
        
        # Initialize selected sizes list
        self.selected_sizes = []
        
        tk.Label(product_frame, text="סוג בד:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=4, pady=4)
        self.order_fabric_type_var = tk.StringVar()
        self.order_fabric_type_combo = ttk.Combobox(
            product_frame, 
            textvariable=self.order_fabric_type_var, 
            state='readonly', 
            width=15, 
            justify='right'
        )
        self.order_fabric_type_combo.grid(row=2, column=1, sticky='w', padx=2, pady=4)
        
        tk.Label(product_frame, text="צבע בד:", font=('Arial', 10, 'bold')).grid(row=2, column=2, sticky='w', padx=4, pady=4)
        self.order_fabric_color_var = tk.StringVar(value="לבן")  # Default to white
        self.order_fabric_color_combo = ttk.Combobox(
            product_frame, 
            textvariable=self.order_fabric_color_var, 
            state='readonly', 
            width=15, 
            justify='right'
        )
        self.order_fabric_color_combo.grid(row=2, column=3, sticky='w', padx=2, pady=4)
        
        # Quantity and packaging
        tk.Label(product_frame, text="כמות:", font=('Arial', 10, 'bold')).grid(row=3, column=0, sticky='w', padx=4, pady=4)
        self.order_quantity_var = tk.StringVar()
        tk.Entry(product_frame, textvariable=self.order_quantity_var, width=10).grid(row=3, column=1, sticky='w', padx=2, pady=4)
        
        tk.Label(product_frame, text="צורת אריזה:", font=('Arial', 10, 'bold')).grid(row=3, column=2, sticky='w', padx=4, pady=4)
        self.order_packaging_var = tk.StringVar()
        self.order_packaging_combo = ttk.Combobox(
            product_frame, 
            textvariable=self.order_packaging_var, 
            values=['חמישיות', 'שלישיות', 'יחידים'], 
            state='readonly', 
            width=12, 
            justify='right'
        )
        self.order_packaging_combo.grid(row=3, column=3, sticky='w', padx=2, pady=4)
        
        # Add product button
        tk.Button(
            product_frame, 
            text="➕ הוסף מוצר", 
            command=self._add_product_to_order, 
            bg='#27ae60', 
            fg='white'
        ).grid(row=3, column=0, columnspan=4, pady=10)
        
        # Order items table
        items_frame = ttk.LabelFrame(parent, text="פריטי ההזמנה", padding=6)
        items_frame.pack(fill='both', expand=True, padx=10, pady=6)
        
        # Create treeview for order items
        columns = ('product', 'size', 'fabric_type', 'fabric_color', 'quantity', 'packaging', 'total_units')
        self.order_items_tree = ttk.Treeview(items_frame, columns=columns, show='headings', height=8)
        
        headers = {
            'product': 'מוצר',
            'size': 'מידה',
            'fabric_type': 'סוג בד',
            'fabric_color': 'צבע בד',
            'quantity': 'כמות',
            'packaging': 'צורת אריזה',
            'total_units': 'סה"כ יחידות'
        }
        
        widths = {
            'product': 120,
            'size': 80,
            'fabric_type': 100,
            'fabric_color': 100,
            'quantity': 80,
            'packaging': 100,
            'total_units': 100
        }
        
        for col in columns:
            self.order_items_tree.heading(col, text=headers[col])
            self.order_items_tree.column(col, width=widths[col], anchor='center')
        
        # Enable editing for quantity column
        self.order_items_tree.bind('<Double-1>', self._on_item_double_click)
        self.order_items_tree.bind('<Return>', self._on_item_edit)
        
        # Scrollbar for treeview
        items_scrollbar = ttk.Scrollbar(items_frame, orient='vertical', command=self.order_items_tree.yview)
        self.order_items_tree.configure(yscroll=items_scrollbar.set)
        self.order_items_tree.pack(side='left', fill='both', expand=True)
        items_scrollbar.pack(side='right', fill='y')
        
        # Action buttons
        buttons_frame = ttk.Frame(parent)
        buttons_frame.pack(fill='x', padx=10, pady=6)
        
        tk.Button(
            buttons_frame, 
            text="💾 שמור הזמנה", 
            command=self._save_order, 
            bg='#2c3e50', 
            fg='white'
        ).pack(side='right', padx=4)
        
        tk.Button(
            buttons_frame, 
            text="🗑️ מחק פריט נבחר", 
            command=self._remove_selected_order_item, 
            bg='#e67e22', 
            fg='white'
        ).pack(side='right', padx=4)
        
        tk.Button(
            buttons_frame, 
            text="🔄 נקה הזמנה", 
            command=self._clear_order, 
            bg='#95a5a6', 
            fg='white'
        ).pack(side='right', padx=4)
        
        # Initialize order data
        self.current_order_items = []
        self._load_customers_from_file()  # Load customers from file first
        self._load_customers_for_order()
        self._load_products_for_order()
        self._load_all_products_initially()  # Load all products initially
        self._generate_order_number()
    
    def _build_customers_tab(self, parent):
        """Build the customers management tab."""
        # Add customer form
        form_frame = ttk.LabelFrame(parent, text="הוספת לקוח", padding=10)
        form_frame.pack(fill='x', padx=10, pady=6)
        
        # Customer fields
        tk.Label(form_frame, text="שם הלקוח:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=4, pady=4)
        self.customer_name_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.customer_name_var, width=25).grid(row=0, column=1, sticky='w', padx=2, pady=4)
        
        tk.Label(form_frame, text="טלפון:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky='w', padx=4, pady=4)
        self.customer_phone_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.customer_phone_var, width=15).grid(row=0, column=3, sticky='w', padx=2, pady=4)
        
        tk.Label(form_frame, text="כתובת:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=4, pady=4)
        self.customer_address_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.customer_address_var, width=40).grid(row=1, column=1, columnspan=2, sticky='w', padx=2, pady=4)
        
        tk.Label(form_frame, text="הערות:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky='w', padx=4, pady=4)
        self.customer_notes_var = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.customer_notes_var, width=40).grid(row=2, column=1, columnspan=2, sticky='w', padx=2, pady=4)
        
        # Action buttons
        tk.Button(
            form_frame, 
            text="➕ הוסף לקוח", 
            command=self._add_customer, 
            bg='#27ae60', 
            fg='white'
        ).grid(row=3, column=0, pady=10)
        
        tk.Button(
            form_frame, 
            text="✏️ עדכן לקוח", 
            command=self._update_customer, 
            bg='#3498db', 
            fg='white'
        ).grid(row=3, column=1, pady=10)
        
        tk.Button(
            form_frame, 
            text="🗑️ מחק לקוח", 
            command=self._delete_customer, 
            bg='#e67e22', 
            fg='white'
        ).grid(row=3, column=2, pady=10)
        
        # Customers table
        customers_frame = ttk.LabelFrame(parent, text="רשימת לקוחות", padding=6)
        customers_frame.pack(fill='both', expand=True, padx=10, pady=6)
        
        # Create treeview for customers
        customer_columns = ('id', 'name', 'phone', 'address', 'notes', 'created_at')
        self.customers_tree = ttk.Treeview(customers_frame, columns=customer_columns, show='headings', height=10)
        
        customer_headers = {
            'id': 'ID',
            'name': 'שם הלקוח',
            'phone': 'טלפון',
            'address': 'כתובת',
            'notes': 'הערות',
            'created_at': 'נוצר'
        }
        
        customer_widths = {
            'id': 50,
            'name': 150,
            'phone': 120,
            'address': 200,
            'notes': 150,
            'created_at': 140
        }
        
        for col in customer_columns:
            self.customers_tree.heading(col, text=customer_headers[col])
            self.customers_tree.column(col, width=customer_widths[col], anchor='center')
        
        # Scrollbar for customers treeview
        customers_scrollbar = ttk.Scrollbar(customers_frame, orient='vertical', command=self.customers_tree.yview)
        self.customers_tree.configure(yscroll=customers_scrollbar.set)
        self.customers_tree.pack(side='left', fill='both', expand=True)
        customers_scrollbar.pack(side='right', fill='y')
        
        # Load customers
        self._load_customers_from_file()  # Load customers from file first
        self._load_customers_into_tree()
    
    def _build_orders_history_tab(self, parent):
        """Build the orders history tab."""
        # Orders table
        orders_frame = ttk.LabelFrame(parent, text="היסטוריית הזמנות", padding=6)
        orders_frame.pack(fill='both', expand=True, padx=10, pady=6)
        
        # Create treeview for orders
        order_columns = ('id', 'order_number', 'customer', 'date', 'total_items', 'status', 'notes')
        self.orders_tree = ttk.Treeview(orders_frame, columns=order_columns, show='headings', height=12)
        
        order_headers = {
            'id': 'ID',
            'order_number': 'מספר הזמנה',
            'customer': 'לקוח',
            'date': 'תאריך',
            'total_items': 'סך פריטים',
            'status': 'סטטוס',
            'notes': 'הערות'
        }
        
        order_widths = {
            'id': 50,
            'order_number': 120,
            'customer': 150,
            'date': 100,
            'total_items': 100,
            'status': 100,
            'notes': 200
        }
        
        for col in order_columns:
            self.orders_tree.heading(col, text=order_headers[col])
            self.orders_tree.column(col, width=order_widths[col], anchor='center')
        
        # Scrollbar for orders treeview
        orders_scrollbar = ttk.Scrollbar(orders_frame, orient='vertical', command=self.orders_tree.yview)
        self.orders_tree.configure(yscroll=orders_scrollbar.set)
        self.orders_tree.pack(side='left', fill='both', expand=True)
        orders_scrollbar.pack(side='right', fill='y')
        
        # Bind double-click to view order details
        self.orders_tree.bind('<Double-1>', lambda event: self._view_order_details())
        
        # Action buttons
        buttons_frame = ttk.Frame(parent)
        buttons_frame.pack(fill='x', padx=10, pady=6)
        
        tk.Button(
            buttons_frame, 
            text="🔄 רענן", 
            command=self._load_orders_into_tree, 
            bg='#3498db', 
            fg='white'
        ).pack(side='right', padx=4)
        
        tk.Button(
            buttons_frame, 
            text="👁️ צפה בהזמנה", 
            command=self._view_order_details, 
            bg='#9b59b6', 
            fg='white'
        ).pack(side='right', padx=4)
        
        # Load orders
        self._load_orders_from_file()  # Load orders from file first
        self._load_orders_into_tree()
    
    # Order management methods
    def _on_size_select(self):
        """Handle size selection for multi-size support."""
        size = self.order_size_var.get().strip()
        if size and size not in self.selected_sizes:
            self.selected_sizes.append(size)
            self.order_sizes_display_var.set(', '.join(self.selected_sizes))
        self.order_size_combo.set('')
    
    def _clear_sizes(self):
        """Clear selected sizes."""
        self.selected_sizes.clear()
        self.order_sizes_display_var.set('')
    
    def _add_product_to_order(self):
        """Add a product to the current order."""
        main_category = self.order_main_category_var.get().strip()
        product = self.order_product_var.get().strip()
        fabric_type = self.order_fabric_type_var.get().strip()
        fabric_color = self.order_fabric_color_var.get().strip()
        quantity = self.order_quantity_var.get().strip()
        packaging = self.order_packaging_var.get().strip()
        
        if not all([main_category, product, fabric_type, fabric_color, quantity, packaging]):
            messagebox.showerror("שגיאה", "חובה למלא את כל השדות")
            return
        
        if not self.selected_sizes:
            messagebox.showerror("שגיאה", "חובה לבחור לפחות מידה אחת")
            return
        
        try:
            qty = int(quantity)
            if qty <= 0:
                messagebox.showerror("שגיאה", "הכמות חייבת להיות מספר חיובי")
                return
        except ValueError:
            messagebox.showerror("שגיאה", "הכמות חייבת להיות מספר")
            return
        
        # Calculate total units based on packaging
        if packaging == 'חמישיות':
            total_units = qty * 5
        elif packaging == 'שלישיות':
            total_units = qty * 3
        else:  # יחידים
            total_units = qty
        
        # Add items for each selected size
        added_count = 0
        for size in self.selected_sizes:
            # Check if this combination already exists
            existing_item = None
            for item in self.current_order_items:
                if (item['product'] == product and 
                    item['size'] == size and 
                    item['fabric_type'] == fabric_type and 
                    item['fabric_color'] == fabric_color and 
                    item['packaging'] == packaging):
                    existing_item = item
                    break
            
            if existing_item:
                # Update existing item
                existing_item['quantity'] += qty
                existing_item['total_units'] += total_units
            else:
                # Add new item
                item = {
                    'product': product,
                    'size': size,
                    'fabric_type': fabric_type,
                    'fabric_color': fabric_color,
                    'quantity': qty,
                    'packaging': packaging,
                    'total_units': total_units
                }
                self.current_order_items.append(item)
            
            added_count += 1
        
        # Refresh the tree view
        self._refresh_order_items_tree()
        
        # Clear form
        self.order_quantity_var.set('')
        self._clear_sizes()
        
        messagebox.showinfo("הצלחה", f"נוספו {added_count} וריאנטים של המוצר להזמנה")
    
    def _remove_selected_order_item(self):
        """Remove selected item from order."""
        selection = self.order_items_tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "אנא בחר פריט למחיקה")
            return
        
        # Remove items in reverse order to maintain correct indices
        for item in reversed(selection):
            index = self.order_items_tree.index(item)
            if index < len(self.current_order_items):
                self.current_order_items.pop(index)
        
        # Refresh the tree view
        self._refresh_order_items_tree()
    
    def _clear_order(self):
        """Clear the current order."""
        self.current_order_items.clear()
        self._refresh_order_items_tree()
        self._generate_order_number()
    
    def _refresh_order_items_tree(self):
        """Refresh the order items tree view."""
        # Clear existing items
        for item in self.order_items_tree.get_children():
            self.order_items_tree.delete(item)
        
        # Add all current order items
        for item in self.current_order_items:
            self.order_items_tree.insert('', 'end', values=(
                item['product'],
                item['size'],
                item['fabric_type'],
                item['fabric_color'],
                item['quantity'],
                item['packaging'],
                item['total_units']
            ))
    
    def _save_order(self):
        """Save the current order."""
        if not self.current_order_items:
            messagebox.showerror("שגיאה", "אין פריטים בהזמנה")
            return
        
        customer = self.order_customer_var.get().strip()
        if not customer:
            messagebox.showerror("שגיאה", "חובה לבחור לקוח")
            return
        
        order_number = self.order_number_var.get().strip()
        if not order_number:
            messagebox.showerror("שגיאה", "חובה להזין מספר הזמנה")
            return
        
        # Create order data
        order_data = {
            'id': len(getattr(self.data_processor, 'orders', [])) + 1,
            'order_number': order_number,
            'customer': customer,
            'date': self.order_date_var.get(),
            'items': self.current_order_items.copy(),
            'total_items': len(self.current_order_items),
            'status': 'פתוח',
            'notes': self.order_notes_var.get().strip(),
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Save to data processor
        if not hasattr(self.data_processor, 'orders'):
            self.data_processor.orders = []
        
        self.data_processor.orders.append(order_data)
        
        # Save to file
        self._save_orders_to_file()
        
        messagebox.showinfo("הצלחה", f"הזמנה {order_number} נשמרה בהצלחה")
        
        # Clear order
        self._clear_order()
        
        # Refresh orders history
        self._load_orders_into_tree()
    
    def _load_customers_for_order(self):
        """Load customers into the order customer combo."""
        customers = getattr(self.data_processor, 'customers', [])
        customer_names = [c.get('name', '') for c in customers if c.get('name')]
        self.order_customer_combo['values'] = customer_names
    
    def _load_products_for_order(self):
        """Load products into the order product combo."""
        # Load main categories
        main_categories = getattr(self.data_processor, 'main_categories', [])
        main_category_names = [c.get('name', '') for c in main_categories if c.get('name')]
        self.order_main_category_combo['values'] = main_category_names
        
        # Load attributes
        sizes = [s.get('name', '') for s in getattr(self.data_processor, 'product_sizes', [])]
        fabric_types = [f.get('name', '') for f in getattr(self.data_processor, 'product_fabric_types', [])]
        fabric_colors = [c.get('name', '') for c in getattr(self.data_processor, 'product_fabric_colors', [])]
        
        self.order_size_combo['values'] = sizes
        self.order_fabric_type_combo['values'] = fabric_types
        self.order_fabric_color_combo['values'] = fabric_colors
        
        # Set default fabric color to white
        if "לבן" in fabric_colors:
            self.order_fabric_color_var.set("לבן")
    
    def _load_all_products_initially(self):
        """Load all products initially for the product combo."""
        products = getattr(self.data_processor, 'products_catalog', [])
        all_product_names = list(set([p.get('name', '') for p in products if p.get('name')]))
        self.order_product_combo['values'] = all_product_names
    
    def _on_main_category_select(self):
        """Handle main category selection to filter products."""
        main_category = self.order_main_category_var.get().strip()
        if not main_category:
            self.order_product_combo['values'] = []
            return
        
        # Get products for this main category
        products = getattr(self.data_processor, 'products_catalog', [])
        filtered_products = []
        
        for product in products:
            product_main_category = product.get('main_category', '')
            # Check both exact match and if main_category is in the product's main_category field
            if (product_main_category == main_category or 
                (product_main_category and main_category in product_main_category) or
                (not product_main_category and main_category == 'בגדים')):  # Default fallback
                filtered_products.append(product.get('name', ''))
        
        # Update product combo
        self.order_product_combo['values'] = list(set(filtered_products))
        self.order_product_var.set('')  # Clear current selection
    
    def _on_product_select(self):
        """Handle product selection to filter fabric types and sizes."""
        product_name = self.order_product_var.get().strip()
        if not product_name:
            return
        
        # Get the selected product from catalog
        products = getattr(self.data_processor, 'products_catalog', [])
        selected_product = None
        for product in products:
            if product.get('name', '') == product_name:
                selected_product = product
                break
        
        if not selected_product:
            return
        
        # Get fabric types and sizes for this specific product
        product_fabric_types = selected_product.get('fabric_type', '')
        product_sizes = selected_product.get('size', '')
        
        # Filter fabric types - get all unique fabric types for this product
        all_fabric_types = []
        for product in products:
            if product.get('name', '') == product_name:
                fabric_types = product.get('fabric_type', '')
                if fabric_types:
                    # Split by common separators and clean up
                    fabric_type_list = [ft.strip() for ft in fabric_types.replace(',', ';').split(';') if ft.strip()]
                    all_fabric_types.extend(fabric_type_list)
        
        # Remove duplicates and update combo
        unique_fabric_types = list(set(all_fabric_types))
        if unique_fabric_types:
            self.order_fabric_type_combo['values'] = unique_fabric_types
            self.order_fabric_type_var.set('')  # Clear selection to let user choose
        else:
            # If no specific fabric types found, show all available
            all_available_fabric_types = [f.get('name', '') for f in getattr(self.data_processor, 'product_fabric_types', [])]
            self.order_fabric_type_combo['values'] = all_available_fabric_types
            self.order_fabric_type_var.set('')
        
        # Filter sizes - get all unique sizes for this product
        all_sizes = []
        for product in products:
            if product.get('name', '') == product_name:
                sizes = product.get('size', '')
                if sizes:
                    # Split by common separators and clean up
                    size_list = [s.strip() for s in sizes.replace(',', ';').split(';') if s.strip()]
                    all_sizes.extend(size_list)
        
        # Remove duplicates and update combo
        unique_sizes = list(set(all_sizes))
        if unique_sizes:
            self.order_size_combo['values'] = unique_sizes
        else:
            # If no specific sizes found, show all available
            all_available_sizes = [s.get('name', '') for s in getattr(self.data_processor, 'product_sizes', [])]
            self.order_size_combo['values'] = all_available_sizes
    
    def _generate_order_number(self):
        """Generate a new order number."""
        orders = getattr(self.data_processor, 'orders', [])
        next_number = len(orders) + 1
        self.order_number_var.set(f"ORD-{next_number:04d}")
    
    # Customer management methods
    def _add_customer(self):
        """Add a new customer."""
        name = self.customer_name_var.get().strip()
        phone = self.customer_phone_var.get().strip()
        address = self.customer_address_var.get().strip()
        notes = self.customer_notes_var.get().strip()
        
        if not name:
            messagebox.showerror("שגיאה", "חובה להזין שם לקוח")
            return
        
        # Check if customer already exists
        customers = getattr(self.data_processor, 'customers', [])
        if any(c.get('name', '').lower() == name.lower() for c in customers):
            messagebox.showerror("שגיאה", "לקוח עם שם זה כבר קיים")
            return
        
        # Create customer data
        customer_data = {
            'id': len(customers) + 1,
            'name': name,
            'phone': phone,
            'address': address,
            'notes': notes,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Add to data processor
        if not hasattr(self.data_processor, 'customers'):
            self.data_processor.customers = []
        
        self.data_processor.customers.append(customer_data)
        
        # Save to file
        self._save_customers_to_file()
        
        # Clear form
        self.customer_name_var.set('')
        self.customer_phone_var.set('')
        self.customer_address_var.set('')
        self.customer_notes_var.set('')
        
        # Refresh trees
        self._load_customers_into_tree()
        self._load_customers_for_order()
        
        messagebox.showinfo("הצלחה", f"לקוח '{name}' נוסף בהצלחה")
    
    def _update_customer(self):
        """Update selected customer."""
        selection = self.customers_tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "אנא בחר לקוח לעדכון")
            return
        
        item = selection[0]
        values = self.customers_tree.item(item, 'values')
        customer_id = int(values[0])
        
        # Get updated data
        name = self.customer_name_var.get().strip()
        phone = self.customer_phone_var.get().strip()
        address = self.customer_address_var.get().strip()
        notes = self.customer_notes_var.get().strip()
        
        if not name:
            messagebox.showerror("שגיאה", "חובה להזין שם לקוח")
            return
        
        # Update customer
        customers = getattr(self.data_processor, 'customers', [])
        for customer in customers:
            if customer.get('id') == customer_id:
                customer['name'] = name
                customer['phone'] = phone
                customer['address'] = address
                customer['notes'] = notes
                break
        
        # Save to file
        self._save_customers_to_file()
        
        # Refresh trees
        self._load_customers_into_tree()
        self._load_customers_for_order()
        
        messagebox.showinfo("הצלחה", "הלקוח עודכן בהצלחה")
    
    def _delete_customer(self):
        """Delete selected customer."""
        selection = self.customers_tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "אנא בחר לקוח למחיקה")
            return
        
        item = selection[0]
        values = self.customers_tree.item(item, 'values')
        customer_name = values[1]
        
        if messagebox.askyesno("אישור מחיקה", f"האם אתה בטוח שברצונך למחוק את הלקוח '{customer_name}'?"):
            customer_id = int(values[0])
            
            # Remove from data processor
            customers = getattr(self.data_processor, 'customers', [])
            self.data_processor.customers = [c for c in customers if c.get('id') != customer_id]
            
            # Save to file
            self._save_customers_to_file()
            
            # Refresh trees
            self._load_customers_into_tree()
            self._load_customers_for_order()
            
            messagebox.showinfo("הצלחה", "הלקוח נמחק בהצלחה")
    
    def _load_customers_into_tree(self):
        """Load customers into the customers tree."""
        for item in self.customers_tree.get_children():
            self.customers_tree.delete(item)
        
        customers = getattr(self.data_processor, 'customers', [])
        for customer in customers:
            self.customers_tree.insert('', 'end', values=(
                customer.get('id', ''),
                customer.get('name', ''),
                customer.get('phone', ''),
                customer.get('address', ''),
                customer.get('notes', ''),
                customer.get('created_at', '')
            ))
    
    def _load_orders_into_tree(self):
        """Load orders into the orders tree."""
        for item in self.orders_tree.get_children():
            self.orders_tree.delete(item)
        
        orders = getattr(self.data_processor, 'orders', [])
        print(f"🔍 Loading {len(orders)} orders into tree")
        for order in orders:
            print(f"📋 Order: {order.get('order_number', '')} - {order.get('customer', '')}")
            self.orders_tree.insert('', 'end', values=(
                order.get('id', ''),
                order.get('order_number', ''),
                order.get('customer', ''),
                order.get('date', ''),
                order.get('total_items', 0),
                order.get('status', ''),
                order.get('notes', '')
            ))
    
    def _view_order_details(self):
        """View details of selected order."""
        selection = self.orders_tree.selection()
        if not selection:
            messagebox.showwarning("אזהרה", "אנא בחר הזמנה לצפייה")
            return
        
        item = selection[0]
        values = self.orders_tree.item(item, 'values')
        order_id = int(values[0])
        
        # Find order
        orders = getattr(self.data_processor, 'orders', [])
        order = next((o for o in orders if o.get('id') == order_id), None)
        
        if not order:
            messagebox.showerror("שגיאה", "הזמנה לא נמצאה")
            return
        
        # Create details window
        details_window = tk.Toplevel(self.root)
        details_window.title(f"פרטי הזמנה {order.get('order_number', '')}")
        details_window.geometry("800x600")
        
        # Order header
        header_frame = ttk.LabelFrame(details_window, text="פרטי הזמנה", padding=10)
        header_frame.pack(fill='x', padx=10, pady=6)
        
        tk.Label(header_frame, text=f"מספר הזמנה: {order.get('order_number', '')}", font=('Arial', 12, 'bold')).pack(anchor='w')
        tk.Label(header_frame, text=f"לקוח: {order.get('customer', '')}", font=('Arial', 12, 'bold')).pack(anchor='w')
        tk.Label(header_frame, text=f"תאריך: {order.get('date', '')}", font=('Arial', 12, 'bold')).pack(anchor='w')
        tk.Label(header_frame, text=f"סטטוס: {order.get('status', '')}", font=('Arial', 12, 'bold')).pack(anchor='w')
        if order.get('notes'):
            tk.Label(header_frame, text=f"הערות: {order.get('notes', '')}", font=('Arial', 12, 'bold')).pack(anchor='w')
        
        # Order items
        items_frame = ttk.LabelFrame(details_window, text="פריטי ההזמנה", padding=10)
        items_frame.pack(fill='both', expand=True, padx=10, pady=6)
        
        # Create treeview for order items
        columns = ('product', 'size', 'fabric_type', 'fabric_color', 'quantity', 'packaging', 'total_units')
        items_tree = ttk.Treeview(items_frame, columns=columns, show='headings', height=15)
        
        headers = {
            'product': 'מוצר',
            'size': 'מידה',
            'fabric_type': 'סוג בד',
            'fabric_color': 'צבע בד',
            'quantity': 'כמות',
            'packaging': 'צורת אריזה',
            'total_units': 'סה"כ יחידות'
        }
        
        widths = {
            'product': 120,
            'size': 80,
            'fabric_type': 100,
            'fabric_color': 100,
            'quantity': 80,
            'packaging': 100,
            'total_units': 100
        }
        
        for col in columns:
            items_tree.heading(col, text=headers[col])
            items_tree.column(col, width=widths[col], anchor='center')
        
        # Add items to tree
        for item in order.get('items', []):
            items_tree.insert('', 'end', values=(
                item.get('product', ''),
                item.get('size', ''),
                item.get('fabric_type', ''),
                item.get('fabric_color', ''),
                item.get('quantity', 0),
                item.get('packaging', ''),
                item.get('total_units', 0)
            ))
        
        # Scrollbar for items tree
        items_scrollbar = ttk.Scrollbar(items_frame, orient='vertical', command=items_tree.yview)
        items_tree.configure(yscroll=items_scrollbar.set)
        items_tree.pack(side='left', fill='both', expand=True)
        items_scrollbar.pack(side='right', fill='y')
        
        # Action buttons
        buttons_frame = ttk.Frame(details_window)
        buttons_frame.pack(fill='x', padx=10, pady=6)
        
        tk.Button(
            buttons_frame, 
            text="📊 ייצא לאקסל", 
            command=lambda: self._export_order_to_excel(order), 
            bg='#27ae60', 
            fg='white'
        ).pack(side='right', padx=4)
    
    def _export_order_to_excel(self, order):
        """Export order to Excel file."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from datetime import datetime
            import os
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = f"הזמנה {order.get('order_number', '')}"
            
            # Set RTL direction
            ws.sheet_view.rightToLeft = True
            
            # Header information
            ws['A1'] = f"מספר הזמנה: {order.get('order_number', '')}"
            ws['A2'] = f"לקוח: {order.get('customer', '')}"
            ws['A3'] = f"תאריך: {order.get('date', '')}"
            ws['A4'] = f"סטטוס: {order.get('status', '')}"
            if order.get('notes'):
                ws['A5'] = f"הערות: {order.get('notes', '')}"
            
            # Style header
            header_font = Font(bold=True, size=12)
            for row in range(1, 6):
                ws[f'A{row}'].font = header_font
            
            # Items table headers
            headers = ['מוצר', 'מידה', 'סוג בד', 'צבע בד', 'כמות', 'צורת אריזה', 'סה"כ יחידות']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Add items data
            for row_idx, item in enumerate(order.get('items', []), 8):
                ws.cell(row=row_idx, column=1, value=item.get('product', ''))
                ws.cell(row=row_idx, column=2, value=item.get('size', ''))
                ws.cell(row=row_idx, column=3, value=item.get('fabric_type', ''))
                ws.cell(row=row_idx, column=4, value=item.get('fabric_color', ''))
                ws.cell(row=row_idx, column=5, value=item.get('quantity', 0))
                ws.cell(row=row_idx, column=6, value=item.get('packaging', ''))
                ws.cell(row=row_idx, column=7, value=item.get('total_units', 0))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"exports/orders/order_{order.get('order_number', '')}_{timestamp}.xlsx"
            
            # Create exports directory if it doesn't exist
            os.makedirs(os.path.dirname(filename), exist_ok=True)
            
            wb.save(filename)
            messagebox.showinfo("הצלחה", f"ההזמנה יוצאה בהצלחה לקובץ:\n{filename}")
            
        except ImportError:
            messagebox.showerror("שגיאה", "ספריית openpyxl לא מותקנת. אנא התקן אותה עם: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בייצוא לאקסל: {e}")
    
    # File operations
    def _save_orders_to_file(self):
        """Save orders to JSON file."""
        try:
            orders_file = 'orders.json'
            with open(orders_file, 'w', encoding='utf-8') as f:
                json.dump(getattr(self.data_processor, 'orders', []), f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving orders: {e}")
    
    def _save_customers_to_file(self):
        """Save customers to JSON file."""
        try:
            customers_file = 'customers.json'
            with open(customers_file, 'w', encoding='utf-8') as f:
                json.dump(getattr(self.data_processor, 'customers', []), f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving customers: {e}")
    
    def _load_orders_from_file(self):
        """Load orders from JSON file."""
        try:
            orders_file = 'orders.json'
            if os.path.exists(orders_file):
                with open(orders_file, 'r', encoding='utf-8') as f:
                    orders = json.load(f)
                    if not hasattr(self.data_processor, 'orders'):
                        self.data_processor.orders = []
                    self.data_processor.orders = orders
                    print(f"✅ Loaded {len(orders)} orders from file")
            else:
                print("❌ Orders file not found")
        except Exception as e:
            print(f"❌ Error loading orders: {e}")
    
    def _load_customers_from_file(self):
        """Load customers from JSON file."""
        try:
            customers_file = 'customers.json'
            if os.path.exists(customers_file):
                with open(customers_file, 'r', encoding='utf-8') as f:
                    customers = json.load(f)
                    if not hasattr(self.data_processor, 'customers'):
                        self.data_processor.customers = []
                    self.data_processor.customers = customers
        except Exception as e:
            print(f"Error loading customers: {e}")
    
    # Inline editing methods for order items
    def _on_item_double_click(self, event):
        """Handle double-click on order item for editing."""
        item = self.order_items_tree.selection()[0] if self.order_items_tree.selection() else None
        if not item:
            return
        
        # Get the column that was clicked
        column = self.order_items_tree.identify_column(event.x)
        column_index = int(column.replace('#', '')) - 1
        columns = ('product', 'size', 'fabric_type', 'fabric_color', 'quantity', 'packaging', 'total_units')
        
        # Only allow editing of quantity column
        if column_index == 4:  # quantity column
            self._start_edit_item(item, column_index)
    
    def _on_item_edit(self, event):
        """Handle Enter key for editing."""
        item = self.order_items_tree.selection()[0] if self.order_items_tree.selection() else None
        if not item:
            return
        
        # Start editing quantity column
        self._start_edit_item(item, 4)  # quantity column
    
    def _start_edit_item(self, item, column_index):
        """Start editing an item in the tree."""
        # Get current values
        values = list(self.order_items_tree.item(item, 'values'))
        current_value = values[column_index]
        
        # Get item position
        bbox = self.order_items_tree.bbox(item, column_index)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Create entry widget for editing
        self.edit_entry = tk.Entry(self.order_items_tree, width=10)
        self.edit_entry.place(x=x, y=y, width=width, height=height)
        self.edit_entry.insert(0, str(current_value))
        self.edit_entry.select_range(0, tk.END)
        self.edit_entry.focus()
        
        # Bind events
        self.edit_entry.bind('<Return>', lambda e: self._finish_edit_item(item, column_index))
        self.edit_entry.bind('<Escape>', lambda e: self._cancel_edit_item())
        self.edit_entry.bind('<FocusOut>', lambda e: self._finish_edit_item(item, column_index))
    
    def _finish_edit_item(self, item, column_index):
        """Finish editing an item."""
        if not hasattr(self, 'edit_entry'):
            return
        
        try:
            new_value = self.edit_entry.get().strip()
            
            # Validate quantity
            if column_index == 4:  # quantity column
                new_quantity = int(new_value)
                if new_quantity <= 0:
                    messagebox.showerror("שגיאה", "הכמות חייבת להיות מספר חיובי")
                    self._cancel_edit_item()
                    return
                
                # Update the item in current_order_items
                tree_index = self.order_items_tree.index(item)
                if tree_index < len(self.current_order_items):
                    # Get packaging type to recalculate total units
                    packaging = self.current_order_items[tree_index]['packaging']
                    
                    # Calculate total units based on packaging
                    if packaging == 'חמישיות':
                        total_units = new_quantity * 5
                    elif packaging == 'שלישיות':
                        total_units = new_quantity * 3
                    else:  # יחידים
                        total_units = new_quantity
                    
                    # Update the item
                    self.current_order_items[tree_index]['quantity'] = new_quantity
                    self.current_order_items[tree_index]['total_units'] = total_units
                    
                    # Update the tree display
                    values = list(self.order_items_tree.item(item, 'values'))
                    values[4] = str(new_quantity)  # quantity
                    values[6] = str(total_units)   # total_units
                    self.order_items_tree.item(item, values=values)
            
        except ValueError:
            messagebox.showerror("שגיאה", "הכמות חייבת להיות מספר")
            self._cancel_edit_item()
            return
        except Exception as e:
            messagebox.showerror("שגיאה", f"שגיאה בעדכון: {str(e)}")
            self._cancel_edit_item()
            return
        finally:
            self._cleanup_edit_entry()
    
    def _cancel_edit_item(self):
        """Cancel editing an item."""
        self._cleanup_edit_entry()
    
    def _cleanup_edit_entry(self):
        """Clean up the edit entry widget."""
        if hasattr(self, 'edit_entry'):
            self.edit_entry.destroy()
            delattr(self, 'edit_entry')
