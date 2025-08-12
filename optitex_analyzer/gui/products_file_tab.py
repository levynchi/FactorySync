import tkinter as tk
from tkinter import ttk, messagebox

class ProductsFileTabMixin:
    """Mixin לטאב המציג את קובץ המוצרים שנבחר.

    לוגיקה:
    - קורא את self.products_file (אם נטען) ומציג את הגיליון הראשון.
    - רענון בלחיצה על כפתור.
    - קריאה עם openpyxl (ללא תלות בפנדס). אם לא זמין -> הודעה.
    """
    def _create_products_file_tab(self):
        tab = tk.Frame(self.notebook, bg='#f7f9fa')
        self.notebook.add(tab, text="קובץ מוצרים")

        header = tk.Frame(tab, bg='#f7f9fa'); header.pack(fill='x', padx=12, pady=(10,5))
        tk.Label(header, text="תצוגת קובץ מוצרים (קריאה בלבד)", font=('Arial',16,'bold'), bg='#f7f9fa', fg='#2c3e50').pack(side='right')
        btns = tk.Frame(header, bg='#f7f9fa'); btns.pack(side='left')
        tk.Button(btns, text="🔄 רענן", command=self._refresh_products_file_tab, bg='#3498db', fg='white').pack(side='left', padx=4)
        tk.Button(btns, text="📄 פרטים", command=self._products_file_info, bg='#2980b9', fg='white').pack(side='left', padx=4)

        # Tree container
        tree_frame = tk.Frame(tab, bg='#ffffff', relief='groove', bd=1)
        tree_frame.pack(fill='both', expand=True, padx=12, pady=(0,12))

        self.products_file_tree = ttk.Treeview(tree_frame, show='headings')
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.products_file_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.products_file_tree.xview)
        self.products_file_tree.configure(yscroll=vsb.set, xscroll=hsb.set)
        self.products_file_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        self.products_file_status_var = tk.StringVar(value="אין קובץ מוצרים טעון")
        tk.Label(tab, textvariable=self.products_file_status_var, bg='#34495e', fg='white', anchor='w', padx=10).pack(fill='x', side='bottom')

        self._load_products_into_tree()

    # --- Actions ---
    def _products_file_info(self):
        if not getattr(self, 'products_file', None):
            messagebox.showinfo("מידע", "לא נטען קובץ מוצרים.")
            return
        try:
            import os
            size = os.path.getsize(self.products_file)
            messagebox.showinfo("מידע", f"קובץ: {os.path.basename(self.products_file)}\nגודל: {size/1024:.1f} KB")
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))

    def _refresh_products_file_tab(self):
        self._load_products_into_tree()

    def _load_products_into_tree(self):
        tree = getattr(self, 'products_file_tree', None)
        if not tree:
            return
        # Clear existing
        for col in tree['columns'] if tree['columns'] else []:
            tree.heading(col, text='')
        tree.delete(*tree.get_children())

        if not getattr(self, 'products_file', None):
            self.products_file_status_var.set("לא נטען קובץ מוצרים")
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.products_file_status_var.set("חסרה הספרייה openpyxl להצגת הקובץ")
            return
        try:
            wb = load_workbook(self.products_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            self.products_file_status_var.set(f"שגיאה בפתיחת הקובץ: {e}")
            return
        # Extract header row (first non-empty row)
        header = []
        max_cols = 0
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for i, val in enumerate(row):
                header.append(str(val).strip() if val is not None else f"עמודה {i+1}")
            max_cols = len(header)
        if not header:
            self.products_file_status_var.set("קובץ ללא כותרות")
            return
        # Configure columns
        tree['columns'] = header
        for col in header:
            tree.heading(col, text=col)
            tree.column(col, width=max(90, int(len(col)*12)), anchor='center')
        # Insert rows (limit to avoid freezing UI)
        rows_loaded = 0
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = []
            for c in range(max_cols):
                cell_val = row[c] if c < len(row) else ''
                if cell_val is None:
                    cell_val = ''
                values.append(str(cell_val))
            tree.insert('', 'end', values=values)
            rows_loaded += 1
            if rows_loaded >= 2000:  # safety limit
                break
        self.products_file_status_var.set(f"נטענו {rows_loaded} שורות | עמודות: {max_cols}")

        # Auto-size a bit better for short lists
        if rows_loaded < 50:
            for col in header:
                # heuristic width
                current_vals = [tree.set(iid, col) for iid in tree.get_children()[:200]]
                if not current_vals:
                    continue
                max_len = max(len(v) for v in current_vals + [col])
                tree.column(col, width=min(300, max(70, max_len*9)))

        try:
            wb.close()
        except Exception:
            pass
 
