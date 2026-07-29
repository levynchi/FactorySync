"""מחולל PDF A4 לדוח הובלות לתשלום (עברית RTL)."""
from __future__ import annotations

import os
from typing import Dict, List, Optional

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

try:
    from bidi.algorithm import get_display
except Exception:  # pragma: no cover
    def get_display(s):
        return s

PAGE_W, PAGE_H = A4
MARGIN = 1.5 * cm

_FONT = 'ShipRptHe'
_FONT_BOLD = 'ShipRptHeBold'
_FONTS_READY = False

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
HEEBO_MEDIUM = os.path.join('assets', 'fonts', 'Heebo-Medium.ttf')
HEEBO_REGULAR = os.path.join('assets', 'fonts', 'Heebo-Regular.ttf')


VAT_RATE = 0.18


def _abs_path(rel: str) -> str:
    rel = str(rel or '').strip()
    if not rel:
        return ''
    if os.path.isabs(rel):
        return rel
    p_cwd = os.path.join(os.getcwd(), rel)
    if os.path.exists(p_cwd):
        return p_cwd
    p_root = os.path.join(_PROJECT_ROOT, rel)
    return p_root if os.path.exists(p_root) else p_cwd


def _register_fonts():
    global _FONTS_READY
    if _FONTS_READY:
        return
    heebo_m = _abs_path(HEEBO_MEDIUM)
    heebo_r = _abs_path(HEEBO_REGULAR)
    if heebo_m and os.path.exists(heebo_m):
        try:
            pdfmetrics.registerFont(TTFont(_FONT, heebo_r if (heebo_r and os.path.exists(heebo_r)) else heebo_m))
            pdfmetrics.registerFont(TTFont(_FONT_BOLD, heebo_m))
            _FONTS_READY = True
            return
        except Exception:
            pass
    candidates_reg = [r'C:\Windows\Fonts\arial.ttf', r'C:\Windows\Fonts\tahoma.ttf']
    candidates_bold = [r'C:\Windows\Fonts\arialbd.ttf', r'C:\Windows\Fonts\tahomabd.ttf']
    reg = next((p for p in candidates_reg if os.path.exists(p)), None)
    bold = next((p for p in candidates_bold if os.path.exists(p)), reg)
    if reg:
        pdfmetrics.registerFont(TTFont(_FONT, reg))
        pdfmetrics.registerFont(TTFont(_FONT_BOLD, bold or reg))
    _FONTS_READY = True


def _has_hebrew(text: str) -> bool:
    return any('\u0590' <= ch <= '\u05FF' for ch in (text or ''))


def _rtl(text) -> str:
    text = str(text if text is not None else '')
    if _has_hebrew(text):
        try:
            return get_display(text)
        except Exception:
            return text
    return text


def _draw_right(c: canvas.Canvas, text: str, x_right: float, y: float, font: str, size: float, color=(0.12, 0.16, 0.23)):
    c.setFillColorRGB(*color)
    c.setFont(font, size)
    c.drawRightString(x_right, y, _rtl(text))


def _draw_center(c: canvas.Canvas, text: str, x: float, y: float, font: str, size: float, color=(0.12, 0.16, 0.23)):
    c.setFillColorRGB(*color)
    c.setFont(font, size)
    c.drawCentredString(x, y, _rtl(text))


def read_report_items_from_excel(excel_path: str) -> List[Dict]:
    """קורא שורות פירוט מקובץ האקסל שנשמר עם הדוח."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    header_row = None
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=40):
        vals = [str(cell.value or '').strip() for cell in row]
        if 'מספר תעודה' in vals and 'פריט הובלה' in vals:
            header_row = row[0].row
            for idx, name in enumerate(vals, start=1):
                if name:
                    headers[name] = idx
            break
    if not header_row:
        return []

    def col(name, default=None):
        return headers.get(name, default)

    items = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        rec_id = row[col('מספר תעודה') - 1].value if col('מספר תעודה') else None
        if rec_id in (None, ''):
            continue
        try:
            qty = float(row[col('כמות') - 1].value or 0) if col('כמות') else 0
        except Exception:
            qty = 0
        price_raw = row[col('מחיר ליחידה') - 1].value if col('מחיר ליחידה') else 0
        cost_raw = row[col('עלות') - 1].value if col('עלות') else 0
        try:
            price = float(str(price_raw).replace(',', ''))
        except Exception:
            price = 0.0
        try:
            cost = float(str(cost_raw).replace(',', ''))
        except Exception:
            cost = qty * price
        items.append({
            'rec_id': rec_id,
            'kind': str(row[col('סוג') - 1].value or '') if col('סוג') else '',
            'date': str(row[col('תאריך') - 1].value or '') if col('תאריך') else '',
            'package_type': str(row[col('פריט הובלה') - 1].value or '') if col('פריט הובלה') else '',
            'quantity': int(qty) if float(qty).is_integer() else qty,
            'driver': str(row[col('מוביל') - 1].value or '') if col('מוביל') else '',
            'price': price,
            'cost': cost,
        })
    return items


def _build_summary(items: List[Dict]) -> Dict[str, Dict]:
    """סיכום לפי סוג חבילה: qty, price, cost."""
    summary = {}
    for item in items:
        ptype = item.get('package_type') or 'לא מוגדר'
        if ptype not in summary:
            summary[ptype] = {'qty': 0, 'price': float(item.get('price') or 0), 'cost': 0.0}
        summary[ptype]['qty'] += float(item.get('quantity') or 0)
        summary[ptype]['cost'] += float(item.get('cost') or 0)
        if item.get('price'):
            summary[ptype]['price'] = float(item['price'])
    return summary


def create_shipment_report_pdf(file_path: str, report: Dict, items: List[Dict]) -> str:
    """יוצר PDF A4 ומחזיר את הנתיב."""
    _register_fonts()
    os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)

    c = canvas.Canvas(file_path, pagesize=A4)
    left = MARGIN
    right = PAGE_W - MARGIN
    top = PAGE_H - MARGIN
    bottom = MARGIN + 1.0 * cm
    usable_w = right - left

    driver = report.get('driver', '')
    start_date = report.get('start_date', '')
    end_date = report.get('end_date', '')
    created_at = report.get('created_at', '')
    report_id = report.get('id', '')
    payment_filter = report.get('payment_filter', '')
    total_cost = float(report.get('total_cost') or 0)
    total_packages = report.get('total_packages')
    if total_packages is None:
        total_packages = sum(float(i.get('quantity') or 0) for i in items)

    summary = _build_summary(items)
    if not total_cost and items:
        total_cost = sum(float(i.get('cost') or 0) for i in items)

    vat_rate = float(report.get('vat_rate') or VAT_RATE)
    vat_amount = report.get('vat_amount')
    if vat_amount is None:
        vat_amount = round(total_cost * vat_rate, 2)
    else:
        vat_amount = float(vat_amount)
    total_incl_vat = report.get('total_incl_vat')
    if total_incl_vat is None:
        total_incl_vat = round(total_cost + vat_amount, 2)
    else:
        total_incl_vat = float(total_incl_vat)
    vat_pct = int(round(vat_rate * 100))

    # ---- column layout for detail table (LTR x coords, RTL visual via right-align) ----
    # widths from right to left visually: תעודה, סוג, תאריך, פריט, כמות, מחיר, עלות
    col_defs = [
        ('rec_id', 2.2 * cm),
        ('kind', 2.4 * cm),
        ('date', 2.4 * cm),
        ('package_type', 3.2 * cm),
        ('quantity', 1.6 * cm),
        ('price', 2.2 * cm),
        ('cost', 2.2 * cm),
    ]
    # compute right edges for each column (first col starts at page right)
    col_rights = []
    x = right
    for _key, w in col_defs:
        col_rights.append(x)
        x -= w
    # leftover space goes to package_type column (index 3)
    leftover = x - left
    if leftover > 0:
        # shift columns after package_type leftward... easier: widen package_type
        # Rebuild with extra width on package_type
        widths = [w for _, w in col_defs]
        widths[3] += leftover
        col_rights = []
        x = right
        for w in widths:
            col_rights.append(x)
            x -= w

    headers_he = {
        'rec_id': 'מס׳ תעודה',
        'kind': 'סוג',
        'date': 'תאריך',
        'package_type': 'פריט הובלה',
        'quantity': 'כמות',
        'price': 'מחיר',
        'cost': 'עלות',
    }
    row_h = 0.55 * cm
    header_h = 0.65 * cm

    def draw_page_header(y: float, first_page: bool) -> float:
        if first_page:
            _draw_center(c, 'דוח הובלות לתשלום', PAGE_W / 2, y, _FONT_BOLD, 16)
            y -= 0.7 * cm
            _draw_right(c, f'מוביל: {driver}', right, y, _FONT_BOLD, 12)
            y -= 0.5 * cm
            _draw_right(c, f'תקופה: {start_date} עד {end_date}', right, y, _FONT, 10)
            y -= 0.45 * cm
            meta = f'מספר דוח: {report_id}'
            if created_at:
                meta += f'   |   נוצר: {created_at}'
            if payment_filter:
                meta += f'   |   סינון: {payment_filter}'
            _draw_right(c, meta, right, y, _FONT, 9, color=(0.4, 0.45, 0.5))
            y -= 0.35 * cm
            c.setStrokeColorRGB(0.85, 0.88, 0.92)
            c.setLineWidth(1)
            c.line(left, y, right, y)
            y -= 0.55 * cm

            # summary box
            _draw_right(c, 'סיכום כמויות ועלויות', right, y, _FONT_BOLD, 11)
            y -= 0.5 * cm
            for ptype in sorted(summary.keys()):
                info = summary[ptype]
                qty = info['qty']
                price = info['price']
                cost = info['cost']
                if price:
                    line = f'{ptype}: {int(qty) if float(qty).is_integer() else qty} × {price:.2f} ₪ = {cost:.2f} ₪'
                else:
                    line = f'{ptype}: {int(qty) if float(qty).is_integer() else qty}'
                _draw_right(c, line, right - 0.3 * cm, y, _FONT, 10)
                y -= 0.42 * cm

            y -= 0.15 * cm
            box_h = 1.85 * cm
            c.setFillColorRGB(0.93, 0.97, 0.93)
            c.roundRect(left, y - box_h + 0.45 * cm, usable_w, box_h, 4, fill=1, stroke=0)
            _draw_right(
                c,
                f'סה״כ חבילות: {int(total_packages) if float(total_packages).is_integer() else total_packages}',
                right - 0.3 * cm,
                y + 0.15 * cm,
                _FONT_BOLD,
                10,
                color=(0.13, 0.55, 0.33),
            )
            _draw_right(
                c,
                f'סה״כ לפני מע״מ: {total_cost:.2f} ₪',
                right - 0.3 * cm,
                y - 0.25 * cm,
                _FONT,
                10,
                color=(0.13, 0.55, 0.33),
            )
            _draw_right(
                c,
                f'מע״מ ({vat_pct}%): {vat_amount:.2f} ₪',
                right - 0.3 * cm,
                y - 0.65 * cm,
                _FONT,
                10,
                color=(0.13, 0.55, 0.33),
            )
            _draw_right(
                c,
                f'סה״כ לתשלום כולל מע״מ: {total_incl_vat:.2f} ₪',
                right - 0.3 * cm,
                y - 1.1 * cm,
                _FONT_BOLD,
                12,
                color=(0.13, 0.55, 0.33),
            )
            y -= box_h + 0.25 * cm

            _draw_right(c, 'פירוט הובלות', right, y, _FONT_BOLD, 11)
            y -= 0.45 * cm
        else:
            _draw_right(c, f'דוח הובלות — {driver} (המשך)', right, y, _FONT_BOLD, 11)
            y -= 0.55 * cm
        return y

    def draw_table_header(y: float) -> float:
        c.setFillColorRGB(0.23, 0.35, 0.60)
        c.rect(left, y - header_h + 0.15 * cm, usable_w, header_h, fill=1, stroke=0)
        for i, (key, _) in enumerate(col_defs):
            _draw_right(c, headers_he[key], col_rights[i] - 0.1 * cm, y - 0.35 * cm, _FONT_BOLD, 8, color=(1, 1, 1))
        return y - header_h

    def draw_footer(page_num: int, page_count_hint: Optional[int] = None):
        c.setStrokeColorRGB(0.85, 0.88, 0.92)
        c.line(left, bottom - 0.2 * cm, right, bottom - 0.2 * cm)
        label = f'עמוד {page_num}'
        _draw_center(c, label, PAGE_W / 2, bottom - 0.55 * cm, _FONT, 8, color=(0.5, 0.55, 0.6))

    page_num = 1
    y = draw_page_header(top, first_page=True)
    y = draw_table_header(y)

    if not items:
        _draw_right(c, 'אין שורות פירוט בקובץ הדוח.', right, y - 0.4 * cm, _FONT, 10, color=(0.7, 0.2, 0.2))
        draw_footer(page_num)
        c.save()
        return file_path

    for idx, item in enumerate(items):
        if y - row_h < bottom:
            draw_footer(page_num)
            c.showPage()
            page_num += 1
            y = draw_page_header(top, first_page=False)
            y = draw_table_header(y)

        if idx % 2 == 1:
            c.setFillColorRGB(0.96, 0.97, 0.98)
            c.rect(left, y - row_h + 0.12 * cm, usable_w, row_h, fill=1, stroke=0)

        vals = {
            'rec_id': str(item.get('rec_id', '')),
            'kind': str(item.get('kind', '')),
            'date': str(item.get('date', ''))[:10],
            'package_type': str(item.get('package_type', '')),
            'quantity': str(item.get('quantity', '')),
            'price': f"{float(item.get('price') or 0):.2f}",
            'cost': f"{float(item.get('cost') or 0):.2f}",
        }
        for i, (key, _) in enumerate(col_defs):
            _draw_right(c, vals[key], col_rights[i] - 0.1 * cm, y - 0.32 * cm, _FONT, 8)
        y -= row_h

    draw_footer(page_num)
    c.save()
    return file_path
