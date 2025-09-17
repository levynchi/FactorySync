"""
Excel Formatter - Optitex Analyzer
קובץ עזר לעיצוב קבצי אקסל עם הגדרות עקביות
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image as XLImage
import os


class ExcelFormatter:
    """מחלקה לעיצוב קבצי אקסל עם הגדרות עקביות"""
    
    # הגדרות קבועות
    FONT_SIZE = 16
    FONT_SIZE_HEADER = 16
    MIN_COLUMN_WIDTH = 8
    MAX_COLUMN_WIDTH = 50
    DEFAULT_COLUMN_WIDTH = 12
    
    def __init__(self):
        self.gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.thin_border = Side(style='thin', color='000000')
        self.border = Border(left=self.thin_border, right=self.thin_border, 
                           top=self.thin_border, bottom=self.thin_border)
    
    def setup_page(self, ws, title="דף"):
        """הגדרת הגדרות דף בסיסיות"""
        try:
            ws.title = title
            ws.sheet_view.rightToLeft = True
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.paperSize = 9
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = False
            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
        except Exception:
            pass
    
    def add_logo(self, ws, logo_path, row=1):
        """הוספת לוגו לכותרת"""
        try:
            if logo_path and os.path.exists(logo_path):
                img = XLImage(logo_path)
                # מידות: גובה 4.66 ס"מ, רוחב 15.73 ס"מ
                img.height = 4.66 * 37.8
                img.width = 15.73 * 37.8
                ws.add_image(img, f'A{row}')
                # גובה שורה עבור לוגו
                ws.row_dimensions[row].height = 4.66 * 28.35
                return True
        except Exception:
            pass
        return False
    
    def add_business_header(self, ws, business_name="", business_type="", business_vat="", logo_path=""):
        """הוספת כותרת עסקית"""
        try:
            # נסה להוסיף לוגו
            if self.add_logo(ws, logo_path):
                ws.append([])
                ws.append([])
            elif business_name:
                # כותרת עסקית ללא לוגו
                ws.append([business_name])
                r = ws.max_row
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                cell = ws.cell(row=r, column=1)
                cell.font = Font(size=self.FONT_SIZE_HEADER, bold=True)
                cell.alignment = Alignment(horizontal='right')
            
            # שורת סוג/ח.פ.
            if business_type or business_vat:
                line = f"{(business_type or '').strip()} {(business_vat or '').strip()}".strip()
                ws.append([line])
                r = ws.max_row
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                cell = ws.cell(row=r, column=1)
                cell.font = Font(size=self.FONT_SIZE_HEADER)
                cell.alignment = Alignment(horizontal='right')
            
            if business_name or business_type or business_vat:
                ws.append([])  # רווח
        except Exception:
            pass
    
    def add_document_info(self, ws, doc_type, doc_id, date, supplier=""):
        """הוספת מידע מסמך"""
        try:
            start_row = ws.max_row + 1
            ws.append(["מסמך", f"{doc_type} #{doc_id}"])
            
            # תאריך
            if date:
                ws.append(["תאריך", date])
                # פורמט תאריך
                date_cell = ws.cell(row=ws.max_row, column=2)
                date_cell.number_format = 'd/m/yyyy'
                date_cell.alignment = Alignment(horizontal='right')
            
            if supplier:
                ws.append(["ספק", supplier])
            
            # עיצוב מידע מסמך
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=2):
                try:
                    row[0].font = Font(bold=True, size=self.FONT_SIZE)
                    row[0].alignment = Alignment(horizontal='right')
                    row[1].font = Font(size=self.FONT_SIZE)
                    row[1].alignment = Alignment(horizontal='right')
                except Exception:
                    pass
            
            ws.append([])  # רווח
        except Exception:
            pass
    
    def add_table_headers(self, ws, headers, start_row=None):
        """הוספת כותרות טבלה"""
        try:
            if start_row is None:
                start_row = ws.max_row + 1
            
            ws.append(headers)
            header_row = ws.max_row
            
            # עיצוב כותרות
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.fill = self.gray_fill
                cell.font = Font(bold=True, size=self.FONT_SIZE_HEADER)
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            return header_row
        except Exception:
            return None
    
    def add_table_data(self, ws, data_rows, start_row, num_cols):
        """הוספת נתוני טבלה"""
        try:
            for row_data in data_rows:
                ws.append(row_data)
            
            end_row = ws.max_row
            
            # עיצוב נתונים
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=num_cols):
                for cell in row:
                    try:
                        cell.border = self.border
                        if cell.row > start_row:  # לא כותרת
                            cell.font = Font(size=self.FONT_SIZE)
                    except Exception:
                        pass
                
                # יישור טקסט
                try:
                    for i in range(num_cols):
                        if i < len(row):
                            # עברית מימין, כמויות במרכז
                            if i == num_cols - 1:  # עמודה אחרונה (כמויות)
                                row[i].alignment = Alignment(horizontal='center')
                            else:
                                row[i].alignment = Alignment(horizontal='right')
                except Exception:
                    pass
            
            return end_row
        except Exception:
            return None
    
    def add_total_row(self, ws, total_label, total_value, label_col, value_col):
        """הוספת שורת סה"כ"""
        try:
            # יצור שורה ריקה עם ערכים רק בעמודות הנדרשות
            row_data = [None] * max(label_col, value_col)
            row_data[label_col - 1] = total_label
            row_data[value_col - 1] = total_value
            ws.append(row_data)
            
            # עיצוב שורת סה"כ
            total_row = ws.max_row
            ws.cell(row=total_row, column=label_col).font = Font(bold=True, size=self.FONT_SIZE)
            ws.cell(row=total_row, column=label_col).alignment = Alignment(horizontal='right')
            ws.cell(row=total_row, column=value_col).alignment = Alignment(horizontal='center')
        except Exception:
            pass
    
    def calculate_column_width(self, ws, col_letter):
        """חישוב רוחב עמודה אוטומטי"""
        try:
            max_width = 0
            col_num = ord(col_letter) - ord('A') + 1
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        if val.strip():
                            char_width = 0
                            for char in val:
                                if '\u0590' <= char <= '\u05FF':  # עברית
                                    char_width += 1.2
                                elif char.isdigit() or char.isalpha():  # מספרים ואנגלית
                                    char_width += 0.6
                                else:  # תווים מיוחדים
                                    char_width += 0.8
                            
                            char_width += 2  # רווח לגבולות
                            max_width = max(max_width, char_width)
                    except Exception:
                        pass
            
            return min(max(self.MIN_COLUMN_WIDTH, max_width), self.MAX_COLUMN_WIDTH) if max_width > 0 else self.DEFAULT_COLUMN_WIDTH
        except Exception:
            return self.DEFAULT_COLUMN_WIDTH
    
    def auto_fit_columns(self, ws, columns):
        """התאמה אוטומטית של רוחב עמודות"""
        try:
            for col_letter in columns:
                width = self.calculate_column_width(ws, col_letter)
                ws.column_dimensions[col_letter].width = width
        except Exception:
            pass
    
    def create_delivery_note_excel(self, delivery_data, business_info=None, output_path=None):
        """יצירת תעודת משלוח מעוצבת"""
        try:
            wb = Workbook()
            ws = wb.active
            
            # הגדרות דף
            self.setup_page(ws, "תעודת משלוח")
            
            # כותרת עסקית
            if business_info:
                self.add_business_header(ws, 
                    business_info.get('name', ''),
                    business_info.get('type', ''),
                    business_info.get('vat_id', ''),
                    business_info.get('logo_path', ''))
            
            # מידע מסמך
            self.add_document_info(ws, 
                "תעודת משלוח", 
                delivery_data.get('id', ''),
                delivery_data.get('date', ''),
                delivery_data.get('supplier', ''))
            
            # טבלת מוצרים
            ws.append(["שורות מוצרים"])
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            cell = ws.cell(row=r, column=1)
            cell.font = Font(bold=True, size=self.FONT_SIZE)
            cell.alignment = Alignment(horizontal='right')
            
            # כותרות טבלה
            headers = ["שם הדגם", "מידה", "תיאור", "כמות"]
            header_row = self.add_table_headers(ws, headers)
            
            # נתוני טבלה
            data_rows = []
            for line in delivery_data.get('lines', []):
                model = line.get('product', '')
                size = line.get('size', '')
                fabric_type = line.get('fabric_type', '')
                color = line.get('fabric_color', '')
                print_name = line.get('print_name', '')
                
                parts = [p for p in [fabric_type, color, print_name] if p]
                desc = " | ".join(parts)
                qty = line.get('quantity', '')
                
                data_rows.append([model, size, desc, qty])
            
            self.add_table_data(ws, data_rows, header_row + 1, 4)
            
            # שורת סה"כ
            total_qty = sum(int((l or {}).get('quantity', 0) or 0) for l in delivery_data.get('lines', []))
            self.add_total_row(ws, "סה\"כ כמות", total_qty, 3, 4)
            
            # אביזרי תפירה
            if delivery_data.get('accessories'):
                ws.append([])
                ws.append(["אביזרי תפירה"])
                r = ws.max_row
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                cell = ws.cell(row=r, column=1)
                cell.font = Font(bold=True, size=self.FONT_SIZE)
                cell.alignment = Alignment(horizontal='right')
                
                acc_headers = ["אביזר תפירה", "יחידה", "כמות", ""]
                acc_header_row = self.add_table_headers(ws, acc_headers)
                
                acc_data = []
                for acc in delivery_data.get('accessories', []):
                    acc_data.append([acc.get('accessory', ''), acc.get('unit', ''), acc.get('quantity', ''), ''])
                
                self.add_table_data(ws, acc_data, acc_header_row + 1, 4)
            
            # התאמת רוחב עמודות
            self.auto_fit_columns(ws, ['A', 'B', 'C', 'D'])
            
            # שמירה
            if output_path:
                wb.save(output_path)
            
            return wb
        except Exception as e:
            print(f"שגיאה ביצירת תעודת משלוח: {e}")
            return None


# דוגמה לשימוש
if __name__ == "__main__":
    formatter = ExcelFormatter()
    
    # דוגמה לנתוני תעודת משלוח
    sample_data = {
        'id': 'DN-001',
        'date': '2025-01-15',
        'supplier': 'ספק לדוגמה',
        'lines': [
            {'product': 'חולצה', 'size': 'M', 'fabric_type': 'כותנה', 'fabric_color': 'כחול', 'print_name': 'לוגו', 'quantity': 10},
            {'product': 'מכנסיים', 'size': 'L', 'fabric_type': 'דנים', 'fabric_color': 'שחור', 'print_name': '', 'quantity': 5}
        ],
        'accessories': [
            {'accessory': 'כפתורים', 'unit': 'יחידות', 'quantity': 20},
            {'accessory': 'חוט', 'unit': 'מטר', 'quantity': 50}
        ]
    }
    
    business_info = {
        'name': 'חברת Optitex',
        'type': 'חברה בע"מ',
        'vat_id': '123456789',
        'logo_path': ''  # נתיב ללוגו אם קיים
    }
    
    # יצירת תעודת משלוח
    wb = formatter.create_delivery_note_excel(sample_data, business_info, 'delivery_note_example.xlsx')
    if wb:
        print("תעודת משלוח נוצרה בהצלחה!")
